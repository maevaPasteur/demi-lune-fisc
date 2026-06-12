#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
14_synthese_perte_reelle.py
Synthese data-analyst : ou vont les boissons achetees (3 exercices), avec le
DETAIL de la consommation du personnel (cout d'achat + CA equivalent) et des
offerts, et la CASCADE finale de la perte reelle.

Sorties :
  - synthese_perte_reelle.json   (pour la page /analyses/boissons-disparues)
  - public/documents/pieces-defense/Consommation-personnel-et-offerts.xlsx
"""
import json, os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

ICI = os.path.dirname(os.path.abspath(__file__))
OUT = os.path.normpath(os.path.join(ICI, "..", "..", "..", "public", "documents", "pieces-defense"))
os.makedirs(OUT, exist_ok=True)

base = json.load(open(os.path.join(ICI, "base_disparu_ajuste2.json"), encoding="utf-8"))
softs = json.load(open(os.path.join(ICI, "softs_balance.json"), encoding="utf-8"))
B = {x["nom"]: x for x in base["boissons"]}
JOURS = 662

def prix(nom, sens):
    return B.get(nom, {}).get(f"prix_{sens}_l")

# ---------------------------------------------------------------------------
# 1. CONSOMMATION DU PERSONNEL / OWNER (litres, cout achat, CA equivalent)
# ---------------------------------------------------------------------------
# volumes etablis dans 06/07 (attestes ou bornes) ; CA = litres x prix de revente.
# Picon : pas de prix revente direct (bu en Picon Biere 4cl) -> valorise au Picon Biere ~5,9 €.
PERSO = [
    # nom_affiche, produit_prix, litres, prix_achat, prix_revente, base_calcul
    ("Coca (Chris-Elian 4/j + Ghislaine 2/j)", "Coca", 1311.0, 1.97, 11.82, "6 cannettes/jour x 662 j"),
    ("Limonade (tout le staff, 1/j)", "Limonade", 165.0, 0.94, 15.60, "1 limonade 25cl/jour"),
    ("Jus de fruit (tout le staff, 3/j)", "Jus d'orange", 397.0, 3.10, 15.60, "3 jus 20cl/jour"),
    ("Sirop a l'eau (tout le staff, 3/j)", "Sirop Framboise", 40.0, 7.09, 12.00, "3 sirops 2cl/jour"),
    ("Macvin (chef Thierry, 2,5 verres/j)", "Macvin", 99.0, 18.15, 81.67, "2,5 verres 6cl/jour"),
    ("Picon (chef Thierry, ~1,7 Picon Biere/j)", "Picon", 44.0, 10.38, 147.50, "~1,7 Picon Biere/j x 4 cl de Picon x 662 j (≈44 L)"),
]
perso_rows = []
tot_l = tot_cout = tot_ca = 0.0
for aff, prod, litres, pa, pr, base_calc in PERSO:
    cout = litres * pa
    ca = litres * pr
    perso_rows.append({"poste": aff, "litres": round(litres, 1), "cout_achat_eur": round(cout),
                       "ca_equivalent_eur": round(ca), "base": base_calc})
    tot_l += litres; tot_cout += cout; tot_ca += ca
personnel = {"lignes": perso_rows, "total_litres": round(tot_l), "total_cout_eur": round(tot_cout),
             "total_ca_eur": round(tot_ca)}

# ---------------------------------------------------------------------------
# 2. OFFERTS AUX CLIENTS (non sonnes en caisse) - ESTIMATIONS BORNEES
# ---------------------------------------------------------------------------
offerts = {
    "avertissement": "Estimations bornees (offerts non enregistres en caisse, par nature). A confirmer par l'exploitante.",
    "lignes": [
        {"poste": "Aperitifs offerts (~1/jour)", "litres": round(1 * JOURS * 0.06, 1),
         "cout_achat_eur": round(1 * JOURS * 0.06 * 15), "ca_equivalent_eur": round(1 * JOURS * 5.5),
         "base": "1 aperitif 6cl/jour x 662 j"},
        {"poste": "Cafes offerts (~3/jour)", "litres": None,
         "cout_achat_eur": round(3 * JOURS * 0.25), "ca_equivalent_eur": round(3 * JOURS * 2.20),
         "base": "3 cafes/jour x 662 j (cafe hors achats FCBS)"},
    ],
}
offerts["total_cout_eur"] = sum(l["cout_achat_eur"] for l in offerts["lignes"])
offerts["total_ca_eur"] = sum(l["ca_equivalent_eur"] for l in offerts["lignes"])

# ---------------------------------------------------------------------------
# CORRECTION AVOIRS : notes de credit FCBS stockees en quantite POSITIVE et
# comptees a tort en achat. Verifie au code produit (rapprochement contenants).
# On les retire des achats (traitement conservateur : une seule deduction).
#   Cidre Brut  (avoir 520312 16/07/2024, code 403200, 30 x 75cl)
#   Bourgogne Aligote (avoir 532782 11/12/2024, code 603040, 1 BIB x 10 L)
#   Grand Marnier (avoir 526876 25/09/2024, code 510090, 3 x 70cl)
AVOIRS = {"Cidre Brut": 22.5, "Bourgogne Aligoté maison": 10.0, "Grand Marnier": 2.1}
AVOIR_TOTAL_L = round(sum(AVOIRS.values()), 1)  # 34.6 L

# ---------------------------------------------------------------------------
# 3. CASCADE ALCOOL : ou va l'alcool achete (litres)
# ---------------------------------------------------------------------------
alc = [x for x in base["boissons"] if x.get("conso_complete") is True]
A = sum(x["achats_l"] for x in alc) - AVOIR_TOTAL_L
stock = sum(x["stock_final_l"] for x in alc)
V_verre = sum(x["conso"]["seches_l"] for x in alc)
V_cock = sum(x["conso"]["cocktails_l"] for x in alc)
C = sum(x["conso"]["plats_l"] for x in alc)
M = sum(x["conso"]["menu_moyen_l"] for x in alc)
P = sum((x.get("conso_staff_l") or {}).get("moyen", 0) for x in alc)  # Picon+Macvin
O_alc = round(1 * JOURS * 0.06, 1)  # aperitifs offerts
S = round(0.08 * V_verre, 1)        # sur-versement ~8% du servi au verre (borne basse free-pour)
# Postes arrondis, puis PERTE = achats - somme des postes arrondis (la cascade
# affichee se boucle EXACTEMENT : achats - postes = perte, sans ecart d'arrondi).
ach_r = round(A)
postes = [
    {"poste": "Vendu au verre (caisse)", "litres": round(V_verre)},
    {"poste": "Vendu en cocktails (caisse, biere des cocktails incluse)", "litres": round(V_cock)},
    {"poste": "Cuisine (fondues, babas, flambage)", "litres": round(C)},
    {"poste": "Alcool des menus (non detaille en caisse)", "litres": round(M)},
    {"poste": "Consommation du chef (Picon + Macvin)", "litres": round(P)},
    {"poste": "Aperitifs offerts aux clients", "litres": round(O_alc)},
    {"poste": "Sur-versement au verre (~8 %)", "litres": round(S)},
    {"poste": "Stock final (inventaire)", "litres": round(stock)},
]
perte_r = ach_r - sum(p["litres"] for p in postes)
cascade = {
    "achats_alcool_l": ach_r,
    "postes": postes,
    "perte_reelle_residuelle_l": perte_r,
    "perte_reelle_pct_achats": round(100 * perte_r / ach_r, 1),
    "interpretation": "Perte d'exploitation d'un bar (casse, evaporation, mousse, sur-versement non mesure). "
                      "A rapprocher de l'abattement que l'administration retient elle-meme en reconstitution de bar "
                      "(CAA Paris, 17 mars 2021 : 22 % des achats pour personnel/offerts/pertes/vol, methode validee ; "
                      "n de requete a fiabiliser). AUCUNE vente au noir : CA bancarise, especes 1,3 %.",
}

# ---------------------------------------------------------------------------
# 4. SOFTS : bilan ferme (rappel script 07)
# ---------------------------------------------------------------------------
softs_table = []
for cat, v in softs["bilan_agrege_litres"].items():
    softs_table.append({"categorie": cat, "achat_l": v["achat_l"], "vendu_l": v["vendu_direct_l"],
                        "cocktails_l": v["conso_cocktails_l"], "staff_l": v["staff_l"],
                        "stock_l": v["stock_l"], "residu_l": v["residu_l"]})

out = {
    "description": "Synthese : consommation personnel/offerts chiffree + cascade de la perte reelle de boisson.",
    "personnel": personnel,
    "offerts": offerts,
    "cascade_alcool": cascade,
    "softs": {"lignes": softs_table, "constat": "Bilan softs ferme (ventes recuperees au ticket + staff + stock) : disparu reel ~ 0."},
}
json.dump(out, open(os.path.join(ICI, "synthese_perte_reelle.json"), "w", encoding="utf-8"),
          ensure_ascii=False, indent=1)

# ---------------------------------------------------------------------------
# XLSX piece
# ---------------------------------------------------------------------------
H1 = Font(bold=True, size=14, color="FFFFFF"); FILL1 = PatternFill("solid", fgColor="1F3A5F")
BOLD = Font(bold=True); FILLT = PatternFill("solid", fgColor="E2E8F0")
THIN = Border(*[Side(style="thin", color="CBD5E0")] * 4); RIGHT = Alignment(horizontal="right")
def titre(ws, t, span):
    ws.append([t]); ws.merge_cells(start_row=ws.max_row, start_column=1, end_row=ws.max_row, end_column=span)
    c = ws.cell(ws.max_row, 1); c.font = H1; c.fill = FILL1; ws.row_dimensions[ws.max_row].height = 24
def ent(ws, cols):
    ws.append(cols)
    for c in ws[ws.max_row]: c.font = BOLD; c.fill = FILLT; c.border = THIN
def lig(ws, vals, bold=False):
    ws.append(vals)
    for c in ws[ws.max_row]:
        c.border = THIN
        if bold: c.font = BOLD
        if isinstance(c.value, (int, float)): c.alignment = RIGHT

wb = openpyxl.Workbook()
ws = wb.active; ws.title = "Conso personnel"
for w, c in zip([44, 12, 16, 18, 26], "ABCDE"): ws.column_dimensions[c].width = w
titre(ws, "Consommation du personnel et de l'owner (3 exercices)", 5)
ent(ws, ["Poste", "Litres", "Cout d'achat EUR", "CA equivalent EUR", "Base de calcul"])
for r in perso_rows:
    lig(ws, [r["poste"], r["litres"], r["cout_achat_eur"], r["ca_equivalent_eur"], r["base"]])
lig(ws, ["TOTAL personnel", personnel["total_litres"], personnel["total_cout_eur"], personnel["total_ca_eur"], ""], bold=True)

ws = wb.create_sheet("Offerts clients")
for w, c in zip([34, 12, 16, 18, 30], "ABCDE"): ws.column_dimensions[c].width = w
titre(ws, "Offerts aux clients (non sonnes) - estimations bornees", 5)
ws.append([offerts["avertissement"]])
ent(ws, ["Poste", "Litres", "Cout d'achat EUR", "CA equivalent EUR", "Base"])
for r in offerts["lignes"]:
    lig(ws, [r["poste"], r["litres"], r["cout_achat_eur"], r["ca_equivalent_eur"], r["base"]])
lig(ws, ["TOTAL offerts", "", offerts["total_cout_eur"], offerts["total_ca_eur"], ""], bold=True)

ws = wb.create_sheet("Cascade perte reelle")
for w, c in zip([48, 14], "AB"): ws.column_dimensions[c].width = w
titre(ws, "Ou va l'alcool achete (litres, 3 exercices)", 2)
ent(ws, ["Poste", "Litres"])
lig(ws, ["ACHATS ALCOOL (factures)", cascade["achats_alcool_l"]], bold=True)
for p in cascade["postes"]:
    lig(ws, ["- " + p["poste"], p["litres"]])
lig(ws, [f"= PERTE REELLE RESIDUELLE ({cascade['perte_reelle_pct_achats']} % des achats)",
         cascade["perte_reelle_residuelle_l"]], bold=True)
ws.append([]); ws.append([cascade["interpretation"]])

ws = wb.create_sheet("Softs (bilan ferme)")
for w, c in zip([18, 12, 12, 12, 12, 12, 12], "ABCDEFG"): ws.column_dimensions[c].width = w
titre(ws, "Softs : bilan ferme (ventes au ticket + staff + stock)", 7)
ent(ws, ["Categorie", "Achat L", "Vendu L", "Cocktails L", "Staff L", "Stock L", "Residu L"])
for r in softs_table:
    lig(ws, [r["categorie"], r["achat_l"], r["vendu_l"], r["cocktails_l"], r["staff_l"], r["stock_l"], r["residu_l"]])

wb.save(os.path.join(OUT, "Consommation-personnel-et-offerts.xlsx"))

print("CONSOMMATION PERSONNEL :")
for r in perso_rows:
    print(f"  {r['poste'][:42]:42} {r['litres']:>7} L  cout {r['cout_achat_eur']:>6} EUR  CA {r['ca_equivalent_eur']:>7} EUR")
print(f"  {'TOTAL':42} {personnel['total_litres']:>7} L  cout {personnel['total_cout_eur']:>6} EUR  CA {personnel['total_ca_eur']:>7} EUR")
print()
print("CASCADE ALCOOL :")
print(f"  Achats {cascade['achats_alcool_l']} L")
for p in cascade["postes"]:
    print(f"    - {p['poste'][:44]:44} {p['litres']:>6} L")
print(f"  = PERTE REELLE {cascade['perte_reelle_residuelle_l']} L ({cascade['perte_reelle_pct_achats']} % des achats)")
print("\nsynthese_perte_reelle.json + Consommation-personnel-et-offerts.xlsx ecrits.")
