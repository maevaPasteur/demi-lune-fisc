#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
13_generer_pieces_publiques.py
Genere les PIECES DE DEFENSE (XLSX data-lourd) pour public/documents/pieces-defense/,
a partir des JSON deja calcules (scripts 08-12). Destinees a l'avocat et au fisc :
chaque chiffre est sourcé, reproductible, et renvoie a l'annexe certifiee d'origine.
"""
import json, os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

ICI = os.path.dirname(os.path.abspath(__file__))
OUT = os.path.normpath(os.path.join(ICI, "..", "..", "..", "public", "documents", "pieces-defense"))
os.makedirs(OUT, exist_ok=True)

def J(name):
    return json.load(open(os.path.join(ICI, name), encoding="utf-8"))

# styles
H1 = Font(bold=True, size=14, color="FFFFFF")
H2 = Font(bold=True, size=11, color="FFFFFF")
BOLD = Font(bold=True)
FILL1 = PatternFill("solid", fgColor="1F3A5F")
FILL2 = PatternFill("solid", fgColor="2C5282")
FILLT = PatternFill("solid", fgColor="E2E8F0")
THIN = Border(*[Side(style="thin", color="CBD5E0")] * 4)
RIGHT = Alignment(horizontal="right")
WRAP = Alignment(wrap_text=True, vertical="top")

def titre(ws, txt, span=6):
    ws.append([txt])
    ws.merge_cells(start_row=ws.max_row, start_column=1, end_row=ws.max_row, end_column=span)
    c = ws.cell(ws.max_row, 1); c.font = H1; c.fill = FILL1; c.alignment = Alignment(vertical="center")
    ws.row_dimensions[ws.max_row].height = 24

def sous(ws, txt, span=6):
    ws.append([txt])
    ws.merge_cells(start_row=ws.max_row, start_column=1, end_row=ws.max_row, end_column=span)
    c = ws.cell(ws.max_row, 1); c.font = H2; c.fill = FILL2
    ws.row_dimensions[ws.max_row].height = 18

def entete(ws, cols):
    ws.append(cols)
    for c in ws[ws.max_row]:
        c.font = BOLD; c.fill = FILLT; c.border = THIN

def ligne(ws, vals, bold=False):
    ws.append(vals)
    for c in ws[ws.max_row]:
        c.border = THIN
        if bold: c.font = BOLD
        if isinstance(c.value, (int, float)): c.alignment = RIGHT

def vide(ws): ws.append([])

# ============================================================================
# PIECE 1 : SUPPRESSIONS DE CAISSE (DEL)
# ============================================================================
rec = J("reconciliation_suppressions.json")
ab = J("aberrations_del.json")
dj = J("del_justification.json")
pz = J("reconciliation_par_z.json")

wb = openpyxl.Workbook()

# --- onglet 0 : Sommaire / these ---
ws = wb.active; ws.title = "Synthese"
ws.column_dimensions["A"].width = 70
titre(ws, "PIECE DE DEFENSE - Suppressions de caisse (lignes DEL)", 1)
for t in [
    "",
    "Grief du fisc : 430 763 EUR de lignes supprimees en caisse (3 exercices) presentees comme ventes dissimulees.",
    "",
    "THESE : ces suppressions ne sont PAS de l'occultation. Demonstration en 5 niveaux + retournement de la charge.",
    "",
    "1. ERREURS DE QUANTITE : 9 lignes = 140 221 EUR (33 %) sont des fautes de frappe (prix x quantite absurde).",
    "2. PREUVE PAR SESSION : 9 sessions ou les suppressions DEPASSENT le CA encaisse du jour -> forcement fictives.",
    "3. PREUVE PAR DISTRIBUTION : les prix supprimes epousent ceux des ventes reelles -> vrais articles, re-encaisses.",
    "4. TRIANGULATION : 3 exports certifies (synthese A, liste tickets H, encaissements) convergent a < 0,2 %.",
    "   Les suppressions ne figurent dans AUCUN des trois.",
    "5. DOUBLE VERROU : CA declare = encaissements ; especes = 1,3 % ; pas de canal pour encaisser au noir.",
    "   Et 430 k de ventes cachees exigeraient 430 k d'achats caches : aucun (stock stable).",
    "",
    "CHARGE DE LA PREUVE : pour soutenir l'occultation, le fisc doit exhiber un canal d'encaissement (1,3 % especes,",
    "le reste 100 % bancarise/CB/ticket-resto trace) ET un canal d'approvisionnement (achats reconcilies). Aucun n'existe.",
    "",
    "Sources : ANNEXE-E (journal d'evenements certifie), ANNEXE-A (synthese CA certifiee), ANNEXE-H (liste tickets),",
    "ANNEXE-B (ventes ligne a ligne). Tout est reproductible (scripts 08-12, dossier src/data/incertitudeDisparu).",
]:
    ws.append([t])
    if t.startswith(("1.", "2.", "3.", "4.", "5.", "THESE", "CHARGE")): ws[ws.max_row][0].font = BOLD

# --- onglet 1 : totaux par exercice ---
ws = wb.create_sheet("1-Totaux par exercice")
for w, c in zip([14, 16, 16, 12, 10, 16, 12], "ABCDEFG"): ws.column_dimensions[c].width = w
titre(ws, "Reconciliation par exercice (source ANNEXE-A et E)")
entete(ws, ["Exercice", "CA declare TTC", "Encaissements", "Especes", "% especes", "Suppressions DEL", "Suppr/CA"])
for exo, v in rec["par_exercice"].items():
    ligne(ws, [exo, v["ca_declare_ttc"], v["encaissements_ttc"], v["especes"],
               f'{v["part_especes_pct"]} %', v["suppressions_eur"], f'{v["suppr_sur_ca_pct"]:.0f} %'])
t = rec["totaux"]
ligne(ws, ["TOTAL 3 exercices", t["ca_declare_ttc"], t["encaissements_ttc"], t["especes_total"],
           f'{t["part_especes_pct"]} %', t["suppressions_total"], f'{t["suppr_sur_ca_pct"]:.0f} %'], bold=True)
vide(ws)
ligne(ws, ["CA declare = encaissements (ecart) :", t["ca_declare_ttc"] - t["encaissements_ttc"]])
ligne(ws, ["Plafond d'occultation possible (= especes) :", t["especes_total"]])

# --- onglet 2 : erreurs de quantite ---
ws = wb.create_sheet("2-Erreurs de quantite")
for w, c in zip([12, 8, 14, 30, 14], "ABCDE"): ws.column_dimensions[c].width = w
titre(ws, "Lignes a quantite absurde = fautes de frappe, supprimees aussitot", 5)
ws.append(["Chaque ligne supprimee = 1 produit x sa quantite. 9 lignes ont une quantite invraisemblable."])
entete(ws, ["Date", "", "Montant", "Decomposition (prix x quantite)", ""])
for d in dj["A_erreurs_quantite"]["lignes"]:
    ligne(ws, [d["date"], "", d["montant"], f"{d['prix']} EUR x {d['quantite']}", ""])
vide(ws)
ligne(ws, ["TOTAL erreurs de quantite", "", dj["A_erreurs_quantite"]["somme_eur"],
           f"{dj['A_erreurs_quantite']['pct_du_total_del']} % du total des suppressions", ""], bold=True)
vide(ws)
ws.append([ab["avertissement_honnete"]]); ws[ws.max_row][0].alignment = WRAP
ws.merge_cells(start_row=ws.max_row, start_column=1, end_row=ws.max_row, end_column=5)
ws.row_dimensions[ws.max_row].height = 60

# --- onglet 3 : sessions impossibles ---
ws = wb.create_sheet("3-Sessions impossibles")
for w, c in zip([12, 12, 14, 14, 12], "ABCDE"): ws.column_dimensions[c].width = w
titre(ws, "Sessions ou les suppressions DEPASSENT le CA encaisse du jour", 5)
ws.append(["On ne peut pas supprimer plus de ventes qu'on n'en a faites : ces lignes sont forcement fictives."])
entete(ws, ["Session Z", "Exercice", "CA tickets", "Suppressions", "Nb lignes"])
for z in pz["z_avec_suppr_superieure_au_ca"]["exemples"]:
    ligne(ws, [z["z"], z["exercice"], z["ca_tickets"], z["suppr"], z["nb_suppr"]])
vide(ws)
ligne(ws, [f"Nombre total de sessions concernees : {pz['z_avec_suppr_superieure_au_ca']['n']}", "", "", "", ""], bold=True)

# --- onglet 4 : triangulation ---
ws = wb.create_sheet("4-Triangulation")
for w, c in zip([40, 16], "AB"): ws.column_dimensions[c].width = w
titre(ws, "Trois exports certifies independants donnent le meme CA", 2)
tr = pz["triangulation_3_exports_certifies"]
entete(ws, ["Source certifiee", "CA 3 exercices (EUR)"])
ligne(ws, ["Synthese du CA (ANNEXE-A)", tr["ca_synthese_ANNEXE_A"]])
ligne(ws, ["Liste des tickets (ANNEXE-H)", tr["ca_liste_tickets_ANNEXE_H"]])
ligne(ws, ["Encaissements (ANNEXE-A)", tr["encaissements_ANNEXE_A"]])
vide(ws)
ligne(ws, ["Suppressions (ANNEXE-E) - EN DEHORS des trois", tr["suppressions_ANNEXE_E"]], bold=True)
vide(ws)
ws.append([tr["constat"]]); ws[ws.max_row][0].alignment = WRAP
ws.merge_cells(start_row=ws.max_row, start_column=1, end_row=ws.max_row, end_column=2)
ws.row_dimensions[ws.max_row].height = 45

# --- onglet 5 : distribution prix ---
ws = wb.create_sheet("5-Distribution prix")
for w, c in zip([14, 16, 18], "ABC"): ws.column_dimensions[c].width = w
titre(ws, "Les suppressions epousent la grille de prix des ventes reelles", 3)
ws.append(["= ce sont de vrais articles du menu (re-encaisses/corriges), pas une categorie cachee."])
entete(ws, ["Prix", "% des ventes", "% des suppressions"])
for d in dj["B_distribution_prix"]["top_prix"]:
    ligne(ws, [d["prix"], f'{d["part_ventes_pct"]} %', f'{d["part_suppr_pct"]} %'])

wb.save(os.path.join(OUT, "Defense-Suppressions-de-caisse-DEL.xlsx"))
print("ecrit : Defense-Suppressions-de-caisse-DEL.xlsx (6 onglets)")

# ============================================================================
# PIECE 2 : BOISSONS PRETENDUMENT DISPARUES
# ============================================================================
mc = J("resultats_montecarlo_ajuste2.json")
rd = J("reconciliation_disparu_vs_del.json")
conso = json.load(open(os.path.normpath(os.path.join(ICI, "..", "calculsBoissons", "consoTotaleParBoisson.json")), encoding="utf-8"))

wb2 = openpyxl.Workbook()
ws = wb2.active; ws.title = "Synthese"
ws.column_dimensions["A"].width = 80
titre(ws, "PIECE DE DEFENSE - Boissons pretendument disparues", 1)
disp = mc["disparu_montecarlo_euros_revente"]
for t in [
    "",
    "Grief du fisc : des centaines de bouteilles 'disparues' chaque annee, reconstituees en CA.",
    "",
    f"NOTRE RECALCUL (modele d'incertitude Monte Carlo, {mc['n_iterations']} tirages, reproductible) :",
    f"  - Disparu alcool (au prix de revente) = mediane {disp['p50']:,.0f} EUR ; intervalle 95 % : {disp['p2.5']:,.0f} - {disp['p97.5']:,.0f} EUR.",
    f"  - Soit {mc['disparu_en_pct_des_achats']['p50']:.0f} % des achats = la perte NORMALE d'un bar (15-25 %, norme CHR).",
    f"  - Softs/eaux : trou de donnees corrige (ventes recuperees au ticket) -> disparu reel proche de zero.",
    "",
    "CAUSES DE CONSOMMATION DOCUMENTEES (pas des ventes) : sur-versement au verre, conso du chef (Picon, Macvin),",
    "offerts, cuisine jurassienne (fondues/babas/flambage), pertes techniques biere. Doses validees par l'exploitante.",
    "",
    "RECONCILIATION AVEC LES SUPPRESSIONS : le disparu (perte de conso) et les suppressions (workflow caisse)",
    "mesurent des choses differentes et ne s'additionnent pas (voir onglet 'Reconciliation DEL').",
    "",
    "Sources : factures fournisseur (FCBS), ventes caisse (ANNEXE-B/D), 3 inventaires physiques, cartes des vins.",
]:
    ws.append([t])
    if t.strip().startswith(("NOTRE", "CAUSES", "RECONCILIATION")): ws[ws.max_row][0].font = BOLD

# onglet par boisson (disparu en litres)
ws = wb2.create_sheet("Disparu par boisson")
for w, c in zip([34, 14, 12, 12, 12, 12], "ABCDEF"): ws.column_dimensions[c].width = w
titre(ws, "Disparu par boisson (litres, 3 exercices) - achats factures vs conso caisse", 6)
entete(ws, ["Boisson", "Categorie", "Achats L", "Conso L", "Stock fin L", "Disparu L"])
rows = []
for b in conso["boissons"]:
    taille = b.get("taille_achat_cl") or 0
    ach = sum(x for x in (b.get("achats_litres_par_periode") or {}).values() if x)
    cons = b["total_3_exercices"]["total_l"]["moyen"]
    inv = b.get("inventaire_fin_contenants_par_periode") or {}
    sf = (inv.get("2024-2025") or 0) * taille / 100 if taille else 0
    rows.append([b["nom_canonique"], b["categorie"], round(ach), round(cons), round(sf), round(ach - cons - sf)])
for r in sorted(rows, key=lambda x: -x[5])[:40]:
    ligne(ws, r)

# onglet Monte Carlo
ws = wb2.create_sheet("Modele incertitude")
for w, c in zip([46, 18, 18], "ABC"): ws.column_dimensions[c].width = w
titre(ws, "Resultat du modele d'incertitude (Monte Carlo)", 3)
entete(ws, ["Indicateur", "Mediane", "Intervalle 95 %"])
ligne(ws, ["Disparu alcool - litres", mc["disparu_montecarlo_litres"]["p50"],
           f'{mc["disparu_montecarlo_litres"]["p2.5"]:.0f} - {mc["disparu_montecarlo_litres"]["p97.5"]:.0f}'])
ligne(ws, ["Disparu alcool - EUR (revente)", disp["p50"], f'{disp["p2.5"]:.0f} - {disp["p97.5"]:.0f}'])
ligne(ws, ["Disparu en % des achats", f'{mc["disparu_en_pct_des_achats"]["p50"]:.1f} %',
           f'{mc["disparu_en_pct_des_achats"]["p2.5"]:.1f} - {mc["disparu_en_pct_des_achats"]["p97.5"]:.1f} %'])
ligne(ws, ["(reference perte normale CHR)", "15 - 25 %", ""])

# onglet reconciliation DEL
ws = wb2.create_sheet("Reconciliation DEL")
ws.column_dimensions["A"].width = 100
titre(ws, "Pourquoi disparu boissons et suppressions DEL ne s'additionnent pas", 1)
ws.append([rd["synthese"]]); ws[ws.max_row][0].alignment = WRAP
ws.merge_cells(start_row=ws.max_row, start_column=1, end_row=ws.max_row, end_column=1)
ws.row_dimensions[ws.max_row].height = 70
vide(ws)
for k, a in rd["arguments"].items():
    ws.append([k.replace("_", " ").upper()]); ws[ws.max_row][0].font = BOLD
    ws.append([a.get("principe", "")]); ws[ws.max_row][0].alignment = WRAP
    ws.merge_cells(start_row=ws.max_row, start_column=1, end_row=ws.max_row, end_column=1)
    ws.row_dimensions[ws.max_row].height = 45

wb2.save(os.path.join(OUT, "Defense-Boissons-disparues.xlsx"))
print("ecrit : Defense-Boissons-disparues.xlsx (4 onglets)")
print(f"Dossier : {OUT}")
