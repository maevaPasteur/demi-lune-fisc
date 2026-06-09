#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
15_donnees_page_boissons.py
Produit TOUTES les tables de la page /analyses/boissons-disparues, en donnees
exactes et reproductibles, dans src/data/boissonsPageData.json (importe par la page).
Aucune famille : listes exactes, par boisson / plat / cocktail / menu / periode.
"""
import json, os, re
from collections import defaultdict

ICI = os.path.dirname(os.path.abspath(__file__))
CALC = os.path.normpath(os.path.join(ICI, "..", "calculsBoissons"))
SRCDATA = os.path.normpath(os.path.join(ICI, ".."))

def Jc(name): return json.load(open(os.path.join(CALC, name), encoding="utf-8"))
def Ji(name): return json.load(open(os.path.join(ICI, name), encoding="utf-8"))

conso = Jc("consoTotaleParBoisson.json")
base = {x["nom"]: x for x in Ji("base_disparu.json")["boissons"]}
EXOS = ["2022-2023", "2023-2024", "2024-2025"]
JOURS = {"2022-2023": 222, "2023-2024": 221, "2024-2025": 219}

def prix(nom, sens):
    return base.get(nom, {}).get(f"prix_{sens}_l")

def stock_l(b):
    t = b.get("taille_achat_cl") or 0
    inv = b.get("inventaire_fin_contenants_par_periode") or {}
    return (inv.get("2024-2025") or 0) * t / 100 if t else 0

# ---------------------------------------------------------------------------
# 1. DISPARU PAR BOISSON (liste exacte : cout d'achat + CA pretendument perdu)
# ---------------------------------------------------------------------------
disparu_rows = []
for b in conso["boissons"]:
    nom = b["nom_canonique"]
    if b["categorie"] in ("soda", "jus", "sirop"):
        continue  # softs : bilan ferme a part
    ach = sum(x for x in (b.get("achats_litres_par_periode") or {}).values() if x)
    cons = b["total_3_exercices"]["total_l"]["moyen"]
    disp = ach - cons - stock_l(b)
    pa, pr = prix(nom, "achat"), prix(nom, "revente")
    disparu_rows.append({
        "nom": nom, "categorie": b["categorie"],
        "achat_l": round(ach, 1), "conso_l": round(cons, 1), "disparu_l": round(disp, 1),
        "prix_achat": pa, "prix_revente": pr,
        "cout_disparu": round(disp * pa) if (pa and disp > 0) else 0,
        "ca_disparu": round(disp * pr) if (pr and disp > 0) else 0,
    })
disparu_rows.sort(key=lambda r: -(r["ca_disparu"] or 0))

# ---------------------------------------------------------------------------
# 2. COCKTAILS (composition + quantite caisse + alcool consomme)
# ---------------------------------------------------------------------------
ck = Jc("cocktailsConsoComposition.json")
cocktails = []
for c in ck["cocktails"]:
    alc_l = 0.0
    for ex in EXOS:
        for ing, l in (c["par_periode"].get(ex, {}).get("ingredients_l", {})).items():
            # somme uniquement des ingredients alcool
            rec = next((r for r in c["recette"] if r["nom"] == ing), None)
            if rec and rec.get("categorie") == "alcool":
                alc_l += l
    cocktails.append({
        "cocktail": c["cocktail"], "volume_cl": c["volume_cl"],
        "recette": [{"nom": r["nom"], "cl": r["cl_unitaire"], "alcool": r.get("categorie") == "alcool"} for r in c["recette"]],
        "qte_3ans": round(c.get("total_quantite", 0)),
        "alcool_l_3ans": round(alc_l, 1),
    })
cocktails.sort(key=lambda c: -c["alcool_l_3ans"])

# ---------------------------------------------------------------------------
# 3. CUISINE : contenance alcool de chaque plat/entree/dessert/sauce
# ---------------------------------------------------------------------------
plats = Jc("platsAlcool.json")
sauces = Jc("saucesAlcool.json")
cuisine_rows = []
for typ, key in [("Entrée", "entrees"), ("Plat", "plats"), ("Dessert", "desserts")]:
    for p in plats.get(key, []):
        for a in p.get("alcools", []):
            cuisine_rows.append({"plat": p["nom"], "type": typ, "alcool": a["type"], "dose_cl": a["dose_cl"]})
for s in sauces.get("sauces", []):
    cuisine_rows.append({"plat": s["nom"], "type": "Sauce", "alcool": s["alcool"], "dose_cl": s["dose_cl_par_portion"]})
# total cuisine par alcool (litres, 3 ans) depuis consoTotaleParBoisson (plats_desserts)
cuisine_par_alcool = []
for b in conso["boissons"]:
    lp = sum(v.get("detail_exact_l", {}).get("plats_desserts", 0) for v in b["par_periode"].values())
    if lp > 0.5:
        cuisine_par_alcool.append({"alcool": b["nom_canonique"], "litres_3ans": round(lp, 1),
                                   "cout": round(lp * (prix(b["nom_canonique"], "achat") or 0))})
cuisine_par_alcool.sort(key=lambda r: -r["litres_3ans"])

# ---------------------------------------------------------------------------
# 4. MENUS PAR PERIODE (options alcoolisees, dose, proba %, cl)
# ---------------------------------------------------------------------------
mc = Jc("menusCompositionAlcool.json")
menus_par_periode = {}
for carte, cd in mc["cartes"].items():
    rows = []
    for menu, md in cd.get("menus", {}).items():
        for service, opts in md.get("services", {}).items():
            for o in opts:
                rows.append({
                    "menu": menu, "service": service, "option": o["option"],
                    "obligatoire": o.get("obligatoire", False),
                    "alcool": o["alcool"], "dose_cl": o["dose_cl"],
                    "proba_pct": round(100 * o.get("proba_moyenne", 0)),
                    "cl_moyen": o.get("cl_moyen", 0),
                })
    menus_par_periode[carte] = {"dates": cd.get("dates", ""), "lignes": rows}
# alcool des menus par periode (litres) depuis consoAlcoolMenus
cam = Jc("consoAlcoolMenus.json")
menu_alcool_par_periode = {}
for ex, d in cam["exercices"].items():
    menu_alcool_par_periode[ex] = [{"alcool": a["alcool"], "litres": round(a.get("litres_moyen", a.get("cl_moyen", 0) / 100), 1)}
                                   for a in d["alcools"]]

# ---------------------------------------------------------------------------
# 5. PERSONNEL (depuis synthese_perte_reelle)
# ---------------------------------------------------------------------------
spr = Ji("synthese_perte_reelle.json")
personnel = spr["personnel"]

# ---------------------------------------------------------------------------
# 6. CONSO PAR PERIODE (caisse + cuisine + menu + personnel -> total, cout, CA)
# ---------------------------------------------------------------------------
# personnel ALCOOL uniquement (chef : Macvin + Picon) reparti par periode au prorata
# des jours. Les softs du personnel (Coca/jus/sirop) relevent du bilan softs separe.
PERSO_ALC_L = 99.0 + 44.0           # Macvin + Picon
PERSO_ALC_COUT = 1797 + 457
PERSO_ALC_CA = 8085 + 6490
def perso_periode(ex):
    f = JOURS[ex] / sum(JOURS.values())
    return PERSO_ALC_L * f, PERSO_ALC_COUT * f, PERSO_ALC_CA * f

conso_par_periode = {}
tot = defaultdict(float)
for ex in EXOS:
    caisse_l = cuisine_l = menu_l = cout = ca = 0.0
    for b in conso["boissons"]:
        if b["categorie"] in ("soda", "jus", "sirop"):
            continue
        v = b["par_periode"].get(ex, {})
        de = v.get("detail_exact_l", {})
        seches = de.get("boissons_seches", 0) + de.get("ingredients_cocktails", 0)
        cuis = de.get("plats_desserts", 0)
        menu = v.get("menu_estime_l", {}).get("moyen", 0)
        l = seches + cuis + menu
        caisse_l += seches; cuisine_l += cuis; menu_l += menu
        cout += l * (prix(b["nom_canonique"], "achat") or 0)
        ca += l * (prix(b["nom_canonique"], "revente") or 0)
    pl, pc, pca = perso_periode(ex)
    total_l = caisse_l + cuisine_l + menu_l + pl
    row = {"caisse_l": round(caisse_l), "cuisine_l": round(cuisine_l), "menu_l": round(menu_l),
           "personnel_l": round(pl), "total_l": round(total_l),
           "cout": round(cout + pc), "ca": round(ca + pca)}
    conso_par_periode[ex] = row
    for k, val in row.items(): tot[k] += val
conso_par_periode["total"] = {k: round(v) for k, v in tot.items()}

# ---------------------------------------------------------------------------
# 7. SYNTHESE finale (reconstitution mathematique)
# ---------------------------------------------------------------------------
achat_alcool_l = sum(r["achat_l"] for r in disparu_rows)
achat_alcool_cout = sum((r["achat_l"] * (r["prix_achat"] or 0)) for r in disparu_rows)
disparu_total_l = sum(r["disparu_l"] for r in disparu_rows if r["disparu_l"] > 0)
disparu_total_cout = sum(r["cout_disparu"] for r in disparu_rows)
disparu_total_ca = sum(r["ca_disparu"] for r in disparu_rows)
synthese = {
    "achat_alcool_l": round(achat_alcool_l), "achat_alcool_cout": round(achat_alcool_cout),
    "conso_totale_l": conso_par_periode["total"]["total_l"],
    "disparu_brut_l": round(disparu_total_l), "disparu_brut_cout": round(disparu_total_cout),
    "disparu_brut_ca": round(disparu_total_ca),
    "perte_reelle_l": spr["cascade_alcool"]["perte_reelle_residuelle_l"],
    "perte_reelle_pct": spr["cascade_alcool"]["perte_reelle_pct_achats"],
    "cascade": spr["cascade_alcool"]["postes"],
    "fisc_ca_reconstitue": 575253, "fisc_coef": 3.1, "ca_declare": 435525,
}

out = {
    "description": "Donnees exactes de la page boissons-disparues (script 15, reproductible).",
    "disparuParBoisson": disparu_rows,
    "cocktails": cocktails,
    "cuisine": cuisine_rows,
    "cuisineParAlcool": cuisine_par_alcool,
    "menusParPeriode": menus_par_periode,
    "menuAlcoolParPeriode": menu_alcool_par_periode,
    "personnel": personnel,
    "offerts": spr["offerts"],
    "consoParPeriode": conso_par_periode,
    "synthese": synthese,
}
json.dump(out, open(os.path.join(SRCDATA, "boissonsPageData.json"), "w", encoding="utf-8"),
          ensure_ascii=False, indent=1)

# ---------------------------------------------------------------------------
# XLSX complet (1 onglet par table de la page) pour public/documents
# ---------------------------------------------------------------------------
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
OUTX = os.path.normpath(os.path.join(ICI, "..", "..", "..", "public", "documents", "pieces-defense"))
H1 = Font(bold=True, size=13, color="FFFFFF"); FILL1 = PatternFill("solid", fgColor="1F3A5F")
BOLD = Font(bold=True); FILLT = PatternFill("solid", fgColor="E2E8F0")
THIN = Border(*[Side(style="thin", color="CBD5E0")] * 4); R = Alignment(horizontal="right")
def _t(ws, t, n):
    ws.append([t]); ws.merge_cells(start_row=ws.max_row, start_column=1, end_row=ws.max_row, end_column=n)
    c = ws.cell(ws.max_row, 1); c.font = H1; c.fill = FILL1; ws.row_dimensions[ws.max_row].height = 22
def _h(ws, cols):
    ws.append(cols)
    for c in ws[ws.max_row]: c.font = BOLD; c.fill = FILLT; c.border = THIN
def _l(ws, vals, b=False):
    ws.append(vals)
    for c in ws[ws.max_row]:
        c.border = THIN
        if b: c.font = BOLD
        if isinstance(c.value, (int, float)): c.alignment = R
wb = openpyxl.Workbook()
ws = wb.active; ws.title = "1-Manquant par boisson"
for w, c in zip([32, 10, 10, 10, 14, 18], "ABCDEF"): ws.column_dimensions[c].width = w
_t(ws, "Manquant par boisson : cout d'achat et CA pretendument perdu", 6)
_h(ws, ["Boisson", "Achat L", "Conso L", "Manquant L", "Cout achat EUR", "CA perdu EUR"])
for r in disparu_rows: _l(ws, [r["nom"], r["achat_l"], r["conso_l"], r["disparu_l"], r["cout_disparu"], r["ca_disparu"]])
_l(ws, ["TOTAL", "", "", synthese["disparu_brut_l"], synthese["disparu_brut_cout"], synthese["disparu_brut_ca"]], b=True)

ws = wb.create_sheet("2-Cocktails")
for w, c in zip([22, 40, 8, 12, 14], "ABCDE"): ws.column_dimensions[c].width = w
_t(ws, "Cocktails : recette et alcool consomme", 5)
_h(ws, ["Cocktail", "Recette alcool", "Volume cl", "Vendus 3 ans", "Alcool L"])
for c in cocktails:
    rec = " + ".join(f"{r['nom']} {r['cl']}cl" for r in c["recette"])
    _l(ws, [c["cocktail"], rec, c["volume_cl"], c["qte_3ans"], c["alcool_l_3ans"]])

ws = wb.create_sheet("3-Cuisine")
for w, c in zip([30, 10, 16, 8], "ABCD"): ws.column_dimensions[c].width = w
_t(ws, "Alcool de chaque plat / entree / dessert / sauce", 4)
_h(ws, ["Plat", "Type", "Alcool", "Dose cl"])
for p in cuisine_rows: _l(ws, [p["plat"], p["type"], p["alcool"], p["dose_cl"]])

ws = wb.create_sheet("4-Menus par periode")
for w, c in zip([16, 10, 30, 14, 8, 10, 12], "ABCDEFG"): ws.column_dimensions[c].width = w
_t(ws, "Composition alcool des menus, par periode et probabilite", 7)
for carte in ["C", "B", "A"]:
    cd_ = menus_par_periode.get(carte)
    if not cd_: continue
    _l(ws, [f"--- Carte {carte} ({cd_['dates']}) ---"], b=True)
    _h(ws, ["Menu", "Service", "Option", "Alcool", "Dose", "Choisi %", "Alcool/menu cl"])
    for r in cd_["lignes"]:
        _l(ws, [r["menu"], r["service"], r["option"], r["alcool"], r["dose_cl"], r["proba_pct"], r["cl_moyen"]])

ws = wb.create_sheet("5-Personnel")
for w, c in zip([44, 10, 16, 18], "ABCD"): ws.column_dimensions[c].width = w
_t(ws, "Consommation du personnel et du chef", 4)
_h(ws, ["Poste", "Litres", "Cout achat EUR", "CA equivalent EUR"])
for l in personnel["lignes"]: _l(ws, [l["poste"], l["litres"], l["cout_achat_eur"], l["ca_equivalent_eur"]])
_l(ws, ["TOTAL", personnel["total_litres"], personnel["total_cout_eur"], personnel["total_ca_eur"]], b=True)

ws = wb.create_sheet("6-Conso par periode")
for w, c in zip([34, 14, 14, 14, 14], "ABCDE"): ws.column_dimensions[c].width = w
_t(ws, "Consommation totale par exercice (caisse + cuisine + menus + personnel)", 5)
_h(ws, ["Source", "2022-2023", "2023-2024", "2024-2025", "Total"])
for lab, k in [("Vendu en caisse", "caisse_l"), ("Cuisine", "cuisine_l"), ("Menus", "menu_l"), ("Personnel", "personnel_l"), ("TOTAL consomme L", "total_l"), ("Cout d'achat EUR", "cout"), ("CA revente EUR", "ca")]:
    _l(ws, [lab, conso_par_periode["2022-2023"][k], conso_par_periode["2023-2024"][k], conso_par_periode["2024-2025"][k], conso_par_periode["total"][k]], b=(lab.startswith("TOTAL") or "EUR" in lab))
wb.save(os.path.join(OUTX, "Boissons-detail-complet.xlsx"))
print("Boissons-detail-complet.xlsx ecrit (6 onglets).")

print("boissonsPageData.json ecrit.")
print(f"  disparu par boisson : {len(disparu_rows)} lignes")
print(f"  cocktails : {len(cocktails)} | cuisine : {len(cuisine_rows)} plats")
print(f"  conso/periode total : {conso_par_periode['total']}")
print(f"  synthese disparu brut : {synthese['disparu_brut_l']} L / {synthese['disparu_brut_ca']:,} € revente ; perte reelle {synthese['perte_reelle_l']} L ({synthese['perte_reelle_pct']}%)")
