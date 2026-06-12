#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
17_avoirs_factures.py
Piece fiscale : detail LIGNE A LIGNE des 199 factures fournisseur (FCBS), avec
numero, date, exercice et la MENTION de chaque ligne speciale (avoir, retour,
deconsigne, article manquant, consigne...) et le TRAITEMENT applique aux achats.
But : prouver au fisc que les avoirs et retours sont correctement traites, sans
gonfler les achats. Sortie : public/documents/pieces-defense/Factures-avoirs-et-retours.xlsx
"""
import json, os, re
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

ICI = os.path.dirname(os.path.abspath(__file__))
SRCDATA = os.path.normpath(os.path.join(ICI, ".."))
OUT = os.path.normpath(os.path.join(ICI, "..", "..", "..", "public", "documents", "pieces-defense"))
os.makedirs(OUT, exist_ok=True)

fact = json.load(open(os.path.join(SRCDATA, "factures-fournisseur.json"), encoding="utf-8"))

EXOS = ["2022-2023", "2023-2024", "2024-2025"]
def exercice_de(ds):
    if not ds:
        return "hors-periode"
    m = re.match(r"(\d{2})/(\d{2})/(\d{4})", str(ds))
    if not m:
        return "hors-periode"
    dd, mo, y = int(m.group(1)), int(m.group(2)), int(m.group(3))
    deb = y if mo >= 4 else y - 1
    e = f"{deb}-{deb+1}"
    return e if e in EXOS else "hors-periode"

# Produits dont l'avoir a ete RETIRE des achats (cf. scripts 14/15)
AVOIRS_CORRIGES = {  # code -> (nom, litres)
    "403200": ("Cidre Brut", 22.5),
    "603040": ("Bourgogne Aligoté maison", 10.0),
    "510090": ("Grand Marnier", 2.1),
}

def mention_et_traitement(f, ln):
    """Retourne (mention, traitement) pour une ligne de facture."""
    des = (ln.get("designation") or "")
    desU = des.upper()
    condU = (ln.get("conditionnement") or "").upper()
    q = ln.get("quantite") or 0
    ht = ln.get("montantHT")
    code = str(ln.get("code"))
    if f.get("type") == "avoir":
        if "CONSIGNE" in desU or "CONSIGNE" in condU:
            return "AVOIR (consigne de fûts)", "Emballage repris, sans volume de boisson"
        if code in AVOIRS_CORRIGES:
            return "AVOIR (note de crédit)", "RETIRÉ des achats (corrigé)"
        return "AVOIR (note de crédit)", "Hors champ boissons (ou produit non listé)"
    if q < 0 or (ht is not None and ht < 0):
        return "RETOUR / REPRISE", "Déjà déduit (quantité négative)"
    if ht is None or ht == 0:
        return "NON FACTURÉ (manquant/refusé/échantillon)", "Non compté (ni volume ni coût)"
    if "DECONSIGNE" in desU or "DÉCONSIGNE" in desU:
        return "DÉCONSIGNE", "Emballage, sans volume de boisson"
    if "CONSIGNE" in desU or "CONSIGNE" in condU:
        return "CONSIGNE", "Emballage, sans volume de boisson"
    if "MANQUANT" in desU:
        return "ARTICLE MANQUANT (annoté)", "Non reçu : non compté si HT = 0"
    if "REPRISE" in desU:
        return "REPRISE", "Reprise fournisseur"
    return "", "Achat comptabilisé"

# ----------------------------- styles XLSX -----------------------------------
H1 = Font(bold=True, size=13, color="FFFFFF"); FILL1 = PatternFill("solid", fgColor="1F3A5F")
BOLD = Font(bold=True); FILLT = PatternFill("solid", fgColor="E2E8F0")
FILL_AV = PatternFill("solid", fgColor="FCE7E7")   # avoir : rouge pale
FILL_RET = PatternFill("solid", fgColor="FEF3C7")  # retour/consigne : ambre pale
FILL_OK = PatternFill("solid", fgColor="E8F3EC")   # corrige : vert pale
THIN = Border(*[Side(style="thin", color="CBD5E0")] * 4)
R = Alignment(horizontal="right"); WRAP = Alignment(wrap_text=True, vertical="top")

def title(ws, t, n):
    ws.append([t]); ws.merge_cells(start_row=ws.max_row, start_column=1, end_row=ws.max_row, end_column=n)
    c = ws.cell(ws.max_row, 1); c.font = H1; c.fill = FILL1; ws.row_dimensions[ws.max_row].height = 22

def head(ws, cols):
    ws.append(cols)
    for c in ws[ws.max_row]:
        c.font = BOLD; c.fill = FILLT; c.border = THIN

def row(ws, vals, fill=None, bold=False):
    ws.append(vals)
    for c in ws[ws.max_row]:
        c.border = THIN
        if fill: c.fill = fill
        if bold: c.font = BOLD
        if isinstance(c.value, (int, float)): c.alignment = R

wb = openpyxl.Workbook()

# ===================== Onglet 1 : TOUTES LES LIGNES ==========================
ws = wb.active; ws.title = "Toutes les lignes"
for w, c in zip([10, 12, 12, 9, 46, 8, 10, 11, 30, 30], "ABCDEFGHIJ"):
    ws.column_dimensions[c].width = w
title(ws, "Détail ligne à ligne des 199 factures fournisseur (FCBS) - mentions et traitement", 10)
head(ws, ["N° facture", "Date", "Exercice", "Code", "Désignation", "Qté", "PU net €", "HT €", "Mention", "Traitement achats"])
n_av = n_ret = n_nonfac = 0
for f in fact["factures"]:
    num = f.get("numero"); date = f.get("dateFacture"); ex = exercice_de(date)
    for ln in f["lignes"]:
        men, trait = mention_et_traitement(f, ln)
        fill = FILL_AV if men.startswith("AVOIR") else (FILL_RET if men in ("RETOUR / REPRISE", "CONSIGNE", "DÉCONSIGNE", "REPRISE") or men.startswith("ARTICLE") or men.startswith("NON") else None)
        if "RETIRÉ" in trait: fill = FILL_OK
        row(ws, [num, date, ex, ln.get("code"), ln.get("designation"),
                 ln.get("quantite"), ln.get("puNet"), ln.get("montantHT"), men, trait], fill=fill)
        if men.startswith("AVOIR"): n_av += 1
        if men == "RETOUR / REPRISE": n_ret += 1
        if men.startswith("NON FACTURÉ"): n_nonfac += 1
ws.freeze_panes = "A3"

# ===================== Onglet 2 : AVOIRS (6) =================================
ws = wb.create_sheet("Avoirs (notes de crédit)")
for w, c in zip([10, 12, 9, 46, 8, 10, 11, 32], "ABCDEFGH"):
    ws.column_dimensions[c].width = w
title(ws, "Les 6 avoirs autonomes (notes de crédit) et leur traitement", 8)
head(ws, ["N° avoir", "Date", "Code", "Désignation", "Qté", "HT €", "Total avoir TTC €", "Traitement"])
for f in fact["factures"]:
    if f.get("type") != "avoir":
        continue
    tot = (f.get("totaux") or {}).get("totalTTC")
    for ln in f["lignes"]:
        men, trait = mention_et_traitement(f, ln)
        fill = FILL_OK if "RETIRÉ" in trait else FILL_AV
        row(ws, [f.get("numero"), f.get("dateFacture"), ln.get("code"), ln.get("designation"),
                 ln.get("quantite"), ln.get("montantHT"), tot, trait], fill=fill)

# ===================== Onglet 3 : RETOURS & DÉCONSIGNES ======================
ws = wb.create_sheet("Retours et déconsignes")
for w, c in zip([10, 12, 9, 46, 8, 11, 30], "ABCDEFG"):
    ws.column_dimensions[c].width = w
title(ws, "Retours (quantités négatives) et lignes annotées - déjà déduits des achats", 7)
head(ws, ["N° facture", "Date", "Code", "Désignation", "Qté", "HT €", "Mention"])
for f in fact["factures"]:
    if f.get("type") == "avoir":
        continue
    for ln in f["lignes"]:
        men, trait = mention_et_traitement(f, ln)
        if men in ("RETOUR / REPRISE", "DÉCONSIGNE", "REPRISE") or men.startswith("ARTICLE") or men.startswith("NON FACTURÉ"):
            row(ws, [f.get("numero"), f.get("dateFacture"), ln.get("code"), ln.get("designation"),
                     ln.get("quantite"), ln.get("montantHT"), men], fill=FILL_RET)

# ===================== Onglet 4 : CORRECTION ACHATS =========================
ws = wb.create_sheet("Correction des achats")
for w, c in zip([28, 12, 14, 16, 16, 14], "ABCDEF"):
    ws.column_dimensions[c].width = w
title(ws, "Correction appliquée : avoirs retirés des achats (volume de boisson)", 6)
head(ws, ["Produit", "N° avoir", "Date avoir", "Achat AVANT (L)", "Avoir retiré (L)", "Achat APRÈS (L)"])
CORR = [("Cidre Brut", "520312", "16/07/2024", 207.0, 22.5),
        ("Bourgogne Aligoté maison", "532782", "11/12/2024", 640.0, 10.0),
        ("Grand Marnier", "526876", "25/09/2024", 10.5, 2.1)]
tot_av = 0.0
for nom, num, dt, avant, av in CORR:
    row(ws, [nom, num, dt, avant, av, round(avant - av, 1)], fill=FILL_OK)
    tot_av += av
row(ws, ["TOTAL retiré des achats", "", "", "", round(tot_av, 1), ""], bold=True)
ws.append([])
ws.append(["Note : les 40 retours intégrés aux factures (quantités négatives) étaient déjà déduits."])
ws.append(["Les lignes de consigne/déconsigne sont des emballages, sans volume de boisson."])

wb.save(os.path.join(OUT, "Factures-avoirs-et-retours.xlsx"))
print(f"Factures-avoirs-et-retours.xlsx écrit : {sum(len(f['lignes']) for f in fact['factures'])} lignes, "
      f"{n_av} avoir, {n_ret} retours, {n_nonfac} non facturés. Avoirs retirés = {round(tot_av,1)} L.")
