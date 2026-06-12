#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
16_generer_xlsx_juridiques.py
Genere un maximum de pieces XLSX (a partir des JSON du dossier) pour la preuve
juridique : tables boissons granulaires (1 par tableau de la page) + dossiers
"methode / preuve de process". Ecrit aussi un manifeste consomme par documents.ts.
"""
import json, os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

ICI = os.path.dirname(os.path.abspath(__file__))
SRCDATA = os.path.normpath(os.path.join(ICI, ".."))
OUT = os.path.normpath(os.path.join(ICI, "..", "..", "..", "public", "documents", "pieces-defense"))
os.makedirs(OUT, exist_ok=True)

def Ji(n): return json.load(open(os.path.join(ICI, n), encoding="utf-8"))
BPD = json.load(open(os.path.join(SRCDATA, "boissonsPageData.json"), encoding="utf-8"))

H1 = Font(bold=True, size=13, color="FFFFFF"); FILL1 = PatternFill("solid", fgColor="1F3A5F")
BOLD = Font(bold=True); FILLT = PatternFill("solid", fgColor="E2E8F0")
THIN = Border(*[Side(style="thin", color="CBD5E0")] * 4); RA = Alignment(horizontal="right")
WRAP = Alignment(wrap_text=True, vertical="top")

def ws_title(ws, t, n=6):
    ws.append([t]); ws.merge_cells(start_row=ws.max_row, start_column=1, end_row=ws.max_row, end_column=n)
    c = ws.cell(ws.max_row, 1); c.font = H1; c.fill = FILL1; ws.row_dimensions[ws.max_row].height = 22

def head(ws, cols):
    ws.append(cols)
    for c in ws[ws.max_row]: c.font = BOLD; c.fill = FILLT; c.border = THIN

def row(ws, vals, b=False):
    ws.append(vals)
    for c in ws[ws.max_row]:
        c.border = THIN
        if b: c.font = BOLD
        if isinstance(c.value, (int, float)): c.alignment = RA

def widths(ws, ws_):
    for col, w in ws_.items(): ws.column_dimensions[col].width = w

def text_block(ws, txt, n=4):
    ws.append([txt]); ws.merge_cells(start_row=ws.max_row, start_column=1, end_row=ws.max_row, end_column=n)
    ws[ws.max_row][0].alignment = WRAP; ws.row_dimensions[ws.max_row].height = 42

# ============================ TABLES BOISSONS (granulaires) ==================
def b1_manquant(wb):
    ws = wb.active; ws.title = "Manquant par boisson"
    widths(ws, {"A": 32, "B": 11, "C": 11, "D": 11, "E": 11, "F": 14, "G": 16})
    ws_title(ws, "Manquant = Achat - Conso - Stock fin (cout + CA pretendument perdu)", 7)
    head(ws, ["Boisson", "Achat L", "Conso L", "Stock fin L", "Manquant L", "Cout achat EUR", "CA perdu EUR"])
    for r in BPD["disparuParBoisson"]:
        row(ws, [r["nom"], r["achat_l"], r["conso_l"], r.get("stock_l", 0), r["disparu_l"], r["cout_disparu"], r["ca_disparu"]])
    s = BPD["synthese"]
    row(ws, ["TOTAL (manquant positif)", "", "", "", s["disparu_brut_l"], s["disparu_brut_cout"], s["disparu_brut_ca"]], b=True)
    row(ws, ["TOTAL NET (lignes negatives incluses)", "", "", "", s.get("disparu_net_l", ""), "", ""], b=True)

def b2_cocktails(wb):
    ws = wb.active; ws.title = "Cocktails"
    widths(ws, {"A": 22, "B": 46, "C": 9, "D": 13, "E": 13})
    ws_title(ws, "Cocktails : composition complete et alcool consomme", 5)
    head(ws, ["Cocktail", "Recette complete", "Volume cl", "Vendus 3 ans", "Alcool L"])
    for c in BPD["cocktails"]:
        rec = " + ".join(f"{r['nom']} {r['cl']}cl" for r in c["recette"])
        row(ws, [c["cocktail"], rec, c["volume_cl"], c["qte_3ans"], c["alcool_l_3ans"]])

def b3_cuisine(wb):
    ws = wb.active; ws.title = "Cuisine par plat"
    widths(ws, {"A": 32, "B": 10, "C": 16, "D": 9})
    ws_title(ws, "Alcool de chaque plat / entree / dessert / sauce", 4)
    head(ws, ["Plat", "Type", "Alcool", "Dose cl"])
    for p in BPD["cuisine"]:
        row(ws, [p["plat"], p["type"], p["alcool"], p["dose_cl"]])
    ws2 = wb.create_sheet("Total par alcool")
    widths(ws2, {"A": 22, "B": 12, "C": 14})
    ws_title(ws2, "Ce que la cuisine consomme au total (3 ans)", 3)
    head(ws2, ["Alcool", "Litres", "Cout achat EUR"])
    for r in BPD["cuisineParAlcool"]:
        row(ws2, [r["alcool"], r["litres_3ans"], r["cout"]])

def b4_menus(wb):
    ws = wb.active; ws.title = "Menus par periode"
    widths(ws, {"A": 16, "B": 10, "C": 32, "D": 14, "E": 8, "F": 10, "G": 13})
    ws_title(ws, "Composition alcool des menus par periode et probabilite de choix", 7)
    for carte in ["C", "B", "A"]:
        cd = BPD["menusParPeriode"].get(carte)
        if not cd: continue
        row(ws, [f"--- Carte {carte} ({cd['dates']}) ---"], b=True)
        head(ws, ["Menu", "Service", "Option", "Alcool", "Dose cl", "Choisi %", "Alcool/menu cl"])
        for r in cd["lignes"]:
            row(ws, [r["menu"], r["service"], r["option"], r["alcool"], r["dose_cl"], r["proba_pct"], r["cl_moyen"]])

def b5_conso_periode(wb):
    ws = wb.active; ws.title = "Conso par periode"
    widths(ws, {"A": 34, "B": 14, "C": 14, "D": 14, "E": 14})
    ws_title(ws, "Consommation totale par exercice (caisse + cuisine + menus + personnel)", 5)
    head(ws, ["Source", "2022-2023", "2023-2024", "2024-2025", "Total"])
    P = BPD["consoParPeriode"]
    for lab, k in [("Vendu en caisse", "caisse_l"), ("Cuisine", "cuisine_l"), ("Menus", "menu_l"),
                   ("Personnel (alcool)", "personnel_l"), ("TOTAL consomme L", "total_l"),
                   ("Cout d'achat EUR", "cout"), ("CA revente EUR", "ca")]:
        row(ws, [lab, P["2022-2023"][k], P["2023-2024"][k], P["2024-2025"][k], P["total"][k]],
            b=(lab.startswith("TOTAL") or "EUR" in lab))

# ============================ METHODE / PREUVE DE PROCESS ====================
def m1_montecarlo(wb):
    p = Ji("parametres.json"); r = Ji("resultats_montecarlo_ajuste2.json")
    s = Ji("sensibilite.json"); g = Ji("signature.json")
    ws = wb.active; ws.title = "Parametres"
    widths(ws, {"A": 26, "B": 10, "C": 10, "D": 10, "E": 52})
    ws_title(ws, "Modele d'incertitude (Monte Carlo) - parametres d'hypothese", 5)
    text_block(ws, p["avertissement"], 5)
    head(ws, ["Parametre", "Min", "Mode", "Max", "Justification"])
    for k, v in p["parametres"].items():
        row(ws, [k, v.get("min", v.get("champ_min", "")), v.get("mode", v.get("champ_mode", "")),
                 v.get("max", v.get("champ_max", "")), v.get("justification", "")])
    ws2 = wb.create_sheet("Resultats")
    widths(ws2, {"A": 40, "B": 16, "C": 22})
    ws_title(ws2, "Resultats Monte Carlo (apres corrections)", 3)
    head(ws2, ["Indicateur", "Mediane", "IC 95 %"])
    dl, de = r["disparu_montecarlo_litres"], r["disparu_montecarlo_euros_revente"]
    pa = r["disparu_en_pct_des_achats"]
    row(ws2, ["Disparu alcool - litres", dl["p50"], f'{dl["p2.5"]:.0f} - {dl["p97.5"]:.0f}'])
    row(ws2, ["Disparu alcool - EUR revente", de["p50"], f'{de["p2.5"]:.0f} - {de["p97.5"]:.0f}'])
    row(ws2, ["Disparu en % des achats", f'{pa["p50"]:.1f} %', f'{pa["p2.5"]:.1f} - {pa["p97.5"]:.1f} %'])
    ws3 = wb.create_sheet("Sensibilite")
    widths(ws3, {"A": 34, "B": 16})
    ws_title(ws3, "Analyse de sensibilite (tornado)", 2)
    head(ws3, ["Parametre", "Balancement EUR"])
    for t in s["tornado"]:
        row(ws3, [t["label"], t["balancement_eur"]])
    ws4 = wb.create_sheet("Signature")
    widths(ws4, {"A": 50, "B": 14, "C": 12})
    ws_title(ws4, "Signature du disparu par sous-groupe", 3)
    head(ws4, ["Groupe", "Disparu EUR", "% achats"])
    for r2 in g["groupes"]:
        row(ws4, [r2["groupe"], r2["disparu_eur_revente"], r2.get("taux_disparu_pct_achats", "")])

def m2_cubis(wb):
    m = Ji("mapping_audit.json")
    ws = wb.active; ws.title = "Correction Cubis"
    widths(ws, {"A": 48, "B": 14, "C": 14})
    ws_title(ws, "Correction du bug d'attribution 'Cubis de vin'", 3)
    text_block(ws, m["scan_orphelins"]["conclusion_scan"], 3)
    row(ws, [f"Cubis redistribue : {m['scan_orphelins']['cubis_litres_redistribues']} L", "", ""], b=True)
    head(ws, ["Vin maison", "Disparu avant L", "Disparu apres L"])
    for nom in m["disparu_maison_avant_l"]:
        row(ws, [nom, m["disparu_maison_avant_l"][nom], m["disparu_maison_apres_l"][nom]])
    row(ws, ["GAIN", m["gain_litres"], f"{m['gain_eur_revente']} EUR"], b=True)

def m3_softs(wb):
    sb = Ji("softs_balance.json")
    ws = wb.active; ws.title = "Bilan softs"
    widths(ws, {"A": 18, "B": 11, "C": 11, "D": 11, "E": 11, "F": 11, "G": 11})
    ws_title(ws, "Softs : bilan ferme (ventes au ticket + staff + stock)", 7)
    text_block(ws, sb["constat"], 7)
    head(ws, ["Categorie", "Achat L", "Vendu L", "Cocktails L", "Staff L", "Stock L", "% ferme"])
    for cat, v in sb["bilan_agrege_litres"].items():
        row(ws, [cat, v["achat_l"], v["vendu_direct_l"], v["conso_cocktails_l"], v["staff_l"], v["stock_l"], v["pct_ferme"]])
    cc = sb["coca_canettes"]
    ws2 = wb.create_sheet("Coca (canettes)")
    widths(ws2, {"A": 36, "B": 12})
    ws_title(ws2, "Bilan Coca en canettes", 2)
    for lab, k in [("Achat", "achat_canettes"), ("Vendu (tickets)", "vendu_tickets_canettes"),
                   ("Staff (6/jour)", "staff_canettes"), ("Stock final", "stock_final_canettes"),
                   ("Compte nomme", "compte_nomme"), ("Residu (offerts/casse)", "residu_offerts_casse_canettes")]:
        row(ws2, [lab, cc[k]], b=(k == "compte_nomme"))

def m4_staff(wb):
    sc = Ji("synthese_perte_reelle.json")
    ws = wb.active; ws.title = "Conso personnel"
    widths(ws, {"A": 44, "B": 10, "C": 16, "D": 18, "E": 26})
    ws_title(ws, "Consommation du personnel et de l'owner (detail)", 5)
    head(ws, ["Poste", "Litres", "Cout achat EUR", "CA equivalent EUR", "Base de calcul"])
    for r in sc["personnel"]["lignes"]:
        row(ws, [r["poste"], r["litres"], r["cout_achat_eur"], r["ca_equivalent_eur"], r["base"]])
    pr = sc["personnel"]
    row(ws, ["TOTAL", pr["total_litres"], pr["total_cout_eur"], pr["total_ca_eur"], ""], b=True)

def m5_suppr(wb):
    rs = Ji("reconciliation_suppressions.json"); pz = Ji("reconciliation_par_z.json")
    rd = Ji("reconciliation_disparu_vs_del.json")
    ws = wb.active; ws.title = "Reconciliation"
    widths(ws, {"A": 14, "B": 14, "C": 14, "D": 10, "E": 14, "F": 10})
    ws_title(ws, "Suppressions de caisse : CA declare = encaissements", 6)
    head(ws, ["Exercice", "CA declare", "Encaissements", "% especes", "Suppressions", "Suppr/CA"])
    for exo, v in rs["par_exercice"].items():
        row(ws, [exo, v["ca_declare_ttc"], v["encaissements_ttc"], v["part_especes_pct"], v["suppressions_eur"], v["suppr_sur_ca_pct"]])
    t = rs["totaux"]
    row(ws, ["TOTAL", t["ca_declare_ttc"], t["encaissements_ttc"], t["part_especes_pct"], t["suppressions_total"], t["suppr_sur_ca_pct"]], b=True)
    ws2 = wb.create_sheet("Triangulation")
    widths(ws2, {"A": 42, "B": 16})
    ws_title(ws2, "3 exports certifies = meme CA (suppressions en dehors)", 2)
    tr = pz["triangulation_3_exports_certifies"]
    head(ws2, ["Source certifiee", "CA 3 exercices"])
    row(ws2, ["Synthese (ANNEXE-A)", tr["ca_synthese_ANNEXE_A"]])
    row(ws2, ["Liste tickets (ANNEXE-H)", tr["ca_liste_tickets_ANNEXE_H"]])
    row(ws2, ["Encaissements (ANNEXE-A)", tr["encaissements_ANNEXE_A"]])
    row(ws2, ["Suppressions (ANNEXE-E) - hors des 3", tr["suppressions_ANNEXE_E"]], b=True)
    ws3 = wb.create_sheet("Sessions impossibles")
    widths(ws3, {"A": 12, "B": 12, "C": 14, "D": 14, "E": 10})
    ws_title(ws3, "Sessions ou les suppressions depassent le CA du jour", 5)
    head(ws3, ["Session Z", "Exercice", "CA tickets", "Suppressions", "Nb lignes"])
    for z in pz["z_avec_suppr_superieure_au_ca"]["exemples"]:
        row(ws3, [z["z"], z["exercice"], z["ca_tickets"], z["suppr"], z["nb_suppr"]])
    ws4 = wb.create_sheet("Disparu vs DEL")
    widths(ws4, {"A": 100})
    ws_title(ws4, "Pourquoi disparu boissons et suppressions ne s'additionnent pas", 1)
    text_block(ws4, rd["synthese"], 1)
    for k, a in rd["arguments"].items():
        ws4.append([a.get("principe", "")]); ws4[ws4.max_row][0].alignment = WRAP
        ws4.merge_cells(start_row=ws4.max_row, start_column=1, end_row=ws4.max_row, end_column=1)
        ws4.row_dimensions[ws4.max_row].height = 40

def m6_quantite(wb):
    ab = Ji("aberrations_del.json"); dj = Ji("del_justification.json")
    ws = wb.active; ws.title = "Erreurs de quantite"
    widths(ws, {"A": 12, "B": 14, "C": 30})
    ws_title(ws, "Grosses suppressions = fautes de frappe sur la quantite", 3)
    text_block(ws, ab["avertissement_honnete"], 3)
    head(ws, ["Date", "Montant EUR", "Decomposition prix x quantite"])
    for d in dj["A_erreurs_quantite"]["lignes"]:
        row(ws, [d["date"], d["montant"], f"{d['prix']} EUR x {d['quantite']}"])
    row(ws, ["TOTAL", dj["A_erreurs_quantite"]["somme_eur"], f"{dj['A_erreurs_quantite']['pct_du_total_del']} % du total DEL"], b=True)
    ws2 = wb.create_sheet("Distribution prix")
    widths(ws2, {"A": 12, "B": 14, "C": 16})
    ws_title(ws2, "Les suppressions epousent la grille de prix des ventes", 3)
    head(ws2, ["Prix", "% ventes", "% suppressions"])
    for d in dj["B_distribution_prix"]["top_prix"]:
        row(ws2, [d["prix"], d["part_ventes_pct"], d["part_suppr_pct"]])

# ============================ GENERATION + MANIFESTE ========================
DOSSIER = "pieces-defense"
PIECES = [
    # boisson granulaires (CTA par tableau)
    ("DB1", "Boissons-1-manquant-par-boisson.xlsx", "Manquant par boisson", "Liste exacte : achat, conso, manquant, coUt et CA perdu par boisson.", "boisson", b1_manquant),
    ("DB2", "Boissons-2-cocktails.xlsx", "Cocktails (composition)", "Composition complete de chaque cocktail et alcool consomme.", "boisson", b2_cocktails),
    ("DB3", "Boissons-3-cuisine.xlsx", "Cuisine (alcool des plats)", "Alcool de chaque plat/entree/dessert/sauce + total par alcool.", "boisson", b3_cuisine),
    ("DB4", "Boissons-4-menus-par-periode.xlsx", "Menus par periode", "Composition alcool des menus par periode et probabilite de choix.", "boisson", b4_menus),
    ("DB5", "Boissons-5-conso-par-periode.xlsx", "Consommation par exercice", "Caisse + cuisine + menus + personnel par exercice : litres, cout, CA.", "boisson", b5_conso_periode),
    # methode / preuve de process
    ("DM1", "Methode-1-modele-Monte-Carlo.xlsx", "Modele d'incertitude (Monte Carlo)", "Parametres, resultats, sensibilite et signature du disparu.", "methode", m1_montecarlo),
    ("DM2", "Methode-2-correction-mapping-Cubis.xlsx", "Correction d'attribution (Cubis)", "Bug d'attribution du vin generique corrige, gain chiffre.", "methode", m2_cubis),
    ("DM3", "Methode-3-bilan-softs.xlsx", "Bilan des softs", "Softs reconcilies (ventes ticket + staff + stock) : disparu ~ 0.", "methode", m3_softs),
    ("DM4", "Methode-4-conso-personnel-detail.xlsx", "Conso personnel detaillee", "Detail borne et justifie de la consommation du personnel/owner.", "methode", m4_staff),
    ("DM5", "Methode-5-reconciliation-suppressions.xlsx", "Reconciliation des suppressions", "CA=encaissements, triangulation, sessions impossibles, disparu vs DEL.", "methode", m5_suppr),
    ("DM6", "Methode-6-erreurs-quantite-DEL.xlsx", "Erreurs de quantite (DEL)", "Decomposition des grosses suppressions + distribution des prix.", "methode", m6_quantite),
]

manifest = []
for pid, fichier, titre, desc, cat, fn in PIECES:
    wb = openpyxl.Workbook()
    fn(wb)
    wb.save(os.path.join(OUT, fichier))
    manifest.append({"id": pid, "fichier": f"{DOSSIER}/{fichier}", "titre": titre,
                     "description": desc, "categorie": cat})
    print("ecrit :", fichier)

# pieces deja existantes (scripts 13/14/15) a inclure dans le manifeste
EXISTANTES = [
    {"id": "D1", "fichier": f"{DOSSIER}/Defense-Suppressions-de-caisse-DEL.xlsx", "titre": "Defense - Suppressions de caisse (synthese)", "description": "Synthese complete des suppressions DEL (6 onglets).", "categorie": "methode"},
    {"id": "D2", "fichier": f"{DOSSIER}/Defense-Boissons-disparues.xlsx", "titre": "Defense - Boissons disparues (synthese)", "description": "Recalcul Monte Carlo + reconciliation (4 onglets).", "categorie": "boisson"},
    {"id": "D3", "fichier": f"{DOSSIER}/Consommation-personnel-et-offerts.xlsx", "titre": "Conso personnel & offerts + perte reelle", "description": "Conso personnel/offerts + cascade de la perte reelle (4 onglets).", "categorie": "boisson"},
    {"id": "D4", "fichier": f"{DOSSIER}/Boissons-detail-complet.xlsx", "titre": "Boissons - detail complet (toutes les tables)", "description": "Toutes les tables de la page boissons (6 onglets).", "categorie": "boisson"},
]
full = EXISTANTES + manifest
json.dump({"description": "Manifeste des pieces de defense XLSX (public/documents/pieces-defense).",
           "pieces": full},
          open(os.path.join(SRCDATA, "piecesDefenseManifest.json"), "w", encoding="utf-8"),
          ensure_ascii=False, indent=1)
print(f"\nmanifeste : {len(full)} pieces -> src/data/piecesDefenseManifest.json")
