#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""Rapprochement conso (ventes) vs inventaire vs achats, par exercice fiscal.
Tout calculé en VOLUME (cl) puis converti en CONTENANTS (unité d'achat du produit).
disparu = stock_ouverture + achats - conso - stock_cloture.
stock_ouverture = clôture de l'exercice précédent ; 2022-2023 : pas d'ouverture (31/03/2022 indisponible) -> ouverture=0 (signalé).
Cas FÛT : la quantité facturée est en LITRES (pas en nombre de fûts) -> géré.
Enrichit consoTotaleParBoisson.json (inventaire + achats par période) et crée rapprochement-disparu.xlsx.
"""
import json,collections,unicodedata,re
from openpyxl import Workbook
from openpyxl.styles import Font,PatternFill,Alignment,Border,Side
from openpyxl.utils import get_column_letter as Lc
BASE="/Users/maevapasteur/Documents/demi-lune-comptabilite/"; JS=BASE+"src/data/calculsBoissons/"
EXOS=["2022-2023","2023-2024","2024-2025"]; PREV={"2023-2024":"2022-2023","2024-2025":"2023-2024"}
INVDATE={"2023-03-31":"2022-2023","2024-03-31":"2023-2024","2025-03-31":"2024-2025"}
def na(s): return ''.join(c for c in unicodedata.normalize('NFD',str(s)) if unicodedata.category(c)!='Mn').lower()
def size_cl(name):
    if not name: return None
    s=na(name).replace(",","."); best=None
    for m in re.finditer(r"(\d+(?:\.\d+)?)\s*(cl|l)\b",s):
        v=float(m.group(1)); best=v*100 if m.group(2)=="l" else v
    return best
def unite_label(t):
    if t is None: return "n/d"
    if t==1000: return "BIB 10L"
    if t==800: return "fût 8L"
    return f"bouteille {t:g}cl"
STOP=set("75cl 70cl 100cl 50cl 33cl 62cl 37 5cl 10l 8l cl l de du des la le maison bib verre bouteille pichet rouge blanc rose rosé n vp c10 ° et aux".split())
def toks(s): return [t for t in re.split(r"[^a-z0-9]+",na(s)) if t and t not in STOP and len(t)>2 and not t.isdigit()]
def anchors(name): return set(t for t in toks(name) if t not in("cotes","cote","arbois","jura","vin","sirop"))

conso=json.load(open(JS+"consoTotaleParBoisson.json",encoding="utf-8"))
inv=json.load(open(BASE+"public/documents/inventaires/inventaires.json",encoding="utf-8"))
ach=json.load(open(JS+"achatsBoissonsParPeriode.json",encoding="utf-8"))["achats"]

# inventaire (unités) par produit/date
invq=collections.defaultdict(lambda: collections.defaultdict(float))
for i in inv["inventaires"]:
    for l in i["lignes"]:
        if l["categorie"]=="alcool": invq[l["produit"]][i["date"]]+=(l.get("quantite") or 0)
def inv_units(nom):
    if not nom: return None
    base=na(nom); res={e:0.0 for e in EXOS}; hit=False
    for prod,perd in invq.items():
        if na(prod)==base or na(prod).startswith(base):
            hit=True
            for dt,q in perd.items(): res[INVDATE[dt]]+=q
    return res if hit else None

# achats -> volume cl par produit (FÛT = litres), puis match canonique
canon_anchor={b["nom_canonique"]:anchors(b["nom_canonique"]) for b in conso["boissons"]}
ach_vol=collections.defaultdict(lambda:{e:0.0 for e in EXOS}); nonrat=[]
for a in ach:
    isfut="fut" in na(a["produit"]) or "fût" in na(a["produit"])
    sz=100 if isfut else size_cl(a["produit"])   # fût: q en litres -> cl ; sinon q en contenants * taille
    at=set(toks(a["produit"])); best=None;bn=0
    for c,anc in canon_anchor.items():
        ov=len(anc&at)
        if ov>bn: best=c;bn=ov
    if best and bn>=1 and sz:
        for e in EXOS: ach_vol[best][e]+=a["par_periode"][e]["quantite"]*sz
    elif a["categorie"] in("vin/cidre","bière","spiritueux/liqueur") and a["total_quantite"]>0:
        nonrat.append({"produit":a["produit"],"categorie":a["categorie"],"par_periode":{e:a["par_periode"][e]["quantite"] for e in EXOS},"total":a["total_quantite"]})

# enrichir le JSON
for b in conso["boissons"]:
    t=size_cl(b.get("nom_inventaire") or b.get("nom_facture")); b["taille_achat_cl"]=t; b["unite_achat"]=unite_label(t)
    iu=inv_units(b.get("nom_inventaire")); iu=iu or {e:0.0 for e in EXOS}
    b["inventaire_fin_contenants_par_periode"]={e:round(iu[e],1) for e in EXOS}
    av=ach_vol.get(b["nom_canonique"],{e:0.0 for e in EXOS})
    b["achats_litres_par_periode"]={e:round(av[e]/100,1) for e in EXOS}
    b["achats_contenants_par_periode"]={e:round(av[e]/t,1) if t else None for e in EXOS}
conso["note_enrichissement"]="inventaire_fin_contenants = unités en stock à la clôture (31/03) ; achats en litres et en contenants (unité d'achat) par exercice ; taille_achat_cl = contenance d'un contenant. Bilan 'disparu' détaillé dans rapprochement-disparu.xlsx."
conso["achats_non_rattaches"]=sorted(nonrat,key=lambda z:-z["total"])
json.dump(conso,open(JS+"consoTotaleParBoisson.json","w",encoding="utf-8"),ensure_ascii=False,indent=2)
print("JSON enrichi.")

# ---- prix d'achat (factures) et prix de revente (cartes) ----
import xlrd
VB=BASE+"public/documents/vins-boissons/"
def match_canon(label):
    at=set(toks(label)); best=None;bn=0
    for c,anc in canon_anchor.items():
        ov=len(anc&at)
        if ov>bn: best=c;bn=ov
    return best if bn>=1 else None
import importlib.util as _i
_sm=_i.spec_from_file_location("mp",VB+"_mapping_noms.py"); mp=_i.module_from_spec(_sm); _sm.loader.exec_module(mp)
revente=mp.CARTE_PRIX   # canonique -> (serving_cl, serving_prix), mapping explicite
taille_of={b["nom_canonique"]:b["taille_achat_cl"] for b in conso["boissons"]}
cost=collections.defaultdict(lambda:[0.0,0.0])
for a in ach:
    c=match_canon(a["produit"]); t=taille_of.get(c)
    if not c or not t: continue
    isf="fut" in na(a["produit"]); sz=100 if isf else size_cl(a["produit"])
    if not sz: continue
    cost[c][0]+=a.get("total_montant_ht",0); cost[c][1]+=a["total_quantite"]*sz/t
PRIX_ACHAT={c:(v[0]/v[1] if v[1]>0 else None) for c,v in cost.items()}
PRIX_REVENTE={b["nom_canonique"]:((revente[b["nom_canonique"]][1]*b["taille_achat_cl"]/revente[b["nom_canonique"]][0]) if (b["nom_canonique"] in revente and b["taille_achat_cl"]) else None) for b in conso["boissons"]}

# ---- xlsx ----
HEAD=PatternFill("solid",fgColor="1F4E78");HF=Font(bold=True,color="FFFFFF",size=9)
CONS=PatternFill("solid",fgColor="DDEBF7");INVF=PatternFill("solid",fgColor="E2EFDA");ACHF=PatternFill("solid",fgColor="FFF2CC");DISF=PatternFill("solid",fgColor="FCE4D6");VALF=PatternFill("solid",fgColor="F4B183");PXF=PatternFill("solid",fgColor="E2EFDA")
thin=Side("thin",color="BFBFBF");B=Border(thin,thin,thin,thin);ctr=Alignment("center",vertical="center");lft=Alignment("left",vertical="center")
wb=Workbook();ws=wb.active;ws.title="Rapprochement disparu"
SUB=["Conso bas (L)","Conso moyen (L)","Conso haut (L)","Inventaire fin (L)","Achats (L)","Disparu (L)","% disparu","Valeur disparu achat €","Valeur disparu revente €"]
FILLS=[CONS,CONS,CONS,INVF,ACHF,DISF,DISF,VALF,VALF]
INFO=["Boisson","Catégorie","Unité d'achat","Nom inventaire","Prix achat €/L","Prix revente €/L"]
ncol=len(INFO)+len(SUB)*3
ws.merge_cells(start_row=1,start_column=1,end_row=1,end_column=ncol)
ws["A1"]="DEMI LUNE - Rapprochement consommation / inventaire / achats : disparu par exercice (en LITRES)";ws["A1"].font=Font(bold=True,size=13)
ws.merge_cells(start_row=2,start_column=1,end_row=2,end_column=ncol)
ws["A2"]=("Tout en LITRES. Disparu = stock d'ouverture + achats − conso − stock de clôture. "
 "Stock d'ouverture = clôture de l'exercice précédent. 2022-2023 : sans stock d'ouverture (31/03/2022 indisponible) → ouverture=0 (le disparu peut être surévalué si stock initial existait). "
 "% disparu = disparu / (stock d'ouverture + achats), conso MOYENNE. Prix en €/L. Pour convertir en bouteilles : ÷ (taille de l'unité d'achat en L, ex. 0,75).")
ws["A2"].font=Font(italic=True,size=9);ws["A2"].alignment=Alignment(wrap_text=True,vertical="center");ws.row_dimensions[2].height=60
HR=4
for i,t in enumerate(INFO,1):
    ws.cell(HR,i,t);ws.merge_cells(start_row=HR,start_column=i,end_row=HR+1,end_column=i)
col=len(INFO)+1
for e in EXOS:
    ws.merge_cells(start_row=HR,start_column=col,end_row=HR,end_column=col+len(SUB)-1);ws.cell(HR,col,e)
    for k,s in enumerate(SUB): ws.cell(HR+1,col+k,s)
    col+=len(SUB)
for r in (HR,HR+1):
    for c in range(1,ncol+1):
        cc=ws.cell(r,c);cc.fill=HEAD;cc.font=HF;cc.border=B;cc.alignment=Alignment("center",wrap_text=True,vertical="center")
r=HR+2
for b in sorted(conso["boissons"],key=lambda z:-z["total_3_exercices"]["total_l"]["moyen"]):
    t=b["taille_achat_cl"]; pac=PRIX_ACHAT.get(b["nom_canonique"]); prc=PRIX_REVENTE.get(b["nom_canonique"])
    # prix par LITRE (les prix sont calculés par contenant)
    pa=(pac/(t/100)) if (pac is not None and t) else None
    pr=(prc/(t/100)) if (prc is not None and t) else None
    row=[b["nom_canonique"],b["categorie"],b["unite_achat"],b.get("nom_inventaire") or "—",
         round(pa,2) if pa is not None else "n/d",round(pr,2) if pr is not None else "n/d"]
    for e in EXOS:
        cb=b["par_periode"][e]["total_l"]["bas"]; cm=b["par_periode"][e]["total_l"]["moyen"]; ch=b["par_periode"][e]["total_l"]["haut"]
        invL=(b["inventaire_fin_contenants_par_periode"][e]*t/100) if t else None
        acL=b["achats_litres_par_periode"][e]
        if t is None:
            disp=pct=va=vr=None
        else:
            opnL=0.0 if e=="2022-2023" else b["inventaire_fin_contenants_par_periode"][PREV[e]]*t/100
            disp=opnL+acL-cm-invL; den=opnL+acL; pct=(disp/den*100) if den>0 else None
            va=disp*pa if pa is not None else None; vr=disp*pr if pr is not None else None
        vals=[cb,cm,ch,invL,acL,disp,pct,va,vr]
        row+=[round(v,2 if i>=7 else 1) if isinstance(v,(int,float)) else "n/d" for i,v in enumerate(vals)]
    for j,v in enumerate(row,1):
        cc=ws.cell(r,j,v);cc.border=B;cc.alignment=lft if j in(1,2,4) else ctr
        if j in(5,6): cc.fill=PXF
        if j>len(INFO):
            k=(j-len(INFO)-1)%len(SUB); cc.fill=FILLS[k]
            if isinstance(v,(int,float)): cc.number_format="0.00" if k>=7 else "0.0"
        elif j in(5,6) and isinstance(v,(int,float)): cc.number_format="0.00"
    r+=1
ws.freeze_panes=f"{Lc(len(INFO)+1)}{HR+2}";ws.auto_filter.ref=f"A{HR}:{Lc(ncol)}{r-1}"
ws.column_dimensions["A"].width=30;ws.column_dimensions["B"].width=16;ws.column_dimensions["C"].width=14;ws.column_dimensions["D"].width=28
for j in range(len(INFO)+1,ncol+1): ws.column_dimensions[Lc(j)].width=11
wb.save(JS+"rapprochement-disparu.xlsx");print("-> rapprochement-disparu.xlsx")
