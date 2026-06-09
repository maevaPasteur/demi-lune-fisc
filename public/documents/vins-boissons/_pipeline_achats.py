#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""Achats de boissons par exercice fiscal depuis factures-fournisseur.json.
Sortie : src/data/calculsBoissons/achatsBoissonsParPeriode.json + .xlsx
Lancer : /tmp/xlsenv/bin/python _pipeline_achats.py"""
import json,collections,re
from openpyxl import Workbook
from openpyxl.styles import Font,PatternFill,Alignment,Border,Side
from openpyxl.utils import get_column_letter as Lc
BASE="/Users/maevapasteur/Documents/demi-lune-comptabilite/"; JS=BASE+"src/data/calculsBoissons/"
EXOS=["2022-2023","2023-2024","2024-2025"]
F=json.load(open(BASE+"src/data/factures-fournisseur.json",encoding="utf-8"))["factures"]
def period(df):
    j,m,a=df.split("/"); iso=f"{a}-{m}-{j}"
    if iso<"2022-04-01": return None
    if iso<="2023-03-31": return "2022-2023"
    if iso<="2024-03-31": return "2023-2024"
    if iso<="2025-03-31": return "2024-2025"
    return None
def clean(s):
    s=re.sub(r"\s*(Ci-dessous|Articles Manquants|Livré le|Marchandises factur).*$","",s,flags=re.I)
    s=re.sub(r"(\s+C10\b)+.*$","",s); return re.sub(r"\s+"," ",s).strip()
EXCL=["napolitain","milka","cracker","bretzel","chips","cacahu","olive","biscuit","bonbon","carambar","party troopers","croco","galette","galet ","madeleine","funky mix","verres ","verre perdu","bac recyclage","caisse recup","emballage","pailles ","kit ","touille","mini caramel","sucre","detartrant","nivona",
 "stylo","bloc","livre de","cahier","tasse","biscoff","cookie","lotus","bonne maman","mokador","verona","reservation","cappuccino","expresso *","crayon","agenda","ardoise","menu ","carte ","ticket","sommelier","divers","set de table","coutale","limonadier","decapsuleur","tire bouchon","claris","cartouche","filtre"]
def is_bev(low): return not any(k in low for k in EXCL)
BEER=["affligem","fut ","fût","biere","bière","picon biere","rouget","monaco","panach","1664","heineken","mort subite","sziget","hefeweizen","white rabbit","witte","desperados","kronenbourg","carlsberg","leffe","grimbergen","pelforth","corona","stout"," ipa","blonde","brune","mont blanc","weizen","pils","triple"]
SOFT=["coca","schweppes","orangina","fanta","perrier","vittel","san pell","limonade","mortuacienne","jus ","granini","monin","sirop","pulco","routin","fuze","tea","tonic","nectar","norbert","carola","rieme","volvic","tropico","sprite","badoit","evian","contrex","velleminfroy","juicy","oasis","ice tea","gini","minute maid","ginger"]
WINE=["vin ","rouge"," rge","blanc"," blc","rosé","rose ","cremant","crémant","champagne","cidre","macvin","savagnin","chardonnay","aligote","gewur","chablis","beaune","nuits","trousseau","joseph","moulin a vent","veran","macon","pavois","bordeaux","miraval","minuty","jamelles","pive","boheme","paille","jaune","ravelin","gascogne","sauternes","cote rouge","cotes du rhone","syrah","pinot","auxey","duresses","montagny","1er cru","alsace","ruinart"]
HOT=["cafe","café","malongo","the ","thé","infusion","earl grey","peppermint","jasmin","verveine","camomille","tilleul","segafredo","baronny"]
def cat(low,accise):
    if any(k in low for k in BEER): return "bière"
    if any(k in low for k in SOFT): return "sans_alcool"     # sirops Monin, sodas, jus (avant vin pour éviter 'rose')
    if any(k in low for k in HOT): return "boisson_chaude"   # thés (avant vin pour éviter 'rouge')
    if any(k in low for k in WINE): return "vin/cidre"
    if accise: return "spiritueux/liqueur"
    return "autre"
agg=collections.defaultdict(lambda: dict(accise=False,cond="",code="",pe={e:dict(q=0.0,m=0.0) for e in EXOS}))
for f in F:
    p=period(f["dateFacture"])
    if not p: continue
    for ln in f.get("lignes",[]):
        if ln.get("montantHT") is None: continue   # ligne jamais facturée = non livré / avoir noté / annotation (vérifié : aucun puNet) -> pas un achat
        des=clean(ln.get("designation","")); low=des.lower()
        if not des or not is_bev(low): continue
        A=agg[des]
        if ln.get("accise"): A["accise"]=True
        A["cond"]=A["cond"] or (ln.get("conditionnement") or ""); A["code"]=A["code"] or (ln.get("code") or "")
        A["pe"][p]["q"]+=ln.get("quantite") or 0; A["pe"][p]["m"]+=ln.get("montantHT") or 0
achats=[]
for des,A in agg.items():
    tq=sum(A["pe"][e]["q"] for e in EXOS); tm=sum(A["pe"][e]["m"] for e in EXOS)
    if tq==0 and tm==0: continue
    achats.append({"produit":des,"categorie":cat(des.lower(),A["accise"]),"accise":A["accise"],
        "conditionnement":A["cond"],"code":A["code"],"total_quantite":round(tq,2),"total_montant_ht":round(tm,2),
        "par_periode":{e:{"quantite":round(A["pe"][e]["q"],2),"montant_ht":round(A["pe"][e]["m"],2)} for e in EXOS}})
achats.sort(key=lambda z:-z["total_montant_ht"])
json.dump({"description":"Achats de boissons (fournisseur Franche-Comté Boissons Services) par exercice fiscal. Affectation par date de facture à l'exercice (01/04→31/03). Exclus : snacks, verrerie, matériel, kits. Quantité = unités achetées (bouteilles/BIB/fûts selon conditionnement) ; montant HT en €.","source":"factures-fournisseur.json","exercices":EXOS,"nb_produits":len(achats),"achats":achats},open(JS+"achatsBoissonsParPeriode.json","w",encoding="utf-8"),ensure_ascii=False,indent=2)
print("-> achatsBoissonsParPeriode.json |",len(achats),"produits")
for e in EXOS: print(f"   {e}: {sum(a['par_periode'][e]['montant_ht'] for a in achats):.0f} € HT")
import collections as C; print("catégories:",dict(C.Counter(a['categorie'] for a in achats)))

# ---- xlsx ----
HEAD=PatternFill("solid",fgColor="1F4E78");HF=Font(bold=True,color="FFFFFF",size=9);GT=PatternFill("solid",fgColor="F4B183")
thin=Side("thin",color="BFBFBF");B=Border(thin,thin,thin,thin);ctr=Alignment("center",vertical="center");lft=Alignment("left",vertical="center");bold=Font(bold=True)
wb=Workbook();ws=wb.active;ws.title="Achats boissons par période";ncol=4+2*3+2
ws.merge_cells(start_row=1,start_column=1,end_row=1,end_column=ncol);ws["A1"]="DEMI LUNE - Achats de boissons par exercice fiscal (Franche-Comté Boissons)";ws["A1"].font=Font(bold=True,size=13)
ws.merge_cells(start_row=2,start_column=1,end_row=2,end_column=ncol);ws["A2"]="Source factures-fournisseur.json. Affectation par date de facture à l'exercice 01/04→31/03. Qté = unités achetées ; montant HT €. Hors snacks/verrerie/matériel.";ws["A2"].font=Font(italic=True,size=9);ws["A2"].alignment=Alignment(wrap_text=True,vertical="center");ws.row_dimensions[2].height=28
HR=4
for i,t in enumerate(["Produit (facture)","Catégorie","Accise","Conditionnement"],1):
    ws.cell(HR,i,t); ws.merge_cells(start_row=HR,start_column=i,end_row=HR+1,end_column=i)
col=5
for e in EXOS+["TOTAL"]:
    ws.merge_cells(start_row=HR,start_column=col,end_row=HR,end_column=col+1);ws.cell(HR,col,e)
    ws.cell(HR+1,col,"Qté");ws.cell(HR+1,col+1,"Montant HT €");col+=2
for r in (HR,HR+1):
    for c in range(1,ncol+1):
        cc=ws.cell(r,c);cc.fill=HEAD;cc.font=HF;cc.border=B;cc.alignment=Alignment("center",wrap_text=True,vertical="center")
r=HR+2
for a in achats:
    row=[a["produit"],a["categorie"],"oui" if a["accise"] else "",a["conditionnement"]]
    for e in EXOS: row+=[a["par_periode"][e]["quantite"],a["par_periode"][e]["montant_ht"]]
    row+=[a["total_quantite"],a["total_montant_ht"]]
    for j,v in enumerate(row,1):
        cc=ws.cell(r,j,v);cc.border=B;cc.alignment=lft if j in(1,2,4) else ctr
        if j>=5: cc.number_format="0.##" if (j-5)%2==0 else "0.00"
    r+=1
ws.cell(r,1,"TOTAL").font=bold
for j in range(1,ncol+1):
    cc=ws.cell(r,j);cc.fill=GT;cc.border=B;cc.font=bold;cc.alignment=ctr
    if j>=5: cc.value=f"=SUM({Lc(j)}{HR+2}:{Lc(j)}{r-1})";cc.number_format="0.##" if (j-5)%2==0 else "0.00"
ws.cell(r,1).value="TOTAL";ws.cell(r,1).alignment=lft
ws.freeze_panes="E6";ws.auto_filter.ref=f"A{HR}:{Lc(ncol)}{r-1}"
for i,wd in enumerate([46,17,7,22,8,12,8,12,8,12,9,13],1): ws.column_dimensions[Lc(i)].width=wd
wb.save(JS+"achatsBoissonsParPeriode.xlsx");print("-> achatsBoissonsParPeriode.xlsx")
