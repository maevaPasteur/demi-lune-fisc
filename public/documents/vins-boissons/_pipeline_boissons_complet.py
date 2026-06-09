#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Pipeline complet boissons Demi Lune.
Produit (1 xlsx dans public/documents/vins-boissons + JSON dans src/data/calculsBoissons) :
  items-caisse-par-periode.xlsx / itemsCaisse.json   : tous les items caisse, 3 exercices
  cocktailsComposition.json    : recettes exactes des cocktails (cl/ingrédient)
  saucesAlcool.json            : alcool des sauces
  platsAlcool.json             : alcool des plats/entrées/desserts
  menusPrix.json               : prix des menus par carte (C/B/A)
  menusCompositionAlcool.json  : alcool des menus par carte + proba bas/moyen/haut
  menusVentesParPrix.json      : menus vendus, 5 périodes, par prix + prix_modifie
  menusNonModifiesAlcool.json  : menus non modifiés, 5 périodes, options alcool + estim
  consoAlcoolMenus.json        : conso alcool menus, 3 exercices, bas/moyen/haut
  cocktailsConsoComposition.json : ingrédients cocktails consommés, 3 exercices
  boissonsHorsCocktail.json    : boissons hors cocktail, 3 exercices (produit+format)
  consoTotaleParBoisson.json   : cumul final, 3 exercices
"""
import xlrd, json, collections, importlib.util
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter as Lc

BASE="/Users/maevapasteur/Documents/demi-lune-comptabilite/"
CAISSE=BASE+"public/documents/caisse-enregistreuse/"
XL=BASE+"public/documents/vins-boissons/"
JS=BASE+"src/data/calculsBoissons/"
EXOS=["2022-2023","2023-2024","2024-2025"]
D_FILES={"2022-2023":CAISSE+"ANNEXE-D1_synthese-produit_2022-2023.xls","2023-2024":CAISSE+"ANNEXE-D2_synthese-produit_2023-2024.xls","2024-2025":CAISSE+"ANNEXE-D3_synthese-produit_2024-2025.xls"}
C_FILES=[CAISSE+"ANNEXE-C1_detail-tickets_2022-2023.xls",CAISSE+"ANNEXE-C2_detail-tickets_2023-2024.xls",CAISSE+"ANNEXE-C3_detail-tickets_2024-2025.xls"]
def norm(s): return str(s).replace("\xa0"," ").strip()
def w(name,obj):
    json.dump(obj,open(JS+name,"w",encoding="utf-8"),ensure_ascii=False,indent=2)
    print("  ->",name)

# réutilise le mapping alcool existant (FOOD, BEV, build_carte, classify_wine)
spec=importlib.util.spec_from_file_location("pa",XL+"_pipeline_conso_alcool.py"); pa=importlib.util.module_from_spec(spec); spec.loader.exec_module(pa)
_sm=importlib.util.spec_from_file_location("mp",XL+"_mapping_noms.py"); mp=importlib.util.module_from_spec(_sm); _sm.loader.exec_module(mp)

# ============================ RECETTES COCKTAILS ============================
# categorie: alcool / sirop / jus / soda / biere ; estimation: 1 = dose déduite
COCKTAILS={
 "La Vouivre":{"volume_cl":12,"ingredients":[{"nom":"Crémant","categorie":"alcool","cl":8,"estimation":1},{"nom":"Macvin","categorie":"alcool","cl":3,"estimation":1},{"nom":"Crème de cassis","categorie":"alcool","cl":1,"estimation":1}]},
 "Apéritif du Père Grégoire":{"volume_cl":6,"ingredients":[{"nom":"Crémant","categorie":"alcool","cl":1,"estimation":1},{"nom":"Macvin","categorie":"alcool","cl":4,"estimation":1},{"nom":"Liqueur de cerise","categorie":"alcool","cl":1,"estimation":1}]},
 "Chat Perché":{"volume_cl":12,"ingredients":[{"nom":"Macvin","categorie":"alcool","cl":6,"estimation":1},{"nom":"Jus de poire","categorie":"jus","cl":3,"estimation":1},{"nom":"Sirop de châtaigne","categorie":"sirop","cl":3,"estimation":1}]},
 "Kittykir":{"volume_cl":12,"ingredients":[{"nom":"Soho (liqueur de litchi)","categorie":"alcool","cl":2,"estimation":1},{"nom":"Crémant","categorie":"alcool","cl":9,"estimation":1},{"nom":"Sirop de framboise","categorie":"sirop","cl":1,"estimation":1}]},
 "Mambo":{"volume_cl":25,"ingredients":[{"nom":"Jus d'orange","categorie":"jus","cl":8.3,"estimation":1},{"nom":"Jus de fraise","categorie":"jus","cl":8.3,"estimation":1},{"nom":"Limonade","categorie":"soda","cl":8.4,"estimation":1}]},
 "Luna":{"volume_cl":25,"ingredients":[{"nom":"Jus de poire","categorie":"jus","cl":8.3,"estimation":1},{"nom":"Limonade","categorie":"soda","cl":8.3,"estimation":1},{"nom":"Sirop de châtaigne","categorie":"sirop","cl":8.4,"estimation":1}]},
 "Balidou":{"volume_cl":25,"ingredients":[{"nom":"Passoa","categorie":"alcool","cl":4,"estimation":0},{"nom":"Jus d'ananas","categorie":"jus","cl":10.5,"estimation":1},{"nom":"Sirop de framboise","categorie":"sirop","cl":10.5,"estimation":1}]},
 "Rêve Bleu":{"volume_cl":25,"ingredients":[{"nom":"Vodka","categorie":"alcool","cl":4,"estimation":0},{"nom":"Jus de pomme","categorie":"jus","cl":10.5,"estimation":1},{"nom":"Sirop curaçao bleu","categorie":"sirop","cl":10.5,"estimation":1}]},
 "Maëva":{"volume_cl":25,"ingredients":[{"nom":"Soho (liqueur de litchi)","categorie":"alcool","cl":4,"estimation":0},{"nom":"Vodka","categorie":"alcool","cl":4,"estimation":0},{"nom":"Jus de framboise","categorie":"jus","cl":17,"estimation":1}]},
 "Rabasse":{"volume_cl":25,"ingredients":[{"nom":"Pontarlier","categorie":"alcool","cl":2,"estimation":0},{"nom":"Liqueur de sapin","categorie":"alcool","cl":2,"estimation":0},{"nom":"Limonade","categorie":"soda","cl":21,"estimation":1}]},
 "Tequila Sunrise":{"volume_cl":25,"ingredients":[{"nom":"Tequila","categorie":"alcool","cl":4,"estimation":0},{"nom":"Jus d'orange","categorie":"jus","cl":10.5,"estimation":1},{"nom":"Sirop de grenadine","categorie":"sirop","cl":10.5,"estimation":1}]},
 "Kir":{"volume_cl":12,"ingredients":[{"nom":"Aligoté","categorie":"alcool","cl":10,"estimation":1},{"nom":"Crème de cassis","categorie":"alcool","cl":2,"estimation":1}]},
 "Kir Princier":{"volume_cl":12,"ingredients":[{"nom":"Crémant","categorie":"alcool","cl":10,"estimation":1},{"nom":"Crème de cassis","categorie":"alcool","cl":2,"estimation":1}]},
 "Kir Pamplemousse":{"volume_cl":12,"ingredients":[{"nom":"Aligoté","categorie":"alcool","cl":10,"estimation":1},{"nom":"Sirop de pamplemousse","categorie":"sirop","cl":2,"estimation":1}]},
 "Rosé Pamp":{"volume_cl":12,"ingredients":[{"nom":"Côtes de Provence rosé","categorie":"alcool","cl":10,"estimation":1},{"nom":"Sirop de pamplemousse","categorie":"sirop","cl":2,"estimation":1}]},
 "Whisky-Coca":{"volume_cl":25,"ingredients":[{"nom":"Whisky","categorie":"alcool","cl":4,"estimation":0},{"nom":"Coca","categorie":"soda","cl":21,"estimation":1}]},
 "Panaché":{"volume_cl":25,"ingredients":[{"nom":"Bière (fût)","categorie":"biere","cl":12.5,"estimation":0},{"nom":"Limonade","categorie":"soda","cl":12.5,"estimation":0}]},
 "Picon bière 25cl":{"volume_cl":25,"ingredients":[{"nom":"Picon","categorie":"alcool","cl":4,"estimation":0},{"nom":"Bière (fût)","categorie":"biere","cl":21,"estimation":0}]},
 "Picon bière 50cl":{"volume_cl":50,"ingredients":[{"nom":"Picon","categorie":"alcool","cl":8,"estimation":0},{"nom":"Bière (fût)","categorie":"biere","cl":42,"estimation":0}]},
 "Monaco":{"volume_cl":25,"ingredients":[{"nom":"Bière (fût)","categorie":"biere","cl":12.5,"estimation":1},{"nom":"Sirop de grenadine","categorie":"sirop","cl":2,"estimation":1},{"nom":"Limonade","categorie":"soda","cl":10.5,"estimation":1}]},
}
CAISSE_COCKTAIL={"La Vouivre":"La Vouivre","Père Gregoire":"Apéritif du Père Grégoire","Chat perché":"Chat Perché","KITTYKIR":"Kittykir","Mambo":"Mambo","Luna":"Luna","BALIDOU":"Balidou","Rêve Bleu":"Rêve Bleu","MAEVA":"Maëva","Rabasse":"Rabasse","TEQUILA SUNRISE":"Tequila Sunrise","Kir Bourgogne":"Kir","Kir Princier":"Kir Princier","Kir Pamplemousse":"Kir Pamplemousse","Rosé Pamplemousse":"Rosé Pamp","Whisky-Coca":"Whisky-Coca","Panaché":"Panaché","Picon bière":"Picon bière 25cl","Pinte Picon":"Picon bière 50cl","Monaco":"Monaco"}
COCKTAIL_LABELS=set(CAISSE_COCKTAIL)

# ============================ MENUS ============================
MENU_PRICES={  # par carte
 "C":{"dates":"29/03/2021-22/07/2022","Menu":11,"Express":15,"Galette":15.9,"Vegetarien":23,"Bambin":6.9,"Enfant":9.9,"Bourguignon":32,"Franc-Comtois":32,"Demi Lune":45,"Dahu":29.9},
 "B":{"dates":"23/07/2022-16/04/2023","Menu":12.5,"Express":16.5,"Galette":17,"Vegetarien":25,"Bambin":6.9,"Enfant":9.9,"Bourguignon":34,"Franc-Comtois":34,"Demi Lune":45,"Dahu":29.9},
 "A":{"dates":"17/04/2023-17/07/2025","Menu":12.5,"Express":18.6,"Galette":19.5,"Vegetarien":26.5,"Bambin":6.9,"Enfant":9.9,"Bourguignon":34,"Franc-Comtois":34,"Demi Lune":45,"Dahu":29.9},
}
ACCEPT={m:set() for m in ["Menu","Express","Galette","Vegetarien","Bambin","Enfant","Bourguignon","Franc-Comtois","Demi Lune","Dahu"]}
for c in MENU_PRICES.values():
    for m in ACCEPT: ACCEPT[m].add(round(float(c[m]),2))
MENU_LABEL={"Menu 'Demi Lune'":"Demi Lune","Menu du Dahu":"Dahu","Menu Végétarien":"Vegetarien","Menu Bourguignon":"Bourguignon","Menu Franc-Comtois":"Franc-Comtois","Menu 'Galette'":"Galette","Menu 'Enfant'":"Enfant","Menu Bambin":"Bambin","Formule à 11":"Menu","Formule à 12.5":"Menu","Formule Express":"Express","FORMULE EXPRESS":"Express"}
COMPOSED={"Demi Lune","Dahu","Vegetarien","Bourguignon","Franc-Comtois"}
# composition riche par carte : (service,option,nb_options,obligatoire,alcool,dose,p_bas,p_haut)
MENU_COMP={
"A":{"Demi Lune":[("Plat","Poulet ou truite aux morilles",2,1,"Porto",1,1.0,1.0),("Entrée","Feuilleté forestier",3,0,"Porto",1,.33,.60),("Dessert","Coupe glacée Ivresse",4,0,"Baileys",4,.15,.35),("Dessert","Baba liqueur de poire",4,0,"Liqueur de poire",4,.15,.35)],
     "Dahu":[("Plat","Fondue jurassienne",1,1,"Ravelin",10,1.0,1.0)],
     "Vegetarien":[("Plat","Assiette du père Grégoire",2,0,"Calvados",4,.35,.65),("Plat","Galette sarrasin forestière",2,0,"Porto",1,.35,.65)],
     "Bourguignon":[("Dessert","Coupe Dijonnaise",2,0,"Crème de cassis",4,.35,.65),("Dessert","Baba au Baileys",2,0,"Baileys",4,.35,.65)],
     "Franc-Comtois":[("Plat","Poulet à la jurassienne",2,0,"Vin jaune",1,.35,.65),("Plat","Assiette franc-comtoise",2,0,"Macvin",4,.35,.65)]},
"B":{"Demi Lune":[("Plat","Poulet ou truite aux morilles",2,1,"Porto",1,1.0,1.0),("Entrée","Feuilleté forestier",3,0,"Porto",1,.33,.60)],
     "Dahu":[("Plat","Fondue jurassienne",1,1,"Ravelin",10,1.0,1.0)],
     "Vegetarien":[("Plat","Assiette du père Grégoire",2,0,"Calvados",4,.35,.65),("Plat","Galette sarrasin forestière",2,0,"Porto",1,.35,.65)],
     "Bourguignon":[("Dessert","Coupe Dijonnaise",2,0,"Crème de cassis",4,.35,.65),("Dessert","Baba au marc de bourgogne",2,0,"Marc de bourgogne",4,.35,.65)],
     "Franc-Comtois":[("Plat","Poulet à la jurassienne",2,0,"Vin jaune",1,.35,.65),("Plat","Assiette franc-comtoise",2,0,"Macvin",4,.35,.65),("Dessert","Baba au macvin",3,0,"Macvin",4,.20,.45)]},
"C":{"Demi Lune":[("Plat","Poulet aux morilles",1,1,"Porto",1,1.0,1.0),("Entrée","Feuilleté forestier",2,0,"Porto",1,.40,.65)],
     "Dahu":[("Plat","Fondue jurassienne",1,1,"Ravelin",10,1.0,1.0)],
     "Vegetarien":[("Plat","Assiette du père Grégoire",2,0,"Calvados",4,.35,.65),("Plat","Galette sarrasin forestière",2,0,"Porto",1,.35,.65)],
     "Bourguignon":[("Dessert","Coupe Dijonnaise",2,0,"Crème de cassis",4,.35,.65),("Dessert","Baba au marc de bourgogne",2,0,"Marc de bourgogne",4,.35,.65)],
     "Franc-Comtois":[("Plat","Poulet à la jurassienne",2,0,"Vin jaune",1,.35,.65),("Plat","Assiette franc-comtoise",2,0,"Macvin",4,.35,.65),("Dessert","Baba au macvin",3,0,"Macvin",4,.20,.45)]},
}
def per5(d):
    if d<"2022-07-23": return "P1"
    if d<"2023-04-01": return "P2"
    if d<"2023-04-17": return "P3"
    if d<"2024-04-01": return "P4"
    return "P5"
P5_INFO={"P1":("01/04/2022-22/07/2022","C","2022-2023"),"P2":("23/07/2022-31/03/2023","B","2022-2023"),
         "P3":("01/04/2023-16/04/2023","B","2023-2024"),"P4":("17/04/2023-31/03/2024","A","2023-2024"),
         "P5":("01/04/2024-31/03/2025","A","2024-2025")}

# ============================ EXTRACTION ITEMS CAISSE (ANNEXE-D) ============================
def extract_items():
    prods=collections.defaultdict(lambda: dict(section="",qte={e:0.0 for e in EXOS},ca={e:0.0 for e in EXOS},pu={e:0.0 for e in EXOS}))
    for exo,fn in D_FILES.items():
        sh=xlrd.open_workbook(fn).sheet_by_index(0); section=None
        for r in range(1,sh.nrows):
            a=norm(sh.cell_value(r,0)); b=norm(sh.cell_value(r,1))
            if a and not b and a!="Ref_prd":
                section=a
            elif b and b!="Lib_ticket" and section and "TOTAL" not in section.upper() and "RAPPORT" not in section.upper():
                try: q=float(sh.cell_value(r,4))
                except: q=0
                try: ca=float(sh.cell_value(r,5))
                except: ca=0
                try: pu=float(sh.cell_value(r,2))
                except: pu=0
                P=prods[(section,b)]; P["section"]=section
                P["qte"][exo]+=q; P["ca"][exo]+=ca
                if q>P["pu"][exo]*0 and pu: P["pu"][exo]=pu
    return prods

ITEMS=extract_items()

def get_qty(label):  # qty par exo pour un libellé (toutes sections)
    out={e:0.0 for e in EXOS}
    for (sec,lab),P in ITEMS.items():
        if lab==label:
            for e in EXOS: out[e]+=P["qte"][e]
    return out

# ============================ 1. ITEMS CAISSE (json + xlsx) ============================
def build_items():
    rows=[]
    for (sec,lab),P in sorted(ITEMS.items()):
        tot=sum(P["qte"].values())
        if tot<=0: continue
        rec={"section":sec,"produit":lab}
        if lab in mp.CAISSE2CANON:
            canon,fmt,vol=mp.CAISSE2CANON[lab]; rr=mp.ref(canon)
            rec.update({"nom_canonique":canon,"format_service":fmt,"volume_unitaire_cl":vol,"nom_inventaire":rr["nom_inventaire"],"nom_facture":rr["nom_facture"]})
        rec.update({"prix_unitaire":{e:round(P["pu"][e],2) for e in EXOS},
                     "quantite":{e:round(P["qte"][e],1) for e in EXOS},"quantite_totale":round(tot,1),
                     "ca_ttc":{e:round(P["ca"][e],2) for e in EXOS}})
        rows.append(rec)
    w("itemsCaisse.json",{"description":"Tous les items vendus en caisse (ANNEXE-D synthèse produit), quantité et CA par exercice fiscal.","exercices":EXOS,"nb_items":len(rows),"items":rows})
    # xlsx
    wb=Workbook(); ws=wb.active; ws.title="Items caisse"
    HEAD=PatternFill("solid",fgColor="1F4E78"); HF=Font(bold=True,color="FFFFFF",size=10); thin=Side("thin",color="BFBFBF"); B=Border(thin,thin,thin,thin)
    ws.merge_cells("A1:K1"); ws["A1"]="DEMI LUNE - Tous les items vendus en caisse, par exercice fiscal"; ws["A1"].font=Font(bold=True,size=13)
    cols=["Section","Produit","PU 22-23","PU 23-24","PU 24-25","Qté 22-23","Qté 23-24","Qté 24-25","Qté totale","CA 22-23","CA 23-24","CA 24-25"]
    HR=3
    for j,h in enumerate(cols,1):
        c=ws.cell(HR,j,h); c.fill=HEAD; c.font=HF; c.border=B; c.alignment=Alignment("center",wrap_text=True)
    r=HR+1
    for x in sorted(rows,key=lambda z:(z["section"],z["produit"])):
        vals=[x["section"],x["produit"]]+[x["prix_unitaire"][e] for e in EXOS]+[x["quantite"][e] for e in EXOS]+[x["quantite_totale"]]+[x["ca_ttc"][e] for e in EXOS]
        for j,v in enumerate(vals,1):
            c=ws.cell(r,j,v); c.border=B
            if j>=3: c.number_format="0.0" if j<9 or j==9 else "0.00"
            if j>=10: c.number_format="0.00"
        r+=1
    ws.freeze_panes="A4"; ws.auto_filter.ref=f"A{HR}:L{r-1}"
    for i,wd in enumerate([26,30,9,9,9,9,9,9,10,11,11,11],1): ws.column_dimensions[Lc(i)].width=wd
    wb.save(XL+"items-caisse-par-periode.xlsx"); print("  -> items-caisse-par-periode.xlsx")

# ============================ 2. RÉFÉRENCES DEPUIS LE JSON ============================
def build_refs():
    ref=json.load(open(XL+"alcool-plats-cocktails-menus.json",encoding="utf-8"))
    # cocktails (recettes complètes carte) + nom canonique par ingrédient
    cock={}
    for cname,rec in COCKTAILS.items():
        cock[cname]={"volume_cl":rec["volume_cl"],"ingredients":[
            {**ig,"nom_canonique":mp.INGREDIENT2CANON.get(ig["nom"],ig["nom"])} for ig in rec["ingredients"]]}
    w("cocktailsComposition.json",{"description":"Recette exacte des cocktails (cl par ingrédient) + nom_canonique unifié. estimation=1 : répartition déduite (volume total connu, split estimé).","unite":"cl","cocktails":cock})
    # sauces
    w("saucesAlcool.json",{"description":"Alcool contenu dans les sauces.","unite_dose":"cl","convention":"10 cl d'alcool / litre = 1 cl / portion","sauces":ref["sauces"]["liste"]})
    # plats
    w("platsAlcool.json",{"description":"Alcool contenu dans les plats, entrées et desserts à la carte.","unite_dose":"cl","entrees":ref["cartes"]["entrees"],"plats":ref["cartes"]["plats"],"desserts":ref["cartes"]["desserts"]})
    # prix menus
    w("menusPrix.json",{"description":"Prix des menus par période de carte.","cartes":{c:MENU_PRICES[c] for c in ["C","B","A"]}})
    # composition alcool des menus + proba
    out={}
    for c in ["C","B","A"]:
        mc={}
        for menu,opts in MENU_COMP[c].items():
            services=collections.defaultdict(list)
            for (svc,opt,nopt,obl,alc,dose,pb,ph) in opts:
                pm=round((pb+ph)/2,3)
                services[svc].append({"option":opt,"nb_options_service":nopt,"obligatoire":bool(obl),"alcool":alc,"dose_cl":dose,
                    "proba_basse":round(pb,3),"proba_moyenne":pm,"proba_haute":round(ph,3),
                    "cl_bas":round(dose*pb,3),"cl_moyen":round(dose*pm,3),"cl_haut":round(dose*ph,3)})
            mc[menu]={"services":dict(services)}
        out[c]={"dates":MENU_PRICES[c]["dates"],"menus":mc}
    w("menusCompositionAlcool.json",{"description":"Alcool des menus par carte (C/B/A). Pour chaque option alcoolisée : dose + probabilité de choix (basse/moyenne/haute) et cl correspondants. obligatoire=true -> alcool certain (100%).","regle":"proba_moyenne=(basse+haute)/2 ; cl = dose x proba","cartes":out})

# ============================ 3. MENUS VENDUS (ANNEXE-C, 5 périodes) ============================
def scan_menu_sales():
    # (p5,menu_key,prix) -> qte ; + libellés hors catalogue
    sales=collections.defaultdict(float); hors=collections.defaultdict(float)
    for fn in C_FILES:
        sh=xlrd.open_workbook(fn).sheet_by_index(0)
        for r in range(1,sh.nrows):
            lib=str(sh.cell_value(r,10)).strip()
            if lib not in MENU_LABEL:
                if lib.lower().startswith("menu") or lib.lower().startswith("formule"):
                    d=str(sh.cell_value(r,0))[:10]; hors[(per5(d),lib,round(float(sh.cell_value(r,13)),2))]+=float(sh.cell_value(r,11))
                continue
            mk=MENU_LABEL[lib]; d=str(sh.cell_value(r,0))[:10]; p=per5(d)
            prix=round(float(sh.cell_value(r,13)),2); q=float(sh.cell_value(r,11))
            sales[(p,mk,prix)]+=q
    return sales,hors

def build_menu_sales():
    sales,hors=scan_menu_sales()
    # D5 : menusVentesParPrix
    out={}
    for p in ["P1","P2","P3","P4","P5"]:
        dates,carte,exo=P5_INFO[p]
        menus={}
        for (pp,mk,prix),q in sales.items():
            if pp!=p: continue
            attendu=round(float(MENU_PRICES[carte][mk]),2)
            modifie="NON" if round(prix,2) in ACCEPT[mk] else "OUI"
            menus.setdefault(mk,{"prix_attendu":attendu,"lignes":[]})
            menus[mk]["lignes"].append({"prix":prix,"quantite":round(q,1),"prix_modifie":modifie})
        for mk in menus: menus[mk]["lignes"].sort(key=lambda z:z["prix"])
        out[p]={"dates":dates,"carte":carte,"exercice":exo,"menus":menus}
    w("menusVentesParPrix.json",{"description":"Menus vendus en caisse par période (5). Classés par prix avec quantité. prix_modifie=OUI si le prix ne correspond à aucun prix catalogue connu du menu (facture sans détail). prix_attendu = prix de la carte de la période.","note_hors_catalogue":"Les libellés 'Menu du jour' et autres hors liste sont ignorés ici.","periodes":out})
    return sales

# ============================ 4. MENUS NON MODIFIÉS + ALCOOL (5 périodes) ; 5. CONSO MENUS (3 exos)
def build_menu_alcool(sales):
    # qte non modifiée par (p5,menu composé)
    qok=collections.defaultdict(float)
    for (p,mk,prix),q in sales.items():
        if mk in COMPOSED and round(prix,2) in ACCEPT[mk]:
            qok[(p,mk)]+=q
    # D6
    d6={}
    conso_exo=collections.defaultdict(lambda:[0.0,0.0,0.0])  # (exo,alcool)->[bas,moy,haut]
    for p in ["P1","P2","P3","P4","P5"]:
        dates,carte,exo=P5_INFO[p]; menus={}
        for mk in COMPOSED:
            q=qok.get((p,mk),0.0)
            if q<=0: continue
            opts=[]
            for (svc,opt,nopt,obl,alc,dose,pb,ph) in MENU_COMP[carte][mk]:
                pm=(pb+ph)/2
                cb,cm,ch=q*dose*pb,q*dose*pm,q*dose*ph
                opts.append({"service":svc,"option":opt,"alcool":alc,"dose_cl":dose,"obligatoire":bool(obl),
                    "proba_basse":round(pb,3),"proba_moyenne":round(pm,3),"proba_haute":round(ph,3),
                    "cl_bas":round(cb,1),"cl_moyen":round(cm,1),"cl_haut":round(ch,1)})
                conso_exo[(exo,alc)][0]+=cb; conso_exo[(exo,alc)][1]+=cm; conso_exo[(exo,alc)][2]+=ch
            menus[mk]={"quantite_non_modifiee":round(q,1),"options_alcool":opts}
        d6[p]={"dates":dates,"carte":carte,"exercice":exo,"menus":menus}
    w("menusNonModifiesAlcool.json",{"description":"Menus non modifiés (prix catalogue) par période (5), avec options alcoolisées et estimation d'alcool (cl) basse/moyenne/haute = quantité x dose x proba.","periodes":d6})
    # D7
    d7={}
    for exo in EXOS:
        lst=[]
        for (e,alc),(b,m,h) in conso_exo.items():
            if e!=exo: continue
            lst.append({"alcool":alc,"cl_bas":round(b,1),"cl_moyen":round(m,1),"cl_haut":round(h,1),"litres_moyen":round(m/100,2)})
        lst.sort(key=lambda z:-z["cl_moyen"])
        d7[exo]={"alcools":lst}
    w("consoAlcoolMenus.json",{"description":"Consommation d'alcool liée aux menus (non modifiés), par exercice fiscal, estimation basse/moyenne/haute (cl).","exercices":d7})
    return conso_exo

# ============================ 6. COCKTAILS CONSO (ingrédients, 3 exos) ============================
def build_cocktail_conso():
    ing=collections.defaultdict(lambda: {e:0.0 for e in EXOS})
    cocktails=[]
    for caisse_lab,cname in CAISSE_COCKTAIL.items():
        rec=COCKTAILS[cname]; perexo={}; tq=0.0
        for e in EXOS:
            q=get_qty(caisse_lab)[e]; ingl={}
            for ig in rec["ingredients"]:
                clt=q*ig["cl"]; ingl[ig["nom"]]=round(clt/100,2)
                ing[mp.INGREDIENT2CANON.get(ig["nom"],ig["nom"])][e]+=clt
            perexo[e]={"quantite":round(q,1),"ingredients_l":ingl}; tq+=q
        if tq<=0: continue
        cocktails.append({"cocktail":cname,"libelle_caisse":caisse_lab,"volume_cl":rec["volume_cl"],
            "recette":[{"nom":ig["nom"],"nom_canonique":mp.INGREDIENT2CANON.get(ig["nom"],ig["nom"]),"categorie":ig["categorie"],"cl_unitaire":ig["cl"],"estimation":bool(ig["estimation"])} for ig in rec["ingredients"]],
            "total_quantite":round(tq,1),"par_periode":perexo})
    cocktails.sort(key=lambda z:-z["total_quantite"])
    toting=[]
    for canon,pe in ing.items():
        tot=sum(pe.values())
        if tot<=0: continue
        r=mp.ref(canon)
        toting.append({"ingredient":canon,"categorie":r["categorie"],"nom_inventaire":r["nom_inventaire"],"nom_facture":r["nom_facture"],"total_l":round(tot/100,2),"par_periode":{e:round(pe[e]/100,2) for e in EXOS}})
    toting.sort(key=lambda z:-z["total_l"])
    w("cocktailsConsoComposition.json",{"description":"Cocktails enregistrés en caisse, classés par cocktail. Recette avec nom_canonique de chaque ingrédient. 'totaux_par_ingredient' : cumul par ingrédient (nom canonique unifié + nom_inventaire/nom_facture) avec détail par exercice.","exercices":EXOS,"cocktails":cocktails,"totaux_par_ingredient":toting})

# ============================ 7. BOISSONS HORS COCKTAIL (3 exos) ============================
def build_hors_cocktail():
    agg=collections.defaultdict(lambda: {e:0.0 for e in EXOS})   # (canon,fmt,vol) -> exo -> qte
    for lab,(canon,fmt,vol) in mp.CAISSE2CANON.items():
        q=get_qty(lab)
        if sum(q.values())<=0: continue
        for e in EXOS: agg[(canon,fmt,vol)][e]+=q[e]
    boissons=[]
    for (canon,fmt,vol),pe in agg.items():
        r=mp.ref(canon); periodes={}; tq=tv=0.0
        for e in EXOS:
            qt=pe[e]; v=qt*vol
            periodes[e]={"quantite":round(qt,1),"volume_l":round(v/100,2)}
            tq+=qt; tv+=v
        boissons.append({"nom_canonique":canon,"format_service":fmt,"volume_unitaire_cl":vol,"categorie":r["categorie"],
            "nom_inventaire":r["nom_inventaire"],"nom_facture":r["nom_facture"],"flag":r["flag"],
            "total_quantite":round(tq,1),"total_volume_l":round(tv/100,2),"par_periode":periodes})
    boissons.sort(key=lambda z:-z["total_volume_l"])
    w("boissonsHorsCocktail.json",{"description":"Boissons hors cocktail (vins, bières, cidres, alcools secs...), noms unifiés (canonique = inventaire si présent, sinon facture ; + nom_inventaire/nom_facture). Une ligne par (boisson, format de service), taille préservée. volume_l = volume servi (= alcool pour ces boissons). Ventes par exercice fiscal.","exercices":EXOS,"nb_boissons":len(boissons),"boissons":boissons})

# ============================ 8. CUMUL FINAL : 1 ligne par boisson, total par période ============================
def build_final():
    data=collections.defaultdict(lambda: {e:dict(boissons=0.0,cocktails=0.0,plats=0.0,mb=0.0,mm=0.0,mh=0.0) for e in EXOS})
    est=collections.defaultdict(bool)
    # 1) boissons hors cocktail (volume servi = alcool) - clé = canonique (formats fusionnés)
    for lab,(canon,fmt,vol) in mp.CAISSE2CANON.items():
        q=get_qty(lab)
        for e in EXOS: data[canon][e]["boissons"]+=q[e]*vol
    # 2) ingrédients cocktails
    for caisse_lab,cname in CAISSE_COCKTAIL.items():
        q=get_qty(caisse_lab)
        for ig in COCKTAILS[cname]["ingredients"]:
            canon=mp.INGREDIENT2CANON.get(ig["nom"],ig["nom"])
            for e in EXOS: data[canon][e]["cocktails"]+=q[e]*ig["cl"]
            if ig["estimation"]: est[canon]=True
    # 3) alcool plats/desserts à la carte
    rows,_=pa.build_carte()
    for x in rows:
        if x["famille"] not in("Plat","Dessert"): continue
        t=x["typ"]
        if t=="Grand Marnier ou Calvados":                 # flambage : 50% / 50%
            for canon,fr in (("Grand Marnier",0.5),("Calvados",0.5)):
                data[canon][x["exo"]]["plats"]+=float(x["alc"])*fr; est[canon]=True
        else:
            canon=mp.INGREDIENT2CANON.get(t,t); data[canon][x["exo"]]["plats"]+=float(x["alc"])
    # 4) alcool menus (estimé bas/moyen/haut)
    mn=json.load(open(JS+"consoAlcoolMenus.json",encoding="utf-8"))["exercices"]
    for exo in EXOS:
        for a in mn[exo]["alcools"]:
            canon=mp.INGREDIENT2CANON.get(a["alcool"],a["alcool"]); d=data[canon][exo]
            d["mb"]+=a["cl_bas"]; d["mm"]+=a["cl_moyen"]; d["mh"]+=a["cl_haut"]; est[canon]=True
    boissons=[]
    for canon,perexo in data.items():
        r=mp.ref(canon); periodes={}; T=dict(exact=0.0,mb=0.0,mm=0.0,mh=0.0)
        for exo in EXOS:
            d=perexo[exo]; exact=d["boissons"]+d["cocktails"]+d["plats"]
            periodes[exo]={
                "exact_certain_l":round(exact/100,2),
                "detail_exact_l":{"boissons_seches":round(d["boissons"]/100,2),"ingredients_cocktails":round(d["cocktails"]/100,2),"plats_desserts":round(d["plats"]/100,2)},
                "menu_estime_l":{"bas":round(d["mb"]/100,2),"moyen":round(d["mm"]/100,2),"haut":round(d["mh"]/100,2)},
                "total_l":{"bas":round((exact+d["mb"])/100,2),"moyen":round((exact+d["mm"])/100,2),"haut":round((exact+d["mh"])/100,2)}}
            T["exact"]+=exact;T["mb"]+=d["mb"];T["mm"]+=d["mm"];T["mh"]+=d["mh"]
        boissons.append({"nom_canonique":canon,"categorie":r["categorie"],"nom_inventaire":r["nom_inventaire"],"nom_facture":r["nom_facture"],"flag":r["flag"],"estimation":bool(est[canon]),
            "total_3_exercices":{"exact_certain_l":round(T["exact"]/100,2),
                "menu_estime_l":{"bas":round(T["mb"]/100,2),"moyen":round(T["mm"]/100,2),"haut":round(T["mh"]/100,2)},
                "total_l":{"bas":round((T["exact"]+T["mb"])/100,2),"moyen":round((T["exact"]+T["mm"])/100,2),"haut":round((T["exact"]+T["mh"])/100,2)}},
            "par_periode":periodes})
    boissons.sort(key=lambda z:-z["total_3_exercices"]["total_l"]["moyen"])
    w("consoTotaleParBoisson.json",{"description":"Toutes les boissons/ingrédients (noms canoniques unifiés caisse/inventaire/factures). Cumule 4 sources : boissons sèches hors cocktail + ingrédients cocktails + alcool plats/desserts + alcool menus. exact_certain_l (hors menu) ; menu_estime_l (part menus, bas/moyen/haut) ; total_l = exact + menu. Règle : exact_certain + menu_estime.moyen = total.moyen. Litres.","exercices":EXOS,"nb_boissons":len(boissons),"boissons":boissons})

def build_final_xlsx():
    dd=json.load(open(JS+"consoTotaleParBoisson.json",encoding="utf-8")); B=dd["boissons"]
    HEAD=PatternFill("solid",fgColor="1F4E78"); HF=Font(bold=True,color="FFFFFF",size=9)
    EX=PatternFill("solid",fgColor="E2EFDA"); ME=PatternFill("solid",fgColor="FFF2CC"); TO=PatternFill("solid",fgColor="DDEBF7"); GT=PatternFill("solid",fgColor="F4B183")
    thin=Side("thin",color="BFBFBF"); BO=Border(thin,thin,thin,thin); ctr=Alignment("center",vertical="center"); lft=Alignment("left",vertical="center"); bold=Font(bold=True)
    wb=Workbook(); ws=wb.active; ws.title="Conso totale par boisson"
    SUB=["Exact (certain) L","Menu estimé (moyen) L","Total bas (L)","Total moyen (L)","Total haut (L)"]; ncol=3+5*4
    ws.merge_cells(start_row=1,start_column=1,end_row=1,end_column=ncol); ws["A1"]="DEMI LUNE - Consommation totale par boisson et par exercice (Litres)"; ws["A1"].font=Font(bold=True,size=13)
    ws.merge_cells(start_row=2,start_column=1,end_row=2,end_column=ncol)
    ws["A2"]=("Lecture : 'Exact (certain)' = consommation connue hors menu (boissons sèches + ingrédients cocktails + plats/desserts). 'Menu estimé (moyen)' = part d'alcool venant des menus (estimée). TOTAL = Exact + part menu ; bas/moyen/haut = fourchette des menus. Donc Exact + Menu(moyen) = Total moyen ; sans menu, Exact = Total.")
    ws["A2"].font=Font(italic=True,size=9); ws["A2"].alignment=Alignment(wrap_text=True,vertical="center"); ws.row_dimensions[2].height=46
    HR=4; ws.cell(HR,1,"Boisson"); ws.cell(HR,2,"Catégorie"); ws.cell(HR,3,"Estim.")
    for i in (1,2,3): ws.merge_cells(start_row=HR,start_column=i,end_row=HR+1,end_column=i)
    col=4
    for lab in EXOS+["TOTAL 3 exercices"]:
        ws.merge_cells(start_row=HR,start_column=col,end_row=HR,end_column=col+4); ws.cell(HR,col,lab)
        for kk,sub in enumerate(SUB): ws.cell(HR+1,col+kk,sub)
        col+=5
    for r in (HR,HR+1):
        for c in range(1,ncol+1):
            cell=ws.cell(r,c); cell.fill=HEAD; cell.font=HF; cell.border=BO; cell.alignment=Alignment("center",vertical="center",wrap_text=True)
    fills=[EX,ME,TO,TO,TO]; r=HR+2
    for b in B:
        row=[b["nom_canonique"],b["categorie"],"oui" if b["estimation"] else ""]
        def blk(o): return [o["exact_certain_l"] if "exact_certain_l" in o else o["exact_certain_l"], o["menu_estime_l"]["moyen"], o["total_l"]["bas"], o["total_l"]["moyen"], o["total_l"]["haut"]]
        for e in EXOS: row+=blk(b["par_periode"][e])
        row+=blk(b["total_3_exercices"])
        for j,v in enumerate(row,1):
            cell=ws.cell(r,j,v); cell.border=BO; cell.alignment=lft if j<=2 else ctr
            if j>=4: cell.number_format="0.0"; cell.fill=fills[(j-4)%5]
        r+=1
    ws.cell(r,1,"TOTAL").alignment=lft
    for j in range(1,ncol+1):
        cell=ws.cell(r,j); cell.fill=GT; cell.border=BO; cell.alignment=ctr; cell.font=bold
        if j>=4: cell.value=f"=SUM({Lc(j)}{HR+2}:{Lc(j)}{r-1})"; cell.number_format="0.0"
    ws.cell(r,1).value="TOTAL"
    ws.freeze_panes="D6"; ws.auto_filter.ref=f"A{HR}:{Lc(ncol)}{r-1}"
    ws.column_dimensions["A"].width=26; ws.column_dimensions["B"].width=12; ws.column_dimensions["C"].width=7
    for j in range(4,ncol+1): ws.column_dimensions[Lc(j)].width=11
    wb.save(JS+"consoTotaleParBoisson.xlsx"); print("  -> consoTotaleParBoisson.xlsx (src/data/calculsBoissons)")

if __name__=="__main__":
    print("== Items caisse =="); build_items()
    print("== Références (JSON) =="); build_refs()
    print("== Menus vendus / prix =="); sales=build_menu_sales()
    print("== Menus non modifiés + conso menus =="); build_menu_alcool(sales)
    print("== Cocktails conso =="); build_cocktail_conso()
    print("== Boissons hors cocktail =="); build_hors_cocktail()
    print("== Cumul final par boisson =="); build_final(); build_final_xlsx()
    print("FAIT.")
