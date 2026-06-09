#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""rapprochementDisparuTotal.xlsx : bilan PÉRIODE ENTIÈRE (3 exercices cumulés), TOUTES les boissons.
PILOTÉ PAR LES ACHATS : union des boissons VENDUES (consoTotale) et de TOUTES les boissons ACHETÉES (factures),
via le mapping EXPLICITE _mapping_achats.py (millésimes regroupés, premium/perso séparés : Ruinart, Lanson, Aperol,
Sauternes, bières personnel...). Les boissons achetées mais jamais vendues en caisse sont étiquetées « non vendu (perso/offert) ».
disparu = achats - conso - stock_final(31/03/2025) (stock d'ouverture 31/03/2022 = 0, indisponible), planché à 0.
Sort aussi mappingAchatsBoissons.xlsx (table de relecture : chaque produit facture -> canonique)."""
import json,collections,unicodedata,re,importlib.util as _i
from openpyxl import Workbook
from openpyxl.styles import Font,PatternFill,Alignment,Border,Side
from openpyxl.utils import get_column_letter as Lc
BASE="/Users/maevapasteur/Documents/demi-lune-comptabilite/"; JS=BASE+"src/data/calculsBoissons/"; VB=BASE+"public/documents/vins-boissons/"
EXOS=["2022-2023","2023-2024","2024-2025"]
def na(s): return ''.join(c for c in unicodedata.normalize('NFD',str(s)) if unicodedata.category(c)!='Mn').lower()
def size_cl(name):
    if not name: return None
    s=na(name).replace(",","."); best=None
    for m in re.finditer(r"(\d+(?:\.\d+)?)\s*(cl|l)\b",s): best=float(m.group(1))*(100 if m.group(2)=="l" else 1)
    return best

def load(p):
    sp=_i.spec_from_file_location(p,VB+f"_{p}.py"); m=_i.module_from_spec(sp); sp.loader.exec_module(m); return m
mp=load("mapping_noms"); ma=load("mapping_achats")
revente=mp.CARTE_PRIX                      # canonique -> (serving_cl, serving_prix)
canon_achat=ma.canon_achat; NOUVEAUX=ma.NOUVEAUX; NONALC=ma.NONALC
CUISINE={"Absinthe"}                       # acheté, non vendu : usage cuisine (crème brûlée)

# ---- fallback fuzzy (pour boissons NON-alcool vendues : Coca, jus, sirops, limonade) ----
conso=json.load(open(JS+"consoTotaleParBoisson.json",encoding="utf-8"))
STOP=set("75cl 70cl 100cl 50cl 33cl 62cl 37 5cl 10l 8l cl l de du des la le maison bib verre bouteille pichet rouge blanc rose rosé n vp c10 ° et aux".split())
def toks(s): return [t for t in re.split(r"[^a-z0-9]+",na(s)) if t and t not in STOP and len(t)>2 and not t.isdigit()]
def anchors(name): return set(t for t in toks(name) if t not in("cotes","cote","arbois","jura","vin","sirop"))
canon_anchor={b["nom_canonique"]:anchors(b["nom_canonique"]) for b in conso["boissons"]}
def fuzzy(label):
    at=set(toks(label)); best=None;bn=0
    for c,anc in canon_anchor.items():
        ov=len(anc&at)
        if ov>bn: best=c;bn=ov
    return best if bn>=1 else None

# ---- table des canoniques VENDUS (conso + stock + meta) ----
SOLD={}
for b in conso["boissons"]:
    t=b["taille_achat_cl"]
    SOLD[b["nom_canonique"]]=dict(
        cat=b["categorie"],unite=b["unite_achat"],taille=t,
        cb=sum(b["par_periode"][e]["total_l"]["bas"] for e in EXOS),
        cm=sum(b["par_periode"][e]["total_l"]["moyen"] for e in EXOS),
        ch=sum(b["par_periode"][e]["total_l"]["haut"] for e in EXOS),
        stockL=(b["inventaire_fin_contenants_par_periode"]["2024-2025"]*(t or 0)/100))

# ---- achats : tous les produits boissons factures -> canonique (alcool=explicite, non-alcool=fuzzy) ----
ach=json.load(open(JS+"achatsBoissonsParPeriode.json",encoding="utf-8"))["achats"]
ACH=collections.defaultdict(lambda:[0.0,0.0,None,None])   # canon -> [litres, montant, size_repr, cat_facture]
review=[]   # (produit, cat_facture, q, montant, canon, source)
for a in ach:
    q=a["total_quantite"]
    if q<=0: continue
    c=canon_achat(a["produit"]); src="explicite"   # règles explicites sur TOUS les produits (catégorie facture peu fiable)
    if c is None:
        c=fuzzy(a["produit"]); src="fuzzy" if c else "—"
    review.append((a["produit"],a["categorie"],q,a["total_montant_ht"],c,src))
    if not c: continue
    isfut="fut" in na(a["produit"]); sz=100 if isfut else size_cl(a["produit"])
    if not sz: continue
    L=q*sz/100
    ACH[c][0]+=L; ACH[c][1]+=a["total_montant_ht"]
    if ACH[c][0]==L or sz>(ACH[c][2] or 0): ACH[c][2]=sz
    ACH[c][3]=a["categorie"]

UNITE_NEW={"fut":"fût 8L (=800cl)","bib":"BIB","champ":"75cl","biere":"33cl"}
def unite_new(canon,sz,catf):
    n=na(canon)
    if "fut" in n: return "fût 8L (=800cl)"
    if sz and sz>=1000: return f"BIB {sz/100:.0f}L"
    if sz==100: return "100cl"
    return f"{sz:.0f}cl" if sz else "n/d"
CAT_NEW=lambda catf: {"bière":"biere","vin/cidre":"vin","spiritueux/liqueur":"spiritueux"}.get(catf,catf)

# ---- assemblage : UNION ventes + achats ----
def clamp(v): return v if v>0 else 0.0
allcanon=set(SOLD)|set(ACH)
rows=[]
for c in allcanon:
    s=SOLD.get(c)
    al=ACH.get(c,[0.0,0.0,None,None])
    achL=al[0]; mont=al[1]
    if s:
        cat=s["cat"];unite=s["unite"];taille=s["taille"];cb,cm,ch=s["cb"],s["cm"],s["ch"];stockL=s["stockL"];isnew=c in NOUVEAUX
    else:
        isnew=True; cat=CAT_NEW(al[3]); taille=al[2] or 75; unite=unite_new(c,al[2],al[3])
        cb=cm=ch=0.0; stockL=0.0
    pa=(mont/achL) if achL>0 else None                                   # €/L achat
    sv=revente.get(c); pr=(sv[1]*100/sv[0]) if sv else None               # €/L revente
    soft = c in NONALC
    if soft:
        # softs (eau/soda/jus/sirop/chaud) : vendus sous boutons génériques -> disparu non rattachable
        d_bas=d_moy=d_haut=None; pb=pm=ph=None
        statut="Soft (eau/soda/jus/sirop/chaud) - vendu sous bouton générique : disparu non rattachable"
        fiable=False
    else:
        d_bas=clamp(achL-ch-stockL); d_moy=clamp(achL-cm-stockL); d_haut=clamp(achL-cb-stockL)
        pct=lambda d:(d/achL*100) if achL>0 else None
        pb,pm,ph=pct(d_bas),pct(d_moy),pct(d_haut)
        nonvendu = isnew and cm==0
        fiable = (cat not in ("sirop","jus","soda")) and (not nonvendu)
        if c in CUISINE: statut="Non vendu en caisse (usage cuisine : crème brûlée)"
        elif nonvendu:   statut="Non vendu en caisse (perso/offert/événement)"
        elif cat not in ("sirop","jus","soda"): statut="Oui"
        else:            statut="Non (mixer : vente seule non comptée)"
    vA=lambda d:(d*pa if (d is not None and pa is not None) else None)
    vR=lambda d:(d*pr if (d is not None and pr is not None) else None)
    rows.append(dict(b=c,cat=cat,unite=unite,pa=pa,pr=pr,cb=cb,cm=cm,ch=ch,stock=stockL,ach=achL,
        db=d_bas,dm=d_moy,dh=d_haut,pb=pb,pm=pm,ph=ph,
        vab=vA(d_bas),vam=vA(d_moy),vah=vA(d_haut),vrb=vR(d_bas),vrm=vR(d_moy),vrh=vR(d_haut),
        statut=statut,_f=fiable,_new=isnew,_soft=soft))
rows.sort(key=lambda z:(0 if z["_f"] else 1, -((z["vrm"] or 0)+(z["vam"] or 0))))

# ---- xlsx principal ----
HEAD=PatternFill("solid",fgColor="1F4E78");HF=Font(bold=True,color="FFFFFF",size=9);GT=PatternFill("solid",fgColor="F4B183")
CONS=PatternFill("solid",fgColor="DDEBF7");PX=PatternFill("solid",fgColor="E2EFDA");ACHF=PatternFill("solid",fgColor="FFF2CC")
DIS=PatternFill("solid",fgColor="FCE4D6");VAL=PatternFill("solid",fgColor="F4B183");NEWF=PatternFill("solid",fgColor="EDEDED")
thin=Side("thin",color="BFBFBF");B=Border(thin,thin,thin,thin);ctr=Alignment("center",vertical="center");lft=Alignment("left",vertical="center");bold=Font(bold=True)
wb=Workbook();ws=wb.active;ws.title="Disparu total (toutes boissons)"
cols=["Boisson","Catégorie","Unité d'achat","Prix achat €/L","Prix revente €/L",
 "Conso bas (L)","Conso moyen (L)","Conso haut (L)","Stock final 31/03/2025 (L)","Achats (L)",
 "Disparu bas (L)","Disparu moyen (L)","Disparu haut (L)","% disparu bas","% disparu moyen","% disparu haut",
 "Val. achat bas €","Val. achat moyen €","Val. achat haut €","Val. revente bas €","Val. revente moyen €","Val. revente haut €","Statut conso caisse"]
FILL={4:PX,5:PX,6:CONS,7:CONS,8:CONS,9:CONS,10:ACHF,11:DIS,12:DIS,13:DIS,14:DIS,15:DIS,16:DIS,17:VAL,18:VAL,19:VAL,20:VAL,21:VAL,22:VAL}
ncol=len(cols)
ws.merge_cells(start_row=1,start_column=1,end_row=1,end_column=ncol)
ws["A1"]="DEMI LUNE - Rapprochement TOTAL (3 exercices) - TOUTES les boissons : disparu (L) + valorisation achat/revente";ws["A1"].font=Font(bold=True,size=13)
ws.merge_cells(start_row=2,start_column=1,end_row=2,end_column=ncol)
ws["A2"]=("Période 01/04/2022 → 31/03/2025. UNION des boissons VENDUES et de TOUTES les boissons ACHETÉES (mapping explicite des factures). TOUT EN LITRES. "
 "Disparu = achats − conso − stock final (31/03/2025), planché à 0 ; stock d'ouverture (31/03/2022) = 0 (indisponible). "
 "Fourchette : disparu BAS = conso haute, disparu HAUT = conso basse. Prix en €/L. Convertir en bouteilles : ÷ taille unité en L (ex. 0,75). "
 "Statut « Non vendu en caisse » = boisson achetée jamais sonnée (bières personnel, champagnes premium, Aperol...) : le « disparu » = conso personnel/offerts/événements, PAS une recette occultée. Mixers (sirop/jus/soda) : conso incomplète.")
ws["A2"].font=Font(italic=True,size=9);ws["A2"].alignment=Alignment(wrap_text=True,vertical="center");ws.row_dimensions[2].height=72
HR=4
for j,h in enumerate(cols,1):
    cc=ws.cell(HR,j,h);cc.fill=HEAD;cc.font=HF;cc.border=B;cc.alignment=Alignment("center",wrap_text=True,vertical="center")
def fmt(v,d=1): return round(v,d) if isinstance(v,(int,float)) else "n/d"
r=HR+1
for x in rows:
    vals=[x["b"],x["cat"],x["unite"],fmt(x["pa"],2),fmt(x["pr"],2),fmt(x["cb"]),fmt(x["cm"]),fmt(x["ch"]),fmt(x["stock"]),fmt(x["ach"]),
        fmt(x["db"]),fmt(x["dm"]),fmt(x["dh"]),fmt(x["pb"]),fmt(x["pm"]),fmt(x["ph"]),
        fmt(x["vab"],2),fmt(x["vam"],2),fmt(x["vah"],2),fmt(x["vrb"],2),fmt(x["vrm"],2),fmt(x["vrh"],2),x["statut"]]
    for j,v in enumerate(vals,1):
        cc=ws.cell(r,j,v);cc.border=B;cc.alignment=lft if j in(1,2,3,23) else ctr
        if j in FILL: cc.fill=FILL[j]
        if x["_new"] and j==1: cc.fill=NEWF;cc.font=Font(italic=True)
        if isinstance(v,(int,float)): cc.number_format="0.00" if (j in(4,5) or j>=17) else "0.0"
    r+=1
lastdata=r-1
def somme(key,pred): return sum((x[key] or 0) for x in rows if x[key] is not None and pred(x))
TOT=(("TOTAL ALCOOL VENDU (occultation potentielle, conso fiable)",lambda x:x["_f"],GT),
     ("TOTAL boissons NON vendues (perso/offert/événement)",lambda x:x["_new"] and x["statut"].startswith("Non vendu"),PatternFill("solid",fgColor="D9D9D9")),
     ("TOTAL toutes boissons",lambda x:True,PatternFill("solid",fgColor="F8CBAD")))
for lab,pred,fill in TOT:
    ws.cell(r,1,lab).font=bold
    for j,key in ((17,"vab"),(18,"vam"),(19,"vah"),(20,"vrb"),(21,"vrm"),(22,"vrh")):
        ws.cell(r,j,round(somme(key,pred),2)).number_format="0.00"
    for j in range(1,ncol+1): ws.cell(r,j).fill=fill;ws.cell(r,j).border=B;ws.cell(r,j).font=bold;ws.cell(r,j).alignment=ctr if j>=17 else lft
    r+=1
ws.freeze_panes="D5";ws.auto_filter.ref=f"A{HR}:{Lc(ncol)}{lastdata}"
W=[40,12,15,12,13,11,12,11,15,11,11,12,11,11,12,11,12,13,12,12,13,12,30]
for i,wd in enumerate(W,1): ws.column_dimensions[Lc(i)].width=wd
wb.save(JS+"rapprochementDisparuTotal.xlsx")

# ---- xlsx de relecture du mapping ----
wb2=Workbook();w2=wb2.active;w2.title="Mapping factures -> canonique"
h2=["Produit (facture)","Catégorie facture","Quantité","Montant HT €","Canonique attribué","Source mapping","Nouveau (non vendu) ?"]
w2.append(h2)
for j in range(1,len(h2)+1):
    cc=w2.cell(1,j);cc.fill=HEAD;cc.font=HF;cc.border=B;cc.alignment=Alignment("center",wrap_text=True,vertical="center")
review.sort(key=lambda z:(na(z[4] or "~~~"),-z[3]))
for prod,catf,q,m,c,src in review:
    isn = c in NOUVEAUX if c else False
    w2.append([prod,catf,round(q,1),round(m,2),c or "NON MAPPÉ",src,"Oui" if isn else ""])
    if not c:
        for j in range(1,8): w2.cell(w2.max_row,j).fill=PatternFill("solid",fgColor="FFC7CE")
w2.freeze_panes="A2";w2.auto_filter.ref=f"A1:G{w2.max_row}"
for i,wd in enumerate([46,16,10,12,38,13,18],1): w2.column_dimensions[Lc(i)].width=wd
wb2.save(JS+"mappingAchatsBoissons.xlsx")

# ---- report ----
af=[x for x in rows if x["_f"]]; nv=[x for x in rows if x["statut"].startswith("Non vendu")]
print(f"-> rapprochementDisparuTotal.xlsx | {len(rows)} boissons (dont {len(nv)} non vendues ajoutées)")
print(f"-> mappingAchatsBoissons.xlsx | {len(review)} produits facture (NON mappés: {sum(1 for x in review if not x[4])})")
print("ALCOOL VENDU - val revente disparu : bas {:.0f} / moyen {:.0f} / haut {:.0f} €".format(
    sum(x['vrb'] or 0 for x in af),sum(x['vrm'] or 0 for x in af),sum(x['vrh'] or 0 for x in af)))
print("ALCOOL VENDU - val achat disparu   : bas {:.0f} / moyen {:.0f} / haut {:.0f} €".format(
    sum(x['vab'] or 0 for x in af),sum(x['vam'] or 0 for x in af),sum(x['vah'] or 0 for x in af)))
print("NON VENDUES (perso/offert) - val achat : {:.0f} € | {} : {}".format(
    sum(x['vam'] or 0 for x in nv),len(nv),", ".join(x['b'] for x in nv[:30])))
print("Top 8 alcool vendu (val revente moyenne) :")
for x in af[:8]:
    print(f"   {x['b'][:30]:30s} disparu {x['db']:.0f}/{x['dm']:.0f}/{x['dh']:.0f} L | val revente {x['vrm']:.0f} €" )
