#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""Rapprochement carnet manuscrit (anomalies fournisseur) <-> factures.
Artefact de RELECTURE : une ligne par anomalie carnet, avec la facture probable
(fenêtre de dates + mot-clé distinctif) et le traitement déjà appliqué dans le JSON.
Aucune déduction automatique. Sortie : rapprochement-carnet-factures.xlsx + .json
"""
import xlrd,json,collections,unicodedata,re,datetime
from openpyxl import Workbook
from openpyxl.styles import Font,PatternFill,Alignment,Border,Side
from openpyxl.utils import get_column_letter as Lc
BASE="/Users/maevapasteur/Documents/demi-lune-comptabilite/"; JS=BASE+"src/data/calculsBoissons/"
def na(s): return ''.join(c for c in unicodedata.normalize('NFD',str(s)) if unicodedata.category(c)!='Mn').lower()
def d(s):
    try: return datetime.datetime.strptime(str(s).strip(),"%d/%m/%Y").date()
    except: return None
# mots génériques (non distinctifs)
GEN=set("biere bieres vin vins cidre rose sirop jus creme bouteille carton caisse fut futs futes b cl blanc blanche blonde rouge doux brut de la le du des page boites boite verre bib caissette gris neige".split())
SYN={"mortre":"mordue","veran":"veran","jaune":"jaune","bethmale":"bethmale","jamelle":"jamelles","jamelles":"jamelles",
     "gassier":"gassier","sandrin":"sandrin","campbell":"campbell","macon":"macon","mure":"mure","cerise":"cerise"}
def keytokens(art):
    out=[]
    for w in re.split(r"[^a-z0-9]+",na(art)):
        if not w or w in GEN or len(w)<4 or w.isdigit(): continue
        if re.fullmatch(r"\d+(cl|l|kg)?",w): continue   # tokens de format/volume (75cl, 8l...) non distinctifs
        out.append(SYN.get(w,w))
    return out
def parseqty(s):
    s=na(s); nums=[int(x) for x in re.findall(r"\d+",s)]
    if not nums: return None
    fut="fut" in s
    if fut: return nums[0]                       # nb de fûts
    if any(k in s for k in("carton","caisse"," x ","boites","boite")): return nums[-1]  # nb d'unités dans le contenant
    return nums[0]

# carnet
sh=xlrd.open_workbook(BASE+"public/documents/vins-boissons/analyse-manuscrite/analyse-manuscrite-boissons.xls").sheet_by_index(0)
NONRECU={"en manque","facture non livre","reprise","retour","refuse","avoir"}
carnet=[]
for r in range(1,sh.nrows):
    statut=na(sh.cell_value(r,5)); dt=d(sh.cell_value(r,2))
    if statut in NONRECU:
        carnet.append(dict(page=str(sh.cell_value(r,0)).strip(),date=dt,qtetxt=str(sh.cell_value(r,3)).strip(),
            qte=parseqty(sh.cell_value(r,3)),art=str(sh.cell_value(r,4)).strip(),statut=str(sh.cell_value(r,5)).strip(),
            note=str(sh.cell_value(r,6)).strip(),toks=keytokens(str(sh.cell_value(r,4)))))

# factures (lignes avec date)
F=json.load(open("src/data/factures-fournisseur.json".replace("src",BASE+"src")))["factures"]
flines=[]
for f in F:
    fd=d(f.get("dateFacture"))
    for ln in f.get("lignes",[]):
        flines.append(dict(num=f.get("numero"),date=fd,type=f.get("type"),des=ln.get("designation",""),
            ndes=na(ln.get("designation","")),q=ln.get("quantite"),m=ln.get("montantHT")))

def tokmatch(toks,ndes): return any(re.search(r"\b"+re.escape(t)+r"\b",ndes) for t in toks)
def match(c):
    if not c["toks"] or not c["date"]: return [],"sans mot-clé/date -> à vérifier manuellement"
    cand=[fl for fl in flines if fl["date"] and abs((fl["date"]-c["date"]).days)<=21 and tokmatch(c["toks"],fl["ndes"])]
    # avoir : chercher une facture d'avoir correspondante sur fenêtre élargie (crédit souvent émis plus tard)
    avoir_wide=[fl for fl in flines if fl["type"]=="avoir" and fl["date"] and abs((fl["date"]-c["date"]).days)<=60 and tokmatch(c["toks"],fl["ndes"])]
    if not cand and not avoir_wide: return [],"aucune facture trouvée (±21j) -> à vérifier"
    none_lines=[x for x in cand if x["m"] is None and x["type"]=="facture"]
    charged=[x for x in cand if isinstance(x["m"],(int,float)) and x["m"]>0 and x["type"]=="facture"]
    if avoir_wide or [x for x in cand if x["type"]=="avoir"]: return (avoir_wide or cand),"AVOIR présent (déjà crédité)"
    if none_lines: return cand,"DÉJÀ EXCLU (ligne facturée sans montant)"
    if charged: return cand,"FACTURÉ PLEIN (montant présent) -> à examiner / déduire"
    return cand,"trouvé (autre)"

rows=[]
for c in carnet:
    cand,verdict=match(c)
    best=cand[0] if cand else None
    rows.append(dict(c=c,verdict=verdict,best=best,n=len(cand)))
# stats
stat=collections.Counter(r["verdict"].split(" ->")[0].split(" (")[0] for r in rows)
print("Anomalies carnet rapprochées :",len(rows))
for k,v in stat.most_common(): print(f"  {v:3d}  {k}")

# json
out=[]
for r in rows:
    c=r["c"]; b=r["best"]
    out.append({"date":str(c["date"]),"quantite_texte":c["qtetxt"],"quantite_estimee":c["qte"],"article":c["art"],
        "statut":c["statut"],"note":c["note"],"verdict":r["verdict"],"nb_factures_candidates":r["n"],
        "facture":({"numero":b["num"],"date":str(b["date"]),"type":b["type"],"designation":b["des"][:60],"qte":b["q"],"montantHT":b["m"]} if b else None)})
json.dump({"description":"Rapprochement carnet manuscrit (anomalies fournisseur) <-> factures. Artefact de relecture, sans déduction automatique. verdict = traitement déjà appliqué dans achatsBoissonsParPeriode.","nb_anomalies":len(out),"anomalies":out},open(JS+"rapprochement-carnet-factures.json","w",encoding="utf-8"),ensure_ascii=False,indent=2)
print("-> rapprochement-carnet-factures.json")

# xlsx
HEAD=PatternFill("solid",fgColor="1F4E78");HF=Font(bold=True,color="FFFFFF",size=9)
C_OK=PatternFill("solid",fgColor="E2EFDA");C_AV=PatternFill("solid",fgColor="DDEBF7");C_WARN=PatternFill("solid",fgColor="FCE4D6");C_GREY=PatternFill("solid",fgColor="F2F2F2")
thin=Side("thin",color="BFBFBF");B=Border(thin,thin,thin,thin);ctr=Alignment("center",vertical="center");lft=Alignment("left",vertical="center")
wb=Workbook();ws=wb.active;ws.title="Carnet vs factures"
cols=["Date carnet","Qté (texte)","Qté est.","Article (carnet)","Statut","Note","VERDICT (traitement)","Facture n°","Date facture","Type","Désignation facture","Qté fact.","Montant HT"]
ws.merge_cells("A1:M1");ws["A1"]="DEMI LUNE - Rapprochement carnet manuscrit (anomalies fournisseur) ↔ factures";ws["A1"].font=Font(bold=True,size=13)
ws.merge_cells("A2:M2");ws["A2"]=("Une ligne par anomalie du carnet (En manque, Avoir, Reprise, Retour, Refusé, Facturé non livré). VERDICT = traitement déjà appliqué : "
 "'DÉJÀ EXCLU' = ligne facturée sans montant, retirée des achats ; 'AVOIR présent' = crédité ; 'FACTURÉ PLEIN -> à examiner' = montant présent, pas encore déduit (candidat à correction) ; "
 "'aucune facture trouvée' / 'sans mot-clé' = à vérifier manuellement (carnet manuscrit). Appariement par mot-clé + fenêtre ±21 jours : À RELIRE.")
ws["A2"].font=Font(italic=True,size=9);ws["A2"].alignment=Alignment(wrap_text=True,vertical="center");ws.row_dimensions[2].height=56
HR=4
for j,h in enumerate(cols,1):
    cc=ws.cell(HR,j,h);cc.fill=HEAD;cc.font=HF;cc.border=B;cc.alignment=Alignment("center",wrap_text=True,vertical="center")
r=HR+1
for x in rows:
    c=x["c"]; b=x["best"]
    row=[str(c["date"]),c["qtetxt"],c["qte"],c["art"],c["statut"],c["note"],x["verdict"],
         b["num"] if b else "",str(b["date"]) if b else "",b["type"] if b else "",(b["des"][:55] if b else ""),(b["q"] if b else ""),(b["m"] if b else "")]
    fill = C_WARN if "FACTURÉ PLEIN" in x["verdict"] else (C_AV if "AVOIR" in x["verdict"] else (C_OK if "DÉJÀ EXCLU" in x["verdict"] else C_GREY))
    for j,v in enumerate(row,1):
        cc=ws.cell(r,j,v);cc.border=B;cc.alignment=lft if j in(4,6,7,11) else ctr
        if j==7: cc.fill=fill
    r+=1
ws.freeze_panes="A5";ws.auto_filter.ref=f"A{HR}:M{r-1}"
for i,wd in enumerate([12,14,7,22,16,18,34,9,11,8,40,8,10],1): ws.column_dimensions[Lc(i)].width=wd
wb.save(JS+"rapprochement-carnet-factures.xlsx");print("-> rapprochement-carnet-factures.xlsx")
