#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Pipeline consommation d'alcool Demi Lune.
Régénère :
  - conso-exacte-carte.xlsx        (alcool exact à la carte)
  - conso-cumul-carte-menus.xlsx   (cumul exact + estimation menus bas/moyen/haut)
Sources : ANNEXE-D (synthèse produit, qtés) + ANNEXE-C (tickets, menus).
Lancer :  /tmp/xlsenv/bin/python _pipeline_conso_alcool.py
"""
import xlrd, collections
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter as L

CAISSE="/Users/maevapasteur/Documents/demi-lune-comptabilite/public/documents/caisse-enregistreuse/"
OUT="/Users/maevapasteur/Documents/demi-lune-comptabilite/public/documents/vins-boissons/"
D_FILES={"2022-2023":CAISSE+"ANNEXE-D1_synthese-produit_2022-2023.xls","2023-2024":CAISSE+"ANNEXE-D2_synthese-produit_2023-2024.xls","2024-2025":CAISSE+"ANNEXE-D3_synthese-produit_2024-2025.xls"}
C_FILES={"2022-2023":CAISSE+"ANNEXE-C1_detail-tickets_2022-2023.xls","2023-2024":CAISSE+"ANNEXE-C2_detail-tickets_2023-2024.xls","2024-2025":CAISSE+"ANNEXE-C3_detail-tickets_2024-2025.xls"}
EXOS=["2022-2023","2023-2024","2024-2025"]
def norm(s): return str(s).replace("\xa0"," ").strip()
BEV_KEYS=["vins rouge","vins blanc","vins rosé","vins rose","pichet","pétillant","petillant","digestif","apéritifs maison","aperitifs maison","spécialités","specialites","autres apéritifs","autres aperitifs","cockt","bière","biere","autres :"]

# (type, dose_cl, estimation)
FOOD={
 "POULET JURASSIENNE":[("Vin jaune",1,0)],"TRUITE JURASSIENNE":[("Vin jaune",1,0)],
 "Poulet aux Morilles":[("Porto",1,0)],"Truite Morilles":[("Porto",1,0)],
 "ENTRECOTE MORILLES":[("Porto",1,0)],"Entrecôte Forestiére":[("Porto",1,0)],
 "Croûtes Morilles":[("Porto",1,0)],"Feuilleté Forestier":[("Porto",1,0)],
 "PAYSANNE":[("Porto",1,0)],"MACHON GOURMET":[("Porto",1,0)],
 "Fondue Jurassienne":[("Ravelin",10,0)],"FONDUE VIN JAUNE":[("Vin jaune",10,0)],
 "Fondue des Gourmets":[("Vin jaune",10,0)],"Fondue Forestière":[("Ravelin",10,0)],
 "FRELEE":[("Porto",1,0),("Calvados",4,0)],"CANCOINE":[("Porto",1,0)],"PEUTE":[("Vin jaune",1,0)],
 "GALETTE FORESTIERE":[("Porto",1,0)],
 "ASSIETTE FRANC-COMTO":[("Macvin",4,0)],"ASSIETTE PERE GREGOI":[("Calvados",4,0)],
 "CAMEMBERT ROTI":[("Calvados",4,0)],
 "SAUCE JURASSIENNE":[("Vin jaune",1,0)],"Sauce Forestière":[("Porto",1,0)],"Sauce Morilles":[("Porto",1,0)],
 "Baba au Macvin":[("Macvin",4,0)],"Baba":[("Macvin",4,0)],
 "IVRESSE":[("Baileys",4,0)],"FOLIE":[("Crème de cassis",4,0)],
 "GRAPPINS":[("Calvados",4,0)],"VOGEOTTE":[("Grand Marnier",4,0)],"BASILIC":[("Calvados",4,0)],
 "Crêpe Flambée":[("Grand Marnier ou Calvados",4,1)],"Flambée":[("Grand Marnier ou Calvados",4,1)],
}
DESSERT_KW=["Baba","IVRESSE","FOLIE","GRAPPINS","VOGEOTTE","BASILIC","Flamb"]
# Boissons saisies dans la section SOLIDE (cocktail rangé en solide)
SOLIDE_DRINK={"VOUIVRE":("Apéritif maison","Cocktail","Au verre",12,[("Crémant",10,1),("Macvin",1,1),("Crème de cassis",1,1)])}

BEV={
 "La Vouivre":("Apéritif maison","Cocktail","Au verre",12,[("Crémant",10,1),("Macvin",1,1),("Crème de cassis",1,1)]),
 "Père Gregoire":("Apéritif maison","Cocktail","Au verre",6,[("Macvin",4,1),("Liqueur de cerise",2,1)]),
 "Chat perché":("Apéritif maison","Cocktail","Au verre",12,[("Macvin",6,1)]),
 "KITTYKIR":("Apéritif maison","Cocktail","Au verre",12,[("Crémant",9,1),("Liqueur de litchi",2,1)]),
 "BALIDOU":("Cocktail","Cocktail","Au verre",25,[("Passoa",4,0)]),
 "MAEVA":("Cocktail","Cocktail","Au verre",25,[("Liqueur de litchi",4,0),("Vodka",4,0)]),
 "Rabasse":("Cocktail","Cocktail","Au verre",25,[("Pontarlier",2,0),("Liqueur de sapin",2,0)]),
 "Rêve Bleu":("Cocktail","Cocktail","Au verre",25,[("Vodka",4,0)]),
 "TEQUILA SUNRISE":("Cocktail","Cocktail","Au verre",25,[("Tequila",4,0)]),
 "Macvin du Jura":("Spécialité","Verre","Au verre",6,[("Macvin",6,0)]),
 "Vin de Paille":("Spécialité","Verre","Au verre",6,[("Vin de paille",6,0)]),
 "VIN PAILLE":("Spécialité","Verre","Au verre",6,[("Vin de paille",6,0)]),
 "Vin jaune":("Spécialité","Verre","Au verre",12,[("Vin jaune",12,0)]),
 "Pontarlier":("Spécialité","Verre","Au verre",2,[("Pontarlier",2,0)]),
 "Pastis":("Apéritif","Verre","Au verre",2,[("Pastis",2,0)]),"Ricard":("Apéritif","Verre","Au verre",2,[("Ricard",2,0)]),
 "Porto":("Apéritif","Verre","Au verre",6,[("Porto",6,0)]),"Martini Blanc":("Apéritif","Verre","Au verre",6,[("Martini",6,0)]),
 "Kir Bourgogne":("Apéritif","Kir","Au verre",12,[("Aligoté",10,1),("Crème de cassis",2,1)]),
 "Kir Princier":("Apéritif","Kir","Au verre",12,[("Crémant",10,1),("Crème de cassis",2,1)]),
 "Kir Pamplemousse":("Apéritif","Kir","Au verre",12,[("Aligoté",10,1)]),
 "Rosé Pamplemousse":("Apéritif","Kir","Au verre",12,[("Vin rosé",10,1)]),
 "Jack Daniel's":("Apéritif","Verre","Au verre",4,[("Whisky",4,0)]),
 "Whisky L.J (4cl)":("Apéritif","Verre","Au verre",4,[("Whisky",4,0)]),
 "Whisky L.J (2cl) Bab":("Apéritif","Verre","Au verre",2,[("Whisky",2,0)]),
 "Whisky-Coca":("Apéritif","Verre","Au verre",4,[("Whisky",4,0)]),
 "Café  Irlandais":("Autre","Verre","Au verre",4,[("Whisky",4,0)]),
 "Guignolet":("Digestif","Verre","Au verre",4,[("Guignolet",4,1)]),
 "Baileys":("Digestif","Verre","Au verre",4,[("Baileys",4,0)]),"BAILEYS":("Digestif","Verre","Au verre",4,[("Baileys",4,0)]),
 "Calvados":("Digestif","Verre","Au verre",4,[("Calvados",4,0)]),"Calvados (4cl)":("Digestif","Verre","Au verre",4,[("Calvados",4,0)]),
 "Cognac (4cl)":("Digestif","Verre","Au verre",4,[("Cognac",4,0)]),
 "Eau de Vie de Poire":("Digestif","Verre","Au verre",4,[("Eau de vie poire",4,0)]),
 "Eau de vie Framboise":("Digestif","Verre","Au verre",4,[("Eau de vie framboise",4,0)]),
 "Eau de vie Mirabelle":("Digestif","Verre","Au verre",4,[("Eau de vie mirabelle",4,0)]),
 "Grand Marnier":("Digestif","Verre","Au verre",4,[("Grand Marnier",4,0)]),"GRAND MARNIER":("Digestif","Verre","Au verre",4,[("Grand Marnier",4,0)]),
 "Get 27 (4cl)":("Digestif","Verre","Au verre",4,[("Get 27",4,0)]),"Get 31":("Digestif","Verre","Au verre",4,[("Get 31",4,0)]),
 "Génépi":("Digestif","Verre","Au verre",4,[("Génépi",4,0)]),
 "Liqueur de Poire":("Digestif","Verre","Au verre",4,[("Liqueur de poire",4,0)]),
 "Liqueur de Sapin":("Digestif","Verre","Au verre",4,[("Liqueur de sapin",4,0)]),
 "Marc du Jura (4cl)":("Digestif","Verre","Au verre",4,[("Marc du Jura",4,0)]),
 "Vieux Marc Bourgogne":("Digestif","Verre","Au verre",4,[("Marc de bourgogne",4,0)]),
 "TEQUILA":("Digestif","Verre","Au verre",4,[("Tequila",4,0)]),
 "pression":("Bière","Pression","Au verre",25,[("Bière",25,0)]),
 "Pinte":("Bière","Pinte","Au verre",50,[("Bière",50,0)]),
 "Panaché":("Bière","Panaché","Au verre",25,[("Bière",12.5,0)]),
 "Monaco":("Bière","Monaco","Au verre",25,[("Bière",12.5,1)]),
 "Picon bière":("Bière","Demi+Picon","Au verre",25,[("Bière",21,0),("Picon",4,0)]),
 "Pinte Picon":("Bière","Pinte+Picon","Au verre",50,[("Bière",42,0),("Picon",8,0)]),
 "Ambrée":("Bière","Bouteille 33cl","Contenant entier",33,[("Bière",33,0)]),
 "Blanche des plateaux":("Bière","Bouteille 33cl","Contenant entier",33,[("Bière",33,0)]),
 "Crémant":("Pétillant","Verre","Au verre",10,[("Crémant",10,0)]),
 "Crément / Jura VERRE":("Pétillant","Verre","Au verre",10,[("Crémant",10,0)]),
 "Crémant du Jura":("Pétillant","Bouteille","Contenant entier",75,[("Crémant",75,0)]),
 "Crément du Jura":("Pétillant","Bouteille","Contenant entier",75,[("Crémant",75,0)]),
 "Crémant Rosé":("Pétillant","Bouteille","Contenant entier",75,[("Crémant",75,0)]),
 "Champagne":("Pétillant","Verre","Au verre",12,[("Champagne",12,1)]),
 '"Champagne Brut "':("Pétillant","Bouteille","Contenant entier",75,[("Champagne",75,1)]),
 "Cidre Bouché Brut":("Pétillant","Bouteille","Contenant entier",75,[("Cidre",75,0)]),
 "Cidre Bouché Doux":("Pétillant","Bouteille","Contenant entier",75,[("Cidre",75,0)]),
 "cidre la Mordue":("Pétillant","Bouteille 33cl","Contenant entier",33,[("Cidre",33,0)]),
}
UNKNOWN={"Do","Ré","Mi","Fa","Sol","La","Si"}  # à définir, non comptés

def classify_wine(section,label):
    lab=label.lower(); sec=section.lower()
    if "1/2" in lab: fmt,cl,mode=("Demi-bouteille",37.5,"Contenant entier")
    elif "75" in lab: fmt,cl,mode=("Pichet 75cl",75,"Contenant entier")
    elif "pichet" in lab: fmt,cl,mode=("Pichet 50cl",50,"Contenant entier")
    elif "btl" in lab or "bou" in lab or "bouteille" in lab: fmt,cl,mode=("Bouteille",75,"Contenant entier")
    elif "verre" in lab or lab.endswith(" le") or "le ve" in lab or lab.endswith(" ve"): fmt,cl,mode=("Verre",15,"Au verre")
    else: fmt,cl,mode=("Bouteille",75,"Contenant entier")
    if "savagnin" in lab: typ="Savagnin"
    elif "aligot" in lab: typ="Aligoté"
    elif "chusclan" in lab or "chuslan" in lab: typ="Vin rouge"
    elif any(k in lab for k in ["rosé","rose","minuty","miraval","pive","bohème","boheme","clair de ros"]): typ="Vin rosé"
    elif "rouge" in sec: typ="Vin rouge"
    elif "blanc" in sec: typ="Vin blanc"
    elif "rosé" in sec or "rose" in sec: typ="Vin rosé"
    else: typ="Vin (couleur n.p.)"
    return ("Vin",fmt,mode,cl,[(typ,cl,0)])

def build_carte():
    rows=[]; unknown=[]
    for exo,fn in D_FILES.items():
        sh=xlrd.open_workbook(fn).sheet_by_index(0); section=None; kind=None
        for r in range(1,sh.nrows):
            a=norm(sh.cell_value(r,0)); b=norm(sh.cell_value(r,1))
            if a and not b and a!="Ref_prd":
                section=a; low=a.lower()
                if "SOLIDE" in a.upper(): kind="food"
                elif any(k in low for k in BEV_KEYS) and "sans alcool" not in low: kind="bev"
                else: kind=None
            elif kind and b and b!="Lib_ticket":
                try: q=float(sh.cell_value(r,4))
                except: q=0
                if q<=0: continue
                if kind=="food" and b in SOLIDE_DRINK:
                    fam,fmt,mode,drink,alcs=SOLIDE_DRINK[b]; first=True
                    for (typ,dose,est) in alcs:
                        rows.append(dict(exo=exo,section=section,produit=b,famille=fam,fmt=fmt,mode=mode,qte=q,unit=dose,typ=typ,alc=q*dose,est=est,volshow=(q*drink if first else "")))
                        first=False
                elif kind=="food":
                    if b not in FOOD: continue
                    fam="Dessert" if any(k in b for k in DESSERT_KW) else "Plat"
                    for (typ,dose,est) in FOOD[b]:
                        rows.append(dict(exo=exo,section="SOLIDE (à la carte)",produit=b,famille=fam,fmt="Cuisine",mode="Cuisine (plat/dessert)",qte=q,unit=dose,typ=typ,alc=q*dose,est=est,volshow=""))
                else:
                    if b in UNKNOWN: unknown.append(dict(exo=exo,produit=b,qte=q)); continue
                    if b in BEV: fam,fmt,mode,drink,alcs=BEV[b]
                    elif any(k in section.lower() for k in ["vin","pichet","rosé","rose"]): fam,fmt,mode,drink,alcs=classify_wine(section,b)
                    else: unknown.append(dict(exo=exo,produit=f"[NON MAPPÉ] {section} / {b}",qte=q)); continue
                    first=True
                    for (typ,dose,est) in alcs:
                        rows.append(dict(exo=exo,section=section,produit=b,famille=fam,fmt=fmt,mode=mode,qte=q,unit=dose,typ=typ,alc=q*dose,est=est,volshow=(q*drink if first else "")))
                        first=False
    return rows,unknown

# ---------- MENUS ----------
ACCEPT={"Menu 'Demi Lune'":{45.0},"Menu du Dahu":{29.9},"Menu Végétarien":{23.0,25.0,26.5},"Menu Bourguignon":{32.0,34.0},"Menu Franc-Comtois":{32.0,34.0}}
KEY={"Menu 'Demi Lune'":"Demi Lune","Menu du Dahu":"Dahu","Menu Végétarien":"Végétarien","Menu Bourguignon":"Bourguignon","Menu Franc-Comtois":"Franc-Comtois"}
COMP={
"A":{"Demi Lune":[("Porto",1,1,1),("Porto",1,.33,.60),("Baileys",4,.15,.35),("Liqueur de poire",4,.15,.35)],"Dahu":[("Ravelin",10,1,1)],"Végétarien":[("Calvados",4,.35,.65),("Porto",1,.35,.65)],"Bourguignon":[("Crème de cassis",4,.35,.65),("Baileys",4,.35,.65)],"Franc-Comtois":[("Vin jaune",1,.35,.65),("Macvin",4,.35,.65)]},
"B":{"Demi Lune":[("Porto",1,1,1),("Porto",1,.33,.60)],"Dahu":[("Ravelin",10,1,1)],"Végétarien":[("Calvados",4,.35,.65),("Porto",1,.35,.65)],"Bourguignon":[("Crème de cassis",4,.35,.65),("Marc de bourgogne",4,.35,.65)],"Franc-Comtois":[("Vin jaune",1,.35,.65),("Macvin",4,.35,.65),("Macvin",4,.20,.45)]},
"C":{"Demi Lune":[("Porto",1,1,1),("Porto",1,.40,.65)],"Dahu":[("Ravelin",10,1,1)],"Végétarien":[("Calvados",4,.35,.65),("Porto",1,.35,.65)],"Bourguignon":[("Crème de cassis",4,.35,.65),("Marc de bourgogne",4,.35,.65)],"Franc-Comtois":[("Vin jaune",1,.35,.65),("Macvin",4,.35,.65),("Macvin",4,.20,.45)]}}
def per(d): return "C" if d<"2022-07-23" else ("B" if d<"2023-04-17" else "A")
def build_menus():
    menu=collections.defaultdict(lambda:[0.0,0.0,0.0])
    for exo,fn in C_FILES.items():
        sh=xlrd.open_workbook(fn).sheet_by_index(0)
        for r in range(1,sh.nrows):
            lib=str(sh.cell_value(r,10)).strip()
            if lib not in ACCEPT: continue
            prix=round(float(sh.cell_value(r,13)),2)
            if prix not in ACCEPT[lib]: continue
            q=float(sh.cell_value(r,11)); p=per(str(sh.cell_value(r,0))[:10])
            for (typ,dose,pb,ph) in COMP[p][KEY[lib]]:
                pm=(pb+ph)/2
                menu[(exo,typ)][0]+=q*dose*pb; menu[(exo,typ)][1]+=q*dose*pm; menu[(exo,typ)][2]+=q*dose*ph
    return menu

# ---------- STYLES ----------
HEAD=PatternFill("solid",fgColor="1F4E78"); HF=Font(bold=True,color="FFFFFF",size=10)
SUB=PatternFill("solid",fgColor="DDEBF7"); GT=PatternFill("solid",fgColor="F4B183")
thin=Side("thin",color="BFBFBF"); B=Border(thin,thin,thin,thin)
ctr=Alignment("center",vertical="center"); lft=Alignment("left",vertical="center"); bold=Font(bold=True)

def gen_file1(rows,unknown):
    wb=Workbook(); ws=wb.active; ws.title="Conso exacte carte"
    ws.merge_cells("A1:L1"); ws["A1"]="DEMI LUNE - Consommation EXACTE d'alcool enregistrée en caisse (ventes à la carte : boissons, plats et desserts alcoolisés)"; ws["A1"].font=Font(bold=True,size=13)
    ws.merge_cells("A2:L2"); ws["A2"]=("Source : ANNEXE-D synthèse produit (qté vendue par produit/exercice). Hors menus (estimés à part). cl = volume servi pour vins/bières, dose d'alcool pour cocktails/plats. Mode 'Au verre' vs 'Contenant entier'. Estim.=Oui : dose/type à confirmer. Do..Si non comptés.")
    ws["A2"].font=Font(italic=True,size=9); ws["A2"].alignment=Alignment(wrap_text=True,vertical="center"); ws.row_dimensions[2].height=48
    ws["A4"]="Conventions : verre vin 15cl | pichet 50/75cl | 1/2 37,5cl | bouteille 75cl | pression 25cl | pinte 50cl | panaché/monaco 12,5cl bière | bière bouteille 33cl | crémant verre 10cl | vin jaune 12cl | picon bière 21cl+4cl, pinte picon 42cl+8cl | apéritifs/digestifs 2-6cl | cocktails : dose."
    ws.merge_cells("A4:L4"); ws["A4"].font=Font(italic=True,size=8,color="555555"); ws["A4"].alignment=Alignment(wrap_text=True); ws.row_dimensions[4].height=24
    cols=["Exercice","Section caisse","Produit","Famille","Format","Mode de service","Qté vendue","Dose/contenance unit. (cl)","Type d'alcool","Estim.","Alcool total (cl)","Volume boisson total (cl)"]
    HR=6
    for j,h in enumerate(cols,1):
        c=ws.cell(HR,j,h); c.fill=HEAD; c.font=HF; c.alignment=Alignment("center",vertical="center",wrap_text=True); c.border=B
    ws.row_dimensions[HR].height=30
    rows.sort(key=lambda x:(x["exo"],x["famille"],x["typ"],x["produit"]))
    r=HR+1
    for x in rows:
        vals=[x["exo"],x["section"],x["produit"],x["famille"],x["fmt"],x["mode"],x["qte"],x["unit"],x["typ"],"Oui" if x["est"] else "",x["alc"],x["volshow"]]
        for j,v in enumerate(vals,1):
            c=ws.cell(r,j,v); c.border=B; c.alignment=lft if j in (2,3,5,6,9) else ctr
            if j in (7,8,11,12): c.number_format="0.0"
        r+=1
    DF,DL=HR+1,r-1
    ws.freeze_panes=f"A{HR+1}"; ws.auto_filter.ref=f"A{HR}:{L(len(cols))}{DL}"
    r+=2; ws.cell(r,1,"RÉCAPITULATIF PAR TYPE D'ALCOOL (Alcool en cl) - consommation EXACTE à la carte").font=Font(bold=True,size=12); r+=1
    rc=["Type d'alcool","2022-2023","2023-2024","2024-2025","TOTAL (cl)","TOTAL (L)","dont Au verre","dont Contenant","dont Cuisine"]
    for j,h in enumerate(rc,1):
        c=ws.cell(r,j,h); c.fill=HEAD; c.font=HF; c.alignment=Alignment("center",wrap_text=True,vertical="center"); c.border=B
    r+=1
    types=sorted({x["typ"] for x in rows}, key=lambda t:-sum(y["alc"] for y in rows if y["typ"]==t))
    AC=L(11); EC=L(1); TC=L(9); MC=L(6); rf=r
    for t in types:
        ws.cell(r,1,t).alignment=lft; ws.cell(r,1).border=B
        for k,exo in enumerate(EXOS):
            f=f'=SUMIFS(${AC}${DF}:${AC}${DL},${EC}${DF}:${EC}${DL},"{exo}",${TC}${DF}:${TC}${DL},$A{r})'
            c=ws.cell(r,2+k,f); c.number_format="0.0"; c.border=B; c.alignment=ctr
        ws.cell(r,5,f"=SUM(B{r}:D{r})").number_format="0.0"; ws.cell(r,5).border=B; ws.cell(r,5).font=bold; ws.cell(r,5).alignment=ctr
        ws.cell(r,6,f"=E{r}/100").number_format="0.0"; ws.cell(r,6).border=B; ws.cell(r,6).alignment=ctr
        for k,mode in enumerate(["Au verre","Contenant entier","Cuisine (plat/dessert)"]):
            f=f'=SUMIFS(${AC}${DF}:${AC}${DL},${MC}${DF}:${MC}${DL},"{mode}",${TC}${DF}:${TC}${DL},$A{r})'
            c=ws.cell(r,7+k,f); c.number_format="0.0"; c.border=B; c.alignment=ctr
        r+=1
    rl=r-1; ws.cell(r,1,"TOTAL").font=bold; ws.cell(r,1).fill=GT; ws.cell(r,1).border=B
    for j in range(2,10):
        ws.cell(r,j,f"=SUM({L(j)}{rf}:{L(j)}{rl})"); ws.cell(r,j).number_format="0.0"; ws.cell(r,j).fill=GT; ws.cell(r,j).border=B; ws.cell(r,j).font=bold; ws.cell(r,j).alignment=ctr
    r+=2; ws.cell(r,1,"PRODUITS NON COMPTÉS (composition inconnue - voir A_VERIFIER_produits-inconnus.md)").font=Font(bold=True,size=11,color="C00000"); r+=1
    un=collections.defaultdict(lambda: collections.defaultdict(float))
    for u in unknown: un[u["produit"]][u["exo"]]+=float(u["qte"])
    ws.cell(r,1,"Produit").font=bold
    for k,e in enumerate(EXOS): ws.cell(r,2+k,e).font=bold
    ws.cell(r,5,"Qté totale").font=bold
    for j in range(1,6): ws.cell(r,j).fill=SUB; ws.cell(r,j).border=B
    r+=1
    for p in sorted(un):
        ws.cell(r,1,p).border=B; tt=0
        for k,e in enumerate(EXOS):
            q=un[p].get(e,0); ws.cell(r,2+k,q).border=B; ws.cell(r,2+k).alignment=ctr; tt+=q
        ws.cell(r,5,tt).border=B; ws.cell(r,5).alignment=ctr; r+=1
    for i,w in enumerate([11,22,22,10,15,20,9,12,18,7,13,16],1): ws.column_dimensions[L(i)].width=w
    wb.save(OUT+"conso-exacte-carte.xlsx")
    return {(x["exo"],x["typ"]):0 for x in rows}

def gen_file2(rows,menu):
    carte=collections.defaultdict(float)
    for x in rows: carte[(x["exo"],x["typ"])]+=float(x["alc"])
    alltypes=sorted(set(t for (_,t) in carte)|set(t for (_,t) in menu))
    wb=Workbook(); ws=wb.active; ws.title="Cumul exact + menus"
    ws.merge_cells("A1:J1"); ws["A1"]="DEMI LUNE - Consommation TOTALE d'alcool : exact (à la carte) + estimé (menus, fourchette bas/moyen/haut)"; ws["A1"].font=Font(bold=True,size=13)
    ws.merge_cells("A2:J2"); ws["A2"]=("EXACT = à la carte (composition connue). MENUS = estimation (options variables) bas/moyen/haut, recette selon carte A/B/C à la date. TOTAL = Exact + Menus. Menus à prix personnalisé exclus. cl.")
    ws["A2"].font=Font(italic=True,size=9); ws["A2"].alignment=Alignment(wrap_text=True,vertical="center"); ws.row_dimensions[2].height=40
    cols=["Type d'alcool","Exercice","Exact carte (cl)","Menus bas (cl)","Menus moyen (cl)","Menus haut (cl)","TOTAL bas (cl)","TOTAL moyen (cl)","TOTAL haut (cl)","TOTAL moyen (L)"]
    HR=4
    for j,h in enumerate(cols,1):
        c=ws.cell(HR,j,h); c.fill=HEAD; c.font=HF; c.alignment=Alignment("center",wrap_text=True,vertical="center"); c.border=B
    ws.row_dimensions[HR].height=30; r=HR+1; gtot=[0.0]*7
    for t in alltypes:
        tl=[0.0]*7
        for exo in EXOS:
            ca=carte.get((exo,t),0.0); mb,mm,mh=menu.get((exo,t),[0,0,0])
            vals=[t,exo,ca,mb,mm,mh,ca+mb,ca+mm,ca+mh,(ca+mm)/100]
            for j,v in enumerate(vals,1):
                c=ws.cell(r,j,v); c.border=B; c.alignment=lft if j<=2 else ctr
                if j>=3: c.number_format="0.0"
            for k,v in enumerate([ca,mb,mm,mh,ca+mb,ca+mm,ca+mh]): tl[k]+=v
            r+=1
        for k in range(7): gtot[k]+=tl[k]
        vals=[t,"TOTAL "+t]+tl+[tl[5]/100]
        for j,v in enumerate(vals,1):
            c=ws.cell(r,j,v); c.border=B; c.fill=SUB; c.font=bold; c.alignment=lft if j<=2 else ctr
            if j>=3: c.number_format="0.0"
        r+=1
    vals=["TOTAL GÉNÉRAL",""]+gtot+[gtot[5]/100]
    for j,v in enumerate(vals,1):
        c=ws.cell(r,j,v); c.border=B; c.fill=GT; c.font=Font(bold=True,size=11); c.alignment=lft if j<=2 else ctr
        if j>=3: c.number_format="0.0"
    for i,w in enumerate([20,12,15,14,15,15,14,15,15,14],1): ws.column_dimensions[L(i)].width=w
    ws.freeze_panes="A5"; wb.save(OUT+"conso-cumul-carte-menus.xlsx")
    return gtot

if __name__=="__main__":
    rows,unknown=build_carte()
    menu=build_menus()
    gen_file1(rows,unknown)
    g=gen_file2(rows,menu)
    tot=sum(x["alc"] for x in rows)
    print(f"OK. Exact carte: {tot/100:.0f} L | TOTAL avec menus: bas {g[4]/100:.0f} / moyen {g[5]/100:.0f} / haut {g[6]/100:.0f} L")
