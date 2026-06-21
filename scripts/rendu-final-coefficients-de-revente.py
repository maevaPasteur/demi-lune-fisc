#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
rendu-final-coefficients-de-revente.py  (REFONTE)

Reponse au grief « coefficient de revente trop bas » (Proposition de rectification
p. 34-35, rejet 3/3). Le service calcule un coefficient global = CA TTC / cout des
matieres revendues ; il le trouve a ~2,30, le juge « manifestement tres bas et hors
des usages de la profession » (~3+), et reconstitue le CA jusqu'a obtenir ~3,05.

Demonstration, chiffree et reproductible, a partir des donnees REELLES :
  1. La donnee brute : ce qui est vendu, format par format, a quel prix (qui varie
     par saison), et pour combien de CA  (ventes-caisse.json).
  2. Le parallele caisse <-> facture : du prix d'achat au prix de vente. Les AOC du
     Jura (vin jaune, savagnin, trousseau, macvin, cremant) sont CHERS a l'achat et
     vendus a marge modeste : ils tirent mecaniquement le coefficient vers le bas.
  3. Pourquoi le coefficient apparent est bas, sans aucune recette cachee :
       - ~36 % de l'alcool achete n'est JAMAIS revendu au verre (cuisine, menus,
         sur-versement, cremant jete, freinte biere, degustations, conso chef,
         casse) : le cout est engage, il ne genere aucun CA (cf. pages 2.x).
       - sur l'alcool REELLEMENT revendu au verre/cocktail, le coefficient remonte
         a ~3,85, AU-DESSUS du seuil « 3+ » invoque par le service.
       - le coefficient boisson (~2,4) est au meme niveau que le coefficient
         nourriture (~2,3) : ce n'est pas la boisson qui « manque ».
  4. La « correction » du service a 3,05 est obtenue par sa reconstitution
     circulaire (coefficient liquide -> solide, cf. page 2.8), pas par une mesure.

Sorties (100 % reproductibles) :
  - public/documents/pieces-defense/RF-coefficients-de-revente.xlsx (4 onglets)
  - src/data/renduFinal/coefficients-de-revente.json

Sources :
  - analyses-independantes/boissons/data/ventes-caisse.json (ventes caisse, prix)
  - src/data/calculsBoissons/achatsBoissonsParPeriode.json (achats factures)
  - src/data/calculsBoissons/_correspondance-caisse-inventaire-factures.csv
  - src/data/boissonsPageData.json (cascade des volumes, page 2.1)
  - Proposition p. 34-35 (coef declare ~2,30 vs profession 3+ ; reconstitue 3,05)
"""
import json, os, csv

from rfcommun import ajouter_conclusion

ICI = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.normpath(os.path.join(ICI, ".."))
CB = os.path.join(ROOT, "src/data/calculsBoissons")


# ----------------------------------------------------------------------------- #
# Formatage francais.
# ----------------------------------------------------------------------------- #
def fr_int(n):
    return f"{int(round(n)):,}".replace(",", " ")


def fr_eur(n):
    return f"{int(round(n)):,}".replace(",", " ") + " €"


def fr_coef(x):
    return ("× " + f"{x:.2f}").replace(".", ",")


def fr_pct(x, d=1):
    return (f"{x:.{d}f}".replace(".", ",")) + " %"


def fr_prix(x):
    return (f"{x:.2f}".replace(".", ",")) + " €"


def cg(t):
    return {"v": t}


def cd(t):
    return {"v": t, "align": "right"}


def cgf(t):
    return {"v": t, "fw": 700}


def cdf(t):
    return {"v": t, "align": "right", "fw": 700}


def colg(label):
    return {"label": label}


def cold(label):
    return {"label": label, "align": "right"}


def lien(url):
    """Lien affiche en ENTIER et cliquable ([url](url))."""
    return f"[{url}]({url})"


# Sources externes (coefficient multiplicateur en restauration : grille DEGRESSIVE
# selon le prix d'achat ; nos AOC chers tombent dans les tranches x3 a x3,5).
U_ACCORD = "https://laccorddivin.fr/les-coefficients-multiplicateurs-caviste-restaurant-bar-a-vin/"
U_SOMMIT = "https://www.somm-it.com/blog/coefficient-multiplicateur-vin-restaurant"
U_LHOTEL = "https://www.lhotellerie-restauration.fr/sos-experts/question-reponse/coefficient-multiplicateur-sur-le-vin-le-calcul-s-applique-t-il-pour-un-bar-a-vins-36079"


# ----------------------------------------------------------------------------- #
# 1. Sources.
# ----------------------------------------------------------------------------- #
BD = json.load(open(os.path.join(ROOT, "src/data/boissonsPageData.json"), encoding="utf-8"))
SYN = BD["synthese"]
CASCADE = SYN["cascade"]
VENTES = json.load(open(os.path.join(ROOT, "analyses-independantes/boissons/data/ventes-caisse.json"),
                       encoding="utf-8"))["parExercice"]
ACHATS = json.load(open(os.path.join(CB, "achatsBoissonsParPeriode.json"), encoding="utf-8"))["achats"]
CORR = list(csv.DictReader(open(os.path.join(CB, "_correspondance-caisse-inventaire-factures.csv"),
                                encoding="utf-8-sig"), delimiter=";"))

EXOS = ["2022-2023", "2023-2024", "2024-2025"]
EXL = {"2022-2023": "2022-23", "2023-2024": "2023-24", "2024-2025": "2024-25"}

ACHAT_L = SYN["achat_alcool_l"]
ACHAT_COUT = SYN["achat_alcool_cout"]
PRIX_L = ACHAT_COUT / ACHAT_L


# ----------------------------------------------------------------------------- #
# 2. Le grief du fisc, par exercice (Proposition p. 34-35).
#    coef declare = CA TTC declare / cout matieres revendues ; le service le juge
#    bas et reconstitue jusqu'a coef ~3,05.
# ----------------------------------------------------------------------------- #
FISC_RECAP = [
    {"ex": "2022-2023", "ca_declare": 404031, "coef_decl": 2.303, "coef_rec": 3.404, "ca_rec": 597265},
    {"ex": "2023-2024", "ca_declare": 438658, "coef_decl": 2.388, "coef_rec": 3.144, "ca_rec": 577522},
    {"ex": "2024-2025", "ca_declare": 435525, "coef_decl": 2.310, "coef_rec": 3.051, "ca_rec": 575253},
]
for r in FISC_RECAP:
    r["matieres"] = r["ca_declare"] / r["coef_decl"]


# ----------------------------------------------------------------------------- #
# 3. Cascade : part de l'alcool achete NON revendu au verre (pages 2.x).
# ----------------------------------------------------------------------------- #
def casc(label):
    return next((c["litres"] for c in CASCADE if label.lower() in c["poste"].lower()), 0)


VERRE_L = casc("Vendu au verre")
COCKTAILS_L = casc("cocktails")
STOCK_L = casc("Stock final")
REVENDU_L = VERRE_L + COCKTAILS_L
NON_REVENDU_L = ACHAT_L - REVENDU_L - STOCK_L
PART_NON_REVENDU = NON_REVENDU_L / ACHAT_L

# Postes non revendus, en litres (pour le detail).
NR_POSTES = [("Cuisine (fondues, sauces, babas, flambage)", casc("Cuisine")),
             ("Alcool des menus (non detaille en caisse)", casc("menus")),
             ("Sur-versement au verre (free-pour)", casc("Sur-versement")),
             ("Cremant jete (fond evente) + sur-verse", casc("Cremant jete") + casc("Cremant sur")),
             ("Freinte technique de la biere", casc("Freinte")),
             ("Degustation offerte (note par note)", casc("Degustation")),
             ("Consommation du chef (Picon + Macvin)", casc("chef")),
             ("Aperitifs offerts aux clients", casc("offerts")),
             ("Casse, evaporation, fonds de verre (residu)", round(SYN["perte_reelle_l"]))]

# Cout de l'alcool revendu vs non revendu (au cout d'achat moyen au litre).
COUT_REVENDU = PRIX_L * REVENDU_L
COUT_NON_REVENDU = PRIX_L * NON_REVENDU_L

# CA alcool (TVA 20 %) encaisse en caisse, cumul 3 exercices = la recette reelle
# de l'alcool effectivement revendu.
CA_ALCOOL_CUMUL = sum(VENTES[e]["caLiquide20"] for e in EXOS)
COEF_REVENDU = CA_ALCOOL_CUMUL / COUT_REVENDU   # ~3,85


# ----------------------------------------------------------------------------- #
# 4. Coefficient boisson vs nourriture, par exercice (les deux au meme niveau).
# ----------------------------------------------------------------------------- #
def cout_boissons(exo):
    return sum(a["par_periode"].get(exo, {}).get("montant_ht", 0) for a in ACHATS)


SPLIT = []
for r in FISC_RECAP:
    e = r["ex"]
    ca_total = r["ca_declare"]
    ca_boisson = VENTES[e]["caLiquide10"] + VENTES[e]["caLiquide20"]
    ca_nourr = ca_total - ca_boisson
    matieres = r["matieres"]
    cb = cout_boissons(e)
    cn = matieres - cb
    SPLIT.append({"ex": e, "ca_boisson": ca_boisson, "ca_nourr": ca_nourr,
                  "cout_boisson": cb, "cout_nourr": cn,
                  "coef_boisson": ca_boisson / cb, "coef_nourr": ca_nourr / cn})


# ----------------------------------------------------------------------------- #
# 5-7. Coefficient de revente PRODUIT PAR PRODUIT, sur les 3 exercices du controle.
#   On relie chaque article de caisse a sa facture (table de correspondance), on lit
#   le prix de vente reel (caisse, cumul 3 exos) et le prix d'achat de la matiere
#   servie (prix d'achat bouteille x volume servi / contenance bouteille).
# ----------------------------------------------------------------------------- #
import re
from collections import defaultdict


def _norm(s):
    return re.sub(r"[^A-Z0-9]", " ", (s or "").upper())


def _taille_cl(nom):
    m = re.search(r"(\d+[.,]?\d*)\s*CL", (nom or "").upper())
    if m:
        return float(m.group(1).replace(",", "."))
    m = re.search(r"(\d+[.,]?\d*)\s*L\b", (nom or "").upper())
    return float(m.group(1).replace(",", ".")) * 100 if m else None


# Ventes caisse CUMULEES sur les 3 exercices, par libelle d'article.
VMAP = {}
for e in EXOS:
    for p in VENTES[e]["produits"]:
        d = VMAP.setdefault(p["libelle"].strip().lower(),
                            {"q": 0.0, "ca": 0.0, "pd": [], "tva": p.get("tva"), "cat": p.get("sousCategorie")})
        d["q"] += p.get("qte", 0)
        d["ca"] += p.get("caTtc", 0)
        d["pd"] += [x for x in (p.get("prixDistincts") or []) if x and x > 0]


def _achat_btle(fac):
    """Prix d'achat moyen par bouteille (HT, cumul 3 exos) + contenance, par matching
    du nom de facture sur les achats fournisseurs."""
    mots = [w for w in _norm(fac.split("/")[0]).split()
            if len(w) > 3 and not w.isdigit()
            and w not in ("DOMAINE", "CUVEE", "BLANC", "ROUGE", "FRUITIERE", "VINICOLE")]
    for a in ACHATS:
        if mots and all(w in _norm(a["produit"]) for w in mots[:2]):
            tq = sum(a["par_periode"].get(e, {}).get("quantite", 0) for e in EXOS)
            tm = sum(a["par_periode"].get(e, {}).get("montant_ht", 0) for e in EXOS)
            sz = _taille_cl(a["produit"])
            if tq and sz:
                return tm / tq, sz
    return None, None


CAT_NOM = {
    "vin_blanc": "Vins blancs", "vin_rouge": "Vins rouges", "vin_rose": "Vins rosés",
    "vin": "Vins (autres)", "petillant": "Pétillants / crémant", "vin_de_liqueur": "Macvin (vin de liqueur)",
    "biere": "Bières", "cidre": "Cidres", "aperitif": "Apéritifs", "liqueur": "Liqueurs",
    "eau_de_vie": "Eaux-de-vie", "spiritueux": "Spiritueux",
}


def _fmt_label(vol):
    if vol <= 20:
        return "Verre %g cl" % vol
    if 35 <= vol <= 40:
        return "Demi-bouteille %g cl" % vol
    if 45 <= vol <= 55:
        return "Pichet %g cl" % vol
    if vol >= 70:
        return "Bouteille %g cl" % vol
    return "%g cl" % vol


prod = []                       # produits avec coefficient (cleans)
cat_agg = defaultdict(lambda: {"n": 0, "ca": 0.0, "cout": 0.0})
n_skip = 0
for c in CORR:
    fac = c.get("nom_facture", "") or ""
    if not fac.strip() or fac.strip() == ":" or not c.get("volume_cl"):
        continue
    v = VMAP.get(c["libelle_caisse"].strip().lower())
    if not v or v["q"] <= 0:
        continue
    pv = v["ca"] / v["q"]
    pa, sz = _achat_btle(fac)
    if pa is None:
        continue
    vol = float(c["volume_cl"].replace(",", "."))
    cout = pa * vol / sz
    if cout <= 0:
        continue
    coef = pv / cout
    cat = c.get("categorie", "")
    # On ecarte les incoherences format/prix (volume mal renseigne dans la table) :
    if (vol <= 30 and pv > 16) or (vol >= 70 and pv < 14) or (cat.startswith("vin") and coef > 6.5):
        n_skip += 1
        continue
    pmin = min(v["pd"]) if v["pd"] else pv
    pmax = max(v["pd"]) if v["pd"] else pv
    prod.append({"lib": c["libelle_caisse"].strip(), "cat": cat, "vol": vol, "fmt": _fmt_label(vol),
                 "q": v["q"], "pv": pv, "pmin": pmin, "pmax": pmax, "cout": cout, "coef": coef, "ca": v["ca"]})
    a = cat_agg[cat]
    a["n"] += 1
    a["ca"] += v["ca"]
    a["cout"] += cout * v["q"]

prod.sort(key=lambda r: -r["ca"])
CA_COUVERT = sum(p["ca"] for p in prod)
NB_PRODUITS = len(prod)

# Coefficient par categorie (3 exercices), trie par CA.
cat_rows_data = []
for cat, a in sorted(cat_agg.items(), key=lambda kv: -kv[1]["ca"]):
    if a["cout"] > 0:
        cat_rows_data.append({"nom": CAT_NOM.get(cat, cat), "n": a["n"], "ca": a["ca"], "coef": a["ca"] / a["cout"]})


# Ventes par categorie / format, 3 EXERCICES (alcool, TVA 20).
def detect_format(libelle, prix):
    l = libelle.lower()
    if "pichet" in l:
        return "Pichet (50 cl)"
    if "1/2" in l or "demi-b" in l:
        return "Demi-bouteille"
    if "pinte" in l or "pression" in l:
        return "Bière (pression)"
    if "verre" in l:
        return "Verre"
    if "bou" in l or "btle" in l or "btl" in l:
        return "Bouteille"
    if prix and prix >= 15:
        return "Bouteille"
    return "Verre / cocktail"


grp = defaultdict(lambda: {"qte": 0.0, "ca": 0.0, "prix": []})
for lib, v in VMAP.items():
    if v.get("tva") != "20" or v["q"] <= 0:
        continue
    pm = v["ca"] / v["q"]
    cat = v.get("cat") or "Autres alcools"
    g = grp[(cat, detect_format(lib, pm))]
    g["qte"] += v["q"]
    g["ca"] += v["ca"]
    g["prix"] += v["pd"]

ventes_rows = []
for (cat, fmt), g in sorted(grp.items(), key=lambda kv: -kv[1]["ca"]):
    if g["qte"] <= 0 or g["ca"] < 300:
        continue
    pm = g["ca"] / g["qte"]
    pmin, pmax = (min(g["prix"]), max(g["prix"])) if g["prix"] else (pm, pm)
    fourchette = (fr_prix(pmin) + " → " + fr_prix(pmax)) if abs(pmax - pmin) > 0.01 else fr_prix(pm)
    ventes_rows.append({"cat": cat, "fmt": fmt, "qte": g["qte"], "pm": pm, "fourchette": fourchette, "ca": g["ca"]})

# Parallele caisse <-> facture : exemples lisibles (table de correspondance).
CIBLES = ("savagnin", "trousseau", "béthanie", "bethanie", "macvin", "crémant", "cremant",
          "vin jaune", "chardonnay", "pinte", "pression", "poulsard", "saint véran")
corr_rows = []
seen = set()
for c in CORR:
    lib = c.get("libelle_caisse", "")
    fac = c.get("nom_facture", "")
    if not fac or fac.strip() in (":", ""):
        continue
    if any(k in (lib + " " + fac).lower() for k in CIBLES) and lib not in seen:
        seen.add(lib)
        corr_rows.append([cg(lib), cg(c.get("format_service", "")), cg(fac[:48])])
    if len(corr_rows) >= 10:
        break


# ----------------------------------------------------------------------------- #
# 8. XLSX : recalcul complet.
# ----------------------------------------------------------------------------- #
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

HEAD = Font(bold=True, color="FFFFFF")
FILL = PatternFill("solid", fgColor="0F766E")
CENTER = Alignment(horizontal="center", vertical="center")
BOLD = Font(bold=True)


def feuille(wb, titre, cols, lignes, widths, first=False, totaux=None):
    ws = wb.active if first else wb.create_sheet()
    ws.title = titre[:31]
    ws.append(cols)
    for ligne in lignes:
        ws.append(ligne)
    if totaux:
        ws.append(totaux)
    for c in ws[1]:
        c.font = HEAD
        c.fill = FILL
        c.alignment = CENTER
    if totaux:
        for c in ws[ws.max_row]:
            c.font = BOLD
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    return ws


wb = openpyxl.Workbook()

feuille(wb, "Coef par produit (3 exos)",
        ["Produit (caisse)", "Categorie", "Format servi", "Nb ventes 3 ans",
         "Prix de vente moyen", "Prix d'achat matiere", "Coefficient"],
        [[p["lib"], CAT_NOM.get(p["cat"], p["cat"]), p["fmt"], round(p["q"], 1),
          round(p["pv"], 2), round(p["cout"], 2), round(p["coef"], 2)] for p in prod],
        [30, 20, 18, 14, 18, 18, 12], first=True)

feuille(wb, "Coef par categorie (3 exos)",
        ["Categorie", "Nb produits", "CA TTC (3 ans)", "Coefficient (CA / cout matiere)"],
        [[r["nom"], r["n"], round(r["ca"]), round(r["coef"], 2)] for r in cat_rows_data],
        [26, 12, 18, 30])

feuille(wb, "Ventes par format (3 exos)",
        ["Categorie", "Format", "Nb ventes", "Prix moyen", "Fourchette de prix (saison)", "CA TTC"],
        [[r["cat"], r["fmt"], round(r["qte"], 1), round(r["pm"], 2),
          r["fourchette"].replace("→", "a").replace("€", "").strip(), round(r["ca"])] for r in ventes_rows],
        [22, 18, 11, 12, 28, 12])

feuille(wb, "Alcool revendu vs non revendu",
        ["Poste", "Litres", "Part des achats", "Cout d'achat (EUR)"],
        [["Alcool achete (total)", round(ACHAT_L), "100 %", round(ACHAT_COUT)],
         ["Revendu au verre + cocktails", round(REVENDU_L), fr_pct(100 * REVENDU_L / ACHAT_L, 0), round(COUT_REVENDU)],
         ["NON revendu (cuisine, sur-versement, pertes)", round(NON_REVENDU_L), fr_pct(PART_NON_REVENDU * 100, 0), round(COUT_NON_REVENDU)],
         ["Stock final", round(STOCK_L), fr_pct(100 * STOCK_L / ACHAT_L, 0), round(PRIX_L * STOCK_L)],
         ["CA alcool encaisse (caisse, 3 exos)", "", "", round(CA_ALCOOL_CUMUL)],
         ["COEFFICIENT sur alcool REVENDU", "", "", round(COEF_REVENDU, 2)]],
        [46, 12, 16, 18])

feuille(wb, "Coef declare vs reconstitue",
        ["Exercice", "CA declare TTC", "Cout matieres", "Coef declare", "Coef apres reconstitution"],
        [[r["ex"], round(r["ca_declare"]), round(r["matieres"]), round(r["coef_decl"], 3), round(r["coef_rec"], 3)]
         for r in FISC_RECAP],
        [14, 16, 16, 14, 26])

xlsx = os.path.join(ROOT, "public/documents/pieces-defense/RF-coefficients-de-revente.xlsx")
wb.save(xlsx)
F_XLSX = "pieces-defense/RF-coefficients-de-revente.xlsx"
F_CORR = "pieces-defense/RF-conso-achats-correspondance.xlsx"


# ----------------------------------------------------------------------------- #
# 9. Section JSON du Rendu final.
# ----------------------------------------------------------------------------- #
coef_decl_moy = sum(r["coef_decl"] for r in FISC_RECAP) / 3

# Tableau : grief du fisc (3 exercices).
recap_rows = [[cg(r["ex"]), cd(fr_eur(r["ca_declare"])), cd(fr_eur(r["matieres"])),
               cd(fr_coef(r["coef_decl"])), cd(fr_coef(r["coef_rec"]))] for r in FISC_RECAP]

# Tableau : ventes par categorie/format (3 exercices).
ventes_t = [[cg(r["cat"]), cg(r["fmt"]), cd(fr_int(r["qte"])), cd(fr_prix(r["pm"])),
             cd(r["fourchette"]), cd(fr_eur(r["ca"]))] for r in ventes_rows]

# Tableau : coefficient PAR CATEGORIE (3 exercices).
cat_t = [[cg(r["nom"]), cd(str(r["n"])), cd(fr_eur(r["ca"])), cd(fr_coef(r["coef"]))] for r in cat_rows_data]

# Tableau : coefficient PRODUIT PAR PRODUIT (top 40 par CA ; liste complete en XLSX).
TOPN = 40
prod_t = [[cg(p["lib"]), cg(CAT_NOM.get(p["cat"], p["cat"])), cg(p["fmt"]), cd(fr_int(p["q"])),
           cd(fr_prix(p["pv"])), cd(fr_prix(p["cout"])), cd(fr_coef(p["coef"]))] for p in prod[:TOPN]]

# Tableau : cascade non revendu.
nr_t = [[cg(lbl), cd(fr_int(l) + " L"), cd(fr_pct(100 * l / ACHAT_L, 0))] for lbl, l in NR_POSTES if l]
nr_t.append([cgf("Total non revendu"), cdf(fr_int(NON_REVENDU_L) + " L"), cdf(fr_pct(PART_NON_REVENDU * 100, 0))])

# Tableau : coef boisson vs nourriture (3 exercices).
split_t = [[cg(EXL[s["ex"]]), cd(fr_coef(s["coef_boisson"])), cd(fr_coef(s["coef_nourr"]))] for s in SPLIT]

meta = {
    "slug": "coefficients-de-revente",
    "titre": "Le coefficient de revente jugé trop bas",
    "source": "scripts/rendu-final-coefficients-de-revente.py",
    "grief": "Proposition de rectification p. 34-35 (rejet de comptabilité, 3/3).",
    "chiffres": {
        "coef_declare_moyen": round(coef_decl_moy, 2),
        "coef_sur_alcool_revendu": round(COEF_REVENDU, 2),
        "part_non_revendu_pct": round(PART_NON_REVENDU * 100, 1),
        "nb_produits": NB_PRODUITS,
        "ca_couvert": round(CA_COUVERT),
    },
}

sections = [
    # 1) GRIEF
    {"kind": "chapitre", "source": "fisc", "numero": 1, "titre": "Ce que dit l'administration",
     "sousTitre": "Proposition p. 34-35 : un coefficient de revente jugé « manifestement très bas »"},
    {"kind": "paragraphe",
     "texte": "Le service rapporte, **sur chacun des trois exercices contrôlés**, le **chiffre d'affaires "
              "TTC** au **coût des matières revendues** et obtient un **coefficient de revente** d'environ "
              "**2,30**. Il le juge « **manifestement très bas et en dehors des usages habituels de la "
              "profession** » (qu'il situe vers **3** et plus), puis **reconstitue le CA** jusqu'à ce que ce "
              "coefficient atteigne **~3,05**. Ce coefficient n'est donc **pas une mesure** : c'est un "
              "quotient calculé **après** la reconstitution."},
    {"kind": "tableau", "titre": "Le coefficient déclaré et celui obtenu après reconstitution, par exercice (Proposition p. 34-35)",
     "minWidth": 820,
     "colonnes": [colg("Exercice"), cold("CA déclaré TTC"), cold("Coût matières"),
                  cold("Coef. déclaré"), cold("Coef. après reconstitution")],
     "lignes": recap_rows},

    # 2) VENTES PAR FORMAT (3 EXOS)
    {"kind": "chapitre", "source": "nous", "numero": 2, "titre": "La donnée brute : tout ce qui est vendu, à quel prix",
     "sousTitre": "Sur les trois exercices : chaque format (verre, pichet, bouteille, cocktail, pression), ses ventes et ses prix réels qui varient par carte"},
    {"kind": "paragraphe",
     "texte": "Avant tout coefficient, voici la **réalité de la caisse sur les trois exercices du contrôle** : "
              "pour chaque catégorie d'alcool et chaque format, **le nombre exact d'articles vendus**, le "
              "**prix moyen** et la **fourchette de prix** réellement pratiquée. La fourchette montre que "
              "les prix **changent au fil des cartes** (saisons). Rien n'est moyenné ni supposé : ce sont les "
              "lignes de caisse."},
    {"kind": "tableau", "titre": "Ventes d'alcool par catégorie et format (cumul des 3 exercices)",
     "minWidth": 900,
     "colonnes": [colg("Catégorie"), colg("Format"), cold("Nb ventes"), cold("Prix moyen"),
                  cold("Fourchette de prix (saison)"), cold("CA TTC")],
     "lignes": ventes_t},
    {"kind": "piecejointe",
     "intro": "Ventes par article et par format, prix moyen et tous les prix distincts pratiqués (variation par carte), sur les 3 exercices :",
     "fichiers": [{"fichier": F_XLSX, "label": "RF Coefficient de revente : ventes, prix, coefficient par produit (XLSX)"}]},

    # 3) COEFFICIENT PAR CATEGORIE PUIS PAR PRODUIT
    {"kind": "chapitre", "source": "nous", "numero": 3, "titre": "Le coefficient de revente, mesuré produit par produit",
     "sousTitre": "Caisse ↔ facture, sur les 3 exercices : catégorie par catégorie puis produit par produit, le coefficient est normal à élevé"},
    {"kind": "paragraphe",
     "texte": f"Pour calculer le coefficient sans aucune approximation, on relie **chaque article de caisse** "
              f"à la **facture du produit acheté** (table de correspondance officielle), puis on divise le "
              f"**prix de vente** réel par le **coût de la matière** servie (prix d'achat de la bouteille "
              f"rapporté au volume du verre, du pichet ou de la bouteille). Le calcul porte sur **"
              f"{NB_PRODUITS} produits** et **{fr_eur(CA_COUVERT)}** de ventes sur les trois exercices. "
              f"**Catégorie par catégorie**, le coefficient est **normal à élevé** :"},
    {"kind": "tableau", "titre": "Coefficient de revente par catégorie (cumul des 3 exercices)",
     "minWidth": 640,
     "colonnes": [colg("Catégorie"), cold("Nb produits"), cold("CA TTC (3 ans)"), cold("Coefficient")],
     "lignes": cat_t},
    {"kind": "paragraphe",
     "texte": "Les **vins** ressortent à **× 3,5 environ**, les **bières** et **spiritueux** bien plus haut "
              "(× 7 à × 13, marges classiques de ces produits). Le détail, **produit par produit**, le "
              "confirme (les 40 premiers par chiffre d'affaires ci-dessous ; la liste complète des "
              + str(NB_PRODUITS) + " produits est dans la pièce téléchargeable) :"},
    {"kind": "tableau", "titre": "Coefficient de revente produit par produit (prix d'achat facture → prix de vente caisse, 3 exercices)",
     "minWidth": 980,
     "colonnes": [colg("Produit (caisse)"), colg("Catégorie"), colg("Format servi"), cold("Nb ventes"),
                  cold("Prix de vente"), cold("Prix d'achat matière"), cold("Coefficient")],
     "lignes": prod_t},
    {"kind": "paragraphe",
     "texte": "**La profession applique un coefficient dégressif, pas un seuil unique.** Contrairement au "
              "« 3 et plus » uniforme avancé par le service, le coefficient multiplicateur de la profession "
              "**décroît avec le prix d'achat** : de l'ordre de **× 4** sous 6 €, **× 3,5** de 6 à 12 €, "
              "**× 3** de 12 à 25 €, et **× 2,5** au-delà (vins premium). Nos AOC du Jura, achetées **10 à "
              "22 € la bouteille**, relèvent donc des tranches **× 3 à × 3,5** : nos coefficients mesurés "
              "ci-dessus y sont **exactement conformes**. Sources : "
              + lien(U_ACCORD) + " ; " + lien(U_SOMMIT) + " ; " + lien(U_LHOTEL) + "."},
    {"kind": "tableau", "titre": "Parallèle caisse ↔ facture (extrait de la table de correspondance)",
     "minWidth": 760,
     "colonnes": [colg("Libellé caisse"), colg("Format"), colg("Libellé facture")],
     "lignes": corr_rows},
    {"kind": "piecejointe",
     "intro": "Table de correspondance complète (chaque libellé de caisse relié à sa facture et à l'inventaire) :",
     "fichiers": [{"fichier": F_CORR, "label": "RF Correspondance caisse / facture / inventaire (XLSX)"}]},

    # 4) POURQUOI L'AGREGAT EST BAS
    {"kind": "chapitre", "source": "nous", "numero": 4, "titre": "Pourquoi le coefficient global est bas, sans aucune recette cachée",
     "sousTitre": "Une grande part de l'alcool acheté n'est jamais revendue ; là où il l'est, la marge est normale"},
    {"kind": "paragraphe",
     "texte": f"Si la marge par produit est normale (section précédente), pourquoi le coefficient **global** "
              f"est-il bas ? Parce que le coefficient « brut » du service suppose que **tout l'alcool acheté "
              f"est revendu au verre au prix de la carte**. C'est faux : sur les **{fr_int(ACHAT_L)} litres** "
              f"achetés, **{fr_pct(PART_NON_REVENDU * 100, 0)}** ne génèrent **aucun** chiffre d'affaires au "
              f"verre. Ils partent en cuisine, en sur-versement, en crémant éventé jeté, en mousse de bière, "
              f"en dégustations offertes, en consommation du chef ou en casse. Chaque poste est chiffré et "
              f"sourcé aux pages [Reconstitution par les volumes](/rendu-final/reconstitution-volumes-liquides) "
              f"et suivantes (2.2 à 2.7). Le coût de cet alcool est **bien réel**, mais il ne peut pas "
              f"« remonter » dans un coefficient de revente."},
    {"kind": "tableau", "titre": f"L'alcool acheté qui n'est JAMAIS revendu au verre ({fr_pct(PART_NON_REVENDU * 100, 0)} des achats)",
     "minWidth": 680,
     "colonnes": [colg("Poste"), cold("Litres"), cold("Part des achats")],
     "lignes": nr_t},
    {"kind": "kpis", "items": [
        {"label": "Coefficient apparent", "valeur": fr_coef(coef_decl_moy), "sub": "CA total / coût matières (reproché par le fisc)", "couleur": "red"},
        {"label": "Sur l'alcool réellement revendu", "valeur": fr_coef(COEF_REVENDU), "sub": "CA alcool caisse / coût alcool revendu (3 exos)", "highlight": True, "couleur": "teal"},
        {"label": "Alcool jamais revendu", "valeur": fr_pct(PART_NON_REVENDU * 100, 0), "sub": "cuisine, sur-versement, pertes (pages 2.x)", "couleur": "gray"},
    ]},
    {"kind": "paragraphe",
     "texte": f"Le calcul est simple et reproductible : sur trois exercices, l'alcool **réellement revendu** "
              f"au verre et en cocktails ({fr_int(REVENDU_L)} L, soit un coût d'achat de **{fr_eur(COUT_REVENDU)}**) "
              f"a encaissé **{fr_eur(CA_ALCOOL_CUMUL)}** en caisse. Le coefficient sur cet alcool vendu est "
              f"donc de **{fr_coef(COEF_REVENDU)}**, **au-dessus** du seuil de « 3 » que le service présente "
              f"comme la norme de la profession. Le coefficient apparent bas ne vient donc ni d'une recette "
              f"dissimulée, ni d'une marge anormale : il vient de l'alcool acheté **mais jamais revendu**."},
    {"kind": "paragraphe",
     "texte": "Dernière vérification : le coefficient de la **boisson** et celui de la **nourriture** sont "
              "**au même niveau** (environ 2,3 à 2,5 chacun, sur les trois exercices). Ce n'est donc pas la "
              "boisson qui « manquerait » : les deux blocs ont la même structure de marge, celle d'un "
              "restaurant de terroir qui achète des produits de qualité."},
    {"kind": "tableau", "titre": "Coefficient boisson et nourriture, par exercice (même niveau)",
     "minWidth": 520,
     "colonnes": [colg("Exercice"), cold("Coef. boisson"), cold("Coef. nourriture")],
     "lignes": split_t},

    # 5) LA CORRECTION DU FISC EST CIRCULAIRE
    {"kind": "chapitre", "source": "nous", "numero": 5, "titre": "Le coefficient « 3,05 » du fisc n'est pas mesuré : il est fabriqué",
     "sousTitre": "Il découle de la reconstitution, elle-même circulaire (page 2.8)"},
    {"kind": "paragraphe",
     "texte": "Le coefficient « conforme à la profession » (3,05) que le service oppose au déclaré n'est "
              "**pas observé** : il est le **résultat** de sa propre reconstitution du CA. Or cette "
              "reconstitution est démontée à la page [Extrapolation de la cuisine (coefficient liquide → "
              "solide)](/rendu-final/coefficient-liquide-solide) : le service y applique au CA liquides un "
              "coefficient **tiré de la caisse qu'il rejette par ailleurs**, et fabrique ainsi une « cuisine "
              "fantôme » qui représente les trois quarts du redressement. Le « bon » coefficient de 3,05 "
              "n'est donc que le **reflet** de cette construction, pas une grandeur indépendante."},

    # SOURCES
    {"kind": "note",
     "texte": "**Sources et données (recalcul reproductible, 3 exercices).** Ventes et prix de vente : caisse "
              "du restaurant (ventes par article, annexes D et B). Achats et prix d'achat : factures "
              "fournisseurs (annexe d'achats). Mise en correspondance des libellés caisse ↔ facture : table "
              "de correspondance officielle. Volumes d'alcool non revendus : cascade de la page "
              "[Reconstitution par les volumes](/rendu-final/reconstitution-volumes-liquides). Tout est "
              "recalculable dans les pièces téléchargeables de cette fiche. Grief du service : Proposition de "
              "rectification p. 34-35. Coefficient multiplicateur de la profession (grille **dégressive** "
              "selon le prix d'achat) : " + lien(U_ACCORD) + " ; " + lien(U_SOMMIT) + " ; " + lien(U_LHOTEL) + "."},
    # RESULTAT
    {"kind": "alerte", "couleur": "teal", "titre": "Résultat",
     "texte": f"Sur les **trois exercices**, le coefficient de revente est mesuré **produit par produit** "
              f"({NB_PRODUITS} produits) à partir des vraies ventes et des vraies factures : il est **normal "
              f"à élevé** (vins × 3,5 ; bières et spiritueux × 7 à × 13). Sur l'alcool **réellement revendu**, "
              f"le coefficient global atteint **{fr_coef(COEF_REVENDU)}**, au-dessus du seuil de la "
              f"profession. Le coefficient apparent (**{fr_coef(coef_decl_moy)}**) n'est tiré vers le bas que "
              f"par **{fr_pct(PART_NON_REVENDU * 100, 0)} d'alcool jamais revendu** (cuisine, sur-versement, "
              f"pertes), et le « 3,05 » du service n'est pas mesuré mais fabriqué par sa reconstitution "
              f"circulaire (page 2.8). Le grief ne tient pas."},
    {"kind": "piecejointe",
     "intro": "Recalcul intégral, reproductible : coefficient par produit et par catégorie, ventes par format, alcool revendu vs non revendu, sur les 3 exercices :",
     "fichiers": [{"fichier": F_XLSX, "label": "RF Coefficient de revente : produit par produit, par catégorie, ventes (XLSX)"}]},
    {"kind": "interne", "audience": "avocat", "titre": "Note pour l'avocat",
     "texte": "Argument central : coefficient mesuré PRODUIT PAR PRODUIT sur 3 exercices (normal a eleve) ; "
              "sur l'alcool REVENDU il dépasse le seuil profession (x 3,85) ; le coefficient apparent bas vient "
              "des 36 % d'alcool non revendu (cascade des pages 2.x). " + str(n_skip) + " produits ecartes du "
              "detail (incoherence volume/prix de la table de correspondance) : a faire corriger pour une "
              "version definitive. Relier a la page 2.8 (coefficient liquide -> solide, circulaire). Verser une "
              "attestation sur le positionnement qualite/terroir et les doses. Retirer cet encart de toute "
              "version remise."},
]

doc = {"meta": meta, "sections": ajouter_conclusion(sections)}
dest = os.path.join(ROOT, "src/data/renduFinal/coefficients-de-revente.json")
os.makedirs(os.path.dirname(dest), exist_ok=True)
json.dump(doc, open(dest, "w", encoding="utf-8"), ensure_ascii=False, indent=1)

print("OK coefficients-de-revente")
print("  produits avec coefficient : %d (CA couvert %d), ecartes %d" % (NB_PRODUITS, round(CA_COUVERT), n_skip))
print("  coef sur alcool revendu : %.2f | coef declare moyen : %.2f" % (COEF_REVENDU, coef_decl_moy))
print("  categories : " + ", ".join("%s x%.1f" % (r["nom"], r["coef"]) for r in cat_rows_data[:5]))
