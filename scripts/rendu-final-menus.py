#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
rendu-final-menus.py
Détail des PRIX DES MENUS PAR PÉRIODE de carte : pour chaque menu, le prix
catalogue, le nombre de fois vendu AU prix catalogue, le nombre de fois vendu
HORS catalogue (prix personnalisé = facture sans détail), et le détail des prix.
Source : src/data/calculsBoissons/menusVentesParPrix.json (issu des tickets C).

Sorties : complète src/data/renduFinalCalculs.json (clé "menus_par_periode")
          + public/documents/pieces-defense/Menus-prix-par-periode.xlsx
"""
import json, os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

ICI = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.normpath(os.path.join(ICI, ".."))
m = json.load(open(os.path.join(ROOT, "src/data/calculsBoissons/menusVentesParPrix.json"), encoding="utf-8"))

periodes = []
for pk, pd in m["periodes"].items():
    rows = []
    for menu, md in pd.get("menus", {}).items():
        pa = md.get("prix_attendu")
        q_cat = q_cus = eur_cus = 0.0
        prix_cus = []
        for l in md.get("lignes", []):
            q = l.get("quantite", 0) or 0
            prix = l.get("prix", 0) or 0
            if str(l.get("prix_modifie", "")).upper() == "OUI":
                q_cus += q
                eur_cus += q * prix
                prix_cus.append((round(prix, 2), q))
            else:
                q_cat += q
        rows.append({
            "menu": menu, "prix_catalogue": pa,
            "qte_catalogue": round(q_cat), "qte_hors_catalogue": round(q_cus),
            "nb_prix_custom": len(prix_cus),
            "eur_hors_catalogue": round(eur_cus),
            "prix_custom": sorted(prix_cus),
        })
    rows.sort(key=lambda r: -r["qte_hors_catalogue"])
    periodes.append({"periode": pk, "carte": pd.get("carte"), "exercice": pd.get("exercice"),
                     "dates": pd.get("dates"), "menus": rows})

# JSON
dest = os.path.join(ROOT, "src/data/renduFinalCalculs.json")
data = json.load(open(dest, encoding="utf-8"))
data["menus_par_periode"] = periodes
json.dump(data, open(dest, "w", encoding="utf-8"), ensure_ascii=False, indent=1)

# XLSX
H1 = Font(bold=True, size=12, color="FFFFFF"); FILL1 = PatternFill("solid", fgColor="1F3A5F")
BOLD = Font(bold=True); FILLT = PatternFill("solid", fgColor="E2E8F0")
THIN = Border(*[Side(style="thin", color="CBD5E0")] * 4); RA = Alignment(horizontal="right")
wb = openpyxl.Workbook()
first = True
for p in periodes:
    title = f"{p['periode']} {p['carte']} {p['exercice']}"
    ws = wb.active if first else wb.create_sheet()
    ws.title = title[:31]
    first = False
    ws.append([f"Menus — période {p['periode']} (carte {p['carte']}, {p['exercice']}, {p['dates']})"])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)
    ws.cell(1, 1).font = H1; ws.cell(1, 1).fill = FILL1
    ws.append(["Menu", "Prix catalogue", "Vendus AU prix catalogue", "Vendus HORS catalogue", "€ hors catalogue", "Prix personnalisés (prix × qté)"])
    for c in ws[2]:
        c.font = BOLD; c.fill = FILLT; c.border = THIN
    for r in p["menus"]:
        pc = "  ;  ".join(f"{px}€×{int(q)}" for px, q in r["prix_custom"][:30]) or "—"
        ws.append([r["menu"], r["prix_catalogue"], r["qte_catalogue"], r["qte_hors_catalogue"], r["eur_hors_catalogue"], pc])
        for c in ws[ws.max_row]:
            c.border = THIN
            if isinstance(c.value, (int, float)):
                c.alignment = RA
    for i, w in enumerate([16, 13, 22, 21, 15, 80], 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

out = os.path.join(ROOT, "public/documents/pieces-defense/Menus-prix-par-periode.xlsx")
wb.save(out)
print("écrit:", out)
print("périodes:", [p["periode"] for p in periodes])
for p in periodes:
    cat = sum(r["qte_catalogue"] for r in p["menus"])
    cus = sum(r["qte_hors_catalogue"] for r in p["menus"])
    print(f"  {p['periode']} {p['exercice']}: {cat} au catalogue / {cus} hors catalogue")
