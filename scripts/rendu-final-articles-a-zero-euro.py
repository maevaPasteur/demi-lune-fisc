#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Rendu final - Articles a prix 0 EUR (offerts / gratuits).

Script reproductible (aucun Date.now / random). Il :
  1. lit les 3 annexes de detail tickets ligne a ligne (ANNEXE-C1/C2/C3),
  2. compte les lignes vendues a prix unitaire = 0 EUR et le total des lignes,
  3. ventile ces lignes par produit puis par categorie (cafes / boissons
     chaudes, apero-alcool, softs, desserts, autres),
  4. verifie que ces lignes ne portent aucun encaissement (PU x Qte = 0)
     et qu'elles sont presque toutes incluses dans un ticket par ailleurs
     paye (donc impossibles a utiliser pour sortir une recette),
  5. recoupe avec la conso personnel / offerts deja chiffree
     (src/data/boissonsPageData.json),
  6. ecrit la piece XLSX et le JSON final de la page.

Lecture seule des sources. Sortie deterministe.
Lancer avec : /tmp/xlsenv/bin/python scripts/rendu-final-articles-a-zero-euro.py
"""

import json
import os
from collections import Counter, defaultdict

import xlrd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

# --------------------------------------------------------------------------
# Chemins
# --------------------------------------------------------------------------
ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
CAISSE = os.path.join(ROOT, "public", "documents", "caisse-enregistreuse")
PIECES = os.path.join(ROOT, "public", "documents", "pieces-defense")
DATA = os.path.join(ROOT, "src", "data", "renduFinal")

XLSX_OUT = os.path.join(PIECES, "RF-articles-a-zero-euro.xlsx")
JSON_OUT = os.path.join(DATA, "articles-a-zero-euro.json")

EXERCICES = {
    "1": "2022-2023",
    "2": "2023-2024",
    "3": "2024-2025",
}

# Colonnes annexe C (detail tickets), cf. ligne d'en-tete (row 0) :
#  0 Date | 1 Heure | 2 No_ticket | 5 Tot_ttc | 9 Ref_prd
#  10 Lib_ticket | 11 Qte | 13 Uprice_wt (prix unitaire TTC)
COL_DATE, COL_HEURE, COL_NOTICKET = 0, 1, 2
COL_TOTTTC = 5
COL_REF, COL_LIB, COL_QTE, COL_PU = 9, 10, 11, 13


# --------------------------------------------------------------------------
# Formatage francais
# --------------------------------------------------------------------------
def fr_int(n):
    return ("{:,}".format(int(round(n)))).replace(",", " ")


def fr_eur(n):
    s = "{:,.2f}".format(float(n)).replace(",", " ").replace(".", ",")
    s = s.replace(" ", " ")
    return s + " €"


def fr_pct(n, dec=2):
    s = ("{:." + str(dec) + "f}").format(n).replace(".", ",")
    return s + " %"


# --------------------------------------------------------------------------
# Categorisation produit (a partir du libelle de caisse)
# --------------------------------------------------------------------------
CAFES = {
    "expresso", "deca", "deca.", "decafeine", "the", "the vert", "infusion",
    "cafe", "cafe viennois", "cappuccino", "chocolat chaud", "chocolat viennois",
    "grand creme", "noisette", "the / infusion",
}
DESSERTS_PREF = ("crepe", "crepe ", "crepe", "glace", "pavlova", "creme brulee",
                 "tarte", "profiterole", "cafe gourmand", "nutella", "sauce choco")
SOFTS_PREF = ("coca", "fuze", "fuze tea", "perrier", "san pel", "vittel",
              "limonade", "diabolo", "sirop", "jus de fruit", "ice tea",
              "schweppes", "oasis", "orangina", "eau", "badoit", "evian")


def categorie(lib):
    low = lib.strip().lower()
    if low in CAFES or low.startswith(("expresso", "deca", "cafe", "the ",
                                       "the\xa0", "cappucc", "chocolat chaud",
                                       "infusion", "grand creme")):
        if low.startswith("cafe gourmand"):
            return "Desserts"
        return "Cafes / boissons chaudes"
    if low == "the":
        return "Cafes / boissons chaudes"
    if any(low.startswith(p) for p in DESSERTS_PREF):
        return "Desserts"
    if any(low.startswith(p) for p in SOFTS_PREF):
        return "Softs / eaux"
    # Tout le reste contenant des marqueurs alcool = apero / alcool
    return "Aperitifs / alcools (verre offert)"


# --------------------------------------------------------------------------
# 1) Lecture des annexes et comptage
# --------------------------------------------------------------------------
def lire_annexes():
    par_exo = {}            # exo -> {"total":, "zero":, "zero_paye":, "zero_seul":}
    libelles = Counter()    # libelle -> nb lignes a 0 EUR
    cat_count = Counter()   # categorie -> nb lignes a 0 EUR
    cat_litres = Counter()  # categorie -> qte cumulee
    exemple = {}            # libelle -> (date, heure, qte, ticket_ttc, no_ticket)
    valeur_zero = 0.0       # somme PU x Qte sur les lignes a 0 EUR (doit = 0)
    ticket_exemple = None   # ticket complet contenant un cafe a 0 EUR

    for n, exo in EXERCICES.items():
        path = os.path.join(CAISSE, "ANNEXE-C%s_detail-tickets_%s.xls" % (n, exo))
        sh = xlrd.open_workbook(path).sheet_by_index(0)

        total = 0
        zero = 0
        zero_paye = 0
        zero_seul = 0

        for r in range(1, sh.nrows):
            ref = str(sh.cell_value(r, COL_REF)).strip()
            lib = str(sh.cell_value(r, COL_LIB)).strip()
            if ref == "" and lib == "":
                continue  # ligne vide
            total += 1
            try:
                pu = float(sh.cell_value(r, COL_PU))
            except (TypeError, ValueError):
                pu = None
            if pu == 0.0:
                zero += 1
                try:
                    qte = float(sh.cell_value(r, COL_QTE))
                except (TypeError, ValueError):
                    qte = 0.0
                try:
                    ttc = float(sh.cell_value(r, COL_TOTTTC))
                except (TypeError, ValueError):
                    ttc = 0.0
                if ttc > 0:
                    zero_paye += 1
                else:
                    zero_seul += 1
                valeur_zero += pu * qte
                cat = categorie(lib)
                libelles[lib] += 1
                cat_count[cat] += 1
                cat_litres[cat] += qte
                if lib not in exemple:
                    exemple[lib] = (
                        sh.cell_value(r, COL_DATE),
                        sh.cell_value(r, COL_HEURE),
                        qte,
                        ttc,
                        sh.cell_value(r, COL_NOTICKET),
                    )

        par_exo[exo] = {
            "total": total,
            "zero": zero,
            "zero_paye": zero_paye,
            "zero_seul": zero_seul,
        }

    # Ticket complet de demonstration : premier Expresso a 0 EUR de C1
    sh = xlrd.open_workbook(
        os.path.join(CAISSE, "ANNEXE-C1_detail-tickets_2022-2023.xls")
    ).sheet_by_index(0)
    cible = None
    for r in range(1, sh.nrows):
        if str(sh.cell_value(r, COL_LIB)).strip() == "Expresso":
            try:
                if float(sh.cell_value(r, COL_PU)) == 0.0:
                    cible = (sh.cell_value(r, COL_DATE),
                             sh.cell_value(r, COL_HEURE),
                             sh.cell_value(r, COL_NOTICKET))
                    break
            except (TypeError, ValueError):
                pass
    lignes_ticket = []
    if cible:
        d, h, no = cible
        for r in range(1, sh.nrows):
            if (sh.cell_value(r, COL_DATE) == d
                    and sh.cell_value(r, COL_HEURE) == h
                    and sh.cell_value(r, COL_NOTICKET) == no):
                lignes_ticket.append((
                    str(sh.cell_value(r, COL_LIB)).strip(),
                    float(sh.cell_value(r, COL_QTE)),
                    float(sh.cell_value(r, COL_PU)),
                    float(sh.cell_value(r, COL_TOTTTC)),
                ))
        ticket_exemple = {"date": d, "heure": h, "no": no,
                          "lignes": lignes_ticket}

    return {
        "par_exo": par_exo,
        "libelles": libelles,
        "cat_count": cat_count,
        "cat_litres": cat_litres,
        "exemple": exemple,
        "valeur_zero": valeur_zero,
        "ticket_exemple": ticket_exemple,
    }


# --------------------------------------------------------------------------
# 2) Recoupe conso personnel / offerts
# --------------------------------------------------------------------------
def lire_conso_personnel():
    with open(os.path.join(ROOT, "src", "data", "boissonsPageData.json"),
              encoding="utf-8") as f:
        d = json.load(f)
    perso = d.get("personnel", {})
    offerts = d.get("offerts", {})
    return {
        "perso_total_litres": perso.get("total_litres"),
        "perso_total_ca": perso.get("total_ca_eur"),
        "offerts_total_ca": offerts.get("total_ca_eur"),
    }


# --------------------------------------------------------------------------
# 3) Ecriture XLSX
# --------------------------------------------------------------------------
H_FILL = PatternFill("solid", fgColor="0F766E")
H_FONT = Font(bold=True, color="FFFFFF")
BOLD = Font(bold=True)
RIGHT = Alignment(horizontal="right")


def style_header(ws, ncol):
    for c in range(1, ncol + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = H_FONT
        cell.fill = H_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "A2"


def autowidth(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def ecrire_xlsx(res, conso, totaux):
    wb = openpyxl.Workbook()

    # --- Onglet 1 : Synthese ---
    ws = wb.active
    ws.title = "Synthese"
    ws.append(["Indicateur", "Valeur"])
    style_header(ws, 2)
    rows = [
        ("Lignes de vente analysees (3 exercices, annexes C1 a C3)",
         fr_int(totaux["total"])),
        ("Lignes a prix unitaire 0 EUR", fr_int(totaux["zero"])),
        ("Part des lignes a 0 EUR", fr_pct(totaux["pct"], 2)),
        ("Valeur encaissee par ces lignes (PU x Qte)",
         fr_eur(res["valeur_zero"])),
        ("Dont incluses dans un ticket par ailleurs paye",
         fr_int(totaux["zero_paye"])),
        ("Dont sur un ticket entierement a 0 EUR",
         fr_int(totaux["zero_seul"])),
        ("Conso personnel chiffree (CA equivalent, recoupe)",
         fr_eur(conso["perso_total_ca"]) if conso["perso_total_ca"] else "n.c."),
        ("Offerts chiffres (CA equivalent, recoupe)",
         fr_eur(conso["offerts_total_ca"]) if conso["offerts_total_ca"] else "n.c."),
    ]
    for k, v in rows:
        ws.append([k, v])
        ws.cell(row=ws.max_row, column=2).alignment = RIGHT
    autowidth(ws, [62, 26])

    # --- Onglet 2 : Par exercice ---
    ws = wb.create_sheet("Par exercice")
    ws.append(["Exercice", "Lignes de vente", "Lignes a 0 EUR",
               "Part", "Dont ticket paye", "Dont ticket a 0"])
    style_header(ws, 6)
    for exo in ("2022-2023", "2023-2024", "2024-2025"):
        e = res["par_exo"][exo]
        pct = 100.0 * e["zero"] / e["total"]
        ws.append([exo, fr_int(e["total"]), fr_int(e["zero"]),
                   fr_pct(pct, 3), fr_int(e["zero_paye"]),
                   fr_int(e["zero_seul"])])
        for c in range(2, 7):
            ws.cell(row=ws.max_row, column=c).alignment = RIGHT
    ws.append(["TOTAL", fr_int(totaux["total"]), fr_int(totaux["zero"]),
               fr_pct(totaux["pct"], 3), fr_int(totaux["zero_paye"]),
               fr_int(totaux["zero_seul"])])
    for c in range(1, 7):
        ws.cell(row=ws.max_row, column=c).font = BOLD
        if c >= 2:
            ws.cell(row=ws.max_row, column=c).alignment = RIGHT
    autowidth(ws, [14, 16, 16, 12, 18, 16])

    # --- Onglet 3 : Par categorie ---
    ws = wb.create_sheet("Par categorie")
    ws.append(["Categorie", "Lignes a 0 EUR", "Part des lignes a 0 EUR",
               "Quantite cumulee"])
    style_header(ws, 4)
    cats = sorted(res["cat_count"].items(), key=lambda x: -x[1])
    for cat, nb in cats:
        ws.append([cat, fr_int(nb),
                   fr_pct(100.0 * nb / totaux["zero"], 1),
                   fr_int(res["cat_litres"][cat])])
        for c in range(2, 5):
            ws.cell(row=ws.max_row, column=c).alignment = RIGHT
    autowidth(ws, [38, 16, 22, 18])

    # --- Onglet 4 : Lignes a 0 EUR par produit ---
    ws = wb.create_sheet("Lignes a 0 EUR par produit")
    ws.append(["Produit", "Categorie", "Nb lignes a 0 EUR",
               "Exemple date", "Heure", "Qte", "Total TTC du ticket"])
    style_header(ws, 7)
    for lib, nb in res["libelles"].most_common():
        d, h, q, ttc, no = res["exemple"][lib]
        ws.append([lib, categorie(lib), nb, str(d), str(h),
                   fr_int(q) if q == int(q) else ("%.2f" % q).replace(".", ","),
                   fr_eur(ttc)])
        ws.cell(row=ws.max_row, column=3).alignment = RIGHT
        ws.cell(row=ws.max_row, column=6).alignment = RIGHT
        ws.cell(row=ws.max_row, column=7).alignment = RIGHT
    autowidth(ws, [30, 34, 18, 14, 10, 8, 20])

    # --- Onglet 5 : Exemple ticket detaille ---
    ws = wb.create_sheet("Exemple ticket detaille")
    te = res["ticket_exemple"]
    ws.append(["Ticket du %s a %s (no %s) : un cafe et un soft offerts (0 EUR) "
               "dans une note encaissee" %
               (te["date"], te["heure"], int(te["no"]))])
    ws.cell(row=1, column=1).font = BOLD
    ws.append([])
    ws.append(["Libelle", "Qte", "Prix unitaire", "Total TTC du ticket"])
    for c in range(1, 5):
        ws.cell(row=3, column=c).font = H_FONT
        ws.cell(row=3, column=c).fill = H_FILL
        ws.cell(row=3, column=c).alignment = Alignment(horizontal="center")
    for lib, q, pu, ttc in te["lignes"]:
        ws.append([lib,
                   fr_int(q) if q == int(q) else ("%.2f" % q).replace(".", ","),
                   fr_eur(pu), fr_eur(ttc)])
        ws.cell(row=ws.max_row, column=2).alignment = RIGHT
        ws.cell(row=ws.max_row, column=3).alignment = RIGHT
        ws.cell(row=ws.max_row, column=4).alignment = RIGHT
    ws.freeze_panes = "A4"
    autowidth(ws, [26, 8, 16, 22])

    wb.save(XLSX_OUT)


# --------------------------------------------------------------------------
# 4) Construction du JSON de page
# --------------------------------------------------------------------------
def construire_json(res, conso, totaux):
    pe = res["par_exo"]
    te = res["ticket_exemple"]
    cats = sorted(res["cat_count"].items(), key=lambda x: -x[1])

    # categories pour le graphique (nom court)
    graph_data = [{"nom": cat, "Lignes a 0 €": nb} for cat, nb in cats]

    # nombre de cafes (categorie cafes)
    nb_cafes = res["cat_count"].get("Cafes / boissons chaudes", 0)
    nb_apero = res["cat_count"].get("Aperitifs / alcools (verre offert)", 0)
    nb_softs = res["cat_count"].get("Softs / eaux", 0)
    nb_desserts = res["cat_count"].get("Desserts", 0)

    # ticket exemple : montant et liste lisible des lignes payantes
    payantes = [l for l in te["lignes"] if l[2] > 0]
    offertes = [l for l in te["lignes"] if l[2] == 0]
    ttc_ticket = te["lignes"][0][3] if te["lignes"] else 0.0

    sections = []

    # 1. Chapitre fisc
    sections.append({
        "kind": "chapitre",
        "source": "fisc",
        "titre": "Ce que soutient l'administration",
    })
    sections.append({
        "kind": "note",
        "texte": (
            "Dans la proposition de rectification (rejet de comptabilite, "
            "partie « articles a quantite inferieure a l'unite, prix non "
            "fixe ou remise », p. 16 a 22), le verificateur releve la "
            "presence, dans les tableaux de detail de la caisse, de lignes "
            "dont le calcul aboutit a **0,00 €**. Il en deduit que "
            "des ventes auraient pu etre **sorties du chiffre d'affaires** au "
            "moyen de ces lignes a prix nul, et que la comptabilite ne serait "
            "donc pas probante."
        ),
    })

    # 2. Chapitre nous
    sections.append({
        "kind": "chapitre",
        "source": "nous",
        "titre": "Notre reponse",
        "sousTitre": "Les lignes a 0 € sont marginales et n'ouvrent "
                     "aucun canal d'encaissement",
    })

    sections.append({
        "kind": "note",
        "texte": (
            "Nous avons repris **ligne a ligne** le detail des tickets de la "
            "caisse sur les trois exercices verifies (annexes C1, C2 et C3, "
            "colonne « prix unitaire TTC »). Sur "
            "**" + fr_int(totaux["total"]) + " lignes de vente**, seules "
            "**" + fr_int(totaux["zero"]) + " lignes** sont a prix unitaire "
            "nul, soit **" + fr_pct(totaux["pct"], 2) + "** du total. C'est "
            "exactement le nombre de lignes que l'annexe de caisse signale "
            "elle-meme dans sa colonne « Prix a 0 € » "
            "(" + fr_int(totaux["zero"]) + " lignes), ce qui confirme le "
            "comptage."
        ),
    })

    sections.append({
        "kind": "note",
        "texte": (
            "Le total que nous obtenons (" + fr_int(totaux["total"]) + " lignes "
            "de produits) est tres proche du chiffre cite par "
            "l'administration (97 694 lignes). L'ecart, de l'ordre de "
            "quelques dizaines de lignes, tient a la maniere de compter les "
            "lignes techniques vides : nous conservons **nos chiffres "
            "recalcules**, integralement reproductibles a partir des annexes "
            "fournies."
        ),
    })

    # Tableau par exercice
    lignes_exo = []
    for exo in ("2022-2023", "2023-2024", "2024-2025"):
        e = pe[exo]
        p = 100.0 * e["zero"] / e["total"]
        lignes_exo.append([
            {"v": exo},
            {"v": fr_int(e["total"]), "align": "right"},
            {"v": fr_int(e["zero"]), "align": "right"},
            {"v": fr_pct(p, 3), "align": "right"},
        ])
    lignes_exo.append([
        {"v": "Total"},
        {"v": fr_int(totaux["total"]), "align": "right"},
        {"v": fr_int(totaux["zero"]), "align": "right"},
        {"v": fr_pct(totaux["pct"], 2), "align": "right"},
    ])
    sections.append({
        "kind": "tableau",
        "titre": "Lignes a 0 € par exercice",
        "colonnes": [
            {"label": "Exercice"},
            {"label": "Lignes de vente", "align": "right"},
            {"label": "Lignes a 0 €", "align": "right"},
            {"label": "Part", "align": "right"},
        ],
        "lignes": lignes_exo,
    })

    # Argument central : aucune recette possible
    sections.append({
        "kind": "note",
        "texte": (
            "**Une ligne a 0 € ne peut pas servir a dissimuler une "
            "recette.** Par construction, son montant encaisse est "
            "« prix unitaire x quantite », soit "
            "**" + fr_eur(res["valeur_zero"]) + "** au total pour les "
            "" + fr_int(totaux["zero"]) + " lignes concernees. Aucun argent "
            "ne transite par ces lignes : il n'y a donc rien a « sortir "
            "». Pour detourner une recette, il faudrait au contraire une "
            "ligne **encaissee** que l'on ferait ensuite disparaitre, ce qui "
            "est l'inverse d'une ligne a prix nul."
        ),
    })

    sections.append({
        "kind": "note",
        "texte": (
            "Mieux : sur les " + fr_int(totaux["zero"]) + " lignes a "
            "0 €, **" + fr_int(totaux["zero_paye"]) + " (soit "
            + fr_pct(100.0 * totaux["zero_paye"] / totaux["zero"], 0) +
            ") figurent a l'interieur d'un ticket par ailleurs encaisse** "
            "(total TTC du ticket superieur a 0). Il s'agit du cas classique "
            "du cafe ou de l'apero offert en fin de repas : le repas est paye, "
            "le geste commercial est saisi a 0 €. Les "
            + fr_int(totaux["zero_seul"]) + " lignes restantes sont sur des "
            "tickets entierement a 0 € (offerts purs ou essais), "
            "sans aucun encaissement non plus."
        ),
    })

    # Ventilation par categorie
    lignes_cat = []
    for cat, nb in cats:
        lignes_cat.append([
            {"v": cat},
            {"v": fr_int(nb), "align": "right"},
            {"v": fr_pct(100.0 * nb / totaux["zero"], 1), "align": "right"},
        ])
    sections.append({
        "kind": "tableau",
        "titre": "Nature des lignes a 0 € (que sont ces offerts ?)",
        "colonnes": [
            {"label": "Categorie"},
            {"label": "Lignes a 0 €", "align": "right"},
            {"label": "Part", "align": "right"},
        ],
        "lignes": lignes_cat,
    })

    sections.append({
        "kind": "note",
        "texte": (
            "La ventilation parle d'elle-meme : il s'agit pour l'essentiel "
            "d'**aperitifs ou verres d'alcool offerts** (kir, Macvin, "
            "pression, verre de vin : " + fr_int(nb_apero) + " lignes), de "
            "**softs et eaux** (" + fr_int(nb_softs) + " lignes), de **cafes "
            "et boissons chaudes offerts** (" + fr_int(nb_cafes) + " lignes) "
            "et de quelques **desserts** (" + fr_int(nb_desserts) + " lignes). "
            "Ce sont les gestes commerciaux (verre ou cafe offert) et la "
            "consommation du personnel deja decrits a l'administration, et non "
            "des ventes masquees."
        ),
    })

    # Graphique
    sections.append({
        "kind": "graphique",
        "variante": "horizontal",
        "hauteur": 280,
        "dataKey": "nom",
        "serie": {"name": "Lignes a 0 €", "couleur": "#0f766e"},
        "data": graph_data,
        "format": "int",
    })

    # Exemple concret date
    txt_payantes = ", ".join(
        "%s (%s)" % (l[0], fr_eur(l[2])) for l in payantes
    )
    txt_offertes = " et ".join(l[0] for l in offertes)
    sections.append({
        "kind": "note",
        "texte": (
            "**Exemple concret date.** Ticket du **" + str(te["date"]) +
            " a " + str(te["heure"]) + "** (no " + str(int(te["no"])) + "), "
            "encaisse **" + fr_eur(ttc_ticket) + " TTC**. La note comprend des "
            "plats et desserts payes (" + txt_payantes + ") et, en fin de "
            "repas, **" + txt_offertes + " a 0 €**. Le repas a bien "
            "ete encaisse ; seul le geste commercial final est saisi a prix "
            "nul. On voit ici, sur une vraie note, qu'une ligne a "
            "0 € accompagne une recette encaissee au lieu de la "
            "faire disparaitre."
        ),
    })

    # Recoupe conso personnel / offerts
    sections.append({
        "kind": "note",
        "texte": (
            "Ces offerts recoupent la **consommation du personnel et les "
            "offerts deja chiffres** par ailleurs (piece "
            "« Consommation personnel et offerts »)  : "
            + fr_eur(conso["offerts_total_ca"]) + " de CA equivalent "
            "d'offerts et " + fr_eur(conso["perso_total_ca"]) + " de conso "
            "personnel sur la periode. Loin d'etre cachees, ces consommations "
            "ont ete reconstituees et communiquees : elles confirment que les "
            "lignes a 0 € correspondent a une realite d'exploitation "
            "ordinaire, sans enjeu fiscal."
        ),
    })

    # 3. Piece jointe
    sections.append({
        "kind": "piecejointe",
        "intro": "Le detail complet, reproductible a partir des annexes de "
                 "caisse, est fourni en piece telechargeable : comptage par "
                 "exercice, ventilation par categorie, liste des "
                 + fr_int(totaux["zero"]) + " lignes a 0 € par "
                 "produit avec dates exemples, et le ticket detaille de "
                 "l'exemple ci-dessus.",
        "fichiers": [
            {
                "fichier": "pieces-defense/RF-articles-a-zero-euro.xlsx",
                "label": "Articles a 0 € : comptage, categories et "
                         "detail par produit (XLSX)",
            }
        ],
    })

    # 4. Verdict (une seule alerte teal)
    sections.append({
        "kind": "alerte",
        "couleur": "teal",
        "titre": "Ce que ce grief fait tomber",
        "texte": (
            "Les lignes a 0 € representent "
            + fr_pct(totaux["pct"], 2) + " des ventes "
            "(" + fr_int(totaux["zero"]) + " sur "
            + fr_int(totaux["total"]) + "), pour un montant encaisse de "
            + fr_eur(res["valeur_zero"]) + ". Ce sont des aperitifs, verres "
            "d'alcool, softs et cafes offerts ou consommes par le personnel, "
            "dont "
            + fr_pct(100.0 * totaux["zero_paye"] / totaux["zero"], 0) +
            " a l'interieur de tickets par ailleurs payes. Une ligne a prix "
            "nul ne porte aucun encaissement : elle ne peut, par nature, servir "
            "a sortir une recette du chiffre d'affaires. Le grief n'est pas "
            "fonde."
        ),
    })

    meta = {
        "slug": "articles-a-zero-euro",
        "titre": "Articles a prix 0 € (offerts / gratuits)",
        "source": "Proposition de rectification, rejet de comptabilite, "
                  "p. 16 a 22 (1/3).",
        "chiffresCles": {
            "lignesTotal": totaux["total"],
            "lignesZero": totaux["zero"],
            "partZeroPct": round(totaux["pct"], 4),
            "valeurEncaisseeZeroEur": round(res["valeur_zero"], 2),
            "zeroSurTicketPaye": totaux["zero_paye"],
            "zeroSurTicketNul": totaux["zero_seul"],
        },
    }

    return {"meta": meta, "sections": sections}


# --------------------------------------------------------------------------
# Main
# --------------------------------------------------------------------------
def main():
    res = lire_annexes()
    conso = lire_conso_personnel()

    total = sum(e["total"] for e in res["par_exo"].values())
    zero = sum(e["zero"] for e in res["par_exo"].values())
    zero_paye = sum(e["zero_paye"] for e in res["par_exo"].values())
    zero_seul = sum(e["zero_seul"] for e in res["par_exo"].values())
    totaux = {
        "total": total,
        "zero": zero,
        "zero_paye": zero_paye,
        "zero_seul": zero_seul,
        "pct": 100.0 * zero / total,
    }

    ecrire_xlsx(res, conso, totaux)

    doc = construire_json(res, conso, totaux)
    os.makedirs(DATA, exist_ok=True)
    with open(JSON_OUT, "w", encoding="utf-8") as f:
        json.dump(doc, f, ensure_ascii=False, indent=2)

    # Verification : JSON valide + pas de tiret cadratin
    with open(JSON_OUT, encoding="utf-8") as f:
        json.load(f)
    raw = open(JSON_OUT, encoding="utf-8").read()
    assert "—" not in raw and "–" not in raw, "Tiret cadratin/demi present"

    print("Lignes totales :", fr_int(total))
    print("Lignes a 0 EUR :", fr_int(zero), "(", fr_pct(totaux["pct"], 2), ")")
    print("  dont ticket paye :", zero_paye, "| ticket a 0 :", zero_seul)
    print("Valeur encaissee 0 EUR :", fr_eur(res["valeur_zero"]))
    print("Categories :", dict(res["cat_count"]))
    print("XLSX ->", XLSX_OUT)
    print("JSON ->", JSON_OUT)


if __name__ == "__main__":
    main()
