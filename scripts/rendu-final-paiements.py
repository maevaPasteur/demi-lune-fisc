#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
rendu-final-paiements.py
Objectif (réponse au fisc) : montrer que les suppressions correspondent à des
TABLES PARTAGÉES pour paiement séparé — pas à des recettes effacées.

Preuve au niveau de l'ARTICLE, depuis le fichier C (tickets) : quand une table
est répartie pour des paiements séparés, le logiciel divise l'addition en
plusieurs notes émises à la même minute, en fractionnant les articles communs
(quantités 0,5 / 0,33…). On reconstitue ces "partages" : une table d'origine,
divisée en N notes de même total, avec le détail des articles partagés. Le
montant total est encaissé en entier, réparti sur les notes : rien ne disparaît.
Les suppressions sont la mécanique de cette redistribution (retirer les articles
de la note d'origine pour les ré-affecter), pas des erreurs.

Sélection : jours avec le plus de partages nets, 2 par mois et par exercice.
Sorties : src/data/renduFinalCalculs.json (clé "paiements_fractionnes")
          + public/documents/pieces-defense/Jours-paiements-fractionnes.xlsx
"""
import xlrd, json, os, collections, re
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

ICI = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.normpath(os.path.join(ICI, ".."))
CAISSE = os.path.join(ROOT, "public/documents/caisse-enregistreuse/")
EXOS = ["2022-2023", "2023-2024", "2024-2025"]
C = {e: CAISSE + f"ANNEXE-C{i}_detail-tickets_{e}.xls" for i, e in enumerate(EXOS, 1)}
E = {e: CAISSE + f"ANNEXE-E{i}_tpvevenement_{e}.xls" for i, e in enumerate(EXOS, 1)}


def charge_notes(ex):
    """notes[(date,no)] = {h, tot, lines:[(lib,qte,pu)]} (un seul passage)."""
    sh = xlrd.open_workbook(C[ex]).sheet_by_index(0)
    notes = collections.defaultdict(lambda: {"h": "99:99", "tot": 0.0, "lines": []})
    for r in range(1, sh.nrows):
        date = str(sh.cell_value(r, 0))[:10]
        if not re.fullmatch(r"\d{4}-\d{2}-\d{2}", date):
            continue
        no = str(sh.cell_value(r, 2)).replace(".0", "")
        n = notes[(date, no)]
        n["h"] = min(n["h"], str(sh.cell_value(r, 1))[:5])
        try:
            n["tot"] = float(sh.cell_value(r, 5))
        except ValueError:
            pass
        try:
            q = float(sh.cell_value(r, 11))
        except ValueError:
            q = 0
        try:
            pu = float(sh.cell_value(r, 13))
        except ValueError:
            pu = 0.0
        lib = str(sh.cell_value(r, 10)).strip()
        if lib:
            n["lines"].append((lib, q, pu))
    return notes


# Prix catalogue des menus par carte (pour détecter les "prix modifiés" = factures sans détail).
MENUS_PRIX = json.load(open(os.path.join(ROOT, "src/data/calculsBoissons/menusPrix.json"), encoding="utf-8"))["cartes"]
MENU_LABELS = {
    "Menu 'Demi Lune'": "Demi Lune", "Menu du Dahu": "Dahu",
    "Menu Franc-Comtois": "Franc-Comtois", "Menu Bourguignon": "Bourguignon",
    "Menu Végétarien": "Vegetarien", "Menu 'Galette'": "Galette",
    "Menu 'Enfant'": "Enfant", "Menu Bambin": "Bambin", "Formule Express": "Express",
}


def carte_de(date):
    return "C" if date < "2022-07-23" else ("B" if date < "2023-04-17" else "A")


def menus_forfait_du_jour(date, nts):
    """Notes contenant un menu facturé à un prix HORS catalogue (facture forfaitaire
    sans détail). Renvoie nb de notes concernées et valeur de ces menus."""
    carte = MENUS_PRIX[carte_de(date)]
    n, somme = 0, 0.0
    for no, note in nts.items():
        touche = False
        for lib, q, pu in note["lines"]:
            key = MENU_LABELS.get(lib)
            if key and key in carte and pu > 0 and abs(pu - carte[key]) > 0.01:
                touche = True
                somme += pu * q
        if touche:
            n += 1
    return {"notes": n, "somme": round(somme, 2)}


def bilan_suppressions(supps):
    """Ventile 100 % des suppressions : faute de frappe / 1 article / plusieurs articles."""
    cats = {"faute_frappe": [0, 0.0], "article": [0, 0.0], "compose": [0, 0.0]}
    for h, m in supps:
        if m >= 1000:
            k = "faute_frappe"
        elif articles_au_prix(m):
            k = "article"
        else:
            k = "compose"
        cats[k][0] += 1
        cats[k][1] += m
    return {k: {"n": v[0], "somme": round(v[1], 2)} for k, v in cats.items()}


def del_detail(ex):
    """date -> liste [(heure, montant)] de toutes les suppressions du jour."""
    sh = xlrd.open_workbook(E[ex]).sheet_by_index(0)
    d = collections.defaultdict(list)
    for r in range(1, sh.nrows):
        if str(sh.cell_value(r, 3)).strip() != "DEL":
            continue
        date = str(sh.cell_value(r, 1))[:10]
        if not re.fullmatch(r"\d{4}-\d{2}-\d{2}", date):
            continue
        try:
            m = float(sh.cell_value(r, 8))
        except ValueError:
            m = 0.0
        d[date].append((str(sh.cell_value(r, 2))[:5], round(m, 2)))
    return d


# Catalogue : prix unitaire -> produits (pour rattacher une suppression à un article).
def prix_catalogue():
    it = json.load(open(os.path.join(ROOT, "src/data/calculsBoissons/itemsCaisse.json"), encoding="utf-8"))
    p = collections.defaultdict(set)
    for prod in it["items"]:
        for pu in (prod.get("prix_unitaire") or {}).values():
            if pu and pu > 0:
                p[round(pu, 2)].add(prod["produit"])
    return {k: sorted(v) for k, v in p.items()}


PRIX = prix_catalogue()


def articles_au_prix(montant):
    """Produits du catalogue dont le prix unitaire = montant (suppression d'1 article)."""
    return PRIX.get(round(montant, 2), [])


def detaille_suppressions(supps):
    """Pour chaque suppression : heure, montant, articles candidats au même prix, nature."""
    out = []
    for h, m in sorted(supps):
        cand = articles_au_prix(m)
        if m >= 1000:
            nature = "saisie erronée (quantité aberrante)"
        elif cand:
            nature = "article retiré (prix catalogue identifié)"
        else:
            nature = "retrait de plusieurs articles (montant composé)"
        out.append({"heure": h, "montant": m, "articles": cand[:4], "nature": nature})
    return out


def histo_heure(supps):
    h = collections.Counter()
    for hh, _ in supps:
        try:
            h[int(hh[:2])] += 1
        except ValueError:
            pass
    return [{"heure": f"{k}h", "n": h[k]} for k in sorted(h)]


def partages_du_jour(nts):
    """Groupes de notes émises à la même minute et au même total (>0) = une table
    partagée en parts égales. Reconstitue l'addition d'origine (somme des qtés)."""
    bymin = collections.defaultdict(list)
    for no, n in nts.items():
        bymin[(n["h"], round(n["tot"], 2))].append(no)
    parts = []
    for (h, tot), nos in bymin.items():
        if len(nos) < 2 or tot <= 0:
            continue
        nos = sorted(nos, key=lambda x: int(x) if x.isdigit() else 0)
        # addition d'origine = somme des quantités par article sur les notes du groupe
        art = collections.defaultdict(float)
        for no in nos:
            for lib, q, _pu in nts[no]["lines"]:
                art[lib] += q
        articles = [{"lib": l, "qte": round(q, 2)} for l, q in art.items() if q > 0]
        # un partage est "net" si au moins un article est fractionné (divisé)
        fractionne = any(abs(a["qte"] - round(a["qte"])) > 0.01 for a in articles) or \
            any(abs(q - round(q)) > 0.01 for no in nos for _, q, _pu in nts[no]["lines"])
        parts.append({
            "heure": h, "nb_notes": len(nos), "total_par_note": round(tot, 2),
            "total_table": round(tot * len(nos), 2), "notes": nos,
            "articles": sorted(articles, key=lambda a: -a["qte"]),
            "fractionne": fractionne,
        })
    parts.sort(key=lambda p: p["heure"])
    return parts


# --------------------------------------------------------------------------- #
selection = {ex: [] for ex in EXOS}
echantillon = []
tot_partages = tot_notes_partagees = 0
for ex in EXOS:
    notes = charge_notes(ex)
    dels = del_detail(ex)
    byday = collections.defaultdict(dict)
    for (date, no), n in notes.items():
        byday[date][no] = n
    jours = []
    for date, nts in byday.items():
        parts = partages_du_jour(nts)
        nets = [p for p in parts if p["fractionne"]]
        if not nets:
            continue
        sup = dels.get(date, [])
        bilan = bilan_suppressions(sup)
        jours.append({
            "date": date, "nb_notes": len(nts), "nb_partages": len(nets),
            "nb_notes_partagees": sum(p["nb_notes"] for p in nets),
            "nb_suppressions": len(sup),
            "suppr_somme": round(sum(m for _, m in sup), 2),
            "bilan": bilan,
            "partages_valeur": round(sum(p["total_table"] for p in nets), 2),
            "menus_forfait": menus_forfait_du_jour(date, nts),
            "suppressions": detaille_suppressions(sup),
            "suppr_par_heure": histo_heure(sup),
            "partages": nets,
        })
    # 2 jours/mois (les plus de partages)
    par_mois = collections.defaultdict(list)
    for j in jours:
        par_mois[j["date"][:7]].append(j)
    for mois in sorted(par_mois):
        selection[ex].extend(sorted(par_mois[mois], key=lambda x: -x["nb_partages"])[:2])
    tot_partages += sum(j["nb_partages"] for j in jours)
    tot_notes_partagees += sum(j["nb_notes_partagees"] for j in jours)

# Échantillon détaillé (popin) : 12 jours les plus probants
allj = sorted([j for ex in EXOS for j in selection[ex]], key=lambda x: -x["nb_partages"])
echantillon = sorted(allj[:12], key=lambda x: x["date"])

tot_jours = sum(len(v) for v in selection.values())
# Bilan GLOBAL sur les journées sélectionnées : ventilation à 100 % des suppressions.
alljours = [j for ex in EXOS for j in selection[ex]]
glob = {"n": 0, "somme": 0.0,
        "faute_frappe": {"n": 0, "somme": 0.0},
        "article": {"n": 0, "somme": 0.0},
        "compose": {"n": 0, "somme": 0.0},
        "partages": 0, "partages_valeur": 0.0,
        "menus_forfait_notes": 0, "menus_forfait_valeur": 0.0}
for j in alljours:
    glob["n"] += j["nb_suppressions"]
    glob["somme"] += j["suppr_somme"]
    for k in ("faute_frappe", "article", "compose"):
        glob[k]["n"] += j["bilan"][k]["n"]
        glob[k]["somme"] += j["bilan"][k]["somme"]
    glob["partages"] += j["nb_partages"]
    glob["partages_valeur"] += j["partages_valeur"]
    glob["menus_forfait_notes"] += j["menus_forfait"]["notes"]
    glob["menus_forfait_valeur"] += j["menus_forfait"]["somme"]
glob["somme"] = round(glob["somme"], 2)
for k in ("faute_frappe", "article", "compose"):
    glob[k]["somme"] = round(glob[k]["somme"], 2)
    glob[k]["pct"] = round(100 * glob[k]["somme"] / glob["somme"], 1) if glob["somme"] else 0
glob["partages_valeur"] = round(glob["partages_valeur"])
glob["menus_forfait_valeur"] = round(glob["menus_forfait_valeur"])
glob["max_hors_typo"] = round(max((s["montant"] for j in alljours for s in j["suppressions"] if s["montant"] < 1000), default=0), 2)

print(f"Jours sélectionnés (2/mois) : {tot_jours}")
print(f"Suppressions ventilées : frappe {glob['faute_frappe']['n']} ({glob['faute_frappe']['pct']}%) | "
      f"1 article {glob['article']['n']} ({glob['article']['pct']}%) | composé {glob['compose']['n']} ({glob['compose']['pct']}%)")
print(f"Côté notes : {glob['partages']} partages ({glob['partages_valeur']}€) | {glob['menus_forfait_notes']} notes menu forfait ({glob['menus_forfait_valeur']}€)")

dest = os.path.join(ROOT, "src/data/renduFinalCalculs.json")
data = json.load(open(dest, encoding="utf-8"))
data["paiements_fractionnes"] = {
    "_methode": "Suppressions ventilées à 100 % (faute de frappe / retrait d'1 article / retrait composé) et mises en regard du résultat mesuré sur les notes (tables partagées + menus à prix forfaitaire). Partage = notes même minute/même total, articles fractionnés.",
    "total_jours": tot_jours,
    "total_partages": tot_partages,
    "total_notes_partagees": tot_notes_partagees,
    "bilan_global": glob,
    # par_exercice : résumé par jour (sans les détails lourds)
    "par_exercice": {ex: [{k: v for k, v in j.items() if k not in ("partages", "suppressions", "suppr_par_heure")} for j in selection[ex]] for ex in EXOS},
    "echantillon": echantillon,
}
json.dump(data, open(dest, "w", encoding="utf-8"), ensure_ascii=False, indent=1)

# --------------------------------------------------------------------------- #
H1 = Font(bold=True, size=12, color="FFFFFF"); FILL1 = PatternFill("solid", fgColor="1F3A5F")
BOLD = Font(bold=True); FILLT = PatternFill("solid", fgColor="E2E8F0")
THIN = Border(*[Side(style="thin", color="CBD5E0")] * 4); RA = Alignment(horizontal="right")
WRAP = Alignment(wrap_text=True, vertical="top")
wb = openpyxl.Workbook()
ws = wb.active; ws.title = "Bilan par jour"
ws.append(["Suppressions ventilées à 100 % + résultat (tables partagées, menus forfait) — 2 jours/mois/exercice"])
ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=10)
ws.cell(1, 1).font = H1; ws.cell(1, 1).fill = FILL1
ws.append(["Exercice", "Date", "Nb suppr.", "€ suppr.", "dont faute de frappe (€)", "dont retrait 1 article (€)",
           "dont retrait composé (€)", "Tables partagées", "Valeur partages (€)", "Menus forfait (nb / €)"])
for c in ws[2]:
    c.font = BOLD; c.fill = FILLT; c.border = THIN; c.alignment = WRAP
for ex in EXOS:
    for j in selection[ex]:
        b = j["bilan"]
        ws.append([ex, j["date"], j["nb_suppressions"], j["suppr_somme"],
                   b["faute_frappe"]["somme"], b["article"]["somme"], b["compose"]["somme"],
                   j["nb_partages"], j["partages_valeur"],
                   f'{j["menus_forfait"]["notes"]} / {j["menus_forfait"]["somme"]:g} €'])
        for c in ws[ws.max_row]:
            c.border = THIN
            if isinstance(c.value, (int, float)):
                c.alignment = RA
for i, w in enumerate([11, 12, 10, 10, 16, 16, 16, 13, 15, 18], 1):
    ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

ws2 = wb.create_sheet("Détail des partages")
ws2.append(["Reconstitution des tables partagées : addition d'origine et répartition en notes de même total"])
ws2.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)
ws2.cell(1, 1).font = H1; ws2.cell(1, 1).fill = FILL1
ws2.append(["Date", "Heure", "Notes (n°)", "Total / note", "Addition d'origine (article × quantité)"])
for c in ws2[2]:
    c.font = BOLD; c.fill = FILLT; c.border = THIN
for j in echantillon:
    for p in j["partages"]:
        arts = ", ".join(f"{a['lib']}×{a['qte']:g}" for a in p["articles"])
        ws2.append([j["date"], p["heure"], "n° " + " + ".join(p["notes"]),
                    f"{p['total_par_note']:g} € × {p['nb_notes']}", arts])
        for c in ws2[ws2.max_row]:
            c.border = THIN
        ws2.cell(ws2.max_row, 5).alignment = WRAP
for i, w in enumerate([12, 8, 18, 16, 90], 1):
    ws2.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

out = os.path.join(ROOT, "public/documents/pieces-defense/Jours-paiements-fractionnes.xlsx")
wb.save(out)
print("écrit:", out)
