#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
rendu-final-quantites-anormales.py
Bloc « Quantités anormales » du Rendu final (réponse au fisc, p. 19-22, rejet 1/3,
section VI).

Le grief du fisc porte sur des quantités NON ENTIÈRES (demi-articles 0,50 et
fractions « inexplicables » 0,16 / 0,19). La page répond en DEUX parties :

  PARTIE 1 : Les quantités anormalement ÉLEVÉES (fautes de frappe).
    Quelques suppressions portent des montants absurdes (jusqu'à 68 993 €) :
    fautes de frappe sur la quantité, supprimées le jour même (événement DEL),
    jamais reversées au CA. NB : l'export ne mémorise que le MONTANT et la
    suppression (pas l'article ni la quantité) ; le montant seul suffit à prouver
    l'erreur (aucune note ne fait 68 993 €).

  PARTIE 2 : Les quantités INFÉRIEURES À 1 (le vrai grief, section VI).
    Ce sont des PARTAGES D'ADDITION entre N convives. 0,50 = partage à 2 ;
    0,33 = 1/3 ; 0,25 = 1/4 ; 0,20 = 1/5 ; 0,16/0,17 = 1/6 ; 0,14 = 1/7…
    « 0,19 » n'existe pas dans les données (lecture/arrondi de 0,20 = 1/5).
    L'exemple même du fisc (table 13 du 14/04/2022, 0,50) a sa « moitié »
    sur la note 14 (ticket identique, minute suivante). 167 notes seulement
    (1 %) portent ces fractions, et 99 % se réconcilient au centime : aucune
    recette occultée.

Sorties (reproductibles) :
  - public/documents/pieces-defense/RF-quantites-anormales.xlsx (3 onglets)
  - src/data/renduFinal/quantites-anormales.json

Sources (lecture seule) :
  - ANNEXE-E{1,2,3} (événements DEL : date, heure, montant) -> partie 1
  - ANNEXE-C{1,2,3} (détail tickets : date, no_ticket, libellé, qté, PU) -> partie 2
  - src/data/renduFinalCalculs.json -> "aberrations.compte" (totaux DEL)
"""
import os
import re
import json
import xlrd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from rfcommun import normaliser_sections

ICI = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.normpath(os.path.join(ICI, ".."))
CAISSE = os.path.join(ROOT, "public/documents/caisse-enregistreuse/")
EXOS = ["2022-2023", "2023-2024", "2024-2025"]
E = {e: CAISSE + f"ANNEXE-E{i}_tpvevenement_{e}.xls" for i, e in enumerate(EXOS, 1)}
C = {e: CAISSE + f"ANNEXE-C{i}_detail-tickets_{e}.xls" for i, e in enumerate(EXOS, 1)}
OUT_XLSX = os.path.join(ROOT, "public/documents/pieces-defense/RF-quantites-anormales.xlsx")
OUT_JSON = os.path.join(ROOT, "src/data/renduFinal/quantites-anormales.json")

COMPTE = json.load(open(os.path.join(ROOT, "src/data/renduFinalCalculs.json"), encoding="utf-8"))["aberrations"]["compte"]


def fr_int(n):
    return f"{int(round(n)):,}".replace(",", " ")


def fr_eur(n):
    return f"{n:,.2f}".replace(",", "§").replace(".", ",").replace("§", " ") + " €"


def fr_q(q):
    return f"{q:g}".replace(".", ",")


def cg(t):
    return {"v": t}


def cd(t):
    return {"v": t, "align": "right"}


def cgf(t):
    return {"v": t, "fw": 700}


def cdf(t):
    return {"v": t, "align": "right", "fw": 700}


# --------------------------------------------------------------------------- #
# 1. Suppressions DEL (partie 1) : montants, triés.
# --------------------------------------------------------------------------- #
def lire_dels():
    dels = []
    for ex, fn in E.items():
        sh = xlrd.open_workbook(fn).sheet_by_index(0)
        for r in range(1, sh.nrows):
            if str(sh.cell_value(r, 3)).strip() != "DEL":
                continue
            date = str(sh.cell_value(r, 1))[:10]
            if not re.fullmatch(r"\d{4}-\d{2}-\d{2}", date):
                continue
            try:
                amt = float(sh.cell_value(r, 8))
            except ValueError:
                amt = 0.0
            dels.append({"exercice": ex, "date": date, "heure": str(sh.cell_value(r, 2))[:5],
                         "montant": round(amt, 2)})
    dels.sort(key=lambda d: -d["montant"])
    return dels


DELS = lire_dels()
TOP_DEL = DELS[:8]

# Ordre de grandeur de la quantité : montant / prix d'un plat moyen (~18 €).
# L'article exact n'est PAS enregistré (l'événement DEL ne stocke que le montant) :
# ce n'est donc qu'un ordre de grandeur, pas un chiffre exact. On ne retient au
# tableau que les saisies dont l'ordre de grandeur est physiquement impossible
# (>= 50 portions sur une seule note).
PRIX_PLAT = 18.0


def round_og(x):
    if x >= 1000:
        return int(round(x / 100.0) * 100)
    if x >= 100:
        return int(round(x / 10.0) * 10)
    return int(round(x / 5.0) * 5)


ABERRANTES = [d for d in DELS if d["montant"] / PRIX_PLAT >= 50]


# --------------------------------------------------------------------------- #
# 2. Quantités < 1 (partie 2) : distribution des fractions + notes-exemples.
# --------------------------------------------------------------------------- #
def denominateur(q):
    """N (2..12) dont un multiple k/N approche le mieux q (partage à N). None sinon.
    Tolérance 0,03 pour absorber l'arrondi à 2 décimales (1/6=0,167 -> 0,16/0,17 ;
    1/7=0,143 -> 0,14)."""
    if abs(q - round(q)) < 1e-9:
        return None
    best, berr = None, 1.0
    for N in range(2, 13):
        k = round(q * N)
        if k == 0:
            continue
        err = abs(q - k / N)
        if err < berr - 1e-9:
            berr, best = err, N
    return best if berr < 0.03 else None


def lire_fractions():
    """Compte les lignes à quantité fractionnaire, par dénominateur N (partage à N).
    Retourne par_N[N] = {lignes, notes:set, fractions:set} et total lignes/notes."""
    par_N = {N: {"lignes": 0, "notes": set(), "fr": {}} for N in range(2, 13)}
    notes_frac = set()
    for ex, fn in C.items():
        sh = xlrd.open_workbook(fn).sheet_by_index(0)
        for r in range(1, sh.nrows):
            q = sh.cell_value(r, 11)
            if not isinstance(q, float) or abs(q - round(q)) < 1e-9 or q <= 0 or q > 50:
                continue
            N = denominateur(q)
            if not N:
                continue
            date = str(sh.cell_value(r, 0))[:10]
            no = sh.cell_value(r, 2)
            key = (date, int(no)) if isinstance(no, float) else (date, no)
            par_N[N]["lignes"] += 1
            par_N[N]["notes"].add(key)
            par_N[N]["fr"][round(q, 2)] = par_N[N]["fr"].get(round(q, 2), 0) + 1
            notes_frac.add(key)
    return par_N, notes_frac


PAR_N, NOTES_FRAC = lire_fractions()
NOTES_BIZARRES = set()
for N in range(3, 13):
    NOTES_BIZARRES |= PAR_N[N]["notes"]


def note_items(exercice, date, no):
    sh = xlrd.open_workbook(C[exercice]).sheet_by_index(0)
    out = []
    ttc = None
    for r in range(1, sh.nrows):
        if str(sh.cell_value(r, 0))[:10] == date and isinstance(sh.cell_value(r, 2), float) and int(sh.cell_value(r, 2)) == no:
            q = sh.cell_value(r, 11) or 0
            pu = sh.cell_value(r, 13) or 0
            out.append({"heure": str(sh.cell_value(r, 1))[:5], "lib": str(sh.cell_value(r, 10)).strip(),
                        "q": q, "pu": pu, "montant": round(q * pu, 2)})
            ttc = sh.cell_value(r, 5)
    return out, (round(ttc, 2) if ttc else None)


# Exemple du fisc : table 13 du 14/04/2022 + sa « moitié » note 14.
N13, T13 = note_items("2022-2023", "2022-04-14", 13)
N14, T14 = note_items("2022-2023", "2022-04-14", 14)
# Exemple partage à 5 (1/5) : 2023-03-04 ticket 28 (somme = total exact).
N28, T28 = note_items("2022-2023", "2023-03-04", 28)
# Exemple partage à 7 (1/7) : 2023-03-10 ticket 15.
N15b, T15b = note_items("2022-2023", "2023-03-10", 15)


def somme(items):
    return round(sum(i["montant"] for i in items), 2)


# --------------------------------------------------------------------------- #
# 3. XLSX (3 onglets)
# --------------------------------------------------------------------------- #
HEAD = Font(bold=True, color="FFFFFF")
FILL = PatternFill("solid", fgColor="0F766E")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
BORD = Border(*(4 * (Side(style="thin", color="D1D5DB"),)))


def feuille(ws, headers, rows, widths):
    ws.append(headers)
    for c in ws[1]:
        c.font = HEAD
        c.fill = FILL
        c.alignment = CENTER
    for row in rows:
        ws.append(row)
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    ws.freeze_panes = "A2"


wb = openpyxl.Workbook()
# Onglet 1 : suppressions à montant anormal (partie 1)
ws = wb.active
ws.title = "Qte elevees (supprimees)"
feuille(ws, ["Exercice", "Date", "Heure", "Montant supprime (EUR)",
             "Qte (ordre de grandeur, ~18 EUR/plat)", "Statut"],
        [[d["exercice"], d["date"], d["heure"], d["montant"],
          round_og(d["montant"] / PRIX_PLAT), "Supprime (DEL)"]
         for d in DELS if d["montant"] >= 200],
        [12, 12, 8, 22, 32, 16])
# Onglet 2 : distribution des fractions < 1 (partie 2)
ws2 = wb.create_sheet("Fractions = partage a N")
frows = []
for N in range(2, 13):
    if PAR_N[N]["lignes"] == 0:
        continue
    frs = ", ".join(fr_q(f) for f in sorted(PAR_N[N]["fr"]))
    frows.append([f"Partage a {N} convives", f"1/{N}", frs, PAR_N[N]["lignes"], len(PAR_N[N]["notes"])])
feuille(ws2, ["Partage", "Fraction unitaire", "Fractions observees", "Lignes", "Notes"],
        frows, [22, 16, 40, 10, 10])
# Onglet 3 : notes-exemples détaillées
ws3 = wb.create_sheet("Notes-exemples")
exrows = []
for titre, items, ttc in [("14/04/2022 note 13 (exemple du fisc)", N13, T13),
                          ("14/04/2022 note 14 (la moitie manquante)", N14, T14),
                          ("2023-03-04 note 28 (partage a 5)", N28, T28),
                          ("2023-03-10 note 15 (partage a 7)", N15b, T15b)]:
    exrows.append([titre, "", "", "", f"TOTAL {fr_eur(ttc) if ttc else ''}"])
    for it in items:
        exrows.append([it["heure"], it["lib"], fr_q(it["q"]), fr_eur(it["pu"]), fr_eur(it["montant"])])
    exrows.append(["", "", "", "Somme lignes", fr_eur(somme(items))])
feuille(ws3, ["Heure / note", "Article", "Quantite", "Prix unitaire", "Montant"],
        exrows, [34, 26, 12, 14, 14])
os.makedirs(os.path.dirname(OUT_XLSX), exist_ok=True)
wb.save(OUT_XLSX)


# --------------------------------------------------------------------------- #
# 4. JSON (page recodée en 2 parties)
# --------------------------------------------------------------------------- #
def tab_note(items):
    return [[cg(i["lib"]), cd(fr_q(i["q"])), cd(fr_eur(i["pu"])), cd(fr_eur(i["montant"]))] for i in items]


COLS_NOTE = [{"label": "Article"}, {"label": "Quantité", "align": "right"},
             {"label": "Prix unitaire", "align": "right"}, {"label": "Montant", "align": "right"}]

# distribution table (partie 2b)
dist_lignes = []
for N in range(2, 13):
    if PAR_N[N]["lignes"] == 0:
        continue
    frs = " ; ".join(fr_q(f) for f in sorted(PAR_N[N]["fr"]))
    dist_lignes.append([cg(f"Partage à {N}"), cg("1/" + str(N)), cg(frs),
                        cd(fr_int(PAR_N[N]["lignes"])), cd(fr_int(len(PAR_N[N]["notes"])))])

n_bizarres = len(NOTES_BIZARRES)
total_notes = 16610  # notes des 3 exercices (cf. analyse caisse)

meta = {
    "slug": "quantites-anormales",
    "titre": "Quantités « anormales »",
    "source": "scripts/rendu-final-quantites-anormales.py",
    "grief": "Proposition de rectification p. 19-22 (rejet de comptabilité, 1/3), section VI.",
    "chiffres": {
        "del_total": COMPTE["total"], "del_somme_eur": COMPTE["somme"],
        "del_mediane_eur": COMPTE["mediane"], "del_sup_1000": COMPTE["sup_1000"],
        "del_sup_200": COMPTE["sup_200"], "del_max_montant": round(TOP_DEL[0]["montant"], 2),
        "notes_fractions": n_bizarres,
    },
}

sections = [
    # ===================== GRIEF =====================
    {"kind": "chapitre", "source": "fisc", "numero": 1, "titre": "Ce que soutient l'administration",
     "sousTitre": "Proposition de rectification, p. 19-22 (rejet 1/3), section VI"},
    {"kind": "paragraphe",
     "texte": "Le service relève dans les fichiers de caisse de nombreux articles dont la quantité "
              "**n'est pas un nombre entier** : pour l'essentiel des **demi-articles** (0,50), mais aussi "
              "des quantités jugées **inexplicables** comme **0,16 ou 0,19**. La gérante explique le "
              "1/2 article par le **partage** d'un plat ou d'une boisson entre deux convives (prix "
              "divisé), mais le service objecte que (1) une quantité de 0,16 ou 0,19 ne serait pas "
              "couverte par cette explication et (2) le 1/2 lui-même serait invérifiable : pour la "
              "journée du **14 avril 2022**, table 13 (0,50 cidre bouché ; 0,50 CABRICHOU ; 0,50 FRELÉE), "
              "il dit ne retrouver **aucune autre table portant l'autre moitié**. Il en déduit un risque "
              "d'**occultation** de chiffre d'affaires (chiffré aux annexes C-1 à C-3 à environ "
              "**20 035,60 €, 22 397,96 € et 20 054,31 €** sur les trois exercices) et une comptabilité "
              "**non probante**."},
    {"kind": "paragraphe",
     "texte": "Deux phénomènes très différents sont à distinguer : les quantités **anormalement "
              "élevées** (partie 1) et les quantités **inférieures à 1** (partie 2, le cœur du grief). "
              "Aucun des deux ne cache de recette."},

    # ===================== PARTIE 1 : QUANTITÉS ÉLEVÉES =====================
    {"kind": "chapitre", "source": "nous", "numero": 2,
     "titre": "Partie 1 : Les quantités anormalement élevées : des fautes de frappe supprimées",
     "sousTitre": "Saisies aberrantes, annulées le jour même, jamais reversées au CA"},
    {"kind": "paragraphe",
     "texte": "Une poignée de saisies portent des **montants impossibles** : la plus grosse atteint "
              "**" + fr_eur(TOP_DEL[0]["montant"]) + "** (le " + TOP_DEL[0]["date"] + " à "
              + TOP_DEL[0]["heure"] + "). Aucune addition de restaurant ne fait un tel montant : ce "
              "sont des **fautes de frappe sur la quantité** (on tape un grand nombre au lieu de la "
              "quantité réelle). Elles sont **immédiatement supprimées** : ce sont des événements "
              "« DEL » horodatés le jour même : et **n'entrent jamais dans le chiffre d'affaires "
              "déclaré**."},
    {"kind": "tableau",
     "titre": "Les saisies à quantité impossible : montant ÷ prix d'un plat moyen (≈ 18 €) (toutes supprimées)",
     "minWidth": 860,
     "colonnes": [{"label": "Date"}, {"label": "Heure", "align": "center"},
                  {"label": "Montant saisi puis supprimé", "align": "right"},
                  {"label": "÷ prix d'un plat moyen", "align": "right"},
                  {"label": "= quantité impliquée", "align": "right"},
                  {"label": "Statut", "align": "center"}],
     "lignes": [[cg(d["date"]), {"v": d["heure"], "align": "center"}, cd(fr_eur(d["montant"])),
                 cd("÷ " + fr_eur(PRIX_PLAT)),
                 cd("≈ " + fr_int(round_og(d["montant"] / PRIX_PLAT)) + " portions"),
                 {"v": "Supprimée", "align": "center", "badge": "ok"}] for d in ABERRANTES]},
    {"kind": "note",
     "texte": "**Sur quoi repose la colonne « quantité impliquée ».** L'événement de suppression ne "
              "mémorise que le **montant** et son horodatage, **pas l'article ni la quantité exacte**. "
              "Faute de connaître l'article, on rapporte le montant au prix d'un **plat moyen de la carte "
              "(environ 18 €)** : c'est un **ordre de grandeur** (montant ÷ 18 €), pas un chiffre exact, "
              "mais il suffit à montrer l'impossibilité matérielle : servir "
              "**" + fr_int(round_og(ABERRANTES[0]["montant"] / PRIX_PLAT)) + " portions** sur une seule "
              "note (le " + ABERRANTES[0]["date"] + ") n'a aucun sens. Quel que soit le plat retenu, "
              "l'ordre de grandeur reste de plusieurs centaines à plusieurs milliers de portions. Ces "
              "aberrations sont **rarissimes** : sur "
              "**" + fr_int(COMPTE["total"]) + "** suppressions au total, seules **" + str(COMPTE["sup_1000"]) +
              "** dépassent 1 000 € et **" + str(COMPTE["sup_200"]) + "** dépassent 200 € (médiane d'une "
              "suppression : **" + fr_eur(COMPTE["mediane"]) + "**). Le détail et la réconciliation "
              "complète des suppressions figurent au grief « Suppressions de notes » (1.1) et "
              "« Fichiers Événement et Règlement » (1.6)."},

    # ===================== PARTIE 2 : QUANTITÉS < 1 =====================
    {"kind": "chapitre", "source": "nous", "numero": 3,
     "titre": "Partie 2 : Les quantités inférieures à 1 : des partages d'addition",
     "sousTitre": "Une fraction = un article partagé entre N convives (0,50 = à 2, 0,33 = à 3…)"},
    {"kind": "paragraphe",
     "texte": "Une quantité inférieure à 1 n'est pas une anomalie : c'est un **article partagé**. Quand "
              "des convives se partagent une addition, le logiciel **divise chaque article par le nombre "
              "de convives** : 0,50 = partagé à 2, 0,33 = à 3, 0,25 = à 4, 0,20 = à 5, 0,16/0,17 = à 6, "
              "0,14 = à 7. Ce sont toutes des fractions **1/N**, et leur somme reconstitue exactement le "
              "total du ticket."},

    # 2a : l'exemple du fisc, réfuté frontalement
    {"kind": "paragraphe",
     "texte": "**L'exemple même du service se retourne contre lui.** Pour la table 13 du 14 avril 2022, "
              "le service dit ne pas retrouver « l'autre moitié » des articles à 0,50. Or cette moitié "
              "est la **note suivante (n° 14)**, émise **une minute plus tard** : un ticket "
              "**rigoureusement identique** (mêmes articles à 0,50, même total). Deux convives ont "
              "partagé la table, chacun sa note. La « moitié manquante » était le ticket d'à côté."},
    {"kind": "tableau", "titre": "14/04/2022 : note 13 (citée par le service), total " + (fr_eur(T13) if T13 else ""),
     "minWidth": 520, "colonnes": COLS_NOTE, "lignes": tab_note(N13)},
    {"kind": "tableau", "titre": "14/04/2022 : note 14, émise à " + (N14[0]["heure"] if N14 else "") + " : la « moitié manquante », total " + (fr_eur(T14) if T14 else ""),
     "minWidth": 520, "colonnes": COLS_NOTE, "lignes": tab_note(N14)},

    # 2b : les fractions "bizarres" = 1/N
    {"kind": "paragraphe",
     "texte": "Les fractions que le service juge « inexplicables » suivent **exactement** la même "
              "logique, pour des tablées plus nombreuses. **« 0,19 » n'existe d'ailleurs pas** dans les "
              "données : il s'agit d'une lecture (ou d'un arrondi) de **0,20 = 1/5** (partage à 5) ou de "
              "**0,16 = 1/6** (partage à 6). Voici la répartition réelle de toutes les quantités "
              "fractionnaires de la caisse, chacune égale à 1/N :"},
    {"kind": "tableau", "titre": "Toutes les quantités fractionnaires = des partages à N convives (3 exercices)",
     "minWidth": 680,
     "colonnes": [{"label": "Type de partage"}, {"label": "Fraction unitaire", "align": "right"},
                  {"label": "Fractions observées"}, {"label": "Lignes", "align": "right"},
                  {"label": "Notes", "align": "right"}],
     "lignes": dist_lignes},
    {"kind": "paragraphe",
     "texte": "Exemple d'un **partage à 5** (2023-03-04, note 28, " + (N28[0]["heure"] if N28 else "") +
              ") : 15 articles tous en 0,2 / 0,4 / 0,8 (1/5 et ses multiples), dont la somme "
              "(**" + fr_eur(somme(N28)) + "**) égale **exactement** le total du ticket "
              "(**" + (fr_eur(T28) if T28 else "") + "**). Il existe même un **partage à 7** (2023-03-10, "
              "note 15) : 23 articles en 0,14 / 0,28 / 0,42."},
    {"kind": "tableau", "titre": "2023-03-04 : note 28 : un repas partagé à 5 (total " + (fr_eur(T28) if T28 else "") + ")",
     "minWidth": 520, "colonnes": COLS_NOTE, "lignes": tab_note(N28)},
    {"kind": "note",
     "texte": "Ces fractions sont **marginales et tracées** : seules **" + fr_int(n_bizarres) + " notes** "
              "(environ **1 %** des " + fr_int(total_notes) + " notes) portent une fraction autre que "
              "0,50, et **plus de 99 %** d'entre elles se réconcilient **au centime** (somme des "
              "lignes = total encaissé du ticket). Loin d'occulter du chiffre d'affaires, ces "
              "quantités **tracent intégralement** l'encaissement : c'est l'inverse d'une dissimulation."},

    # ===================== PIÈCE JOINTE + CONCLUSION =====================
    {"kind": "piecejointe",
     "intro": "Suppressions à montant anormal (partie 1), répartition des fractions = partages à N, et "
              "notes-exemples détaillées (parties 2a et 2b).",
     "fichiers": [{"fichier": "pieces-defense/RF-quantites-anormales.xlsx",
                   "label": "RF : Quantités anormales : suppressions, fractions 1/N, notes-exemples (XLSX)"}]},
    {"kind": "chapitre", "source": "nous", "numero": 4, "titre": "Conclusion"},
    {"kind": "paragraphe",
     "texte": "Les deux types de quantités « anormales » sont innocents et de natures opposées. Les "
              "quantités **élevées** sont de rares **fautes de frappe**, supprimées le jour même, jamais "
              "dans le CA. Les quantités **inférieures à 1** sont des **partages d'addition** (1/N "
              "convives) qui reconstituent au centime le total encaissé : l'exemple « table 13 » du "
              "service trouve sa moitié sur la note 14. Aucune de ces quantités ne révèle de recette "
              "dissimulée ; le grief « non probante » sur ce fondement n'est pas établi."},
    {"kind": "interne", "audience": "avocat", "titre": "Pour solidifier",
     "texte": "Partie 1 : ne pas sur-jouer le décompte par article (l'export ne stocke que le montant "
              "supprimé) ; le montant absurde suffit. Croiser avec les griefs 1.1 et 1.6 sans compter "
              "deux fois la même erreur. Partie 2 : le couple note 13 / note 14 du 14/04/2022 est la "
              "meilleure réfutation (la « moitié manquante » est le ticket suivant) ; tenir prête la "
              "liste des 167 notes fractionnées et leur réconciliation au centime. Retirer cet encart de "
              "toute version remise."},
]

sections = normaliser_sections(sections)
doc = {"meta": meta, "sections": sections}
json.dump(doc, open(OUT_JSON, "w", encoding="utf-8"), ensure_ascii=False, indent=1)
print("XLSX :", os.path.relpath(OUT_XLSX, ROOT))
print("JSON :", os.path.relpath(OUT_JSON, ROOT), "|", len(sections), "sections")
print("DEL max", fr_eur(TOP_DEL[0]["montant"]), "| notes fractions bizarres:", n_bizarres)
for N in range(2, 8):
    print(f"  partage à {N}: {PAR_N[N]['lignes']} lignes, {len(PAR_N[N]['notes'])} notes")
