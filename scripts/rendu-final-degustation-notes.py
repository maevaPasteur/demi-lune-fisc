#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
RENDU FINAL - Degustation offerte, NOTE PAR NOTE (preuve au ticket).
Demande : lister toutes les notes (additions) comportant des vins NOMMES vendus
au verre ou au pichet (hors cubis = type non precise), avec quantite et prix,
et compter 2 cl de degustation UNE SEULE FOIS par vin et par note (peu importe
la quantite). Total general des litres offerts en bas.

Source : public/documents/caisse-enregistreuse/ANNEXE-C{1,2,3}_detail-tickets_*.xls
  col 0 = date ticket, col 1 = heure, col 2 = n. ticket (note), col 10 = libelle,
  col 11 = quantite, col 13 = prix unitaire TTC.
Lecture seule, reproductible. Sortie :
  public/documents/pieces-defense/RF-degustation-par-note.xlsx

Regle d'inclusion (hors cubis) : un libelle est compte SI il nomme une
appellation/cepage precis servi au verre/pichet. Sont EXCLUS :
  - les generiques cubis : "Verre de vin", "Pichet 50cl vin", "Pichet 75cl vin"
    (la caisse ne precise pas le type -> vin de cubis) ;
  - les bouteilles (vendues scellees, pas de degustation) ;
  - l'effervescent (cremant, champagne : traite a part - cremant jete).
"""
import os
import json
import collections
import xlrd

ICI = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.normpath(os.path.join(ICI, ".."))
CAISSE = os.path.join(ROOT, "public/documents/caisse-enregistreuse/")
PIECES = os.path.join(ROOT, "public/documents/pieces-defense/")
RENDU = os.path.join(ROOT, "src/data/renduFinal/")
EXOS = ["2022-2023", "2023-2024", "2024-2025"]
C = {e: CAISSE + f"ANNEXE-C{i}_detail-tickets_{e}.xls" for i, e in enumerate(EXOS, 1)}
OUT = os.path.join(PIECES, "RF-degustation-par-note.xlsx")
JSON_OUT = os.path.join(RENDU, "degustation-notes.json")
DOSE_DEG = 2.0  # cl offerts par degustation

# --- Libelles caisse -> vin nomme (verre/pichet, hors cubis/bouteille) --------
# Base = mapping valide de scripts/rendu-final-calculs.py, complete du verre
# "Beaujolais Moulin a" (oubli : seul le PICHET y figurait).
TASTE = {
    "Savagnin verre": "Savagnin", "PICHET SAVAGNIN": "Savagnin",
    "Arbois Trousseau Le": "Arbois Trousseau", "PiCHET TROUSSEAU": "Arbois Trousseau",
    "Saint Véran Verre": "Saint Véran", "PICHET SAINT VERAN": "Saint Véran",
    "PICHET ALIGOTE 50cL": "Aligoté", "Verre ALIGOTE": "Aligoté",
    "Pichet CHUSLAN": "Chusclan", "Pichet CdR CHUSCLAN": "Chusclan",
    "VERRE CHUSCLAN": "Chusclan", "Verre CHUSCLAN": "Chusclan",
    "Arbois Chardonnay Le": "Arbois Chardonnay",
    "Chablis Le Verre": "Chablis", "Pichet Chablis": "Chablis",
    "Gewurztraminer Le ve": "Gewurztraminer", "PICHET GEWURZTRAMINE": "Gewurztraminer",
    "MACON VERRE": "Mâcon", "PICHET MACON": "Mâcon",
    "H.C.DE BEAUNE  VERRE": "HC de Beaune", "VERRE H.C.DE BEAUNE": "HC de Beaune",
    "PICHET C.DE BEAUNE R": "HC de Beaune", "PICHET C.DE BEAUNE B": "HC de Beaune",
    "PICHET HCB": "HC de Beaune",
    "Beaujolais Moulin à": "Moulin à Vent", "PICHET MOULIN A VENT": "Moulin à Vent",
    "pichet Saint Joseph": "Saint Joseph",
}


def fmt_service(lib):
    return "Pichet" if "pich" in lib.lower() else "Verre"


def fr_eur(x):
    return f"{x:.2f}".replace(".", ",") + " €"


def lire_lignes():
    """Toutes les lignes de vin nomme au verre/pichet, par note."""
    lignes = []  # (exo, date, heure, note, lib, vin, format, qte, prix)
    for exo, fn in C.items():
        sh = xlrd.open_workbook(fn).sheet_by_index(0)
        for r in range(1, sh.nrows):
            lib = str(sh.cell_value(r, 10)).strip()
            vin = TASTE.get(lib)
            if not vin:
                continue
            try:
                qte = float(sh.cell_value(r, 11))
            except ValueError:
                qte = 0.0
            if qte <= 0:
                continue
            try:
                prix = float(sh.cell_value(r, 13))
            except ValueError:
                prix = 0.0
            date = str(sh.cell_value(r, 0))[:10]
            heure = str(sh.cell_value(r, 1))
            note = str(sh.cell_value(r, 2))
            lignes.append((exo, date, heure, note, lib, vin, fmt_service(lib), qte, prix))
    # tri : exercice, date, note, vin (pour regrouper visuellement par note)
    lignes.sort(key=lambda t: (t[0], t[1], t[3], t[5], t[4]))
    return lignes


def ecrire_xlsx(lignes):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    white = Font(bold=True, color="FFFFFF")
    bold = Font(bold=True)
    head = PatternFill("solid", fgColor="0F766E")
    sub = PatternFill("solid", fgColor="E6F4F1")
    deg_fill = PatternFill("solid", fgColor="FEF3C7")  # ambre : la degustation comptee
    thin = Side(style="thin", color="D8DEE4")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    right = Alignment(horizontal="right")

    # ---- Feuille 1 : Degustation par note --------------------------------
    ws = wb.active
    ws.title = "Dégustation par note"
    cols = ["Exercice", "Date", "Heure", "N° note", "Libellé caisse", "Vin nommé",
            "Service", "Quantité", "Prix unit.", "Dégustation (cl)"]
    ws.append(["Dégustation offerte - preuve note par note (vins nommés au verre/pichet, hors cubis)"])
    ws["A1"].font = Font(bold=True, size=13)
    ws.append(["2 cl offerts UNE SEULE FOIS par vin et par note (quelle que soit la quantité). "
               "Génériques « Verre de vin / Pichet vin » (cubis), bouteilles et effervescent exclus."])
    ws["A2"].font = Font(italic=True, size=9, color="64748B")
    ws.append([])
    hrow = 4
    ws.append(cols)
    for c in range(1, len(cols) + 1):
        cell = ws.cell(row=hrow, column=c)
        cell.font = white
        cell.fill = head
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    seen = set()          # (exo, date, note, vin) deja degustes
    total_deg = 0
    for (exo, date, heure, note, lib, vin, serv, qte, prix) in lignes:
        key = (exo, date, note, vin)
        premier = key not in seen
        if premier:
            seen.add(key)
            total_deg += 1
        deg = DOSE_DEG if premier else None
        ws.append([exo, date, heure, note, lib, vin, serv, qte,
                   round(prix, 2), deg])
        r = ws.max_row
        for c in range(1, len(cols) + 1):
            ws.cell(row=r, column=c).border = border
        ws.cell(row=r, column=8).alignment = right
        ws.cell(row=r, column=9).number_format = '#,##0.00 €'
        ws.cell(row=r, column=9).alignment = right
        ws.cell(row=r, column=10).alignment = right
        if premier:
            ws.cell(row=r, column=10).fill = deg_fill
            ws.cell(row=r, column=10).font = bold

    total_l = round(total_deg * DOSE_DEG / 100, 1)
    ws.append([])
    tr = ws.max_row + 1
    ws.cell(row=tr, column=1, value="TOTAL")
    ws.cell(row=tr, column=6, value=f"{total_deg} dégustations (1 / note / vin)")
    ws.cell(row=tr, column=10, value=total_deg * DOSE_DEG)
    ws.cell(row=tr + 1, column=1, value="TOTAL DÉGUSTATION (litres)")
    ws.cell(row=tr + 1, column=10, value=total_l)
    for rr in (tr, tr + 1):
        for c in range(1, len(cols) + 1):
            ws.cell(row=rr, column=c).font = bold
            ws.cell(row=rr, column=c).fill = sub
    ws.cell(row=tr, column=10).number_format = '#,##0 "cl"'
    ws.cell(row=tr + 1, column=10).number_format = '#,##0.0 "L"'

    widths = [12, 12, 8, 9, 22, 18, 9, 10, 11, 14]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A5"

    # ---- Feuille 2 : Synthese par vin ------------------------------------
    ws2 = wb.create_sheet("Synthèse par vin")
    par_vin = collections.Counter()
    for (exo, date, h, note, lib, vin, serv, qte, prix) in lignes:
        par_vin[(exo, date, note, vin)] += 0  # init
    # nb degustations par vin = nb de (note,vin) distincts
    degu_par_vin = collections.Counter()
    for (exo, date, note, vin) in {(e, d, n, v) for (e, d, h, n, l, v, s, q, p) in lignes}:
        degu_par_vin[vin] += 1
    ws2.append(["Synthèse par vin - nombre de dégustations et litres offerts (3 exercices)"])
    ws2["A1"].font = Font(bold=True, size=13)
    ws2.append([])
    ws2.append(["Vin nommé", "Dégustations (notes distinctes)", "Litres offerts"])
    for c in range(1, 4):
        cell = ws2.cell(row=3, column=c)
        cell.font = white
        cell.fill = head
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    for vin, n in degu_par_vin.most_common():
        ws2.append([vin, n, round(n * DOSE_DEG / 100, 1)])
        r = ws2.max_row
        for c in range(1, 4):
            ws2.cell(row=r, column=c).border = border
        ws2.cell(row=r, column=3).number_format = '#,##0.0 "L"'
    tot = sum(degu_par_vin.values())
    ws2.append(["TOTAL", tot, round(tot * DOSE_DEG / 100, 1)])
    for c in range(1, 4):
        ws2.cell(row=ws2.max_row, column=c).font = bold
        ws2.cell(row=ws2.max_row, column=c).fill = sub
    ws2.cell(row=ws2.max_row, column=3).number_format = '#,##0.0 "L"'
    for i, w in enumerate([22, 30, 16], 1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    os.makedirs(PIECES, exist_ok=True)
    wb.save(OUT)
    return total_deg, total_l, degu_par_vin


def ecrire_json(total_deg, total_l, par_vin, n_lignes):
    doc = {
        "_source": "scripts/rendu-final-degustation-notes.py (ANNEXE-C, niveau note)",
        "methode": "2 cl offerts une seule fois par vin nommé et par note (date+n° ticket) ; cubis et bouteilles exclus.",
        "dose_cl": DOSE_DEG,
        "lignes_vin": n_lignes,
        "total_degustations": total_deg,
        "total_offert_l": total_l,
        "par_vin": [{"vin": v, "degustations": n, "offert_l": round(n * DOSE_DEG / 100, 1)}
                    for v, n in par_vin.most_common()],
    }
    os.makedirs(RENDU, exist_ok=True)
    with open(JSON_OUT, "w", encoding="utf-8") as f:
        json.dump(doc, f, ensure_ascii=False, indent=1)


def main():
    lignes = lire_lignes()
    total_deg, total_l, par_vin = ecrire_xlsx(lignes)
    ecrire_json(total_deg, total_l, par_vin, len(lignes))
    print("RENDU FINAL - Degustation par note")
    print("-" * 50)
    print(f"  lignes de vin nomme (verre/pichet) : {len(lignes)}")
    print(f"  degustations (note x vin distinct) : {total_deg}")
    print(f"  DOSE                               : {DOSE_DEG} cl")
    print(f"  TOTAL degustation                  : {total_l} L")
    print("  par vin :")
    for vin, n in par_vin.most_common():
        print(f"    {vin:22s} {n:5d} deg  {round(n*DOSE_DEG/100,1):5} L")
    print("-" * 50)
    print(f"  XLSX : {OUT}")


if __name__ == "__main__":
    main()
