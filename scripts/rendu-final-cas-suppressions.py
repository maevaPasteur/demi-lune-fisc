#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
rendu-final-cas-suppressions.py
Suppressions présentées PAR CAS, chaque exemple en FICHE détaillée :
à gauche les suppressions (date/heure/montant), à droite l'encaissement
correspondant (note = table, heure, contenu détaillé).

DÉFINITIONS (validées) :
  A. Conversion au forfait — rafale de suppressions dont la somme = total EXACT
     d'UNE note dont TOUTES les lignes sont des menus (facture sans détail).
  B. Paiement séparé / partage — ≥2 notes de même total à la même minute, avec
     articles divisés (qté fractionnaire). À gauche : uniquement les vraies
     lignes DEL dont le montant = prix d'un article divisé.
  C. Article déplacé — rafale d'AU MOINS 4 suppressions dont la somme = total
     EXACT d'une note qui contient des articles DÉTAILLÉS (≥1 ligne non-menu) :
     les plats ont été déplacés vers la bonne table.

Sélection : ~1 exemple par mois et par année (≈36 par cas).
Sortie : src/data/renduFinalCasSuppressions.json.
"""
import xlrd, json, os, collections
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

ICI = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.normpath(os.path.join(ICI, ".."))
CAISSE = os.path.join(ROOT, "public/documents/caisse-enregistreuse/")
EXOS = ["2022-2023", "2023-2024", "2024-2025"]
C = {e: CAISSE + f"ANNEXE-C{i}_detail-tickets_{e}.xls" for i, e in enumerate(EXOS, 1)}
E = {e: CAISSE + f"ANNEXE-E{i}_tpvevenement_{e}.xls" for i, e in enumerate(EXOS, 1)}


def is_menu(lib):
    return lib.lower().startswith(("menu", "formule"))


def tm(h):
    try:
        return int(h[:2]) * 60 + int(h[3:5])
    except Exception:
        return -1


def items_of(n):
    return [{"lib": l, "qte": round(q, 2), "pu": round(pu, 2), "montant": round(q * pu, 2)}
            for l, q, pu in n["lines"]]


def ym(d):
    return d[:7]


# Un exemple par (année-mois) et par cas. On garde le premier qui qualifie.
A, Cc = {}, {}
PART_CANDS = []  # tous les candidats « partage », triés/filtrés ensuite
# Listes EXHAUSTIVES (pour la pièce jointe XLSX).
FA, FB, FC = [], [], []
for ex in EXOS:
    shc = xlrd.open_workbook(C[ex]).sheet_by_index(0)
    notes_by = collections.defaultdict(lambda: collections.defaultdict(lambda: {"h": "", "tot": 0, "lines": []}))
    for r in range(1, shc.nrows):
        d = str(shc.cell_value(r, 0))[:10]
        if not d.startswith("20"):
            continue
        no = str(shc.cell_value(r, 2)).replace(".0", "")
        n = notes_by[d][no]
        n["h"] = str(shc.cell_value(r, 1))[:5]
        try:
            n["tot"] = round(float(shc.cell_value(r, 5)), 2)
        except ValueError:
            pass
        try:
            q = float(shc.cell_value(r, 11)); pu = round(float(shc.cell_value(r, 13) or 0), 2)
        except ValueError:
            q, pu = 0, 0
        n["lines"].append((str(shc.cell_value(r, 10)).strip(), q, pu))
    she = xlrd.open_workbook(E[ex]).sheet_by_index(0)
    dels_by = collections.defaultdict(list)
    for r in range(1, she.nrows):
        d = str(she.cell_value(r, 1))[:10]
        if d.startswith("20") and str(she.cell_value(r, 3)).strip() == "DEL":
            dels_by[d].append((str(she.cell_value(r, 2))[:5], round(float(she.cell_value(r, 8) or 0), 2)))

    for d in sorted(notes_by):
        k = ym(d)
        notes = notes_by[d]
        dels = dels_by.get(d, [])
        burst = collections.defaultdict(list)
        for h, m in dels:
            burst[h].append(m)

        # --- A : FORFAIT (somme = total, note 100 % menus) -------------------
        for h, ms in sorted(burst.items()):
            if len(ms) < 2:
                continue
            s = round(sum(ms), 2)
            for no, n in notes.items():
                if (abs(n["tot"] - s) < 0.05 and abs(tm(n["h"]) - tm(h)) <= 8
                        and n["lines"] and all(is_menu(l) for l, q, pu in n["lines"])):
                    menus = "; ".join(f"{l} {round(q,2)}×{round(pu,2)}" for l, q, pu in n["lines"])
                    FA.append((d, h, len(ms), round(n["tot"], 2), no, menus))
                    A.setdefault(k, {"date": d,
                                     "suppressions": [{"heure": h, "montant": round(x, 2)} for x in sorted(ms)],
                                     "notes": [{"label": f"Note n°{no}", "heure": n["h"], "total": round(n["tot"], 2), "items": items_of(n)}]})
                    break

        # --- C : DÉPLACEMENT (≥4 suppr, somme = total, note avec détail) -----
        for h, ms in sorted(burst.items()):
            if len(ms) < 4:
                continue
            s = round(sum(ms), 2)
            for no, n in notes.items():
                if (abs(n["tot"] - s) < 0.05 and abs(tm(n["h"]) - tm(h)) <= 8
                        and n["lines"] and any(not is_menu(l) for l, q, pu in n["lines"])):
                    arts = "; ".join(f"{l} {round(q,2)}×{round(pu,2)}" for l, q, pu in n["lines"])
                    FC.append((d, h, len(ms), round(n["tot"], 2), no, arts))
                    Cc.setdefault(k, {"date": d,
                                      "suppressions": [{"heure": h, "montant": round(x, 2)} for x in sorted(ms)],
                                      "notes": [{"label": f"Note n°{no}", "heure": n["h"], "total": round(n["tot"], 2), "items": items_of(n)}]})
                    break

        # --- B : PARTAGE / paiement séparé (lié aux suppressions DEL) --------
        # Preuve : une suppression DEL dont le PRIX réapparaît en quantité
        # FRACTIONNAIRE réparti sur ≥2 notes encaissées dans la foulée, la somme
        # des fractions valant un ENTIER (l'article supprimé a été re-saisi divisé
        # entre les convives — on ne vend pas un demi-poulet). On collecte large,
        # puis (après la boucle) on retient pour chaque mois la fiche la plus
        # serrée via une escalade de paliers (fenêtre courte d'abord).
        nosort = lambda x: int(x) if x.isdigit() else 0
        FEN, SPREAD = 60, 30  # bornes larges ; l'escalade privilégie le plus serré
        for h, ms in sorted(burst.items()):
            T = tm(h)
            cand = [(no, n) for no, n in notes.items() if -FEN <= tm(n["h"]) - T <= FEN]
            if len(cand) < 2:
                continue
            pu_notes = collections.defaultdict(set)      # prix -> notes (qté fractionnaire)
            pu_lib = collections.defaultdict(set)        # prix -> libellés
            pu_fracsum = collections.defaultdict(float)  # prix -> somme des qtés fractionnaires
            for no, n in cand:
                for l, q, pu in n["lines"]:
                    if abs(q - round(q)) > 0.01 and pu > 0:
                        pu_notes[round(pu, 2)].add(no)
                        pu_lib[round(pu, 2)].add(l)
                        pu_fracsum[round(pu, 2)] += q
            # un montant DEL « correspond » si son prix est réparti fractionné sur
            # ≥2 notes ET si la somme des fractions reconstitue un entier de parts.
            def matched_ok(m):
                m = round(m, 2)
                s = pu_fracsum.get(m, 0)
                return len(pu_notes.get(m, ())) >= 2 and abs(s - round(s)) < 0.02 and round(s) >= 1
            matched = sorted({round(m, 2) for m in ms if matched_ok(m)})
            if not matched:
                continue
            impl = sorted({no for m in matched for no in pu_notes[m]}, key=nosort)
            if len(impl) < 2:
                continue
            hs = [tm(notes[x]["h"]) for x in impl]
            spread = max(hs) - min(hs)
            if spread > SPREAD:
                continue
            dist = max(abs(t - T) for t in hs)  # éloignement max à la rafale
            mset = set(matched)
            lib_of = lambda m: " / ".join(sorted(pu_lib[m]))

            def items_hi(n):
                return [{"lib": l, "qte": round(q, 2), "pu": round(pu, 2), "montant": round(q * pu, 2),
                         "highlight": abs(q - round(q)) > 0.01 and round(pu, 2) in mset}
                        for l, q, pu in n["lines"]]
            sup = [{"heure": h, "montant": m, "devient": lib_of(m),
                    "tables": [f"n°{x}" for x in sorted(pu_notes[m], key=nosort)]}
                   for m in matched]
            PART_CANDS.append({
                "date": d,
                "dist": dist, "spread": spread, "n_del": len(ms),
                "n_matched": len(matched), "n_tables": len(impl),
                "suppressions": sup,
                "notes": [{"label": f"Note n°{x}", "heure": notes[x]["h"],
                           "total": round(notes[x]["tot"], 2), "items": items_hi(notes[x])} for x in impl]})
            FB.append((d, h, ", ".join(f"n°{x}" for x in impl), len(matched), len(impl),
                       "; ".join(f"{lib_of(m)} ×{round(pu_fracsum[m])} ({m:.2f} €)" for m in matched)))

A = [A[k] for k in sorted(A)]
Cc = [Cc[k] for k in sorted(Cc)]
# Partage : pour chaque mois, on retient la fiche la plus DÉFENDABLE via une
# escalade de paliers (du plus strict au plus large). On ne descend d'un palier
# que pour les mois qu'aucun palier plus strict n'a couverts ; à l'intérieur d'un
# palier, on prend la fiche la plus probante (+ d'articles correspondants, + de
# tables, fenêtre la plus serrée). Chaque palier garde la preuve : suppression(s)
# DEL dont le prix se reconstitue en fractions sommant à un entier sur ≥2 tables.
PALIERS = [  # (n_del_min, dist_max, spread_max, n_matched_min)
    (2, 15, 12, 2), (2, 20, 15, 2), (2, 30, 15, 2),
    (1, 20, 15, 1), (2, 45, 20, 2), (2, 60, 30, 2), (1, 60, 30, 1),
]
_best = {}
for (nd, dmax, smax, mmin) in PALIERS:
    for c in sorted(PART_CANDS, key=lambda x: (x["n_matched"], x["n_tables"], -x["dist"]), reverse=True):
        k = c["date"][:7]
        if k in _best:
            continue
        if c["n_del"] >= nd and c["dist"] <= dmax and c["spread"] <= smax and c["n_matched"] >= mmin:
            _best[k] = c
B = [_best[k] for k in sorted(_best)]
for c in B:
    for f in ("dist", "spread", "n_del", "n_matched", "n_tables"):
        c.pop(f, None)

cas = [
    {"id": "forfait", "titre": "Conversion au forfait (facture sans détail)",
     "preuve": "Preuve exacte : la somme des suppressions = le total d’une note composée uniquement de menus",
     "description": "À la demande d’une tablée (groupe, comité), le détail à la carte est retiré et l’addition est re-facturée en menu(s) à un prix négocié. La recette n’a pas disparu : la somme des lignes supprimées égale, au centime, le total de la note encaissée, qui ne contient plus que des menus.",
     "exemples": A},
    {"id": "partage", "titre": "Paiement séparé / partage de table",
     "preuve": "Une suppression dont le prix réapparaît, divisé en fractions sommant à un entier, réparti sur au moins 2 notes encaissées dans la foulée",
     "description": "Des convives d’une même table paient séparément. Le logiciel n’acceptant qu’un règlement par note, l’article commun est **supprimé** puis ré-enregistré **divisé** sur plusieurs notes (0,5 chacun, ou 0,33/0,33/0,34 à trois). La suppression n’efface donc aucune recette : son **prix réapparaît**, en quantités fractionnaires réparties sur ≥2 notes encaissées dans la foulée, et la **somme des fractions reconstitue un nombre entier** de parts. À gauche, chaque suppression et l’article qu’elle redevient ; à droite, les notes qui se le partagent (article surligné).",
     "exemples": B},
    {"id": "deplacement", "titre": "Articles déplacés vers la bonne table",
     "preuve": "Preuve exacte : une rafale d’au moins 4 suppressions dont la somme = total d’une note détaillée encaissée juste après",
     "description": "Des plats saisis sur la mauvaise table sont supprimés en bloc puis ré-enregistrés sur la bonne. La somme des suppressions égale, au centime, le total d’une note qui contient le détail des plats déplacés (au moins une ligne hors menu).",
     "exemples": Cc},
]
json.dump({"cas": cas}, open(os.path.join(ROOT, "src/data/renduFinalCasSuppressions.json"), "w", encoding="utf-8"), ensure_ascii=False, indent=1)
print(f"exemples (1/mois/an) : forfait {len(A)} | partage {len(B)} | déplacement {len(Cc)}")
for nom, arr in (("forfait", A), ("partage", B), ("déplacement", Cc)):
    print(f"  {nom:12s} mois couverts :", ", ".join(e["date"][:7] for e in arr))

# --- Pièce jointe XLSX : listes exhaustives, 3 onglets -----------------------
HEAD = Font(bold=True, color="FFFFFF")
FILL = PatternFill("solid", fgColor="1F2933")
wb = openpyxl.Workbook()
wb.remove(wb.active)
ONGLETS = [
    ("Conversion forfait", ["Date", "Heure", "Nb suppr.", "Somme = Total (€)", "Note n°", "Menus encaissés"], FA),
    ("Paiement separe", ["Date", "Heure rafale", "Tables", "Nb articles correspondants", "Nb tables", "Articles répartis (prix)"], FB),
    ("Articles deplaces", ["Date", "Heure", "Nb suppr.", "Somme = Total (€)", "Note n°", "Articles encaissés"], FC),
]
for nom, cols, rows in ONGLETS:
    ws = wb.create_sheet(nom)
    ws.append(cols)
    for c in ws[1]:
        c.font = HEAD; c.fill = FILL; c.alignment = Alignment(horizontal="center")
    for row in sorted(rows):
        ws.append(list(row))
    for i, w in enumerate((12, 8, 11, 18, 9, 70), 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
out = os.path.join(ROOT, "public/documents/pieces-defense/Cas-suppressions.xlsx")
wb.save(out)
print(f"XLSX : forfait {len(FA)} | partage {len(FB)} | déplacement {len(FC)} lignes -> {os.path.relpath(out, ROOT)}")
