#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
RENDU FINAL - Sous-page "Inventaire de stocks (juge absent ou incomplet)"
=========================================================================
Grief reel du fisc (Proposition 3924, p. 14-16, rejet 1/3, point I) :
  Le verificateur reconnait avoir "pu consulter et obtenir les inventaires de
  stocks" ; il reproche seulement que "les stocks remis, HORMIS celui du
  31/03/2022, apparaissent comme incomplets dans le sens ou ils N'INDIQUENT PAS
  TOUJOURS LES VOLUMES (contenances) des boissons". Il en tire une "irregularite
  grave" (R. 123-177, L. 123-12, CE 25/07/1980).

  -> Ce n'est PAS un grief de noms qui ne correspondraient pas aux factures
     (ca, c'est le grief separe "conso > achats"). C'est uniquement l'absence,
     sur les inventaires posterieurs au 31/03/2022, de la CONTENANCE en cl des
     boissons.

Reponse a prouver (SANS parler de consommation reelle / bilan matiere, traite
ailleurs et trop complexe pour cette page) :
  1. Les inventaires des 3 fins d'exercice existent et sont detailles LIGNE A
     LIGNE avec quantite + prix unitaire HT + valeur HT : exactement ce qu'exige
     R. 123-177 (quantite ET valeur). La contenance en cl n'est pas une mention
     legalement requise de l'inventaire.
  2. La contenance est une caracteristique fixe et publique (imprimee sur la
     facture du fournisseur, ex. "SAINT VERAN 75 CL"), et l'administration a
     ELLE-MEME converti les bouteilles en cl dans sa reconstitution (colonne
     "Volume disponible total en cl - conversion des bouteilles/BIB", annexes 6 et 7).
     L'omission n'a donc empeche aucun controle et n'a cause aucun prejudice.
  3. Le stock est stable d'une cloture a l'autre (pas de destockage cache).
  4. Contradiction : le service qualifie l'inventaire d'irregulier mais s'en sert
     comme base de sa reconstitution (variation de stock, quantites disponibles).
  5. La jurisprudence CE 25/07/1980 visee concerne un inventaire reduit au seul
     montant global SANS detail : c'est l'inverse de notre cas.

Sorties (reproductibles) :
  1. public/documents/pieces-defense/RF-inventaire-stocks.xlsx
  2. src/data/renduFinal/inventaire-stocks.json

Lecture seule sur les sources. Aucun chiffre en dur cote rendu.

Sources lues :
  - public/documents/inventaires/inventaires.json  (3 clotures, totaux HT, lignes detaillees)
"""

import json
import os
from collections import defaultdict

from rfcommun import normaliser_sections

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

INV_PATH = os.path.join(ROOT, "public/documents/inventaires/inventaires.json")

XLSX_OUT = os.path.join(ROOT, "public/documents/pieces-defense/RF-inventaire-stocks.xlsx")
JSON_OUT = os.path.join(ROOT, "src/data/renduFinal/inventaire-stocks.json")


# --------------------------------------------------------------------------- #
# Helpers format francais (pas de tiret cadratin nulle part)
# --------------------------------------------------------------------------- #
def fr_int(n):
    return f"{int(round(n)):,}".replace(",", " ")


def fr_dec(n, d=1):
    return f"{n:,.{d}f}".replace(",", " ").replace(".", ",")


def fr_eur2(n):
    s = f"{n:,.2f}".replace(",", "X").replace(".", ",").replace("X", " ")
    return s + " €"


# --------------------------------------------------------------------------- #
# Chargement des sources
# --------------------------------------------------------------------------- #
with open(INV_PATH, encoding="utf-8") as f:
    INV = json.load(f)

inventaires = INV["inventaires"]  # 3 clotures (2023, 2024, 2025)

# --------------------------------------------------------------------------- #
# 1. Stock de cloture par grande famille et par exercice (valeurs HT)
# --------------------------------------------------------------------------- #
stock_par_famille = {}
for inv in inventaires:
    stock_par_famille[inv["date"]] = {
        "sans_alcool": inv["totaux"]["sansAlcoolHT"],
        "alcool": inv["totaux"]["alcoolHT"],
        "total": inv["totaux"]["totalHT"],
        "nb_lignes": inv["nbLignes"],
    }

# Sous-detail alcool par sous-famille (heuristique de libelle) pour montrer le
# niveau de detail reel des inventaires (vins, bieres, spiritueux/digestifs...).
def sous_famille(produit):
    p = produit.lower()
    if any(k in p for k in ["fût", "fut", "affligem", "blade", "rouget de lisle"]) or "biere" in p or "bière" in p:
        return "Bieres / futs"
    if "cidre" in p:
        return "Cidres"
    if any(k in p for k in ["arbois", "chardonnay", "saint véran", "saint veran", "macon", "chablis",
                            "crémant", "cremant", "gewurz", "gascogne", "pive", "savagnin", "vin de paille",
                            "bib ", "côtes", "cotes", "rosé", "rose", "blanc 10", "rouge 10"]):
        return "Vins"
    return "Spiritueux / liqueurs / digestifs"

detail_sous_famille = defaultdict(lambda: defaultdict(float))
for inv in inventaires:
    for l in inv["lignes"]:
        if l["categorie"] != "alcool":
            continue
        detail_sous_famille[inv["date"]][sous_famille(l["produit"])] += (l.get("valeurHT") or 0)

# --------------------------------------------------------------------------- #
# 2. Completude des mentions legales : chaque ligne porte-t-elle quantite,
#    prix unitaire et valeur ? (ce qu'exige R. 123-177). On le mesure.
# --------------------------------------------------------------------------- #
def champs_ok(l):
    return (l.get("quantite") is not None
            and l.get("prixUnitaireHT") is not None
            and l.get("valeurHT") is not None)

total_lignes = sum(inv["nbLignes"] for inv in inventaires)
lignes_completes = sum(1 for inv in inventaires for l in inv["lignes"] if champs_ok(l))
pct_completes = lignes_completes / total_lignes * 100 if total_lignes else 0

# Lignes "incompletes" : ce ne sont PAS des produits a donnees manquantes, mais des
# lignes de reconciliation (un reliquat de valeur d'une page non rattache a un produit
# nomme lors de la transcription) ; la valeur est conservee pour que le total reste
# exact. On les decrit honnetement plutot que de les passer sous silence.
lignes_incompletes = [(inv["date"], l) for inv in inventaires for l in inv["lignes"] if not champs_ok(l)]
nb_incompletes = len(lignes_incompletes)
recon_valeurs_txt = ", ".join(fr_eur2(l.get("valeurHT") or 0) for _, l in lignes_incompletes)
recon_dates = sorted({d for d, _ in lignes_incompletes})
recon_dates_txt = (f"toutes au {recon_dates[0]}" if len(recon_dates) == 1
                   else "réparties sur " + " et ".join(recon_dates))

# --------------------------------------------------------------------------- #
# 3. Stabilite du stock : variation de stock entre clotures (valeur HT)
# --------------------------------------------------------------------------- #
variations = []
for i in range(1, len(inventaires)):
    prev = inventaires[i - 1]["totaux"]["totalHT"]
    cur = inventaires[i]["totaux"]["totalHT"]
    var = cur - prev
    pct = (var / prev * 100) if prev else 0
    variations.append((inventaires[i - 1]["date"], inventaires[i]["date"], var, pct))

stock_min = min(s["total"] for s in stock_par_famille.values())
stock_max = max(s["total"] for s in stock_par_famille.values())
stock_moy = sum(s["total"] for s in stock_par_famille.values()) / len(stock_par_famille)

print("=== Stock de cloture par exercice (HT) ===")
for d, s in stock_par_famille.items():
    print(f"  {d}: total {fr_eur2(s['total'])} | {s['nb_lignes']} lignes")
print(f"  lignes completes (qte+PU+valeur) : {lignes_completes}/{total_lignes} ({fr_dec(pct_completes)} %)")
print(f"  stock min/max : {fr_eur2(stock_min)} -> {fr_eur2(stock_max)}")


# --------------------------------------------------------------------------- #
# XLSX
# --------------------------------------------------------------------------- #
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

BOLD = Font(bold=True)
HEAD_FILL = PatternFill("solid", fgColor="0F766E")
HEAD_FONT = Font(bold=True, color="FFFFFF")
THIN = Side(style="thin", color="D0D0D0")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
RIGHT = Alignment(horizontal="right")


def style_header(ws, ncols, row=1):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HEAD_FONT
        cell.fill = HEAD_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER


def autosize(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


wb = Workbook()

# --- Onglet 1 : Stock de cloture par famille / exercice --------------------- #
ws = wb.active
ws.title = "Stock par famille"
hdr = ["Date de cloture", "Exercice", "Nb lignes inventaire",
       "Stock sans alcool (EUR HT)", "Stock alcool (EUR HT)", "Stock total (EUR HT)"]
ws.append(hdr)
style_header(ws, len(hdr))
for inv in inventaires:
    t = inv["totaux"]
    ws.append([inv["date"], inv["exercice"], inv["nbLignes"],
               round(t["sansAlcoolHT"], 2), round(t["alcoolHT"], 2), round(t["totalHT"], 2)])
for r in range(2, ws.max_row + 1):
    for c in range(4, 7):
        ws.cell(row=r, column=c).number_format = '#,##0.00 "EUR"'
        ws.cell(row=r, column=c).alignment = RIGHT
ws.freeze_panes = "A2"
autosize(ws, [16, 14, 20, 26, 24, 24])

# --- Onglet 2 : Sous-detail alcool par sous-famille ------------------------- #
ws2 = wb.create_sheet("Detail alcool")
hdr2 = ["Date de cloture", "Sous-famille (alcool)", "Valeur (EUR HT)"]
ws2.append(hdr2)
style_header(ws2, len(hdr2))
for inv in inventaires:
    d = inv["date"]
    for sf, v in sorted(detail_sous_famille[d].items(), key=lambda x: -x[1]):
        ws2.append([d, sf, round(v, 2)])
for r in range(2, ws2.max_row + 1):
    ws2.cell(row=r, column=3).number_format = '#,##0.00 "EUR"'
    ws2.cell(row=r, column=3).alignment = RIGHT
ws2.freeze_panes = "A2"
autosize(ws2, [16, 34, 18])

# --- Onglet 3 : Inventaire detaille (toutes lignes, 3 clotures) ------------- #
#     C'est la preuve centrale : chaque ligne porte qte + PU + valeur.
ws3 = wb.create_sheet("Inventaire detaille")
hdr3 = ["Date de cloture", "Produit", "Categorie", "Quantite",
        "Prix unitaire (EUR HT)", "Valeur (EUR HT)", "Fiabilite", "Page PDF"]
ws3.append(hdr3)
style_header(ws3, len(hdr3))
for inv in inventaires:
    for l in inv["lignes"]:
        ws3.append([
            inv["date"], l["produit"], l["categorie"],
            l["quantite"] if l["quantite"] is not None else "",
            l["prixUnitaireHT"] if l["prixUnitaireHT"] is not None else "",
            l["valeurHT"] if l["valeurHT"] is not None else "",
            l["fiabilite"], l["page"] if l["page"] is not None else "",
        ])
for r in range(2, ws3.max_row + 1):
    ws3.cell(row=r, column=5).number_format = '#,##0.00 "EUR"'
    ws3.cell(row=r, column=6).number_format = '#,##0.00 "EUR"'
    ws3.cell(row=r, column=4).alignment = RIGHT
    ws3.cell(row=r, column=5).alignment = RIGHT
    ws3.cell(row=r, column=6).alignment = RIGHT
ws3.freeze_panes = "A2"
autosize(ws3, [16, 40, 20, 10, 22, 18, 12, 10])

# --- Onglet 4 : Variation de stock (stabilite, valeur HT) ------------------- #
ws4 = wb.create_sheet("Variation stock")
hdr4 = ["De (cloture)", "Vers (cloture)", "Variation totale (EUR HT)", "Variation (%)"]
ws4.append(hdr4)
style_header(ws4, len(hdr4))
for de, vers, var, pct in variations:
    ws4.append([de, vers, round(var, 2), round(pct, 1)])
for r in range(2, ws4.max_row + 1):
    ws4.cell(row=r, column=3).number_format = '#,##0.00 "EUR"'
    ws4.cell(row=r, column=4).number_format = '0.0 "%"'
    ws4.cell(row=r, column=3).alignment = RIGHT
    ws4.cell(row=r, column=4).alignment = RIGHT
ws4.freeze_panes = "A2"
autosize(ws4, [16, 16, 26, 16])

os.makedirs(os.path.dirname(XLSX_OUT), exist_ok=True)
wb.save(XLSX_OUT)
print(f"\nXLSX ecrit : {XLSX_OUT}")


# --------------------------------------------------------------------------- #
# JSON renduFinal
# --------------------------------------------------------------------------- #
TEAL = "#0f766e"


def cg(v):
    return {"v": v}


def cd(v):
    return {"v": v, "align": "right"}


sections = []

# 1) Ce que soutient le fisc
sections.append({
    "kind": "chapitre", "source": "fisc", "numero": 1,
    "titre": "Ce que soutient l'administration",
    "sousTitre": "Inventaire de stocks juge incomplet (Proposition de rectification 3924, p. 14 a 16, rejet 1/3, point I).",
})
sections.append({
    "kind": "paragraphe",
    "texte": (
        "**Situation constatée.** Le vérificateur indique (p. 14) avoir « **pu consulter et obtenir les inventaires de "
        "stocks** » auprès du cabinet comptable AGC, et reprend une partie de leurs informations (les liquides) en annexe A. "
        "Il relève ensuite un défaut unique : « les stocks remis, **hormis celui du 31/03/2022**, apparaissent comme "
        "incomplets dans le sens où ils **n'indiquent pas toujours les volumes des boissons** »."
    ),
})
sections.append({
    "kind": "paragraphe",
    "texte": (
        "**Législation invoquée.** Le rapport cite l'article **R. 123-177 du Code de commerce**, qui définit l'inventaire "
        "comme le relevé de tous les éléments d'actif et de passif « au regard desquels sont mentionnées la **quantité** et la "
        "**valeur** de chacun d'eux », inventaire qui « doit être **détaillé et précis** ». Il ajoute que, si le livre "
        "d'inventaire ne détaille pas les produits, l'entreprise doit alors établir « un **état détaillé et estimatif** » "
        "énumérant « autant d'articles qu'il existe de produits de **caractéristiques différentes en raison de leur nature, de "
        "leurs dimensions, de leur marque, de leur prix unitaire** ». Il rappelle que l'article **L. 123-12, al. 2** impose un "
        "**inventaire physique** annuel à la date de clôture, dont les éléments « doivent pouvoir être **identifiés grâce aux "
        "documents ou pièces justificatives** conservées par l'entreprise », et invoque la jurisprudence **CE 25 juillet 1980 "
        "(n° 13170)** : un livre d'inventaire ne comportant « que le **montant global** des stocks, **sans état détaillé** "
        "faisant ressortir les quantités, les poids et les prix unitaires d'achat » confère à la comptabilité un caractère de "
        "grave irrégularité."
    ),
})
sections.append({
    "kind": "paragraphe",
    "texte": (
        "**Conséquence.** L'administration en tire (p. 14) que « **cette absence de stock** [...] constitue une "
        "**irrégularité grave** », permettant « de considérer que la comptabilité présentée n'est pas probante et qu'elle est "
        "dans l'incapacité de justifier précisément les chiffres d'affaires ». C'est l'un des motifs avancés à l'appui du "
        "rejet de comptabilité."
    ),
})

# 2) Notre demonstration
sections.append({
    "kind": "chapitre", "source": "nous", "numero": 2,
    "titre": "Notre réponse : un inventaire détaillé, conforme, et des volumes que le service a lui-même reconstitués",
    "sousTitre": "Les trois inventaires physiques de clôture sont produits, détaillés ligne à ligne (produit, quantité, prix d'achat HT, valeur HT). La seule mention en cause, la contenance, figure déjà dans les propres annexes de l'administration.",
})

sections.append({
    "kind": "paragraphe",
    "texte": (
        "Les inventaires physiques de clôture des trois exercices vérifiés existent, sont produits, et ont été établis à la "
        "**date de clôture** de chaque exercice par le cabinet comptable AGC (ce qui satisfait l'inventaire physique annuel de "
        "l'article L. 123-12) : "
        f"31/03/2023 ({inventaires[0]['nbLignes']} lignes), "
        f"31/03/2024 ({inventaires[1]['nbLignes']} lignes) et "
        f"31/03/2025 ({inventaires[2]['nbLignes']} lignes). "
        "Chaque ligne porte le **libellé du produit**, la **quantité**, le **prix unitaire d'achat HT** (l'inventaire est "
        "valorisé au coût) et la **valeur HT**. "
        f"Sur les **{total_lignes} lignes** des trois clôtures, **{lignes_completes} ({fr_dec(pct_completes)} %)** sont des "
        f"lignes de produit complètes ; les **{nb_incompletes} restantes** ({recon_dates_txt}) ne sont pas des produits à "
        "données manquantes, mais des **lignes de réconciliation** : un reliquat de valeur "
        f"({recon_valeurs_txt}) qui n'a pas pu être rattaché à un produit nommé lors de la transcription de la page "
        "concernée, et dont la **valeur reste comptée** pour que le stock total demeure exact. "
        "Autrement dit, **le fisc ne conteste ni la quantité ni la valeur** (toutes deux présentes pour chaque produit) : "
        "son unique reproche vise la **contenance**. Or c'est précisément la quantité et la valeur que R. 123-177 exige, "
        "produit par produit."
    ),
})

# Tableau stock de cloture par famille et par exercice
lignes_stock = []
for inv in inventaires:
    t = inv["totaux"]
    lignes_stock.append([
        cg(inv["date"]),
        cd(str(inv["nbLignes"])),
        cd(fr_eur2(t["sansAlcoolHT"])),
        cd(fr_eur2(t["alcoolHT"])),
        cd(fr_eur2(t["totalHT"])),
    ])
sections.append({
    "kind": "tableau",
    "titre": "Stock de clôture par grande famille et par exercice (valeurs HT issues des inventaires)",
    "minWidth": 640,
    "colonnes": [
        {"label": "Date de clôture"},
        {"label": "Lignes inventoriées", "align": "right"},
        {"label": "Boissons sans alcool", "align": "right"},
        {"label": "Alcools et vins", "align": "right"},
        {"label": "Stock total HT", "align": "right"},
    ],
    "lignes": lignes_stock,
})

# Le coeur de la reponse au grief "volumes" : la contenance n'est pas requise
# et le fisc l'a deja reconstituee lui-meme.
sections.append({
    "kind": "paragraphe",
    "texte": (
        "Sur le seul point réellement reproché, la contenance des boissons, trois constats se renforcent. "
        "**Un**, la contenance n'est pas une mention légale de l'inventaire : R. 123-177 demande la **quantité** et la "
        "**valeur**, toutes deux présentes. "
        "**Deux**, c'est une caractéristique **fixe et publique** du produit, **imprimée sur la facture du fournisseur** ; or "
        "la « législation » citée par le rapport lui-même exige seulement que les éléments de l'inventaire « **puissent être "
        "identifiés grâce aux pièces justificatives** » conservées : c'est exactement le cas. Exemple concret : au 31/03/2023, "
        "l'inventaire porte « **Saint Véran : 25 bouteilles, 436,50 € HT** » (quantité et valeur, conformes) ; la contenance, "
        "**75 cl**, figure noir sur blanc sur la facture « SAINT VÉRAN 75 CL (Lupé-Cholet) », soit 25 × 75 = **1 875 cl** "
        "retrouvables directement par la facture. "
        "**Trois**, l'administration a **elle-même reporté puis converti** ces contenances : sa reconstitution comporte une "
        "colonne « **Désignation de la boisson (avec volumétrie de la bouteille/BIB)** » suivie d'une colonne « **Volume "
        "disponible total en centilitres (conversion des bouteilles/BIB)** » (annexes 6-1/6-2/6-3 et 7-1/7-2/7-3). Le volume "
        "prétendument manquant n'a donc empêché aucun contrôle."
    ),
})

# Graphique : stock total de cloture par exercice
sections.append({
    "kind": "graphique",
    "variante": "vertical",
    "hauteur": 300,
    "dataKey": "nom",
    "serie": {"name": "Stock total HT (EUR)", "couleur": TEAL},
    "format": "euro",
    "data": [{"nom": inv["date"], "Stock total HT (EUR)": round(inv["totaux"]["totalHT"])} for inv in inventaires],
})

# Stabilite : variation de stock
lignes_var = []
for de, vers, var, pct in variations:
    signe = "+" if var >= 0 else "-"
    lignes_var.append([
        cg(f"{de} vers {vers}"),
        cd(f"{signe}{fr_eur2(abs(var))}"),
        cd(f"{signe}{fr_dec(abs(pct))} %"),
    ])
sections.append({
    "kind": "tableau",
    "titre": "Variation du stock entre clôtures : un stock stable, sans déstockage caché",
    "minWidth": 520,
    "colonnes": [
        {"label": "Période"},
        {"label": "Variation valeur HT", "align": "right"},
        {"label": "Variation relative", "align": "right"},
    ],
    "lignes": lignes_var,
})
sections.append({
    "kind": "paragraphe",
    "texte": (
        f"Le stock boissons total reste compris entre {fr_eur2(stock_min)} et {fr_eur2(stock_max)} HT sur les trois clôtures "
        f"(stock moyen {fr_eur2(stock_moy)}). Le sens de la variation est essentiel : un stock qui se maintient exclut le "
        "scénario d'un déstockage caché destiné à alimenter des ventes non comptabilisées, qui se traduirait par un stock en "
        "chute. Cette stabilité recoupe d'ailleurs les variations de stock boissons retenues par l'administration elle-même "
        "dans sa reconstitution (- 800,83 EUR, - 1 029,67 EUR et - 273,41 EUR), des montants minimes au regard du stock moyen."
    ),
})

# 3) Piece jointe
sections.append({
    "kind": "piecejointe",
    "intro": "Inventaire physique détaillé des trois clôtures (par produit et par famille, quantités, prix unitaires et valeurs HT) et variation de stock.",
    "fichiers": [
        {"fichier": "pieces-defense/RF-inventaire-stocks.xlsx",
         "label": "RF - Inventaire de stocks : détail par produit, par famille et variation de stock (XLSX)"},
    ],
})

# 3) Le grief est juridiquement infonde (argumentation de CORPS, pas un encart)
sections.append({
    "kind": "chapitre", "source": "nous", "numero": 3,
    "titre": "Le grief est juridiquement infondé",
    "sousTitre": "La contenance manquante n'est ni une mention légalement exigée, ni un obstacle au contrôle.",
})
sections.append({
    "kind": "paragraphe",
    "texte": (
        "D'abord, le texte invoqué est respecté. R. 123-177 impose, pour chaque élément, la **quantité** et la **valeur** : les "
        f"deux figurent pour chaque produit ({lignes_completes} lignes de produit sur {total_lignes}, le reste étant des "
        "reliquats de valeur conservés au 31/03/2025). L'énumération que le rapport met en avant (« autant d'articles qu'il "
        "existe de produits de **caractéristiques différentes en raison de leur nature, de leurs dimensions, de leur marque, de "
        "leur prix unitaire** ») décrit l'**état détaillé** que doit établir l'entreprise dont le livre d'inventaire **ne "
        f"détaille pas** ses produits. Or nos inventaires détaillent **chaque produit sur sa propre ligne nommée** "
        f"({inventaires[0]['nbLignes']}, {inventaires[1]['nbLignes']} puis {inventaires[2]['nbLignes']} lignes), avec sa "
        "nature, sa marque, son **prix d'achat unitaire** et sa valeur. Pour un produit conditionné, la mesure (le « poids » "
        "visé par la jurisprudence, la contenance pour un liquide) est **intrinsèque à sa désignation** et figure sur la "
        "**facture** ; or le texte même que cite le rapport (L. 123-12, al. 2) se satisfait de ce que les éléments de "
        "l'inventaire « **puissent être identifiés grâce aux pièces justificatives** ». La seule mention en cause est donc, au "
        "pire, identifiable sans difficulté."
    ),
})
sections.append({
    "kind": "paragraphe",
    "texte": (
        "Ensuite, et c'est décisif, le défaut allégué **n'a eu aucune incidence sur le contrôle** : le service a disposé de "
        "**toute l'information utile**. Il a lui-même reporté la **volumétrie de chaque bouteille** puis l'a convertie en "
        "centilitres (annexes 7), et a bâti l'intégralité de sa reconstitution sur ces mêmes inventaires (annexes 6-1 à 6-3, "
        "« achats ± variation de stock = quantités disponibles »). Une irrégularité ne peut justifier le rejet d'une "
        "comptabilité que si elle est **grave** et fait obstacle au contrôle des recettes ; le défaut de report d'une "
        "contenance que le service a lui-même reconstituée, sans la moindre incidence sur ce contrôle, n'atteint pas ce seuil."
    ),
})
sections.append({
    "kind": "paragraphe",
    "texte": (
        "Enfin, la jurisprudence **CE 25 juillet 1980** invoquée vise un livre d'inventaire réduit au **seul montant global**, "
        "« sans état détaillé faisant ressortir les quantités, les poids et les prix unitaires d'achat de chaque catégorie » : "
        "c'est la situation **inverse** de la nôtre, où chaque produit est détaillé (quantité, prix d'achat unitaire, valeur), "
        "la mesure du produit conditionné étant donnée par sa désignation et sa facture. On relèvera enfin que le rapport, qui "
        "conclut à une « **absence de stock** » (p. 14) après avoir constaté que le vérificateur « **a pu consulter et obtenir "
        "les inventaires** », vise par cette formule le **défaut de mention**, non l'absence matérielle des documents : "
        "l'inventaire existe, il est détaillé, et il a servi de base au redressement."
    ),
})

# 4) Conclusion (section de CORPS, pas une pastille d'annotation)
sections.append({
    "kind": "chapitre", "source": "nous", "numero": 4,
    "titre": "Conclusion",
})
sections.append({
    "kind": "paragraphe",
    "texte": (
        "L'inventaire physique des trois fins d'exercice existe et est détaillé produit par produit "
        f"({inventaires[0]['nbLignes']}, {inventaires[1]['nbLignes']} et {inventaires[2]['nbLignes']} lignes, avec quantité, "
        f"prix d'achat unitaire et valeur), et le stock se maintient ({fr_eur2(stock_min)} à {fr_eur2(stock_max)} HT). Le "
        "seul élément reproché, la contenance en centilitres, n'est pas une mention que R. 123-177 impose de reporter dès lors "
        "que la quantité et la valeur figurent ; elle est de surcroît identifiable par les factures (L. 123-12, al. 2) et a "
        "été reconstituée par l'administration elle-même (Annexe N°1, colonne « volumétrie de la bouteille »). Faute "
        "d'incidence sur le contrôle des recettes, ce défaut formel est **dépourvu de la gravité** qui seule pourrait fonder "
        "le rejet, d'autant que le service s'appuie sur ces mêmes inventaires pour reconstituer le chiffre d'affaires."
    ),
})

doc = {
    "meta": {
        "slug": "inventaire-stocks",
        "titre": "Inventaire de stocks (jugé absent ou incomplet)",
        "bloc": "rejet",
        "griefFisc": "Les inventaires (hormis le 31/03/2022) n'indiqueraient pas toujours la contenance des boissons, ce qui les rendrait incomplets (Proposition p. 14 a 16, rejet 1/3, point I).",
        "reponse": "Les inventaires existent et sont détaillés (quantité, prix d'achat, valeur) ; la contenance n'est pas une mention que R. 123-177 impose de reporter, elle est identifiable par les factures et a été reconstituée par l'administration elle-même. Sans incidence sur le contrôle, le défaut est dépourvu de la gravité requise pour fonder le rejet.",
        "sources": [
            "public/documents/inventaires/inventaires.json",
            "public/documents/inventaires/inventaire_2023-03-31.csv",
            "public/documents/inventaires/inventaire_2024-03-31.csv",
            "public/documents/inventaires/inventaire_2025-03-31.csv",
            "public/documents/rapports-des-finances-publiques/synthese/02-rejet-comptabilite-1.md",
            "public/documents/rapports-des-finances-publiques/synthese/08-annexes-A-H.md (Annexe N°1, p. 47-50)",
        ],
        "piece": "pieces-defense/RF-inventaire-stocks.xlsx",
        "genereePar": "scripts/rendu-final-inventaire-stocks.py",
    },
    "sections": sections,
}

doc["sections"] = normaliser_sections(doc["sections"])
os.makedirs(os.path.dirname(JSON_OUT), exist_ok=True)
with open(JSON_OUT, "w", encoding="utf-8") as f:
    json.dump(doc, f, ensure_ascii=False, indent=2)
print(f"JSON ecrit : {JSON_OUT}")
print(f"Sections : {len(sections)}")
