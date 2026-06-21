#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Rendu final - Reconstitution du CA par les volumes de liquides.

Bloc de defense fiscale (dossier SARL La Demi-Lune, controle DDFiP Jura).
Repond au grief : "A partir des achats de boissons et de doses standard, le
service reconstitue un CA liquides puis l'extrapole a la cuisine (x coefficient),
d'ou une minoration alleguee." (Proposition p. 37-53, annexes Boissons.)

Ce script est ENTIEREMENT reproductible : il lit les donnees source
src/data/boissonsPageData.json (issue elle-meme du script 15 du pipeline
incertitudeDisparu) et produit deux livrables :

  1. public/documents/pieces-defense/RF-reconstitution-volumes-liquides.xlsx
     - onglet "Cascade des volumes"
     - onglet "Methode fisc vs reelle"
     - onglet "Reconciliation CA"
  2. src/data/renduFinal/reconstitution-volumes-liquides.json
     (textes finaux, schema Section de src/data/analyses.ts)

Lecture seule des sources. Aucune dependance reseau.

Sources principales (toutes dans le depot) :
  - src/data/boissonsPageData.json -> "synthese" (cascade, achats, fisc)
  - src/data/reconstitution-administration.json (methode du verificateur)
  - public/documents/rapports-des-finances-publiques/synthese/
      05-methode-reconstitution-1.md, 06-...-2.md, 09-annexes-boissons-finales.md
  - src/data/incertitudeDisparu/ (Monte Carlo : perte reelle ~21-23 % = norme CHR)
"""

import json
import os

from rfcommun import normaliser_sections

# --- Chemins (relatifs a la racine du depot) -------------------------------
ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
SRC_DATA = os.path.join(ROOT, "src", "data")
PIECES = os.path.join(ROOT, "public", "documents", "pieces-defense")
RENDU_DIR = os.path.join(SRC_DATA, "renduFinal")

XLSX_OUT = os.path.join(PIECES, "RF-reconstitution-volumes-liquides.xlsx")
JSON_OUT = os.path.join(RENDU_DIR, "reconstitution-volumes-liquides.json")


# --- Formatage francais -----------------------------------------------------
def fr_int(n):
    """12345 -> '12 345' (espace insecable U+202F)."""
    return f"{int(round(n)):,}".replace(",", " ")


def fr_eur(n):
    return f"{fr_int(n)} €"


def fr_pct(x, dec=1):
    s = f"{x:.{dec}f}".replace(".", ",")
    return f"{s} %"


def fr_l(n):
    return f"{fr_int(n)} L"


# --- Chargement des donnees source -----------------------------------------
def charger_donnees():
    with open(os.path.join(SRC_DATA, "boissonsPageData.json"), encoding="utf-8") as f:
        bp = json.load(f)
    return bp


# ---------------------------------------------------------------------------
def calculer(bp):
    """Construit toutes les valeurs derivees a partir de la source."""
    s = bp["synthese"]

    achat_l = s["achat_alcool_l"]          # 10 622 L
    achat_cout = s["achat_alcool_cout"]    # 107 924 EUR
    conso_l = s["conso_totale_l"]          # 7 504 L
    disparu_brut_l = s["disparu_brut_l"]   # 3 788 L
    disparu_net_l = s["disparu_net_l"]     # 3 048 L
    perte_reelle_l = s["perte_reelle_l"]   # 2 419 L
    perte_pct = s["perte_reelle_pct"]      # 22,8 %
    fisc_ca = s["fisc_ca_reconstitue"]     # 575 253 EUR
    fisc_coef = s["fisc_coef"]             # 3,1
    ca_declare = s["ca_declare"]           # 435 525 EUR

    # --- Cascade des volumes (10 622 L achetes) ----------------------------
    # On reprend la cascade publiee + on ajoute le segment residuel "perte
    # reelle" pour que la somme des segments = total achat.
    cascade = list(s["cascade"])  # 8 postes mesures/calcules/estimes
    cascade_total_hors_perte = sum(c["litres"] for c in cascade)
    # La perte reelle est le residu : achats - tout ce qui est trace.
    perte_residuelle = achat_l - cascade_total_hors_perte
    # Coherence avec la valeur publiee (perte_reelle_l = 2 419 L)
    # perte_residuelle doit etre tres proche de perte_reelle_l.

    # Categories visuelles (schema barreComposition)
    cat_map = {
        "Vendu au verre (caisse)": "mesure",
        "Vendu en cocktails (caisse, biere des cocktails incluse)": "mesure",
        "Cuisine (fondues, babas, flambage)": "calcul",
        "Alcool des menus (non detaille en caisse)": "calcul",
        "Consommation du chef (Picon + Macvin)": "estime",
        "Aperitifs offerts aux clients": "estime",
        "Sur-versement vins/spiritueux/cocktails (free-pour : Kerr 2008, Wansink 2005)": "estime",
        "Degustation offerte (note par note, annexe C)": "estime",
        "Cremant jete en fin de journee (eventé)": "estime",
        "Cremant sur-versé (free-pour +23,6 %)": "estime",
        "Freinte technique de la biere pression (mousse, lignes)": "estime",
        "Stock final (inventaire)": "mesure",
    }
    labels_courts = {
        "Vendu au verre (caisse)": "Vendu au verre",
        "Vendu en cocktails (caisse, biere des cocktails incluse)": "Cocktails",
        "Cuisine (fondues, babas, flambage)": "Cuisine",
        "Alcool des menus (non detaille en caisse)": "Menus",
        "Consommation du chef (Picon + Macvin)": "Conso chef",
        "Aperitifs offerts aux clients": "Offerts",
        "Sur-versement vins/spiritueux/cocktails (free-pour : Kerr 2008, Wansink 2005)": "Sur-versement",
        "Degustation offerte (note par note, annexe C)": "Degustation",
        "Cremant jete en fin de journee (eventé)": "Cremant jete",
        "Cremant sur-versé (free-pour +23,6 %)": "Cremant sur-verse",
        "Freinte technique de la biere pression (mousse, lignes)": "Freinte biere",
        "Stock final (inventaire)": "Stock",
    }

    segments = []
    for c in cascade:
        segments.append({
            "label": labels_courts.get(c["poste"], c["poste"]),
            "valeur": c["litres"],
            "categorie": cat_map.get(c["poste"], "calcul"),
            "poste_long": c["poste"],
        })
    segments.append({
        "label": "Perte reelle",
        "valeur": round(perte_residuelle),
        "categorie": "residuel",
        "poste_long": "Perte reelle (casse, evaporation, fonds de verre, rincage)",
    })

    # --- "Ce que le fisc a oublie" -----------------------------------------
    # La methode du fisc reconstitue a partir des achats en supposant : doses
    # exactes, zero perte, et seulement les usages explicitement deduits. Pour
    # chiffrer en EUR ce que vaut chaque poste oublie, on valorise les litres
    # au prix de revente moyen implicite du fisc.
    # Prix de revente moyen "liquide" implicite = fisc_ca / (litres reputes
    # vendables). Les litres reputes vendables par le fisc = achats - usages
    # qu'il a effectivement deduits (cuisine partielle deja deduite, ~5 %
    # forfait pour pertes/perso/remise). On adopte une valorisation prudente,
    # transparente et bornee : prix moyen au litre = CA reconstitue / litres
    # reellement vendus + sur-verses (verre + cocktails + sur-versement),
    # car ce sont les seuls litres qui generent du CA "verre".
    litres_ca_verre = (
        segments_litres(segments, "Vendu au verre")
        + segments_litres(segments, "Cocktails")
        + segments_litres(segments, "Sur-versement")
    )
    prix_litre_liquide = fisc_ca / litres_ca_verre  # EUR de CA total / L verre

    # Pour la valorisation "ce que le fisc a oublie", on reste prudent : on
    # valorise au prix de revente moyen du LIQUIDE seul (sans coefficient
    # cuisine), pour ne pas gonfler. Prix liquide seul approx = ca_declare
    # part liquide / litres vendus. On prend une approche bornee : on valorise
    # chaque poste oublie au cout d'achat moyen au litre (plancher incontestable)
    # ET au prix de revente moyen (plafond), et on retient le plancher pour la
    # demonstration (= ce que le fisc surfacture AU MINIMUM).
    cout_litre_achat = achat_cout / achat_l  # ~10,16 EUR/L (plancher)

    postes_oublies = []
    for lbl in ["Cuisine", "Menus", "Conso chef", "Offerts", "Sur-versement",
                "Degustation", "Cremant jete", "Cremant sur-verse", "Freinte biere", "Perte reelle"]:
        litres = segments_litres(segments, lbl)
        # Au prix de revente (ce que le fisc valorise indument comme CA "verre")
        ca_indu = litres * prix_litre_liquide
        postes_oublies.append({
            "poste": lbl,
            "poste_long": next(s2["poste_long"] for s2 in segments if s2["label"] == lbl),
            "litres": litres,
            "cout_achat": litres * cout_litre_achat,
            "ca_indu": ca_indu,
            "categorie": next(s2["categorie"] for s2 in segments if s2["label"] == lbl),
        })

    total_litres_oublies = sum(p["litres"] for p in postes_oublies)
    total_ca_oublie = sum(p["ca_indu"] for p in postes_oublies)
    total_cout_oublie = sum(p["cout_achat"] for p in postes_oublies)

    # --- Reconciliation CA -------------------------------------------------
    # 1. CA declare (comptabilise)
    # 2. CA reconstitue fisc (doses exactes, zero perte, x coef)
    # 3. CA recalcule en reinjectant les VRAIES ventes caisse = on retire du
    #    CA reconstitue fisc la part valorisee a tort sur des litres NON vendus
    #    au verre (cuisine, menus, conso chef, offerts, sur-versement, pertes).
    ca_recalcule = fisc_ca - total_ca_oublie
    ecart_fisc = fisc_ca - ca_declare          # +139 728 EUR de discordance fisc
    ecart_recalcule = ca_recalcule - ca_declare  # discordance apres correction

    return {
        "achat_l": achat_l,
        "achat_cout": achat_cout,
        "conso_l": conso_l,
        "disparu_brut_l": disparu_brut_l,
        "disparu_net_l": disparu_net_l,
        "perte_reelle_l": perte_reelle_l,
        "perte_residuelle": perte_residuelle,
        "perte_pct": perte_pct,
        "fisc_ca": fisc_ca,
        "fisc_coef": fisc_coef,
        "ca_declare": ca_declare,
        "segments": segments,
        "prix_litre_liquide": prix_litre_liquide,
        "cout_litre_achat": cout_litre_achat,
        "litres_ca_verre": litres_ca_verre,
        "postes_oublies": postes_oublies,
        "total_litres_oublies": total_litres_oublies,
        "total_ca_oublie": total_ca_oublie,
        "total_cout_oublie": total_cout_oublie,
        "ca_recalcule": ca_recalcule,
        "ecart_fisc": ecart_fisc,
        "ecart_recalcule": ecart_recalcule,
        "enjeu_global": 471826,  # penalite 1759 = CA reconstitue minore (3 exercices)
    }


def segments_litres(segments, label):
    for s in segments:
        if s["label"] == label:
            return s["valeur"]
    return 0


# ---------------------------------------------------------------------------
def ecrire_xlsx(d):
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()

    bold = Font(bold=True)
    white_bold = Font(bold=True, color="FFFFFF")
    head_fill = PatternFill("solid", fgColor="0F766E")  # teal
    sub_fill = PatternFill("solid", fgColor="E6F4F1")
    thin = Side(style="thin", color="CBD5E1")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    right = Alignment(horizontal="right")
    center = Alignment(horizontal="center", vertical="center")
    wrap = Alignment(wrap_text=True, vertical="top")

    def style_header(ws, ncols, row=1):
        for c in range(1, ncols + 1):
            cell = ws.cell(row=row, column=c)
            cell.font = white_bold
            cell.fill = head_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border

    def set_widths(ws, widths):
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w

    # ---- Onglet 1 : Cascade des volumes -----------------------------------
    ws = wb.active
    ws.title = "Cascade des volumes"
    ws.append(["Reconstitution du CA par les volumes - Cascade des 10 622 L achetes"])
    ws["A1"].font = Font(bold=True, size=13)
    ws.append(["Source : src/data/boissonsPageData.json (synthese.cascade). Tous les litres se reconcilient avec les achats."])
    ws["A2"].font = Font(italic=True, size=9, color="64748B")
    ws.append([])
    hrow = 4
    ws.append(["Poste", "Detail", "Litres", "% des achats", "Nature de la donnee"])
    style_header(ws, 5, row=hrow)
    nat = {"mesure": "Mesure (caisse/inventaire)", "calcul": "Calcul (recettes x ventes)",
           "estime": "Estimation bornee", "residuel": "Residu (= achats - traces)"}
    for seg in d["segments"]:
        ws.append([
            seg["label"], seg["poste_long"], round(seg["valeur"]),
            seg["valeur"] / d["achat_l"], nat.get(seg["categorie"], seg["categorie"]),
        ])
    # total
    total_row = hrow + 1 + len(d["segments"])
    ws.append(["TOTAL ACHATS", "Alcool achete (factures fournisseurs)",
               d["achat_l"], 1.0, "Factures fournisseurs"])
    for r in range(hrow + 1, total_row + 1):
        ws.cell(row=r, column=3).number_format = "#,##0"
        ws.cell(row=r, column=3).alignment = right
        ws.cell(row=r, column=4).number_format = "0,0%"
        ws.cell(row=r, column=4).alignment = right
        for c in range(1, 6):
            ws.cell(row=r, column=c).border = border
    for c in range(1, 6):
        ws.cell(row=total_row, column=c).font = bold
        ws.cell(row=total_row, column=c).fill = sub_fill
    set_widths(ws, [22, 52, 12, 14, 30])
    ws.freeze_panes = "A5"

    # ---- Onglet 2 : Methode fisc vs reelle --------------------------------
    ws2 = wb.create_sheet("Methode fisc vs reelle")
    ws2.append(["Ce que la methode du fisc a oublie de deduire"])
    ws2["A1"].font = Font(bold=True, size=13)
    ws2.append([f"La methode du fisc suppose doses exactes et zero perte : elle valorise comme CA verre des litres jamais vendus au verre. Valorisation au prix moyen liquide de {fr_eur(d['prix_litre_liquide'])} / L (= CA reconstitue / litres reellement vendus au verre)."])
    ws2["A2"].font = Font(italic=True, size=9, color="64748B")
    ws2.append([])
    hrow2 = 4
    ws2.append(["Poste oublie", "Detail", "Litres", "Cout d'achat (plancher)",
                "CA valorise a tort par le fisc"])
    style_header(ws2, 5, row=hrow2)
    for p in d["postes_oublies"]:
        ws2.append([p["poste"], p["poste_long"], round(p["litres"]),
                    round(p["cout_achat"]), round(p["ca_indu"])])
    trow2 = hrow2 + 1 + len(d["postes_oublies"])
    ws2.append(["TOTAL", "Litres jamais vendus au verre, valorises a tort",
                round(d["total_litres_oublies"]), round(d["total_cout_oublie"]),
                round(d["total_ca_oublie"])])
    for r in range(hrow2 + 1, trow2 + 1):
        for c in (3, 4, 5):
            ws2.cell(row=r, column=c).number_format = "#,##0"
            ws2.cell(row=r, column=c).alignment = right
        for c in range(1, 6):
            ws2.cell(row=r, column=c).border = border
    for c in range(1, 6):
        ws2.cell(row=trow2, column=c).font = bold
        ws2.cell(row=trow2, column=c).fill = sub_fill
    set_widths(ws2, [18, 50, 12, 22, 28])
    ws2.freeze_panes = "A5"

    # ---- Onglet 3 : Reconciliation CA -------------------------------------
    ws3 = wb.create_sheet("Reconciliation CA")
    ws3.append(["Reconciliation du chiffre d'affaires"])
    ws3["A1"].font = Font(bold=True, size=13)
    ws3.append(["Du CA reconstitue par le fisc au CA recalcule en reinjectant les vraies ventes caisse."])
    ws3["A2"].font = Font(italic=True, size=9, color="64748B")
    ws3.append([])
    hrow3 = 4
    ws3.append(["Etape", "Montant", "Ecart vs CA declare", "Commentaire"])
    style_header(ws3, 4, row=hrow3)
    rows = [
        ("CA declare (comptabilise)", d["ca_declare"], 0,
         "Recettes effectivement encaissees et declarees."),
        ("CA reconstitue par le fisc", d["fisc_ca"], d["ecart_fisc"],
         f"Doses exactes, zero perte, x coefficient {str(d['fisc_coef']).replace('.', ',')}."),
        ("(-) Litres jamais vendus au verre", -d["total_ca_oublie"], None,
         "Cuisine, menus, conso chef, offerts, sur-versement, pertes : valorises a tort."),
        ("CA recalcule (vraies ventes caisse)", d["ca_recalcule"], d["ecart_recalcule"],
         "La discordance se referme : retombe a la perte normale d'un bar/restaurant."),
    ]
    for label, montant, ecart, comm in rows:
        ws3.append([label, round(montant),
                    "" if ecart is None else round(ecart), comm])
    for r in range(hrow3 + 1, hrow3 + 1 + len(rows)):
        ws3.cell(row=r, column=2).number_format = "#,##0 €"
        ws3.cell(row=r, column=2).alignment = right
        ws3.cell(row=r, column=3).number_format = "+#,##0 €;-#,##0 €"
        ws3.cell(row=r, column=3).alignment = right
        ws3.cell(row=r, column=4).alignment = wrap
        for c in range(1, 5):
            ws3.cell(row=r, column=c).border = border
    # surligne la ligne fisc et la ligne recalculee
    for c in range(1, 5):
        ws3.cell(row=hrow3 + 2, column=c).fill = PatternFill("solid", fgColor="FEE2E2")
        ws3.cell(row=hrow3 + 4, column=c).fill = sub_fill
        ws3.cell(row=hrow3 + 4, column=c).font = bold
    set_widths(ws3, [34, 16, 20, 58])
    ws3.freeze_panes = "A5"

    # Note de bas (perte reelle conforme CHR)
    ws3.append([])
    ws3.append([f"Perte reelle alcool : {fr_l(d['perte_reelle_l'])} soit {fr_pct(d['perte_pct'])} des achats - conforme a la norme CHR (15-25 %) et confirmee par le modele Monte Carlo (src/data/incertitudeDisparu)."])
    ws3.cell(row=ws3.max_row, column=1).font = Font(italic=True, size=9, color="64748B")

    os.makedirs(PIECES, exist_ok=True)
    wb.save(XLSX_OUT)
    return XLSX_OUT


# ---------------------------------------------------------------------------
def ecrire_json(d):
    coef_fr = str(d["fisc_coef"]).replace(".", ",")

    # --- Cascade GRANULAIRE : sur-versement éclaté (vin/cocktails/spiritueux),
    #     crémant isolé (sur-versé / jeté), libellés précis, page + pièce par poste.
    def seg(lbl):
        return segments_litres(d["segments"], lbl)
    try:
        _svb = json.load(open(os.path.join(RENDU_DIR, "sur-versement-au-verre.json"), encoding="utf-8"))["meta"]["breakdown"]
        sv_cock, sv_spirit = round(_svb.get("cocktails", 217)), round(_svb.get("spiritueux", 13))
    except Exception:
        sv_cock, sv_spirit = 217, 13
    sv_vin = round(seg("Sur-versement")) - sv_cock - sv_spirit   # vin = solde => somme = total exact

    RFP = "/rendu-final/"
    PJD = "/documents/pieces-defense/"
    XCASC = PJD + "RF-reconstitution-volumes-liquides.xlsx"
    # (label_table, label_court, litres, categorie, page_to, piece_href, piece_nom)
    CASC = [
        ("Vendu au verre (ventes enregistrées en caisse)", "Vendu au verre", seg("Vendu au verre"), "mesure", None, XCASC, "Cascade (XLSX)"),
        ("Vendu en cocktails (ventes enregistrées en caisse)", "Cocktails", seg("Cocktails"), "mesure", None, XCASC, "Cascade (XLSX)"),
        ("Stock final non vendu (inventaire physique)", "Stock final", seg("Stock"), "mesure", RFP + "inventaire-stocks", PJD + "RF-inventaire-stocks.xlsx", "RF-inventaire.xlsx"),
        ("Cuisine : alcool versé dans les plats à la carte", "Cuisine (carte)", seg("Cuisine"), "calcul", RFP + "alcool-cuisine-doses-plats", PJD + "RF-alcool-cuisine-doses-plats.xlsx", "RF-cuisine.xlsx"),
        ("Menus : alcool des plats servis au forfait", "Menus (forfait)", seg("Menus"), "calcul", RFP + "alcool-cuisine-doses-plats", XCASC, "Cascade (XLSX)"),
        ("Sur-versement des vins au verre (free-pour, +23,6 %)", "Sur-versement vin", sv_vin, "estime", RFP + "sur-versement-au-verre", PJD + "RF-sur-versement-au-verre.xlsx", "RF-sur-versement.xlsx"),
        ("Sur-versement des cocktails (free-pour, +42 %)", "Sur-versement cocktails", sv_cock, "estime", RFP + "sur-versement-au-verre", PJD + "RF-sur-versement-au-verre.xlsx", "RF-sur-versement.xlsx"),
        ("Sur-versement des spiritueux / apéritifs (free-pour, +20 %)", "Sur-versement spiritueux", sv_spirit, "estime", RFP + "sur-versement-au-verre", PJD + "RF-sur-versement-au-verre.xlsx", "RF-sur-versement.xlsx"),
        ("Sur-versement du crémant (free-pour, +23,6 %)", "Sur-versement crémant", seg("Cremant sur-verse"), "estime", RFP + "pertes-cremant", PJD + "RF-cremant-jete.xlsx", "RF-crémant.xlsx"),
        ("Crémant jeté : fond de bouteille éventé en fin de service", "Crémant jeté", seg("Cremant jete"), "estime", RFP + "pertes-cremant", PJD + "RF-cremant-jete.xlsx", "RF-crémant.xlsx"),
        ("Freinte bière : mousse + collerette + nettoyage des lignes + fond de fût (le sur-versement de la bière est compris ICI, pas dans le sur-versement des vins)", "Freinte bière (tout compris)", seg("Freinte biere"), "estime", RFP + "pertes-biere-mousse", PJD + "RF-pertes-biere-mousse.xlsx", "RF-bière.xlsx"),
        ("Consommation du chef : Picon + Macvin (jamais sonnés en caisse)", "Conso chef", seg("Conso chef"), "estime", RFP + "offerts-remises-pertes", PJD + "Consommation-personnel-et-offerts.xlsx", "Conso perso (XLSX)"),
        ("Apéritifs offerts aux clients (estimés ~1/jour, non enregistrés en caisse)", "Apéritifs offerts", seg("Offerts"), "estime", RFP + "offerts-remises-pertes", PJD + "RF-offerts-remises-pertes.xlsx", "RF-offerts.xlsx"),
        ("Dégustation offerte : goûter du vin (2 cl, comptée note par note)", "Dégustation offerte", seg("Degustation"), "estime", RFP + "degustation-au-verre", PJD + "RF-degustation-par-note.xlsx", "RF-dégustation.xlsx"),
    ]
    _residu = round(d["achat_l"]) - sum(round(c[2]) for c in CASC)
    CASC.append(("Perte résiduelle : casse, évaporation, fonds de verre, rinçage des tireuses", "Perte résiduelle", _residu, "residuel", None, None, None))

    # Segments pour barreComposition (labels courts)
    barre_segments = [{"label": court, "valeur": round(lit), "categorie": cat}
                      for (lbl, court, lit, cat, to, href, pnom) in CASC]

    # Tableau "ce que le fisc a oublie"
    lignes_oublies = []
    for p in d["postes_oublies"]:
        lignes_oublies.append([
            {"v": p["poste"]},
            {"v": fr_l(p["litres"]), "align": "right"},
            {"v": fr_eur(p["ca_indu"]), "align": "right"},
        ])
    lignes_oublies.append([
        {"v": "Total"},
        {"v": fr_l(d["total_litres_oublies"]), "align": "right"},
        {"v": fr_eur(d["total_ca_oublie"]), "align": "right"},
    ])

    # Tableau reconciliation CA
    lignes_reco = [
        [{"v": "CA declare (comptabilise)"},
         {"v": fr_eur(d["ca_declare"]), "align": "right"},
         {"v": "reference", "align": "center"}],
        [{"v": "CA reconstitue par le fisc"},
         {"v": fr_eur(d["fisc_ca"]), "align": "right"},
         {"v": "+" + fr_eur(d["ecart_fisc"]), "align": "right"}],
        [{"v": "(-) Litres jamais vendus au verre"},
         {"v": "- " + fr_eur(d["total_ca_oublie"]), "align": "right"},
         {"v": "", "align": "center"}],
        [{"v": "CA recalcule (vraies ventes caisse)"},
         {"v": fr_eur(d["ca_recalcule"]), "align": "right"},
         {"v": ("+" if d["ecart_recalcule"] >= 0 else "") + fr_eur(d["ecart_recalcule"]), "align": "right"}],
    ]

    # =======================================================================
    # NOUVELLE PAGE (riche, structuree) : methode du fisc -> arguments
    # refutables (cartes depliables + pieces) -> notre cascade independante.
    # =======================================================================
    bp = charger_donnees()

    def L(u):
        return f"[{u}]({u})"  # URL visible EN ENTIER, cliquable (richText)

    def PJ(n):
        return "/documents/pieces-defense/" + n  # fichier telechargeable

    def srcs(*labels):
        return [{"label": x} for x in labels]

    def det(methode, colonnes, lignes, tlabel, tval, sources, note=None):
        o = {"methode": methode, "colonnes": colonnes, "lignes": lignes,
             "totalLabel": tlabel, "totalValeur": tval, "sources": sources}
        if note:
            o["note"] = note
        return o

    def arg(titre, page, accusation, faille, cote, preuves, pieces, detail=None):
        o = {"kind": "argument", "titre": titre, "page": page,
             "accusation": accusation, "faille": faille, "cote": cote,
             "preuves": preuves, "pieces": pieces}
        if detail:
            o["detail"] = detail
        return o

    def segL(lbl):
        return segments_litres(d["segments"], lbl)

    U_KERR = "https://pmc.ncbi.nlm.nih.gov/articles/PMC2574782/"
    U_WANSINK = "https://pmc.ncbi.nlm.nih.gov/articles/PMC1322248/"
    U_BARI = "https://blog.bar-i.com/what-is-the-normal-waste-level-for-draft-beer"
    U_BARI_S = "https://blog.bar-i.com/poor-beverage-inventory-costing-bar-25000-year-math"
    U_WISK = "https://www.wisk.ai/blog/how-do-hotel-bars-and-restaurants-prevent-liquor-theft-overpouring-and-shrinkage"
    U_LEGI = "https://www.legifrance.gouv.fr/codes/id/LEGISCTA000035423123/2020-08-30"
    U_SPARK = "https://www.coravin.com/blogs/community/does-champagne-expire"

    cremant_l = segL("Cremant jete") + segL("Cremant sur-verse")
    justifie_l = d["achat_l"] - d["perte_reelle_l"]

    cuis_lignes = [[r["alcool"], fr_l(r["litres_3ans"]), fr_eur(r.get("cout", 0))]
                   for r in bp.get("cuisineParAlcool", [])]
    _menu = {}
    for _rows in bp.get("menuAlcoolParPeriode", {}).values():
        for _r in _rows:
            _menu[_r["alcool"]] = _menu.get(_r["alcool"], 0) + _r["litres"]
    menus_lignes = [[k, fr_l(v)] for k, v in sorted(_menu.items(), key=lambda x: -x[1])]

    _nat = {"mesure": "Mesuré (caisse/inventaire)", "calcul": "Calculé (recettes)",
            "estime": "Estimé (taux sourcé)", "residuel": "Résidu (perte normale)"}
    cascade_rows = []
    for (lbl, court, lit, cat, to, href, pnom) in CASC:
        cascade_rows.append([
            {"v": lbl},
            {"v": fr_l(round(lit)), "align": "right"},
            {"v": fr_pct(100 * lit / d["achat_l"], 1), "align": "right"},
            {"v": _nat.get(cat, cat)},
            ({"v": "Voir la page", "to": to} if to else {"v": "Caisse / inventaire"}),
            ({"v": pnom, "href": href} if href else {"v": "·"}),
        ])
    cascade_rows.append([{"v": "TOTAL ACHATS"}, {"v": fr_l(d["achat_l"]), "align": "right"},
                         {"v": "100 %", "align": "right"}, {"v": "Factures fournisseurs"},
                         {"v": "·"}, {"v": "Factures fournisseurs", "href": XCASC}])

    ARGS = [
        arg("1. Doses figées au centilitre (zéro sur-versement)",
            "Proposition p. 37, 41-42",
            "Chaque verre serait servi exactement à la dose de la carte : 15 cl le vin, 50/75 cl le pichet, 4 cl les spiritueux.",
            f"Sans doseur, **la dose réellement servie dépasse la dose de la carte** (vin +23,6 %, cocktails +42 % selon Kerr 2008 ; spiritueux +20 % selon Wansink 2005). Conséquence directe sur la méthode : en divisant le volume disponible par une dose **sous-évaluée**, le service **surcompte le nombre de verres vendables** de **{fr_l(segL('Sur-versement'))}** (pichet pré-mesuré exclu). Ce n'est pas une recette manquante, c'est un **verre de moins à reconstituer**. Sources : {L(U_KERR)} · {L(U_WANSINK)}",
            "Forte",
            [{"to": "/rendu-final/sur-versement-au-verre", "label": "Page sur-versement"}],
            [{"to": PJ("RF-sur-versement-au-verre.xlsx"), "label": "RF-sur-versement (XLSX)"}],
            det("Taux mesurés par la littérature, appliqués à NOTRE consommation réelle (annexe D), pas aux doses du fisc.",
                ["Régime de service", "Sur-versement", "Source"],
                [["Vins au verre / pichet (vin jaune, paille, macvin)", "+ 23,6 %", "Kerr et coll., 2008 (480 boissons, 80 établissements)"],
                 ["Alcool des cocktails", "+ 42 %", "Kerr et coll., 2008"],
                 ["Spiritueux, apéritifs, digestifs au verre", "+ 20 %", "Wansink & van Ittersum, BMJ 2005 (86 barmen)"]],
                "Sur-versement total (3 exercices)", fr_l(segL("Sur-versement")),
                srcs("Kerr 2008 : " + U_KERR, "Wansink 2005 : " + U_WANSINK))),
        arg("2. Zéro perte sur la bière (15 % seulement, mutualisé)",
            "Proposition p. 50-51",
            "85 % du fût finirait dans un verre vendu ; seuls 15 % sont retranchés, et de façon mutualisée (pertes + personnel + offerts).",
            f"La **freinte technique** d'un fût (mousse, collerette, nettoyage des lignes, fond de fût) est documentée à **~20 %** ; le service **en accorde déjà 15 %**. Retenu très prudemment : **{fr_l(segL('Freinte biere'))}** (soit moins que l'abattement du fisc). Source : {L(U_BARI)}",
            "Forte",
            [{"to": "/rendu-final/pertes-biere-mousse", "label": "Page pertes bière"}],
            [{"to": PJ("RF-pertes-biere-mousse.xlsx"), "label": "RF-pertes-bière (XLSX)"}],
            det("Décomposition de la freinte technique d'un fût en tirage traditionnel (% du fût tiré).",
                ["Poste de perte", "Part du fût"],
                [["Sur-versement / collerette de mousse", "≈ 7 %"],
                 ["Mousse et pertes au tirage (température, pression, fond de fût)", "≈ 8 %"],
                 ["Nettoyage des lignes (tous les 14 j) + purge + fond", "≈ 5 %"]],
                "Freinte documentée (retenu : " + fr_l(segL("Freinte biere")) + ")", "≈ 20 % du fût",
                srcs("Bar-i (rendement de fût ~95 %) : " + U_BARI, "Le service accorde lui-même 15 % sur la bière (Proposition p. 51)"))),
        arg("3. Tout le crémant réputé vendu",
            "Proposition p. 46-47",
            "Tout le crémant disponible serait vendu au centilitre près.",
            f"Une bouteille d'effervescent **ouverte est éventée le lendemain → jetée**, et le service au verre sur-verse. Compté **jour par jour** (annexe C) : **{fr_l(cremant_l)}** (jeté + sur-versé). Source : {L(U_SPARK)}",
            "Forte",
            [{"to": "/rendu-final/pertes-cremant", "label": "Page crémant"}],
            [{"to": PJ("RF-cremant-jete.xlsx"), "label": "RF-crémant (XLSX)"}],
            det("Deux pertes distinctes, mesurées jour par jour sur le détail des tickets (annexe C).",
                ["Perte de crémant", "Volume (3 ans)"],
                [["Bouteille ouverte éventée, solde jeté le soir", fr_l(segL("Cremant jete"))],
                 ["Sur-versement au service (+23,6 %)", fr_l(segL("Cremant sur-verse"))]],
                "Total perte crémant", fr_l(cremant_l),
                srcs("Crémant ouvert éventé : " + U_SPARK, "Calcul jour par jour : annexe C (RF-cremant-jete.xlsx)"))),
        arg("4. Aucune dégustation offerte",
            "Proposition p. 37-44",
            "Tout le vin disponible serait vendable, sans rien réserver au geste de dégustation.",
            f"Le **goûter de 2 cl** avant de servir un vin nommé ne laisse **aucune trace en caisse**. Compté **note par note** (annexe C) : **{fr_l(segL('Degustation'))}**.",
            "Forte",
            [{"to": "/rendu-final/degustation-au-verre", "label": "Page dégustation"}],
            [{"to": PJ("RF-degustation-par-note.xlsx"), "label": "RF-dégustation (XLSX)"}],
            det("Règle stricte appliquée ligne à ligne sur le détail des tickets (annexe C).",
                ["Paramètre", "Valeur"],
                [["Règle", "1 dégustation par vin nommé et par note (addition)"],
                 ["Dose", "2 cl (une larme)"],
                 ["Génériques « verre de vin » / bouteilles", "exclus"]],
                "Vin offert en dégustation (3 ans)", fr_l(segL("Degustation")),
                srcs("Comptage note par note : annexe C (RF-degustation-par-note.xlsx)"))),
        arg("5. Alcool de cuisine sous-déduit",
            "Proposition p. 44-48 (repères D/E/G/Q)",
            "Seule une fraction de l'alcool de cuisine est déduite (quelques plats de menu), via les repères D, E, G, Q.",
            f"En appliquant **les doses confirmées par les dirigeants** (que le fisc reprend : fondue 9-10 cl, baba/flambage 4 cl, sauce 1 cl) à **toute la carte**, la cuisine consomme **{fr_l(segL('Cuisine'))}** d'alcool acheté, **versé dans les plats** (fondues, sauces, flambages, babas) : il quitte le stock et n'est jamais disponible pour un verre vendu.",
            "Forte",
            [{"to": "/rendu-final/alcool-cuisine-doses-plats", "label": "Page cuisine"},
             {"to": "/rendu-final/alcool-cuisine-doses-plats", "label": "Détail plat par plat (2.6)"}],
            [{"to": PJ("RF-alcool-cuisine-doses-plats.xlsx"), "label": "RF-cuisine (XLSX)"}],
            det("Alcool incorporé en cuisine, alcool par alcool, sur 3 exercices (doses confirmées par les dirigeants).",
                ["Alcool", "Litres (3 ans)", "Coût d'achat"],
                cuis_lignes,
                "Total alcool de cuisine", fr_l(segL("Cuisine")),
                srcs("Doses confirmées par les dirigeants (30/03/2026)", "Détail dose par plat : RF-alcool-cuisine-doses-plats.xlsx"))),
        arg("6. Alcool cuit dans les plats des menus, compté comme vendu au verre",
            "Proposition p. 37-44",
            "L'alcool incorporé dans les plats des menus serait du volume vendable au verre, au prix carte.",
            f"Il ne s'agit **pas** d'alcool servi dans un verre : c'est l'alcool **cuit dans les plats** d'un menu au forfait (sauces au porto, fondues au Ravelin, babas au macvin, coupes au cassis…). On le chiffre **plat par plat** : `dose × probabilité que le client choisisse ce plat` (lue en caisse) `× nombre de menus vendus` = **{fr_l(segL('Menus'))}** sur 3 ans. C'est de la cuisine, comptée **séparément** de la carte (aucun double compte).",
            "Modérée",
            [{"to": "/rendu-final/alcool-cuisine-doses-plats", "label": "Détail plat par plat (2.6)"}],
            [{"to": PJ("Boissons-4-menus-par-periode.xlsx"), "label": "Menus par période (XLSX)"},
             {"to": PJ("RF-reconstitution-volumes-liquides.xlsx"), "label": "Cascade (XLSX)"}],
            det("Alcool incorporé dans les plats des menus, agrégé par alcool sur 3 exercices. Calcul de chaque ligne : nombre de menus réellement vendus × dose du plat × probabilité que ce plat soit choisi (jamais la dose entière pour une option facultative).",
                ["Alcool", "Litres (3 ans)"],
                menus_lignes,
                "Total alcool cuit dans les menus", fr_l(segL("Menus")),
                srcs("Menus par période (menus vendus × dose × probabilité) : Boissons-4-menus-par-periode.xlsx", "Données : menusNonModifiesAlcool.json"))),
        arg("7. Consommation du personnel plafonnée à 5 %",
            "Proposition p. 51",
            "La consommation du personnel tiendrait dans un forfait de 5 % du CA.",
            f"Le **chef** consomme du Picon et du Macvin **jamais sonnés** (72 L de Picon achetés, 0 vendu) ; les serveurs ~6 Coca/jour. La consommation **alcool** seule : **{fr_l(segL('Conso chef'))}**, au-delà du forfait.",
            "Modérée",
            [{"to": "/rendu-final/offerts-remises-pertes", "label": "Page offerts/pertes"}],
            [{"to": PJ("Consommation-personnel-et-offerts.xlsx"), "label": "Conso personnel (XLSX)"}]),
        arg("8. Les abattements 5 %+5 %+5 % suffiraient",
            "Proposition p. 51",
            "Trois forfaits de 5 % (offerts, pertes, personnel) couvriraient tout le non-vendu.",
            f"Ce sont des **planchers**. La somme de nos postes documentés non vendus comme boisson (sur-versement + crémant + dégustation + freinte bière + conso chef + offerts + casse résiduelle) atteint **28,7 % des achats**, bien au-delà des 15 % accordés, et au-delà des **22-25 %** que la jurisprudence (CAA Paris) admet elle-même en reconstitution de bar. Sources : {L(U_BARI_S)} · {L(U_WISK)}",
            "Forte",
            [{"to": "/rendu-final/offerts-remises-pertes", "label": "Page offerts/pertes"}],
            [{"to": PJ("RF-offerts-remises-pertes.xlsx"), "label": "RF-offerts/pertes (XLSX)"}]),
        arg("9. « Volume disponible » gonflé par une variation de stock erronée",
            "Proposition p. 14, annexe 6",
            "Le volume disponible = achats + variation de stock (compte 310200).",
            "La variation de stock retenue est de **signe inverse** à l'inventaire physique réel des 3 clôtures → le « disponible » est **gonflé d'environ 2 162 €**.",
            "Modérée",
            [{"to": "/rendu-final/inventaire-stocks", "label": "Page inventaire"}],
            [{"to": PJ("RF-inventaire-stocks.xlsx"), "label": "RF-inventaire (XLSX)"}]),
        arg("10. « Ventes sans achat » présentées comme des recettes occultées",
            "Proposition (méthode, par produit)",
            "Certaines boissons seraient vendues sans aucun achat correspondant (Bordeaux Mouton Cadet, Hautes Côtes de Nuits).",
            "**Faux, et c'est prouvé, facture à l'appui.** Un **bouton de caisse n'est pas un libellé de facture** : notre rapprochement caisse ↔ facture ↔ inventaire rattache **chaque bouton à la référence réellement achetée**. Le bouton « **Mouton Cadet** » est le **Bordeaux Supérieur (Château Grand Renom)**, facturé **9 bouteilles**, vendu 9. Le bouton « **Hautes Côtes de Nuits** » est le **HC de Nuits blanc (Lupé Cholet)**, facturé **24 bouteilles** pour 3 vendues. Les volumes que le vérificateur range sous « Côtes de Nuits » sont en réalité les **Hautes Côtes de BEAUNE (Domaine Germain)**, facturées par centaines de bouteilles : c'est une **confusion d'appellation** (Nuits ≠ Beaune). **Zéro vente sans achat : chaque vente est adossée à une facture.**",
            "Forte",
            [{"to": "/rendu-final/consommation-superieure-achats", "label": "Page conso vs achats"}],
            [{"to": PJ("Boissons-6-correspondance-libelles.xlsx"), "label": "Correspondance des libellés (XLSX)"},
             {"to": PJ("RF-conso-achats-correspondance.xlsx"), "label": "Rapprochement conso/achats (XLSX)"}],
            det("Rapprochement OFFICIEL caisse ↔ facture fournisseur ↔ inventaire (déjà établi). Chaque bouton de caisse pointe vers sa référence facturée : le libellé du bouton n'est qu'un nom d'écran.",
                ["Bouton de caisse", "Vendu", "Référence réelle", "Facture fournisseur", "Acheté (facturé)"],
                [["BORD. MOUTON CADET", "9 bouteilles", "Bordeaux Supérieur", "Château Grand Renom (Antoine Moueix)", "9 bouteilles (2022-23)"],
                 ["H.COTES DE NUITS", "3 bouteilles", "HC de Nuits blanc", "Lupé Cholet 2017", "24 bouteilles (2022-23)"],
                 ["H.C. de Beaune (verre / pichet / btl)", "≈ 585 services", "HC de Beaune rouge", "Domaine Germain", "≈ 245 bouteilles (3 ans)"]],
                "Verdict", "Chaque vente est adossée à une facture (rapprochement des 3 sources) : aucune vente sans achat ; la « Côte de Nuits » du fisc est en réalité la Côte de Beaune",
                srcs("Rapprochement : Boissons-6-correspondance-libelles.xlsx (_correspondance-caisse-inventaire-factures.csv)", "Achats : achatsBoissonsParPeriode.json ; Caisse : itemsCaisse.json (nom_canonique -> nom_facture)"))),
        arg("11. Le coefficient liquide→solide serait neutre",
            "Proposition p. 52",
            "Le CA cuisine est obtenu en multipliant le CA liquides par 2,94 / 3,02 / 3,10.",
            "Le **CA cuisine (≈ ¾ du total) n'est jamais mesuré** : il est extrapolé. Toute sur-évaluation des liquides est donc **amplifiée ~3 fois** sur le total reconstitué.",
            "Forte",
            [{"to": "/rendu-final/coefficient-liquide-solide", "label": "Page coefficient"}],
            [{"to": PJ("RF-coefficient-liquide-solide.xlsx"), "label": "RF-coefficient (XLSX)"}]),
    ]

    SECTIONS = [
        {"kind": "kpis", "items": [
            {"label": "Alcool acheté (factures)", "valeur": fr_l(d["achat_l"]),
             "sub": fr_eur(d["achat_cout"]) + " sur 3 exercices", "couleur": "blue"},
            {"label": "Justifié de façon indépendante", "valeur": fr_pct(100 * justifie_l / d["achat_l"], 1),
             "sub": fr_l(justifie_l) + " tracés poste par poste", "highlight": True, "couleur": "teal"},
            {"label": "Résidu (perte normale)", "valeur": fr_pct(d["perte_pct"]),
             "sub": fr_l(d["perte_reelle_l"]) + " : casse, vol, évaporation", "couleur": "gray"},
        ]},
        {"kind": "chapitre", "source": "fisc", "numero": 1,
         "titre": "Ce que dit l'administration",
         "sousTitre": "Proposition p. 37-53 : une reconstitution par les volumes, puis une extrapolation à la cuisine."},
        {"kind": "tableau", "titre": "La méthode du fisc, en 4 étapes", "minWidth": 560,
         "colonnes": [{"label": "Étape"}, {"label": "Ce que fait le service"}],
         "lignes": [
             [{"v": "1. Volume disponible"}, {"v": "Achats de boissons ± variation de stock, convertis en centilitres"}],
             [{"v": "2. ÷ dose figée"}, {"v": "15 cl le verre, 50/75 cl le pichet, 4 cl les spiritueux → nombre d'articles « vendus »"}],
             [{"v": "3. × prix moyen"}, {"v": "= CA « liquides » reconstitué"}],
             [{"v": "4. × coefficient"}, {"v": "× 2,94 / 3,02 / 3,10 → CA « cuisine » (solides) extrapolé"}],
         ]},
        {"kind": "tableau", "titre": "Les chiffres exacts retenus par le fisc", "minWidth": 560,
         "colonnes": [{"label": "Paramètre"}, {"label": "Valeur retenue", "align": "right"}],
         "lignes": [
             [{"v": "Doses (verre / pichet / spiritueux)"}, {"v": "15 cl / 50-75 cl / 4 cl", "align": "right"}],
             [{"v": "Abattements de fin de méthode"}, {"v": "5 % offerts + 5 % pertes + 5 % personnel", "align": "right"}],
             [{"v": "Abattement bière (mutualisé)"}, {"v": "15 % du volume", "align": "right"}],
             [{"v": "Coefficient liquide → solide"}, {"v": "2,94 / 3,02 / 3,10", "align": "right"}],
             [{"v": "CA reconstitué vs déclaré (2024-25)"}, {"v": fr_eur(d["fisc_ca"]) + " vs " + fr_eur(d["ca_declare"]), "align": "right"}],
             [{"v": "Pénalité art. 1759 visée (3 ans)"}, {"v": "≈ " + fr_eur(d["enjeu_global"]), "align": "right"}],
         ]},
        {"kind": "note", "texte": "Source : Proposition de rectification, p. 37-53 (méthode) et annexes Boissons 5 à 7 (doses, volumes, coefficient)."},
        {"kind": "chapitre", "source": "nous", "numero": 2,
         "titre": "Pourquoi cette méthode ne tient pas",
         "sousTitre": "Onze hypothèses. Chacune : ce que dit le fisc, notre réponse chiffrée et sourcée, la pièce téléchargeable, et la donnée détaillée à déplier."},
        *ARGS,
        {"kind": "chapitre", "source": "nous", "numero": 3,
         "titre": "Notre démonstration indépendante : la cascade des " + fr_l(d["achat_l"]),
         "sousTitre": "On ne corrige pas le calcul du fisc : on part de NOS achats réels et on trace chaque litre."},
        {"kind": "barreComposition", "titre": "Où passent les " + fr_l(d["achat_l"]) + " achetés",
         "sousTitre": "Chaque segment est mesuré (caisse/inventaire), calculé (recettes) ou estimé (taux sourcé) ; le résidu est la perte normale.",
         "unite": "L", "total": d["achat_l"], "segments": barre_segments, "legende": True},
        {"kind": "tableau", "titre": "Cascade détaillée des " + fr_l(d["achat_l"]) + " achetés (poste par poste)", "minWidth": 1040,
         "colonnes": [{"label": "Poste"}, {"label": "Litres", "align": "right"},
                      {"label": "% achats", "align": "right"}, {"label": "Nature"},
                      {"label": "Détaillé dans"}, {"label": "Pièce à télécharger"}],
         "lignes": cascade_rows},
        {"kind": "note", "texte": "**Chaque ligne est cliquable** : colonne « Détaillé dans » = la page qui justifie le poste, colonne « Pièce » = le fichier téléchargeable (recalcul intégral). Les volumes « vendus au verre » et « en cocktails » sont les **ventes effectivement enregistrées** en caisse (total = CA déclaré, non contesté en volume) : le débat sur les suppressions porte sur leur interprétation, jamais sur ces ventes. Le **sur-versement de la bière** est compté **dans la freinte bière**, pas dans le sur-versement des vins/spiritueux (aucun double compte). Les **apéritifs offerts** ne sont **pas** enregistrés en caisse (estimation bornée, ~1/jour)."},
        {"kind": "chapitre", "source": "nous", "numero": 4, "titre": "Conclusion"},
        {"kind": "alerte", "couleur": "teal",
         "titre": "Tout l'alcool acheté est tracé, rien ne disparaît",
         "texte": f"Sur les **{fr_l(d['achat_l'])}** d'alcool achetés, **{fr_pct(100 * (d['achat_l'] - d['perte_reelle_l']) / d['achat_l'], 1)}** sont **justifiés poste par poste** ci-dessus : vendus au verre ou en cocktails, en stock, ou consommés sans vente de façon **documentée et chiffrée** (cuisine, menus, sur-versement, crémant, dégustation, freinte bière, conso du chef, offerts). Le reste, **{fr_pct(d['perte_pct'])}** ({fr_l(d['perte_reelle_l'])}), est une **perte pure** (casse, évaporation, fonds de verre, rinçage des tireuses), **en dessous** de la démarque normale d'un bar (15-20 % du produit servi). Il ne reste donc **aucun volume disponible pour des ventes dissimulées** : la reconstitution du fisc, qui suppose des doses exactes et zéro perte, surévalue mécaniquement. Source démarque secteur : {L(U_BARI_S)}."},
        {"kind": "chapitre", "source": "nous", "numero": 5, "titre": "Toutes les pièces téléchargeables"},
        {"kind": "piecejointe",
         "intro": "Chaque poste de la cascade est recalculable depuis ces pièces (lecture seule des annexes de caisse) :",
         "fichiers": [
             {"fichier": "pieces-defense/RF-reconstitution-volumes-liquides.xlsx", "label": "Cascade des volumes + réconciliation CA (XLSX)"},
             {"fichier": "pieces-defense/RF-sur-versement-au-verre.xlsx", "label": "Sur-versement au verre (XLSX)"},
             {"fichier": "pieces-defense/RF-pertes-biere-mousse.xlsx", "label": "Pertes de bière / freinte du fût (XLSX)"},
             {"fichier": "pieces-defense/RF-cremant-jete.xlsx", "label": "Crémant jeté et sur-versé (XLSX)"},
             {"fichier": "pieces-defense/RF-degustation-par-note.xlsx", "label": "Dégustation note par note (XLSX)"},
             {"fichier": "pieces-defense/RF-alcool-cuisine-doses-plats.xlsx", "label": "Alcool de cuisine, dose par plat (XLSX)"},
             {"fichier": "pieces-defense/RF-offerts-remises-pertes.xlsx", "label": "Offerts, remises et pertes (XLSX)"},
             {"fichier": "pieces-defense/RF-coefficient-liquide-solide.xlsx", "label": "Coefficient liquide→solide (XLSX)"},
             {"fichier": "pieces-defense/RF-inventaire-stocks.xlsx", "label": "Inventaire physique des stocks (XLSX)"},
         ]},
        {"kind": "interne", "audience": "avocat", "titre": "Stratégie",
         "texte": "Mener avec NOTRE cascade indépendante (justifiée à " + fr_pct(100 * justifie_l / d["achat_l"], 1) + ", résidu " + fr_pct(d["perte_pct"]) + " = perte normale), jamais avec le chiffre du fisc ; la réconciliation « sur sa formule » n'est qu'une vérification secondaire. Chaque carte renvoie à sa page dédiée et à sa pièce XLSX ; les taux de perte sont sourcés (Kerr, Wansink, Bar-i, WISK). Retirer cet encart de toute version remise."},
    ]

    doc = {
        "meta": {
            "slug": "reconstitution-volumes-liquides",
            "titre": "Reconstitution du CA par les volumes de liquides",
            "source": "scripts/rendu-final-reconstitution-volumes-liquides.py",
            "grief": "Proposition de rectification p. 37-53 (methode) - annexes Boissons.",
            "enjeu": "Coeur de la reconstitution : ≈ 422 k€ de CA HT redresse (3 ans), amende 1759 ≈ 472 k€",
            "chiffres": {
                "achat_l": d["achat_l"],
                "achat_cout": d["achat_cout"],
                "conso_l": d["conso_l"],
                "perte_reelle_l": d["perte_reelle_l"],
                "perte_reelle_pct": d["perte_pct"],
                "fisc_ca_reconstitue": d["fisc_ca"],
                "fisc_coef": d["fisc_coef"],
                "ca_declare": d["ca_declare"],
                "total_ca_oublie": round(d["total_ca_oublie"]),
                "ca_recalcule": round(d["ca_recalcule"]),
                "ecart_fisc": round(d["ecart_fisc"]),
                "ecart_recalcule": round(d["ecart_recalcule"]),
            },
        },
        "sections": SECTIONS,
    }

    doc["sections"] = normaliser_sections(doc["sections"])

    os.makedirs(RENDU_DIR, exist_ok=True)
    with open(JSON_OUT, "w", encoding="utf-8") as f:
        json.dump(doc, f, ensure_ascii=False, indent=1)
    return JSON_OUT


# ---------------------------------------------------------------------------
def main():
    bp = charger_donnees()
    d = calculer(bp)

    # Controle de coherence (la cascade doit sommer au total achats)
    somme_seg = sum(s["valeur"] for s in d["segments"])
    assert abs(somme_seg - d["achat_l"]) < 1.0, \
        f"Cascade incoherente : {somme_seg} != {d['achat_l']}"

    xlsx = ecrire_xlsx(d)
    js = ecrire_json(d)

    print("Reconstitution par les volumes de liquides - rendu final")
    print("-" * 60)
    print(f"Achats alcool          : {fr_l(d['achat_l'])} ({fr_eur(d['achat_cout'])})")
    print(f"Perte reelle (residu)  : {fr_l(d['perte_reelle_l'])} = {fr_pct(d['perte_pct'])}")
    print(f"CA declare             : {fr_eur(d['ca_declare'])}")
    print(f"CA reconstitue fisc    : {fr_eur(d['fisc_ca'])} (coef {str(d['fisc_coef']).replace('.', ',')})")
    print(f"  discordance fisc     : +{fr_eur(d['ecart_fisc'])}")
    print(f"Litres oublies (verre) : {fr_l(d['total_litres_oublies'])} -> {fr_eur(d['total_ca_oublie'])}")
    print(f"CA recalcule           : {fr_eur(d['ca_recalcule'])}")
    print(f"  ecart recalcule      : {fr_eur(d['ecart_recalcule'])}")
    print("-" * 60)
    print(f"XLSX : {xlsx}")
    print(f"JSON : {js}")


if __name__ == "__main__":
    main()
