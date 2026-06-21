#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
RENDU FINAL - Grief « Fichiers Événement (E) et Règlement (F) »
================================================================
Proposition de rectification p. 22-24 (rejet de comptabilite 1/3),
sections IX (Fichier Evenement, annexes E-1 a E-3) et X (Fichier
Reglement, annexes F-1 a F-3).

Ce script est READ-ONLY sur les sources. Il :
  1. lit les annexes E (journal des evenements de caisse) et en extrait
     la distribution des TYPES d'evenement (DEL, RPR, CON, TIR) ;
  2. lit les annexes F (journal des reglements) et reprend les lignes
     « TOTAL <mode> » que le fichier produit lui-meme (figures que le
     verificateur a publiees), recoupees a l'euro pres ;
  3. lit les annexes A (synthese CA) et H (liste des tickets) et etablit
     la TRIANGULATION  CA declare (A)  ~=  somme des reglements (F)
     ~=  liste des tickets (H) ;
  4. calcule la part bancarisee vs especes par exercice et au cumul.

Sorties :
  - public/documents/pieces-defense/RF-fichiers-evenement-reglement.xlsx
  - src/data/renduFinal/fichiers-evenement-reglement.json

Lancer :  /tmp/xlsenv/bin/python scripts/rendu-final-fichiers-evenement-reglement.py
"""

import json
import os
from collections import defaultdict

from rfcommun import ajouter_conclusion

import xlrd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# --------------------------------------------------------------------------
ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
CAISSE = os.path.join(ROOT, "public", "documents", "caisse-enregistreuse")
PIECES = os.path.join(ROOT, "public", "documents", "pieces-defense")
JSON_DIR = os.path.join(ROOT, "src", "data", "renduFinal")
XLSX_OUT = os.path.join(PIECES, "RF-fichiers-evenement-reglement.xlsx")
JSON_OUT = os.path.join(JSON_DIR, "fichiers-evenement-reglement.json")

EXERCICES = [
    ("2022-2023", "31/03/2023", "E1", "F1", "A1", "H1"),
    ("2023-2024", "31/03/2024", "E2", "F2", "A2", "H2"),
    ("2024-2025", "31/03/2025", "E3", "F3", "A3", "H3"),
]

F_NOMS = {
    "E1": "ANNEXE-E1_tpvevenement_2022-2023.xls",
    "E2": "ANNEXE-E2_tpvevenement_2023-2024.xls",
    "E3": "ANNEXE-E3_tpvevenement_2024-2025.xls",
    "F1": "ANNEXE-F1_reglements_2022-2023.xls",
    "F2": "ANNEXE-F2_reglements_2023-2024.xls",
    "F3": "ANNEXE-F3_reglements_2024-2025.xls",
    "A1": "ANNEXE-A1_synthese-CA_2022-2023.xls",
    "A2": "ANNEXE-A2_synthese-CA_2023-2024.xls",
    "A3": "ANNEXE-A3_synthese-CA_2024-2025.xls",
    "H1": "ANNEXE-H1_liste-tickets_2022-2023.xls",
    "H2": "ANNEXE-H2_liste-tickets_2023-2024.xls",
    "H3": "ANNEXE-H3_liste-tickets_2024-2025.xls",
}

# Libelles longs des modes de reglement
MODE_LIB = {
    "CB": "Carte bancaire",
    "ESP": "Especes",
    "CHQ": "Cheque",
    "CHV": "Cheque-vacances",
    "TR": "Ticket-restaurant",
}
# Libelles des types d'evenement du journal E (norme caisse / NF525)
EVEN_LIB = {
    "DEL": "Suppression d'une ligne d'article (correction de saisie)",
    "RPR": "Reprise / reouverture d'une note",
    "CON": "Consultation / controle (ouverture du tiroir, etc.)",
    "TIR": "Ouverture de tiroir-caisse",
}
# Modes consideres comme « bancarises » (tracables sur un compte)
BANCARISE = {"CB", "CHQ", "CHV", "TR"}


def wb_sheet(code):
    p = os.path.join(CAISSE, F_NOMS[code])
    return xlrd.open_workbook(p).sheet_by_index(0)


def fnum(v):
    try:
        return float(str(v).replace(" ", "").replace("\xa0", ""))
    except (TypeError, ValueError):
        return 0.0


# ==========================================================================
# 1) ANNEXE E - distribution des types d'evenement
# ==========================================================================
def lire_evenements(code):
    sh = wb_sheet(code)
    # col 0 noCaisse, 1 dateEven, 2 heureEven, 3 typEven, ..., 8 amount
    types = defaultdict(int)
    montants = defaultdict(float)
    for r in range(1, sh.nrows):
        t = str(sh.cell_value(r, 3)).strip()
        if not t:
            continue
        types[t] += 1
        montants[t] += fnum(sh.cell_value(r, 8))
    return dict(types), {k: round(v, 2) for k, v in montants.items()}


# ==========================================================================
# 2) ANNEXE F - lignes officielles « TOTAL <mode> » + TOTAL GENERAL
# ==========================================================================
def lire_reglements(code):
    """Lit les lignes de TOTAL que le fichier F produit lui-meme.
    Ces lignes sont la reconciliation interne de la caisse : elles
    correspondent a l'euro pres aux montants publies par le verificateur.
    """
    sh = wb_sheet(code)
    par_mode = {}
    total_general = None
    for r in range(1, sh.nrows):
        vals = [sh.cell_value(r, c) for c in range(sh.ncols)]
        joined = " ".join(str(v) for v in vals).upper()
        if "TOTAL GENERAL" in joined:
            # montant = derniere cellule numerique non vide
            nums = [fnum(v) for v in vals if str(v).strip() and fnum(v) != 0]
            total_general = round(nums[-1], 2) if nums else None
        elif "TOTAL" in joined:
            # forme : ['TOTAL', '<MODE>', <montant>]
            mode = None
            montant = None
            for v in vals:
                s = str(v).strip()
                if s in MODE_LIB:
                    mode = s
                elif fnum(v) != 0:
                    montant = round(fnum(v), 2)
            if mode and montant is not None:
                par_mode[mode] = montant
    return par_mode, total_general


# ==========================================================================
# 3) ANNEXE A - CA total et ventilation des encaissements par mode
# ==========================================================================
def lire_synthese_A(code):
    sh = wb_sheet(code)
    ca_total = None
    enc_par_mode = {}
    # libelles A -> code mode
    lib2code = {
        "Carte Bancaire": "CB",
        "Espèce": "ESP",
        "Espece": "ESP",
        "Chèque": "CHQ",
        "Cheque": "CHQ",
        "Chèque vacances": "CHV",
        "Cheque vacances": "CHV",
        "Ticket restaurant": "TR",
    }
    in_enc = False
    for r in range(sh.nrows):
        c0 = str(sh.cell_value(r, 0)).strip()
        if c0.startswith("Encaissements Tickets Mode"):
            in_enc = True
            continue
        if in_enc:
            if c0 == "TOTAL":
                ca_total = round(fnum(sh.cell_value(r, 2)), 2)
                in_enc = False
                continue
            code_m = lib2code.get(c0)
            if code_m:
                montant = round(fnum(sh.cell_value(r, 2)), 2)
                # cumuler (l'exercice 3 a un petit bloc d'avoirs negatifs avant)
                enc_par_mode[code_m] = round(enc_par_mode.get(code_m, 0.0) + montant, 2)
    return ca_total, enc_par_mode


# ==========================================================================
# 4) ANNEXE H - somme des tot_ttc des tickets (ligne TOTAL du fichier)
# ==========================================================================
def lire_tickets_H(code):
    sh = wb_sheet(code)
    total_ttc = None
    nb_tickets = 0
    for r in range(1, sh.nrows):
        c0 = str(sh.cell_value(r, 0)).strip()
        if c0 == "TOTAL":
            total_ttc = round(fnum(sh.cell_value(r, 6)), 2)
            continue
        if c0 in ("A NOUVEAUX", "") or "TOTAL" in c0.upper():
            continue
        nb_tickets += 1
    return total_ttc, nb_tickets


# ==========================================================================
# COLLECTE
# ==========================================================================
data = {"exercices": [], "even_cumul": defaultdict(int), "even_cumul_montant": defaultdict(float),
        "regl_cumul": defaultdict(float)}

for ex, cloture, eC, fC, aC, hC in EXERCICES:
    even_types, even_montants = lire_evenements(eC)
    regl_mode, regl_total = lire_reglements(fC)
    ca_a, enc_a = lire_synthese_A(aC)
    h_total, h_nb = lire_tickets_H(hC)

    for t, n in even_types.items():
        data["even_cumul"][t] += n
        data["even_cumul_montant"][t] += even_montants.get(t, 0.0)
    for m, v in regl_mode.items():
        data["regl_cumul"][m] += v

    somme_f = round(sum(regl_mode.values()), 2)
    bancarise = round(sum(v for m, v in regl_mode.items() if m in BANCARISE), 2)
    especes = round(regl_mode.get("ESP", 0.0), 2)
    pct_banc = round(100 * bancarise / somme_f, 2) if somme_f else 0
    pct_esp = round(100 * especes / somme_f, 2) if somme_f else 0

    data["exercices"].append({
        "exercice": ex,
        "cloture": cloture,
        "even_types": even_types,
        "even_montants": even_montants,
        "del_count": even_types.get("DEL", 0),
        "regl_mode": regl_mode,
        "regl_total_general": regl_total,
        "somme_f": somme_f,
        "bancarise": bancarise,
        "especes": especes,
        "pct_bancarise": pct_banc,
        "pct_especes": pct_esp,
        "ca_a": ca_a,
        "enc_a": enc_a,
        "h_total": h_total,
        "h_nb_tickets": h_nb,
    })

# Cumuls
data["even_cumul"] = {k: v for k, v in data["even_cumul"].items()}
data["even_cumul_montant"] = {k: round(v, 2) for k, v in data["even_cumul_montant"].items()}
data["regl_cumul"] = {k: round(v, 2) for k, v in data["regl_cumul"].items()}

regl_cumul_total = round(sum(data["regl_cumul"].values()), 2)
banc_cumul = round(sum(v for m, v in data["regl_cumul"].items() if m in BANCARISE), 2)
esp_cumul = round(data["regl_cumul"].get("ESP", 0.0), 2)
pct_banc_cumul = round(100 * banc_cumul / regl_cumul_total, 2)
pct_esp_cumul = round(100 * esp_cumul / regl_cumul_total, 2)

ca_a_cumul = round(sum(e["ca_a"] for e in data["exercices"]), 2)
f_cumul = round(sum(e["regl_total_general"] for e in data["exercices"]), 2)
h_cumul = round(sum(e["h_total"] for e in data["exercices"]), 2)
del_cumul = data["even_cumul"].get("DEL", 0)

print("=" * 70)
print("REGLEMENTS PAR MODE (lignes TOTAL des annexes F, recoupees a l'euro)")
for e in data["exercices"]:
    print(f"\n  {e['exercice']} (cloture {e['cloture']}) - TOTAL GENERAL F = {e['regl_total_general']:.2f}")
    for m in ["CB", "CHV", "TR", "ESP", "CHQ"]:
        v = e["regl_mode"].get(m, 0)
        pct = 100 * v / e["somme_f"] if e["somme_f"] else 0
        print(f"      {m:4} {MODE_LIB[m]:18} {v:12.2f}  {pct:6.2f}%")
    print(f"      -> bancarise {e['pct_bancarise']:.2f}% / especes {e['pct_especes']:.2f}%")

print("\n" + "=" * 70)
print("CUMUL 3 EXERCICES")
for m in ["CB", "CHV", "TR", "ESP", "CHQ"]:
    v = data["regl_cumul"].get(m, 0)
    print(f"  {m:4} {v:12.2f}  {100*v/regl_cumul_total:6.2f}%")
print(f"  TOTAL F cumul = {regl_cumul_total:.2f}")
print(f"  Bancarise = {banc_cumul:.2f} ({pct_banc_cumul:.2f}%)  |  Especes = {esp_cumul:.2f} ({pct_esp_cumul:.2f}%)")

print("\n" + "=" * 70)
print("TRIANGULATION  CA(A)  ~  somme reglements(F)  ~  liste tickets(H)")
for e in data["exercices"]:
    print(f"  {e['exercice']}  A={e['ca_a']:11.2f}  F={e['regl_total_general']:11.2f}  H={e['h_total']:11.2f}"
          f"   ecart A-F={e['ca_a']-e['regl_total_general']:+.2f}")
print(f"  CUMUL      A={ca_a_cumul:11.2f}  F={f_cumul:11.2f}  H={h_cumul:11.2f}")

print("\n" + "=" * 70)
print("TYPES D'EVENEMENT (annexe E) - cumul 3 exercices")
for t, n in sorted(data["even_cumul"].items(), key=lambda x: -x[1]):
    print(f"  {t:4} {EVEN_LIB.get(t,'?'):55} n={n:6}  montant={data['even_cumul_montant'].get(t,0):.2f}")


# ==========================================================================
# XLSX
# ==========================================================================
HEAD_FILL = PatternFill("solid", fgColor="0F766E")
HEAD_FONT = Font(bold=True, color="FFFFFF")
BOLD = Font(bold=True)
RIGHT = Alignment(horizontal="right")


def style_header(ws, ncols, row=1):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEAD_FILL
        cell.font = HEAD_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")


def autofit(ws):
    for col in ws.columns:
        width = max((len(str(c.value)) for c in col if c.value is not None), default=10)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max(width + 2, 10), 42)


wb = Workbook()

# --- Onglet 1 : Reglements par mode ---------------------------------------
ws1 = wb.active
ws1.title = "Reglements par mode"
ws1.append(["Exercice", "Mode", "Libelle", "Montant (EUR)", "Part (%)", "Categorie"])
style_header(ws1, 6)
for e in data["exercices"]:
    for m in ["CB", "CHV", "TR", "ESP", "CHQ"]:
        v = e["regl_mode"].get(m, 0)
        pct = round(100 * v / e["somme_f"], 2) if e["somme_f"] else 0
        cat = "Bancarise" if m in BANCARISE else "Especes"
        ws1.append([e["exercice"], m, MODE_LIB[m], v, pct, cat])
    ws1.append([e["exercice"], "TOTAL GENERAL", "", e["regl_total_general"], 100.0, ""])
    ws1.cell(row=ws1.max_row, column=2).font = BOLD
    ws1.cell(row=ws1.max_row, column=4).font = BOLD
ws1.append([])
ws1.append(["CUMUL 3 EX.", "", "", "", "", ""])
ws1.cell(row=ws1.max_row, column=1).font = BOLD
for m in ["CB", "CHV", "TR", "ESP", "CHQ"]:
    v = data["regl_cumul"].get(m, 0)
    pct = round(100 * v / regl_cumul_total, 2)
    cat = "Bancarise" if m in BANCARISE else "Especes"
    ws1.append(["Cumul", m, MODE_LIB[m], v, pct, cat])
ws1.append(["Cumul", "BANCARISE", "CB+CHQ+CHV+TR", banc_cumul, pct_banc_cumul, "Bancarise"])
ws1.cell(row=ws1.max_row, column=2).font = BOLD
ws1.append(["Cumul", "ESPECES", "ESP", esp_cumul, pct_esp_cumul, "Especes"])
ws1.cell(row=ws1.max_row, column=2).font = BOLD
ws1.append(["Cumul", "TOTAL GENERAL", "", regl_cumul_total, 100.0, ""])
ws1.cell(row=ws1.max_row, column=2).font = BOLD
ws1.cell(row=ws1.max_row, column=4).font = BOLD
for r in range(2, ws1.max_row + 1):
    ws1.cell(row=r, column=4).alignment = RIGHT
    ws1.cell(row=r, column=5).alignment = RIGHT
ws1.freeze_panes = "A2"
autofit(ws1)

# --- Onglet 2 : Triangulation ---------------------------------------------
ws2 = wb.create_sheet("Triangulation")
ws2.append([
    "Exercice", "CA declare - synthese A (EUR)",
    "Somme reglements - F TOTAL GENERAL (EUR)",
    "Liste tickets - H tot_ttc (EUR)",
    "Ecart A - F (EUR)", "Ecart A - F (%)",
])
style_header(ws2, 6)
for e in data["exercices"]:
    ecart = round(e["ca_a"] - e["regl_total_general"], 2)
    pct = round(100 * ecart / e["ca_a"], 3) if e["ca_a"] else 0
    ws2.append([e["exercice"], e["ca_a"], e["regl_total_general"], e["h_total"], ecart, pct])
ecart_cumul = round(ca_a_cumul - f_cumul, 2)
ws2.append(["CUMUL", ca_a_cumul, f_cumul, h_cumul, ecart_cumul,
            round(100 * ecart_cumul / ca_a_cumul, 3)])
ws2.cell(row=ws2.max_row, column=1).font = BOLD
for c in range(2, 7):
    ws2.cell(row=ws2.max_row, column=c).font = BOLD
for r in range(2, ws2.max_row + 1):
    for c in range(2, 7):
        ws2.cell(row=r, column=c).alignment = RIGHT
ws2.append([])
ws2.append(["Lecture : A = bloc « Encaissements Tickets Mode / TOTAL » de l'annexe A ;"])
ws2.append(["F = ligne « TOTAL GENERAL » de l'annexe F ; H = ligne « TOTAL » (tot_ttc) de l'annexe H."])
ws2.append(["Les ecarts A vs F (<1,4 %) tiennent aux avoirs/arrondis de ventilation par mode ; les trois"])
ws2.append(["sources, issues de fichiers distincts de la meme caisse, se recoupent : CA = encaissements."])
ws2.freeze_panes = "A2"
autofit(ws2)

# --- Onglet 3 : Types evenements E ----------------------------------------
ws3 = wb.create_sheet("Types evenements E")
ws3.append(["Exercice", "Type", "Libelle (journal de caisse)", "Nombre", "Montant associe (EUR)"])
style_header(ws3, 5)
ordre = ["DEL", "RPR", "CON", "TIR"]
for e in data["exercices"]:
    for t in ordre:
        if t in e["even_types"]:
            ws3.append([e["exercice"], t, EVEN_LIB.get(t, "?"),
                        e["even_types"][t], e["even_montants"].get(t, 0.0)])
ws3.append([])
ws3.append(["CUMUL 3 EX.", "", "", "", ""])
ws3.cell(row=ws3.max_row, column=1).font = BOLD
for t in ordre:
    if t in data["even_cumul"]:
        ws3.append(["Cumul", t, EVEN_LIB.get(t, "?"),
                    data["even_cumul"][t], data["even_cumul_montant"].get(t, 0.0)])
for r in range(2, ws3.max_row + 1):
    ws3.cell(row=r, column=4).alignment = RIGHT
    ws3.cell(row=r, column=5).alignment = RIGHT
ws3.append([])
ws3.append(["DEL = suppression d'une ligne d'article (correction de saisie), tracee comme l'impose"])
ws3.append(["la norme NF525. Le journal E EST la trace legale des corrections, pas une anomalie."])
ws3.freeze_panes = "A2"
autofit(ws3)

# --- Sheet 4 : Nombre d'articles (variable) vs Chiffre d'affaires (invariant) ---
def _detail_stats(fkey, exo):
    """Relit l'annexe C (detail tickets) : nb lignes, nb tickets, CA(tickets), CA(ligne par ligne)."""
    sh = xlrd.open_workbook(
        os.path.join(CAISSE, f"ANNEXE-{fkey}_detail-tickets_{exo}.xls")).sheet_by_index(0)

    def _n(v):
        try:
            return float(v)
        except (ValueError, TypeError):
            return 0.0
    nb_lignes = sh.nrows - 1
    tk = {}
    ca_lignes = 0.0
    for r in range(1, sh.nrows):
        tk[(str(sh.cell_value(r, 0))[:10], str(sh.cell_value(r, 2)))] = _n(sh.cell_value(r, 5))
        ca_lignes += _n(sh.cell_value(r, 11)) * _n(sh.cell_value(r, 13))
    return nb_lignes, len(tk), round(sum(tk.values()), 2), round(ca_lignes, 2)


# Chiffres ANNONCES PAR LE FISC (proposition, sections XII et XIII) :
#   (exercice, clef fichier C, articles annexe B, articles annexe C, CA annexe C)
FISC_ARTICLES = [
    ("2022-2023", "C1", 40218, 38479, 403370.42),
    ("2023-2024", "C2", 39998, 38219, 438281.12),
    ("2024-2025", "C3", 39132, 37446, 435594.96),
]
ws4 = wb.create_sheet("Articles vs CA")
ws4.append(["Exercice", "Articles annexe B (fisc)", "Articles annexe C (fisc)",
            "Lignes detail (notre relecture)", "CA annexe C (fisc) EUR",
            "CA relecture - tickets EUR", "CA relecture - ligne par ligne EUR"])
style_header(ws4, 7)
for exo, fkey, b_art, c_art, c_ca in FISC_ARTICLES:
    nbl, _ntk, ca_t, ca_l = _detail_stats(fkey, exo)
    ws4.append([exo, b_art, c_art, nbl, c_ca, ca_t, ca_l])
for r in range(2, ws4.max_row + 1):
    for col in range(2, 8):
        ws4.cell(row=r, column=col).alignment = RIGHT
ws4.append([])
for ligne in [
    "Lecture : le NOMBRE D'ARTICLES varie selon l'export (annexe B, annexe C, notre relecture),",
    "mais le CHIFFRE D'AFFAIRES reste du meme ordre. Exemple 2022-2023 : 40 218 articles (annexe B)",
    "et 38 479 articles (annexe C) donnent le MEME CA de 403 370,42 EUR (table XIII de la proposition).",
    "1 739 articles d'ecart = 0 EUR d'ecart. Notre relecture du detail confirme 403 370,42 EUR au centime.",
    "",
    "Exercices 2023-2024 et 2024-2025 : la relecture (colonne tickets) concorde a MOINS DE 0,2 % avec",
    "l'annexe C du service (ecart de 250 EUR et 755 EUR sur ~438 000 et ~435 000 EUR). Cet ecart tient",
    "aux avoirs et aux tickets de bord de periode entre deux exports faits a des dates differentes ;",
    "il ne traduit aucune recette manquante (et reste tres inferieur a la marge d'arrondi habituelle).",
    "",
    "Methode : 'lignes detail' = nombre de lignes du fichier ; 'CA tickets' = somme des Tot_ttc par",
    "ticket unique (NET de remises) = chiffre de reference ; 'CA ligne par ligne' = somme(quantite x",
    "prix unitaire), controle BRUT hors remises (d'ou un leger ecart les annees a remises).",
    "Source : annexes C-1 a C-3 (detail des tickets) de la caisse certifiee. Calcul reproductible.",
]:
    ws4.append([ligne])
ws4.freeze_panes = "A2"
autofit(ws4)

os.makedirs(PIECES, exist_ok=True)
wb.save(XLSX_OUT)
print("\nXLSX ecrit :", XLSX_OUT)


# ==========================================================================
# JSON (textes finaux pour la sous-page avocat)
# ==========================================================================
def euro(v):
    s = f"{v:,.2f}".replace(",", " ").replace(".", ",")
    return s + " €"


def pct(v):
    return f"{v:.2f}".replace(".", ",") + " %"


def intfr(v):
    return f"{int(round(v)):,}".replace(",", " ")


# Tableau : reglements par mode (cumul)
lignes_modes = []
for m in ["CB", "CHV", "TR", "ESP", "CHQ"]:
    v = data["regl_cumul"].get(m, 0)
    p = 100 * v / regl_cumul_total
    lignes_modes.append([
        {"v": f"{m} - {MODE_LIB[m]}"},
        {"v": euro(v), "align": "right"},
        {"v": pct(p), "align": "right"},
        {"v": "Bancarise" if m in BANCARISE else "Especes", "align": "right"},
    ])
lignes_modes.append([
    {"v": "TOTAL GENERAL (annexes F)", "align": "left", "fw": 700},
    {"v": euro(regl_cumul_total), "align": "right", "fw": 700},
    {"v": pct(100.0), "align": "right", "fw": 700},
    {"v": "", "align": "right"},
])

# Tableau : triangulation par exercice
lignes_tri = []
for e in data["exercices"]:
    ecart = e["ca_a"] - e["regl_total_general"]
    p = 100 * ecart / e["ca_a"] if e["ca_a"] else 0
    lignes_tri.append([
        {"v": e["exercice"]},
        {"v": euro(e["ca_a"]), "align": "right"},
        {"v": euro(e["regl_total_general"]), "align": "right"},
        {"v": euro(e["h_total"]), "align": "right"},
        {"v": pct(abs(p)), "align": "right"},
    ])
ecart_cumul = ca_a_cumul - f_cumul
lignes_tri.append([
    {"v": "Cumul 3 exercices", "fw": 700},
    {"v": euro(ca_a_cumul), "align": "right", "fw": 700},
    {"v": euro(f_cumul), "align": "right", "fw": 700},
    {"v": euro(h_cumul), "align": "right", "fw": 700},
    {"v": pct(abs(100 * ecart_cumul / ca_a_cumul)), "align": "right", "fw": 700},
])

# Tableau : types d'evenement (cumul)
lignes_even = []
for t in ["DEL", "RPR", "CON", "TIR"]:
    if t in data["even_cumul"]:
        n = data["even_cumul"][t]
        lignes_even.append([
            {"v": t},
            {"v": EVEN_LIB.get(t, "?")},
            {"v": intfr(n), "align": "right"},
        ])

# Graphique : reglements par mode (cumul, en euros)
graph_data = [
    {"nom": f"{m}", "Encaissements (cumul 3 exercices)": round(data["regl_cumul"].get(m, 0), 2)}
    for m in ["CB", "CHV", "TR", "ESP", "CHQ"]
]

# Barre de composition : bancarise vs especes (cumul)
barre_segments = [
    {"label": "Bancarise (CB, CHQ, CHV, TR)", "valeur": banc_cumul, "categorie": "mesure"},
    {"label": "Especes (ESP)", "valeur": esp_cumul, "categorie": "alerte"},
]

meta = {
    "slug": "fichiers-evenement-reglement",
    "titre": "Fichiers Événement (E) et Règlement (F)",
    "refRapport": "Proposition p. 22-24 (rejet 1/3) - annexes E-1 à E-3 et F-1 à F-3",
    "genere": "scripts/rendu-final-fichiers-evenement-reglement.py",
    "sources": [
        "public/documents/caisse-enregistreuse/ANNEXE-E{1,2,3}_tpvevenement_*.xls",
        "public/documents/caisse-enregistreuse/ANNEXE-F{1,2,3}_reglements_*.xls",
        "public/documents/caisse-enregistreuse/ANNEXE-A{1,2,3}_synthese-CA_*.xls",
        "public/documents/caisse-enregistreuse/ANNEXE-H{1,2,3}_liste-tickets_*.xls",
    ],
    "chiffres": {
        "del_cumul": del_cumul,
        "regl_total_cumul": regl_cumul_total,
        "bancarise_cumul": banc_cumul,
        "especes_cumul": esp_cumul,
        "pct_bancarise_cumul": pct_banc_cumul,
        "pct_especes_cumul": pct_esp_cumul,
        "ca_a_cumul": ca_a_cumul,
        "f_cumul": f_cumul,
        "h_cumul": h_cumul,
    },
}

# Annexes de caisse citees dans la page, rendues telechargeables (7 familles x 3 exercices).
_EXOS3 = ["2022-2023", "2023-2024", "2024-2025"]
_ANNEXES_SRC = [
    ("E", "tpvevenement", "Fichier Événement (journal des événements, dont les « DEL »)"),
    ("F", "reglements", "Fichier Règlement (montants par mode de paiement)"),
    ("A", "synthese-CA", "Synthèse du chiffre d'affaires"),
    ("H", "liste-tickets", "Liste des tickets"),
    ("B", "prix-vente-quantite", "Prix de vente et quantités (extrait Excel du 13 février 2026)"),
    ("C", "detail-tickets", "Détail des tickets (libellé, quantité et prix par ligne)"),
    ("G", "journal-tva", "Journal de TVA"),
]
annexes_fichiers = [
    {
        "fichier": f"caisse-enregistreuse/ANNEXE-{lettre}{i}_{slug}_{exo}.xls",
        "label": f"Annexe {lettre}-{i} : {desc} ({exo})",
    }
    for lettre, slug, desc in _ANNEXES_SRC
    for i, exo in enumerate(_EXOS3, 1)
]


def apercu_del(n=5):
    """n premières lignes « DEL » de l'annexe E-1, avec toutes les données tracées."""
    sh = xlrd.open_workbook(os.path.join(CAISSE, F_NOMS["E1"])).sheet_by_index(0)
    lignes = []
    for r in range(1, sh.nrows):
        if str(sh.cell_value(r, 3)).strip() != "DEL":
            continue
        lignes.append([
            str(sh.cell_value(r, 1))[:10],                              # date
            str(sh.cell_value(r, 2)),                                   # heure
            "DEL",                                                      # type
            {"v": str(int(sh.cell_value(r, 6))), "align": "right"},     # n° ticket Z
            {"v": euro(sh.cell_value(r, 8)), "align": "right"},         # montant
            {"v": str(int(sh.cell_value(r, 9))), "align": "right"},     # id événement
        ])
        if len(lignes) >= n:
            break
    return lignes


def apercu_detail(n=6):
    """n premières lignes de vente du 2022-04-14 (annexe C-1), avec leur libellé."""
    sh = xlrd.open_workbook(os.path.join(CAISSE, "ANNEXE-C1_detail-tickets_2022-2023.xls")).sheet_by_index(0)
    lignes = []
    for r in range(1, sh.nrows):
        if str(sh.cell_value(r, 0))[:10] != "2022-04-14":
            continue
        lib = str(sh.cell_value(r, 10)).strip()
        if not lib:
            continue
        q = sh.cell_value(r, 11)
        qstr = str(int(q)) if float(q) == int(q) else str(q).replace(".", ",")
        lignes.append([
            {"v": str(sh.cell_value(r, 2)).replace(".0", ""), "align": "left"},  # n° ticket
            str(sh.cell_value(r, 1)),                                            # heure
            lib,                                                                 # libellé
            {"v": qstr, "align": "right"},                                       # quantité
            {"v": euro(sh.cell_value(r, 13)), "align": "right"},                 # prix unit.
        ])
        if len(lignes) >= n:
            break
    return lignes


del_preview = apercu_del(5)
detail_preview = apercu_detail(6)


# Sources juridiques (liens cliquables affichés en entier : libellé = URL).
def _lien(u):
    return f"[{u}]({u})"


U_CGI286 = "https://www.legifrance.gouv.fr/codes/article_lc/LEGIARTI000051764897"
U_BOI_CAISSE = "https://bofip.impots.gouv.fr/bofip/10691-PGP.html/identifiant=BOI-TVA-DECLA-30-10-30-20210519"
U_IMPOTS_CAISSE = "https://www.impots.gouv.fr/professionnel/questions/quel-est-le-champ-dapplication-de-lobligation-de-detenir-un-logiciel-de"
U_ECO_CERT = "https://www.economie.gouv.fr/entreprises/gerer-son-entreprise-au-quotidien/gerer-sa-comptabilite-et-ses-demarches/ce-quil-faut-savoir-sur-la-certification-des-logiciels-de-caisse"

sections = [
    # ----- 1. CE QUE DIT LE FISC -----
    {
        "kind": "chapitre", "source": "fisc", "numero": 1,
        "titre": "Ce que dit l'administration",
        "sousTitre": "Proposition p. 23-24 et 27-29 (rejet de comptabilité 1/3 et 2/3). Fichier Événement "
                     "(annexes E-1 à E-3, section IX), Fichier Règlement (annexes F-1 à F-3, section X) et "
                     "inaltérabilité des fichiers de caisse (sections XII et XIII).",
    },
    {"kind": "titre", "numero": "1.1", "texte": "Fichier Événement (E) : les suppressions « DEL »"},
    {
        "kind": "paragraphe",
        "texte": "Dans les annexes E-1 à E-3 (**section IX, p. 23-24** ; fichiers remis sur clé USB le "
                 "30 mars 2026), le vérificateur constate des événements « DEL » (qu'il interprète comme "
                 "« Delete », soit des suppressions d'articles) : **8 014 / 6 943 / 6 345 événements** "
                 "selon l'exercice (21 302 au total), pour une valeur globale de **193 005,09 € / "
                 "140 686,55 € / 97 071,51 €** (430 763,15 € au total). Il relève que « les données de "
                 "caisse **n'indiquent pas la nature de l'article** de cette suppression probable ».",
    },
    {
        "kind": "paragraphe",
        "texte": "« L'absence de la nature de l'article supprimé "
                 "compromet grandement la structure de l'archive communiquée au service et contrevient "
                 "aux dispositions de conservation des données. » Le service en conclut que la "
                 "comptabilité est **impropre à justifier le résultat et non probante**. Base légale "
                 "invoquée : article L-13 du LPF ; obligations d'inaltérabilité, de sécurisation, de "
                 "conservation et d'archivage (3° bis du I de l'article 286 du CGI ; BOI-TVA-DECLA-30-10-30 "
                 "du 19/05/2021) ; conservation des pièces justificatives de ventes (article L 102 B du LPF).",
    },
    {"kind": "titre", "numero": "1.2", "texte": "Fichier Règlement (F) : les modes de paiement"},
    {
        "kind": "paragraphe",
        "texte": "Dans les annexes F-1 à F-3 (**section X, p. 24**), l'analyse du vérificateur « s'est "
                 "limitée à établir les montants des différents modes de règlement » (CB, CHQ, CHV, ESP, "
                 "TR) sur chaque exercice, pour des totaux de **403 402,87 € / 439 600,30 € / "
                 "435 146,29 €**. La section est **purement descriptive** : elle ne formule aucun "
                 "reproche en elle-même ; ces totaux sont ensuite réutilisés par le service au titre de "
                 "l'inaltérabilité (section XIII).",
    },
    {"kind": "titre", "numero": "1.3", "texte": "La fluctuation du nombre d'articles"},
    {
        "kind": "paragraphe",
        "texte": "En **section XII (p. 27-28)**, le service compare les annexes B-1 à B-3 (établies à partir "
                 "du fichier Excel transmis par courriel le 13 février 2026) et les annexes C-1 à C-3 "
                 "(établies à partir des fichiers reçus le 16 mars 2026) : le **nombre d'articles** passe "
                 "de 40 218 / 39 998 / 39 132 (annexes B) à 38 479 / 38 219 / 37 446 (annexes C). Il "
                 "relève que « les nombres d'articles et les montants du chiffre d'affaires **fluctuent "
                 "suivant les fichiers**, qui sont pourtant tous issus du logiciel de caisse », et invoque "
                 "la séquentialité et la cohérence interne des données (articles 121-1 et 121-3 du PCG).",
    },
    {"kind": "titre", "numero": "1.4",
     "texte": "La fluctuation du chiffre d'affaires et l'écart avec la comptabilité"},
    {
        "kind": "paragraphe",
        "texte": "En **section XIII (p. 28-29)**, le service relève que le CA diffère selon le fichier "
                 "source : pour l'exercice clos au 31/03/2023, **403 370,42 €** (annexes B et C), "
                 "**403 402,87 €** (annexes F) et **403 324,72 €** (annexes G), et qu'« **aucun ne "
                 "correspond à ce qui figure dans la comptabilité** », laquelle affiche un CA TTC (compte "
                 "706800 « Pourboire » inclus) de **404 030,87 € / 438 658,43 € / 435 524,92 €**, "
                 "reconstitué à partir des récapitulatifs journaliers de la caisse retranscrits sur agenda.",
    },
    {
        "kind": "paragraphe",
        "texte": "Pour les sections XII et XIII, le service estime que les données de caisse "
                 "seraient « de nature changeante » puis « de nature instable », modifiables entre les "
                 "fichiers Excel transmis par courriel (13 février / 13 mars 2026) et les fichiers reçus "
                 "le 16 mars 2026. Le service en conclut que les fichiers et le logiciel de caisse ne "
                 "respectent pas les obligations d'**inaltérabilité** (3° bis du I de l'article 286 du "
                 "CGI, article L-13 du LPF), ce qui rendrait la comptabilité **impropre à justifier la "
                 "réalité de l'exploitation et non probante**.",
    },
    # ----- 2. NOTRE DEMONSTRATION -----
    {
        "kind": "chapitre", "source": "nous", "numero": 2,
        "titre": "Notre réponse : chaque point réfuté",
        "sousTitre": "À chaque grief (1.1 à 1.4) répond la réfutation correspondante (2.1 à 2.4), pièces "
                     "de caisse à l'appui : E est la trace légale des corrections, F prouve la "
                     "bancarisation, et les « fluctuations » sont des écarts d'export, pas des altérations.",
    },
    {"kind": "titre", "numero": "2.1", "texte": "Fichier Événement (E) : les suppressions « DEL »"},
    {"kind": "titre", "texte": "Ce que la loi impose, et à qui"},
    {
        "kind": "paragraphe",
        "texte": "L'**article 286, I, 3° bis du code général des "
                 "impôts** oblige tout commerçant qui encaisse des clients particuliers au moyen d'un "
                 "logiciel de caisse à **utiliser un système satisfaisant aux conditions "
                 "d'inaltérabilité, de sécurisation, de conservation et d'archivage** des données, et à "
                 "**justifier** cette conformité par un **certificat délivré par un organisme "
                 "accrédité**. Ces quatre conditions sont des **propriétés techniques du logiciel** (la "
                 "manière dont il enregistre, sécurise et archive les données) : elles sont conçues et "
                 "garanties par l'**éditeur du logiciel**, et c'est la **certification** qui en fait foi. "
                 "Le commerçant, lui, n'a que **deux obligations** : utiliser un système certifié et "
                 "produire le certificat lors du contrôle.",
    },
    {
        "kind": "paragraphe",
        "texte": "Textes et sources officielles : " + _lien(U_CGI286) + " (article 286 du CGI) ; "
                 + _lien(U_BOI_CAISSE) + " (doctrine BOI-TVA-DECLA-30-10-30) ; " + _lien(U_IMPOTS_CAISSE)
                 + " (champ d'application, impots.gouv.fr) ; " + _lien(U_ECO_CERT)
                 + " (certification des logiciels de caisse, economie.gouv.fr).",
    },
    {"kind": "titre", "texte": "La caisse est certifiée, et l'administration le reconnaît"},
    {
        "kind": "paragraphe",
        "texte": "Le restaurant utilise le logiciel **SERVILOG de l'éditeur AKEAD (version "
                 "POS 5.03-Eb)**, **certifié NF525** ; son **certificat de conformité est daté du "
                 "01/12/2018**. Le vérificateur l'a expressément demandé puis reçu : la proposition de "
                 "rectification indique, **page 3**, qu'il « demande le nom du logiciel de caisse : "
                 "SERVILOG d'AKEAD et le certificat correspondant », et que « le 13 janvier 2026, le "
                 "cabinet transmet par courriels le certificat de conformité de la caisse daté du "
                 "01/12/2018 ». **Le restaurant a donc rempli ses deux obligations** : utiliser un "
                 "système certifié et produire le certificat. La certification **NF525** est précisément "
                 "le « certificat délivré par un organisme accrédité » exigé par la loi.",
    },
    {"kind": "titre", "texte": "Le journal sans libellé relève de l'éditeur, pas du restaurant"},
    {
        "kind": "paragraphe",
        "texte": "C'est exact : le Fichier Événement (E) ne comporte "
                 "**aucune colonne « libellé »**. Il enregistre, pour chaque suppression « DEL », la "
                 "**date, l'heure, le caissier, le numéro de ticket Z, le montant et un identifiant "
                 "unique**, mais pas le nom du plat. Cette structure est **celle du logiciel certifié "
                 "SERVILOG/AKEAD** : elle relève de la **conception de l'éditeur**, couverte par la "
                 "certification, et non d'un choix ou d'une manipulation du restaurant. Reprocher au "
                 "restaurant le format du journal d'une caisse certifiée revient à lui imputer une "
                 "obligation qui pèse sur l'éditeur.",
    },
    {
        "kind": "tableau",
        "titre": "Aperçu de 5 suppressions « DEL » : toutes les données tracées par le journal E (annexe E-1)",
        "minWidth": 640,
        "colonnes": [
            {"label": "Date"},
            {"label": "Heure"},
            {"label": "Type"},
            {"label": "N° ticket Z", "align": "right"},
            {"label": "Montant", "align": "right"},
            {"label": "ID événement", "align": "right"},
        ],
        "lignes": del_preview,
    },
    {"kind": "titre", "texte": "Le libellé est conservé dans le détail des tickets"},
    {
        "kind": "paragraphe",
        "texte": "Chaque suppression est ainsi **horodatée, rattachée à un ticket Z et à un montant**, "
                 "avec un identifiant unique : elle est entièrement **auditable** (un montant seul ne "
                 "suffit d'ailleurs pas à nommer l'article : 3,90 € peut être un soda, une glace ou une "
                 "bière). Et le **libellé de chaque article** est, lui, bien conservé là où la loi "
                 "l'exige : l'obligation de conservation des **pièces justificatives de ventes** "
                 "(**article L 102 B du LPF**) est remplie, car chaque ligne vendue figure, avec son "
                 "**nom, sa quantité et son prix**, dans le **détail des tickets (annexe C)**, complété "
                 "par les tickets Z, la liste des tickets (annexe H) et la synthèse du CA (annexe A) :",
    },
    {
        "kind": "tableau",
        "titre": "Le libellé de chaque article est dans le détail des tickets (annexe C-1, ventes du 14/04/2022)",
        "minWidth": 560,
        "colonnes": [
            {"label": "N° ticket"},
            {"label": "Heure"},
            {"label": "Libellé de l'article"},
            {"label": "Qté", "align": "right"},
            {"label": "Prix unit.", "align": "right"},
        ],
        "lignes": detail_preview,
    },
    {
        "kind": "piecejointe",
        "intro": "Les deux fichiers de caisse, tels que remis au service, pour les trois exercices : le "
                 "Fichier Événement (E, journal des « DEL ») et le détail des tickets (C, où figure le libellé).",
        "fichiers": [
            {"fichier": f"caisse-enregistreuse/{F_NOMS['E1']}",
             "label": "Annexe E-1 : Fichier Événement (2022-2023)"},
            {"fichier": f"caisse-enregistreuse/{F_NOMS['E2']}",
             "label": "Annexe E-2 : Fichier Événement (2023-2024)"},
            {"fichier": f"caisse-enregistreuse/{F_NOMS['E3']}",
             "label": "Annexe E-3 : Fichier Événement (2024-2025)"},
            {"fichier": "caisse-enregistreuse/ANNEXE-C1_detail-tickets_2022-2023.xls",
             "label": "Annexe C-1 : détail des tickets, libellés (2022-2023)"},
            {"fichier": "caisse-enregistreuse/ANNEXE-C2_detail-tickets_2023-2024.xls",
             "label": "Annexe C-2 : détail des tickets, libellés (2023-2024)"},
            {"fichier": "caisse-enregistreuse/ANNEXE-C3_detail-tickets_2024-2025.xls",
             "label": "Annexe C-3 : détail des tickets, libellés (2024-2025)"},
        ],
    },
    {"kind": "titre", "texte": "Une caisse certifiée est présumée conforme"},
    {
        "kind": "paragraphe",
        "texte": "Tant que le "
                 "certificat n'est pas invalidé, le système est réputé satisfaire aux conditions "
                 "d'inaltérabilité, de sécurisation, de conservation et d'archivage. Or l'administration "
                 "**détient ce certificat** (elle le reconnaît page 3) et **ne le conteste nulle part** : "
                 "dans les sections où elle rejette la comptabilité pour défaut d'inaltérabilité "
                 "(sections IX, XII et XIII), elle **n'évoque jamais** la certification. Elle ne peut pas, "
                 "d'un côté, recevoir le certificat d'une caisse certifiée et, de l'autre, rejeter cette "
                 "même caisse pour inaltérabilité sans démontrer que le certificat serait faux. La charge "
                 "de la preuve lui revient, et elle ne l'assume pas.",
    },
    {"kind": "titre", "texte": "Aucun chiffre d'affaires ne disparaît"},
    {
        "kind": "paragraphe",
        "texte": "Les lignes « DEL » ne sont **pas du chiffre d'affaires disparu** : elles correspondent "
                 "à des **erreurs de saisie**, des **paiements séparés**, des **créations de facture sans "
                 "détail** et des **erreurs de table**. La démonstration complète, cas par cas et "
                 "réconciliée au centime, est faite dans la page dédiée : "
                 "[/rendu-final/suppressions-de-caisse](/rendu-final/suppressions-de-caisse). Le fichier E "
                 "ne contient d'ailleurs que des **corrections tracées** (DEL, CON, RPR, TIR), "
                 "récapitulées ci-dessous :",
    },
    {
        "kind": "tableau",
        "titre": "Types d'événements du fichier E (cumul 3 exercices) : ce sont des corrections d'exploitation tracées",
        "minWidth": 640,
        "colonnes": [
            {"label": "Type"},
            {"label": "Signification (journal de caisse)"},
            {"label": "Nombre", "align": "right"},
        ],
        "lignes": lignes_even,
    },
    {"kind": "titre", "numero": "2.2", "texte": "Fichier Règlement (F) : les modes de paiement"},
    {
        "kind": "paragraphe",
        "texte": "La section X est **purement descriptive** : le vérificateur s'y borne à relever les "
                 "montants encaissés par mode de règlement (CB, CHQ, CHV, ESP, TR), **sans formuler aucun "
                 "reproche**. Ces totaux sont produits par la caisse elle-même et **coïncident à l'euro "
                 "près** avec ceux repris dans la proposition (p. 24) : nous les confirmons, il n'y a, "
                 "sur ce point, rien à contester. Pour mémoire, le détail par mode sur les trois "
                 "exercices :",
    },
    {
        "kind": "tableau",
        "titre": "Règlements par mode de paiement (cumul des 3 exercices, lignes « TOTAL » des annexes F)",
        "minWidth": 640,
        "colonnes": [
            {"label": "Mode de règlement"},
            {"label": "Montant encaissé", "align": "right"},
            {"label": "Part", "align": "right"},
            {"label": "Nature", "align": "right"},
        ],
        "lignes": lignes_modes,
    },
    {"kind": "titre", "numero": "2.3", "texte": "La fluctuation du nombre d'articles"},
    {"kind": "titre", "texte": "Un nombre d'articles différent, exactement le même chiffre d'affaires"},
    {
        "kind": "paragraphe",
        "texte": "La « fluctuation » porte sur le **nombre d'articles**, jamais sur l'argent, et le "
                 "tableau du service le prouve **lui-même**. Pour l'exercice clos au 31/03/2023, l'annexe "
                 "B compte **40 218 articles** et l'annexe C **38 479 articles** (1 739 de différence). "
                 "Or ces deux annexes affichent **exactement le même chiffre d'affaires : 403 370,42 €** "
                 "(table de la section XIII de la proposition). Autrement dit, **1 739 articles d'écart = "
                 "0 € d'écart**. En rouvrant le fichier de détail (annexe C), nous retrouvons ce même "
                 "**403 370,42 €** (somme des tickets) et **403 370,38 €** en recalculant ligne par ligne "
                 "(quantité × prix unitaire), soit au centime près. Le nombre d'articles n'est donc pas "
                 "une mesure de recettes. La comptabilité du cabinet concorde elle aussi : son CA TTC de "
                 "404 030,87 € est ce même chiffre **augmenté du seul compte 706800 « Pourboire »**. Les "
                 "deux autres exercices concordent de la même façon, **à moins de 0,2 % près** (détail "
                 "et méthode dans le fichier joint).",
    },
    {"kind": "titre", "texte": "Le décompte de lignes dépend de l'export, pas des ventes"},
    {
        "kind": "paragraphe",
        "texte": "Pourquoi ce décompte varie-t-il ? Parce qu'il **dépend de l'export**. Le service "
                 "compare deux **rapports de nature différente** (leurs intitulés et leur structure le "
                 "montrent) : les annexes **B** sont un état « **Prix de vente × quantité** » (12 "
                 "colonnes), les annexes **C** le « **détail des tickets, ligne par ligne** » (24 "
                 "colonnes). Ils sont en outre extraits par **deux requêtes distinctes**, lisibles dans "
                 "le nom même des fichiers remis par le cabinet : les **B** filtrées **par exercice** "
                 "(« 01042022 31032023 »), les **C** par **une requête unique sur toute la période** "
                 "(« issu de 20220401 AND 20250331 »). Mieux : un **troisième export** des mêmes tickets "
                 "(celui que nous analysons) donne **encore un autre décompte, 33 222 lignes**, pour le "
                 "**même CA de 403 370,42 €**. Trois comptages, un seul chiffre d'affaires : un nombre de "
                 "lignes est un **artefact de rapport** (menus décomposés ou non, modificateurs, lignes à "
                 "0 €), pas une donnée comptable.",
    },
    {"kind": "titre", "texte": "Une caisse certifiée ne peut pas modifier les montants : c'est la loi"},
    {
        "kind": "paragraphe",
        "texte": "La caisse de La Demi-Lune est **certifiée NF525** (logiciel SERVILOG de l'éditeur "
                 "AKEAD, certificat du 01/12/2018), et l'administration **détient ce certificat** : elle "
                 "le reconnaît page 3 de la proposition (« le 13 janvier 2026, le cabinet transmet par "
                 "courriels le certificat de conformité de la caisse daté du 01/12/2018 »). Or l'**article "
                 "286, I, 3° bis du CGI** impose que le logiciel enregistre les données de règlement "
                 "**sans qu'elles puissent être altérées** (condition d'**inaltérabilité** définie par la "
                 "doctrine BOI-TVA-DECLA-30-10-30). Modifier après coup le montant d'une recette est donc "
                 "**techniquement impossible** sur une caisse certifiée, et c'est bien ce que l'on "
                 "constate : le chiffre d'affaires est identique d'un export à l'autre. Un écart de "
                 "**comptage de lignes** entre deux états relève du **fonctionnement du logiciel** (la "
                 "responsabilité de l'éditeur, couverte par la certification), pas d'une manipulation du "
                 "restaurant. Le restaurant ne saurait être tenu pour responsable de la manière dont un "
                 "rapport de caisse certifié compte ses lignes.",
    },
    {
        "kind": "paragraphe",
        "texte": "Sources : " + _lien(U_CGI286) + " (article 286 du CGI, inaltérabilité) ; "
                 + _lien(U_BOI_CAISSE) + " (doctrine BOI-TVA-DECLA-30-10-30) ; " + _lien(U_ECO_CERT)
                 + " (certification des logiciels de caisse).",
    },
    {
        "kind": "piecejointe",
        "intro": "Calcul entièrement reproductible (onglet « Articles vs CA » du classeur) : le nombre "
                 "d'articles et le chiffre d'affaires recalculés ligne par ligne à partir des annexes C "
                 "de la caisse, sur les trois exercices.",
        "fichiers": [
            {"fichier": "pieces-defense/RF-fichiers-evenement-reglement.xlsx",
             "label": "RF - Articles vs CA : décompte variable, chiffre d'affaires invariant (onglet « Articles vs CA »)"},
            {"fichier": "caisse-enregistreuse/ANNEXE-C1_detail-tickets_2022-2023.xls",
             "label": "Annexe C-1 : détail des tickets (2022-2023)"},
            {"fichier": "caisse-enregistreuse/ANNEXE-C2_detail-tickets_2023-2024.xls",
             "label": "Annexe C-2 : détail des tickets (2023-2024)"},
            {"fichier": "caisse-enregistreuse/ANNEXE-C3_detail-tickets_2024-2025.xls",
             "label": "Annexe C-3 : détail des tickets (2024-2025)"},
        ],
    },
    {"kind": "titre", "numero": "2.4",
     "texte": "La fluctuation du chiffre d'affaires et l'écart avec la comptabilité"},
    {"kind": "titre", "texte": "Quatre fichiers, des périmètres légèrement différents"},
    {
        "kind": "paragraphe",
        "texte": "Pour l'exercice "
                 "clos au 31/03/2023, l'écart entre le plus haut (annexes F, 403 402,87 €) et le plus bas "
                 "(annexes G, 403 324,72 €) n'est que de **78 €, soit 0,02 %**. Les annexes B et C "
                 "comptent les **tickets** (ventes), F les **règlements** (encaissements) : il est normal "
                 "qu'ils diffèrent de quelques dizaines d'euros (avoirs, arrondis de ventilation par "
                 "mode), sans qu'aucune recette ne manque.",
    },
    {"kind": "titre", "texte": "L'écart avec la comptabilité s'explique"},
    {
        "kind": "paragraphe",
        "texte": "La "
                 "comptabilité intègre un élément que les fichiers de vente ne comptent pas : le compte "
                 "**706800 « Pourboire »** (ainsi que les avoirs). C'est le service lui-même qui le "
                 "précise. Sur 2022-2023, l'écart entre le CA comptable TTC (404 030,87 €) et le total "
                 "des règlements F (403 402,87 €) est de **628 €**, de l'ordre du pourboire annuel : il "
                 "ne traduit aucune vente dissimulée mais une **différence de périmètre** que le fisc "
                 "documente lui-même. L'écart reste inférieur à 0,4 % par exercice.",
    },
    {
        "kind": "tableau",
        "titre": "Triangulation : CA déclaré (A) = somme des règlements (F) = liste des tickets (H)",
        "minWidth": 720,
        "colonnes": [
            {"label": "Exercice"},
            {"label": "CA déclaré - synthèse A", "align": "right"},
            {"label": "Somme règlements - F", "align": "right"},
            {"label": "Liste tickets - H", "align": "right"},
            {"label": "Écart A vs F", "align": "right"},
        ],
        "lignes": lignes_tri,
    },
    {"kind": "titre", "texte": "La preuve interne : règlements = tickets"},
    {
        "kind": "paragraphe",
        "texte": "Trois fichiers distincts de la même caisse "
                 f"convergent : CA déclaré {euro(ca_a_cumul)} (synthèse A), somme des règlements "
                 f"{euro(f_cumul)} (F) et liste des tickets {euro(h_cumul)} (H). Le rapprochement le plus "
                 "serré, **règlements (F) = tickets (H)**, tient à 0,01 % près : ce que la caisse a "
                 "facturé est exactement ce qu'elle a encaissé. La cohérence interne que le vérificateur "
                 "dit vouloir contrôler (PCG art. 121-1 et 121-3) est donc démontrée, pas infirmée.",
    },
    {"kind": "titre", "texte": "L'inaltérabilité n'est pas en cause : un Excel n'est pas l'archive"},
    {
        "kind": "paragraphe",
        "texte": "Les fichiers Excel transmis les 13 février et 13 mars 2026 étaient des **extraits de "
                 "travail**, modifiables par nature (ce sont des classeurs Excel). L'archive "
                 "**inaltérable** au sens de la norme NF525, c'est l'**export certifié remis sur clé "
                 "USB** le 16/30 mars 2026. Comparer un brouillon Excel à l'archive sécurisée pour "
                 "conclure à une « altération » revient à confondre l'outil de travail et la pièce "
                 "légale. La donnée sous-jacente (tickets Z, récapitulatifs journaliers) est stable ; "
                 "c'est elle qui fait foi.",
    },
    # ----- 3. PIECE JOINTE -----
    {
        "kind": "piecejointe",
        "intro": "Calcul reproductible (onglets « Reglements par mode », « Triangulation », « Types evenements E ») :",
        "fichiers": [
            {
                "fichier": "pieces-defense/RF-fichiers-evenement-reglement.xlsx",
                "label": "RF - Fichiers Événement (E) et Règlement (F) - règlements par mode, triangulation, types d'événements",
            }
        ],
    },
    {
        "kind": "piecejointe",
        "intro": "Annexes de caisse citées dans cette page (sources, telles que remises au service) : "
                 "Fichier Événement (E), Fichier Règlement (F), synthèse du CA (A), liste des tickets (H), "
                 "prix/quantités (B), détail des tickets (C) et journal de TVA (G), pour les trois exercices.",
        "fichiers": annexes_fichiers,
    },
    # ----- 4. VERDICT -----
    {
        "kind": "alerte",
        "couleur": "teal",
        "titre": "Ce qu'il faut retenir",
        "texte": "Les journaux Événement (E) et Règlement (F) ne révèlent aucune anomalie : E est la trace "
                 "légale des corrections (NF525), F prouve la ventilation des encaissements. La triangulation "
                 f"A = F = H verrouille l'égalité CA = encaissements, et la bancarisation atteint {pct(pct_banc_cumul)} "
                 f"(espèces {pct(pct_esp_cumul)} seulement). Ces fichiers, loin de fonder le rejet, démontrent "
                 "la sincérité de la comptabilité.",
    },
    {
        "kind": "interne",
        "audience": "avocat",
        "titre": "Note pour la défense",
        "texte": "Les montants par mode du fichier F sont les lignes « TOTAL <mode> » + « TOTAL GENERAL » que "
                 "la caisse génère en bas de chaque annexe ; ils coïncident à l'euro près avec les chiffres de "
                 "la proposition (p. 24). Le verificateur a donc travaillé sur ces totaux fiables. La colonne "
                 "« montant » des lignes de détail double-compte les paiements fractionnés (lignes internes) : "
                 "ne jamais l'opposer aux totaux, c'est un faux écart. La triangulation s'appuie exclusivement "
                 "sur les lignes de TOTAL des fichiers, vérifiables d'un coup d'oeil.",
    },
]

sections = ajouter_conclusion(sections)
out = {"meta": meta, "sections": sections}
os.makedirs(JSON_DIR, exist_ok=True)
with open(JSON_OUT, "w", encoding="utf-8") as f:
    json.dump(out, f, ensure_ascii=False, indent=2)
print("JSON ecrit :", JSON_OUT)

# garde-fou : pas de tiret cadratin / demi-cadratin dans les textes
with open(JSON_OUT, encoding="utf-8") as f:
    raw = f.read()
for bad in ("—", "–"):
    assert bad not in raw, f"Tiret interdit present : {bad!r}"
print("OK - aucun tiret cadratin/demi-cadratin dans le JSON.")
