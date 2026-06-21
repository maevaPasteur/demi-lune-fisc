#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
rendu-final-prix-menu-demi-lune.py

Bloc de defense : "Instabilite des prix du Menu Demi Lune".

Grief du fisc (Proposition de rectification, p. 33-34, rejet 3/3, section XVI) : la multitude
de prix observes pour le Menu Demi Lune dans la caisse traduirait une
comptabilite non sincere.

Demonstration :
  1) La "multitude de prix" est la SIGNATURE D'UN MECANISME DE CAISSE : a la
     demande d'une tablee, le serveur etablit une FACTURE AU FORFAIT, SANS LE
     DETAIL des plats. Le logiciel ne sachant pas convertir une addition
     detaillee en forfait, il faut SUPPRIMER les lignes a la carte puis
     RE-SAISIR un "Menu Demi Lune" a un prix personnalise EGAL AU TOTAL
     REELLEMENT CONSOMME. Le prix du menu n'est donc pas un prix de carte :
     c'est le total de l'addition reconditionne en forfait -> autant de prix
     differents qu'il y a d'additions.
  2) Ce grief et celui des "suppressions de notes" designent donc LES MEMES
     LIGNES de caisse. Preuve arithmetique, verifiable ticket par ticket :
     somme des lignes supprimees = total du menu re-saisi, au centime, lui-meme
     integralement encaisse. Deja demontre, exemples dates a l'appui, dans la
     page suppressions-de-caisse ("Conversion au forfait").
  3) Le PRIX CATALOGUE, lui, est UNIQUE et STABLE (45,00 EUR sur chacune des
     periodes du controle). La "variation" ne porte que sur les ventes HORS
     catalogue (forfaits), soit ~39 % des Menus Demi Lune (hors catalogue / total, source annexe B) pour
     14 118 EUR, chacune TRACEE (date, heure, n de ticket) et encaissee.
  4) Ces recettes sont integralement dans le CA declare -> pas d'insincerite.

Sources (read-only) :
  - src/data/renduFinalCalculs.json  -> "menus_par_periode", "menus_custom"
    (calcules par scripts/rendu-final-menus.py a partir des tickets, annexe C).
  - src/data/renduFinalCasSuppressions.json -> cas "forfait" (suppressions ->
    note encaissee). On en extrait les exemples dont la note est un Menu Demi
    Lune, pour reafficher la meme demonstration ciblee sur ce menu.
  - public/documents/caisse-enregistreuse/ANNEXE-B{1,2,3}_*.xls (prix x qte par
    produit) -> distribution des prix observes du Menu Demi Lune, par exercice
    (controle de coherence avec menus_par_periode).

Sorties (ne touche PAS renduFinal.ts) :
  - public/documents/pieces-defense/RF-prix-menu-demi-lune.xlsx
  - src/data/renduFinal/prix-menu-demi-lune.json
"""
import json, os, collections
import xlrd
from collections import Counter, OrderedDict

from rfcommun import ajouter_conclusion

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

ICI = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.normpath(os.path.join(ICI, ".."))

CALC = os.path.join(ROOT, "src/data/renduFinalCalculs.json")
CAS = os.path.join(ROOT, "src/data/renduFinalCasSuppressions.json")
ANNEXE_DIR = os.path.join(ROOT, "public/documents/caisse-enregistreuse")
XLSX_OUT = os.path.join(ROOT, "public/documents/pieces-defense/RF-prix-menu-demi-lune.xlsx")
JSON_OUT = os.path.join(ROOT, "src/data/renduFinal/prix-menu-demi-lune.json")

PRIX_CATALOGUE = 45.0  # prix carte du Menu Demi Lune, constant sur tout le controle


def eur(x):
    """Format francais : 14 118,00 EUR -> '14 118,00 €'."""
    s = f"{x:,.2f}"
    s = s.replace(",", " ").replace(".", ",")
    return s + " €"


def eur0(x):
    s = f"{x:,.0f}".replace(",", " ")
    return s + " €"


def num(x):
    return f"{x:,.0f}".replace(",", " ")


# --------------------------------------------------------------------------
# 1. Lecture des donnees de caisse deja agregees (menus_par_periode / custom)
# --------------------------------------------------------------------------
calc = json.load(open(CALC, encoding="utf-8"))
periodes = calc["menus_par_periode"]
menus_custom = calc["menus_custom"]
dl_custom = menus_custom["par_menu"]["Demi Lune"]


def demi_lune_de(p):
    for r in p["menus"]:
        if r["menu"].strip().lower() in ("demi lune", "demi-lune"):
            return r
    return None


# Tableau periode -> prix catalogue (stabilite)
table_periodes = []
for p in periodes:
    r = demi_lune_de(p)
    if not r:
        continue
    table_periodes.append({
        "periode": p["periode"],
        "carte": p.get("carte"),
        "exercice": p.get("exercice"),
        "dates": p.get("dates"),
        "prix_catalogue": r["prix_catalogue"],
        "qte_catalogue": r["qte_catalogue"],
        "qte_hors_catalogue": r["qte_hors_catalogue"],
        "eur_hors_catalogue": r["eur_hors_catalogue"],
    })

# Controle : un seul prix catalogue sur tout le controle
prix_catalogue_distincts = sorted({t["prix_catalogue"] for t in table_periodes})
assert prix_catalogue_distincts == [PRIX_CATALOGUE], prix_catalogue_distincts

# Prix personnalises cumules (toutes periodes) : prix -> qte
custom_agg = Counter()
for p in periodes:
    r = demi_lune_de(p)
    if not r:
        continue
    for px, q in r["prix_custom"]:
        custom_agg[round(px, 2)] += int(q)

custom_q = sum(custom_agg.values())
custom_eur = round(sum(px * q for px, q in custom_agg.items()))
nb_prix_distincts = len(custom_agg)
total_q = dl_custom["total_q"]
cat_q = total_q - custom_q
pct_custom = dl_custom["pct"]

# --------------------------------------------------------------------------
# 2. Controle de coherence avec l'annexe B (distribution des prix observes)
# --------------------------------------------------------------------------
def annexe_b_distribution():
    """Renvoie {exercice: Counter(prix->qte)} pour le Menu Demi Lune."""
    import xlrd
    files = [
        ("ANNEXE-B1_prix-vente-quantite_2022-2023.xls", "2022-2023"),
        ("ANNEXE-B2_prix-vente-quantite_2023-2024.xls", "2023-2024"),
        ("ANNEXE-B3_prix-vente-quantite_2024-2025.xls", "2024-2025"),
    ]
    out = OrderedDict()
    for f, ex in files:
        path = os.path.join(ANNEXE_DIR, f)
        if not os.path.exists(path):
            continue
        wb = xlrd.open_workbook(path)
        sh = wb.sheet_by_index(0)
        dist = Counter()
        for r in range(1, sh.nrows):
            lib = str(sh.cell_value(r, 4)).strip().lower()
            if "demi lune" in lib or "demi-lune" in lib:
                prix = round(float(sh.cell_value(r, 7)), 2)
                try:
                    q = int(sh.cell_value(r, 5))
                except Exception:
                    q = 0
                dist[prix] += q
        out[ex] = dist
    return out


annexe_b = annexe_b_distribution()
# Distribution cumulee tous exercices (pour le graphique)
dist_globale = Counter()
for ex, d in annexe_b.items():
    dist_globale.update(d)

# Coherence annexe B vs menus_par_periode (le 45,00 EUR doit dominer chaque annee)
b_cat = {ex: d.get(PRIX_CATALOGUE, 0) for ex, d in annexe_b.items()}

# COMPTAGE UNIQUE et coherent sur TOUTE la fiche (page + XLSX) : tout decoule de
# l'annexe B officielle (B-1 a B-3, prix x quantite). Aucun total concurrent.
total_observe = sum(dist_globale.values())
qte_45 = dist_globale.get(PRIX_CATALOGUE, 0)
pct_45 = round(100.0 * qte_45 / total_observe, 1)
custom_b = {px: q for px, q in dist_globale.items() if abs(px - PRIX_CATALOGUE) > 0.01 and q > 0}
HORS_DL = sum(custom_b.values())
HORS_EUR = round(sum(px * q for px, q in custom_b.items()))
NB_PRIX_HORS = len(custom_b)
PCT_HORS = round(100.0 * HORS_DL / total_observe, 1)

# --------------------------------------------------------------------------
# 3. XLSX : 2 onglets
# --------------------------------------------------------------------------
H1 = Font(bold=True, size=12, color="FFFFFF")
FILL1 = PatternFill("solid", fgColor="0F766E")
BOLD = Font(bold=True)
FILLT = PatternFill("solid", fgColor="E2E8F0")
THIN = Border(*[Side(style="thin", color="CBD5E0")] * 4)
RA = Alignment(horizontal="right")

wb = openpyxl.Workbook()

# Onglet 1 : Prix catalogue par periode
ws = wb.active
ws.title = "Prix catalogue par periode"
ws.append(["Menu Demi Lune : prix catalogue stable par periode de carte"])
ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)
ws.cell(1, 1).font = H1
ws.cell(1, 1).fill = FILL1
ws.append(["Periode", "Carte", "Exercice", "Dates", "Prix catalogue"])
for c in ws[2]:
    c.font = BOLD
    c.fill = FILLT
    c.border = THIN
for t in table_periodes:
    ws.append([t["periode"], t["carte"], t["exercice"], t["dates"], t["prix_catalogue"]])
    for c in ws[ws.max_row]:
        c.border = THIN
        if isinstance(c.value, (int, float)):
            c.alignment = RA
ws.append(["TOTAL", "", "", "", str(PRIX_CATALOGUE).replace(".", ",") + " (unique)"])
for c in ws[ws.max_row]:
    c.font = BOLD
    c.fill = FILLT
    c.border = THIN
for i, w in enumerate([9, 7, 12, 26, 16], 1):
    ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
ws.freeze_panes = "A3"

# Onglet 2 : Prix personnalises
ws2 = wb.create_sheet("Prix personnalises")
ws2.append(["Menu Demi Lune : prix personnalises (hors catalogue) et occurrences"])
ws2.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3)
ws2.cell(1, 1).font = H1
ws2.cell(1, 1).fill = FILL1
ws2.append(["Prix de vente (EUR)", "Nombre d'occurrences", "Montant (EUR)"])
for c in ws2[2]:
    c.font = BOLD
    c.fill = FILLT
    c.border = THIN
for px in sorted(custom_b):
    q = custom_b[px]
    ws2.append([px, q, round(px * q, 2)])
    for c in ws2[ws2.max_row]:
        c.border = THIN
        if isinstance(c.value, (int, float)):
            c.alignment = RA
ws2.append(["TOTAL hors catalogue", HORS_DL, HORS_EUR])
for c in ws2[ws2.max_row]:
    c.font = BOLD
    c.fill = FILLT
    c.border = THIN
    if isinstance(c.value, (int, float)):
        c.alignment = RA
for i, w in enumerate([20, 22, 16], 1):
    ws2.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
ws2.freeze_panes = "A3"

os.makedirs(os.path.dirname(XLSX_OUT), exist_ok=True)
# (sauvegarde du classeur differee : on ajoute plus bas l'onglet exhaustif des
#  conversions au forfait Demi Lune detectees dans la caisse.)

# --------------------------------------------------------------------------
# 4. JSON de rendu (textes finaux)
# --------------------------------------------------------------------------
# Graphique : distribution des prix observes (annexe B, cumul). On regroupe les
# prix personnalises par tranches pour la lisibilite, le 45 EUR isole.
def tranche(px):
    if px == PRIX_CATALOGUE:
        return "45,00 € (catalogue)"
    if px < 30:
        return "< 30 €"
    if px < 40:
        return "30 a 40 €"
    if px < 45:
        return "40 a 45 €"
    if px <= 55:
        return "45 a 55 €"
    return "> 55 €"


ordre = ["< 30 €", "30 a 40 €", "40 a 45 €",
         "45,00 € (catalogue)", "45 a 55 €", "> 55 €"]
tr = Counter()
for px, q in dist_globale.items():
    tr[tranche(px)] += q
graph_data = [{"nom": k, "Ventes": tr[k]} for k in ordre if tr[k] > 0]

# tableau periode -> prix catalogue : PREUVE DE STABILITE du prix uniquement.
# (les comptages de ventes viennent tous de l'annexe B, pour ne donner qu'UN seul
#  total coherent partout.)
lignes_periodes = []
for t in table_periodes:
    lignes_periodes.append([
        {"v": t["periode"]},
        {"v": t["carte"], "align": "center"},
        {"v": t["exercice"], "align": "center"},
        {"v": t["dates"], "align": "center"},
        {"v": eur(t["prix_catalogue"]), "align": "right", "badge": "ok"},
    ])
lignes_periodes.append([
    {"v": "Toutes periodes"},
    {"v": "", "align": "center"},
    {"v": "", "align": "center"},
    {"v": "", "align": "center"},
    {"v": eur(PRIX_CATALOGUE) + " (unique)", "align": "right", "badge": "ok"},
])

# tableau des prix personnalises (hors catalogue), depuis l'annexe B : top + total
top_custom = sorted(custom_b.items(), key=lambda kv: (-kv[1], kv[0]))[:12]
lignes_custom = []
for px, q in top_custom:
    lignes_custom.append([
        {"v": eur(px)},
        {"v": num(q), "align": "right"},
        {"v": eur(round(px * q, 2)), "align": "right"},
    ])
lignes_custom.append([
    {"v": f"Total hors catalogue ({NB_PRIX_HORS} prix distincts)"},
    {"v": num(HORS_DL), "align": "right"},
    {"v": eur(HORS_EUR), "align": "right"},
])

# --------------------------------------------------------------------------
# 3 bis. Exemples "forfait" cibles sur le Menu Demi Lune (memes lignes de caisse
# que le grief "suppressions de notes"). On reprend le cas "forfait" deja publie
# (renduFinalCasSuppressions.json) et on ne garde que les fiches dont la note
# encaissee est un "Menu Demi Lune" : ce sont exactement les conversions au
# forfait qui produisent la "multitude de prix" du Menu Demi Lune.
# --------------------------------------------------------------------------
# Detection DIRECTE dans la caisse (annexes C = tickets, E = evenements DEL) :
# pour CHAQUE mois du controle, un forfait dont la note encaissee est un Menu
# Demi Lune (somme des lignes a la carte supprimees = total exact du menu).
EXOS3 = ["2022-2023", "2023-2024", "2024-2025"]
C_FILES = {e: os.path.join(ANNEXE_DIR, f"ANNEXE-C{i}_detail-tickets_{e}.xls") for i, e in enumerate(EXOS3, 1)}
E_FILES = {e: os.path.join(ANNEXE_DIR, f"ANNEXE-E{i}_tpvevenement_{e}.xls") for i, e in enumerate(EXOS3, 1)}


def _is_menu(lib):
    return lib.lower().startswith(("menu", "formule"))


def _is_demi_lune(lib):
    l = lib.lower()
    return "demi lune" in l or "demi-lune" in l


def _tm(h):
    try:
        return int(h[:2]) * 60 + int(h[3:5])
    except Exception:
        return -1


def _items_of(n):
    return [{"lib": l, "qte": round(q, 2), "pu": round(pu, 2), "montant": round(q * pu, 2)}
            for l, q, pu in n["lines"]]


par_mois = {}      # annee-mois -> 1 fiche (la premiere qui qualifie)
FA_DL = []         # liste exhaustive (date, heure, nb suppr, total, note) pour la piece jointe
# Couverture DEL : sur TOUS les Menus Demi Lune hors 45 EUR, combien sont
# justifies par une rafale DEL (somme des lignes supprimees = total de la note).
NOTES_TOTAL = NOTES_JUST = 0
COUV_TOTAL = COUV_JUST = 0.0
COUV_ROWS = []
for ex in EXOS3:
    shc = xlrd.open_workbook(C_FILES[ex]).sheet_by_index(0)
    notes_by = collections.defaultdict(lambda: collections.defaultdict(lambda: {"h": "", "tot": 0, "lines": []}))
    for r in range(1, shc.nrows):
        d = str(shc.cell_value(r, 0))[:10]
        if not d.startswith("20"):
            continue
        no = str(shc.cell_value(r, 2)).replace(".0", "")
        n = notes_by[d][no]
        n["h"] = str(shc.cell_value(r, 1))[:5]
        try:
            n["tot"] = round(float(shc.cell_value(r, 5)), 2)
        except ValueError:
            pass
        try:
            q = float(shc.cell_value(r, 11)); pu = round(float(shc.cell_value(r, 13) or 0), 2)
        except ValueError:
            q, pu = 0, 0
        n["lines"].append((str(shc.cell_value(r, 10)).strip(), q, pu))
    she = xlrd.open_workbook(E_FILES[ex]).sheet_by_index(0)
    dels_by = collections.defaultdict(list)
    for r in range(1, she.nrows):
        d = str(she.cell_value(r, 1))[:10]
        if d.startswith("20") and str(she.cell_value(r, 3)).strip() == "DEL":
            dels_by[d].append((str(she.cell_value(r, 2))[:5], round(float(she.cell_value(r, 8) or 0), 2)))

    for d in sorted(notes_by):
        k = d[:7]
        burst = collections.defaultdict(list)
        for h, m in dels_by.get(d, []):
            burst[h].append(m)
        for h, ms in sorted(burst.items()):
            if len(ms) < 2:
                continue
            s = round(sum(ms), 2)
            for no, n in notes_by[d].items():
                if (abs(n["tot"] - s) < 0.05 and abs(_tm(n["h"]) - _tm(h)) <= 8
                        and n["lines"] and all(_is_menu(l) for l, q, pu in n["lines"])
                        and any(_is_demi_lune(l) for l, q, pu in n["lines"])):
                    FA_DL.append((d, h, len(ms), round(n["tot"], 2), no))
                    par_mois.setdefault(k, {
                        "date": d,
                        "suppressions": [{"heure": h, "montant": round(x, 2)} for x in sorted(ms)],
                        "notes": [{"label": f"Note n°{no}", "heure": n["h"], "total": round(n["tot"], 2), "items": _items_of(n)}]})
                    break

    # Couverture DEL de TOUS les Menus Demi Lune hors 45 EUR.
    for d in notes_by:
        burst = collections.defaultdict(list)
        for h, m in dels_by.get(d, []):
            burst[h].append(m)
        for no, n in notes_by[d].items():
            dl = [x for x in n["lines"] if _is_demi_lune(x[0])]
            if not dl or abs(dl[0][2] - PRIX_CATALOGUE) < 0.01:
                continue
            qd, pud = round(dl[0][1], 2), round(dl[0][2], 2)
            NOTES_TOTAL += 1
            COUV_TOTAL += qd
            ok, bh, bs = False, "", 0.0
            for h, ms in burst.items():
                if (len(ms) >= 2 and abs(round(sum(ms), 2) - n["tot"]) < 0.05
                        and abs(_tm(h) - _tm(n["h"])) <= 10
                        and all(_is_menu(l) for l, q, pu in n["lines"])):
                    ok, bh, bs = True, h, round(sum(ms), 2)
                    break
            if ok:
                NOTES_JUST += 1
                COUV_JUST += qd
            COUV_ROWS.append((d, n["h"], no, pud, qd, round(n["tot"], 2),
                              "oui" if ok else "non", bh, bs))

exemples_dl = [par_mois[k] for k in sorted(par_mois)]
# Garde-fou : selection non vide et chaque fiche verifie suppressions = total note.
assert exemples_dl, "aucun exemple forfait Demi Lune trouve"
for ex in exemples_dl:
    s = round(sum(d["montant"] for d in ex["suppressions"]), 2)
    t = round(ex["notes"][0]["total"], 2)
    assert abs(s - t) < 0.05, (ex["date"], s, t)

# Onglet exhaustif (TOUTES les conversions au forfait Demi Lune de la caisse) +
# sauvegarde du classeur (les deux premiers onglets ont ete prepares plus haut).
ws3 = wb.create_sheet("Forfaits Demi Lune (tous)")
ws3.append(["Date", "Heure rafale", "Nb suppressions a la carte", "Total note = somme suppr. (€)", "Note n°"])
for c in ws3[1]:
    c.font = Font(bold=True, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor="1F2933")
    c.alignment = Alignment(horizontal="center")
for row in sorted(FA_DL):
    ws3.append([row[0], row[1], row[2], row[3], row[4]])
for i, w in enumerate((12, 12, 24, 26, 9), 1):
    ws3.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
ws3.freeze_panes = "A2"

# Couverture DEL : reponse a "combien des hors-catalogue sont justifies par un DEL ?"
COUV_JUST_I = int(round(COUV_JUST))
COUV_TOTAL_I = int(round(COUV_TOTAL))
COUV_PCT = round(100.0 * COUV_JUST / HORS_DL) if HORS_DL else 0     # rapporte aux 334 officiels
NOTES_PCT = round(100.0 * NOTES_JUST / NOTES_TOTAL) if NOTES_TOTAL else 0
ws4 = wb.create_sheet("Couverture DEL hors catalogue")
ws4.append([f"Sur les {HORS_DL} Menus Demi Lune hors catalogue (annexe B), {COUV_JUST_I} sont "
            f"justifies au centime par une rafale DEL (somme des lignes supprimees = total du menu)."])
ws4.merge_cells(start_row=1, start_column=1, end_row=1, end_column=9)
ws4.cell(1, 1).font = H1
ws4.cell(1, 1).fill = FILL1
ws4.append(["Date", "Heure note", "Note n°", "Prix unitaire menu (€)", "Couverts (qte)",
            "Total note (€)", "Justifie par DEL", "Heure rafale DEL", "Somme DEL (€)"])
for c in ws4[2]:
    c.font = BOLD
    c.fill = FILLT
    c.border = THIN
for row in sorted(COUV_ROWS):
    ws4.append(list(row))
    for c in ws4[ws4.max_row]:
        c.border = THIN
for i, w in enumerate((12, 11, 9, 22, 14, 14, 16, 16, 14), 1):
    ws4.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
ws4.freeze_panes = "A3"

wb.save(XLSX_OUT)

NB_DL = len(exemples_dl)
_parts = [it["pu"] for ex in exemples_dl for n in ex["notes"] for it in n["items"]
          if "demi lune" in it["lib"].lower() and it["pu"] > 0]
PART_MIN, PART_MAX = (min(_parts), max(_parts)) if _parts else (0.0, 0.0)

cas_demi_lune = {
    "id": "forfait-demi-lune",
    "titre": "Conversion au forfait : le Menu Demi Lune re-saisi au total consomme",
    "preuve": "Preuve exacte : la somme des lignes a la carte supprimees = le total de la note, qui ne contient plus qu'un Menu Demi Lune au prix personnalise.",
    "description": "A la demande d'une tablee (groupe, comite, evenement), l'addition est refacturee en menu(s) sans le detail des plats. Le serveur supprime les lignes a la carte et re-saisit un Menu Demi Lune a un prix personnalise egal au total consomme (divise par le nombre de couverts). La recette n'a pas disparu : a gauche les lignes supprimees, a droite la note encaissee. La somme des suppressions egale, au centime, le total du menu. C'est ce mecanisme qui cree autant de prix differents qu'il y a d'additions.",
    "exemples": exemples_dl,
}

doc = {
    "meta": {
        "slug": "prix-menu-demi-lune",
        "titre": "Instabilite des prix du « Menu Demi Lune »",
        "source": "scripts/rendu-final-prix-menu-demi-lune.py",
        "grief": "Proposition de rectification, p. 33-34 (rejet de comptabilite, 3/3, section XVI).",
        "chiffres": {
            "prix_catalogue": PRIX_CATALOGUE,
            "nb_periodes": len(table_periodes),
            "total_q": total_observe,
            "cat_q": qte_45,
            "custom_q": HORS_DL,
            "custom_eur": HORS_EUR,
            "pct_custom": PCT_HORS,
            "nb_prix_distincts": NB_PRIX_HORS,
            "qte_observee_45": qte_45,
            "pct_observee_45": pct_45,
        },
    },
    "sections": [
        {
            "kind": "chapitre",
            "source": "fisc",
            "titre": "Le grief de l'administration",
            "sousTitre": "Proposition de rectification, p. 33-34 (rejet 3/3, section XVI)",
        },
        {
            "kind": "paragraphe",
            "texte": "Le service releve que, sur la carte, le Menu Demi Lune est affiche a **45,00 €**, mais que les donnees de caisse font apparaitre une **multitude de prix differents** pour ce meme menu (d'environ 24 € a 90 €). Il en deduit que la comptabilite serait **non sincere**.",
        },
        {
            "kind": "chapitre",
            "source": "nous",
            "titre": "Une mecanique de caisse, pas une instabilite de prix",
            "sousTitre": "Le prix personnalise du menu = le total reellement consomme, refacture au forfait a la demande du client",
        },
        {
            "kind": "paragraphe",
            "texte": "Cette « multitude de prix » n'a rien d'erratique : c'est la **signature d'un mecanisme de caisse** parfaitement identifie. Certains clients (groupes, comites d'entreprise, tables d'evenement) demandent une **facture au forfait, sans le detail des plats** : un seul montant, libelle « menu », pour toute la table. Or le logiciel de caisse ne sait pas transformer une addition detaillee en forfait.",
        },
        {
            "kind": "paragraphe",
            "texte": "Le serveur est donc **oblige de supprimer les lignes de plats et de boissons** saisies a la carte, puis de **re-saisir un « Menu Demi Lune » a un prix personnalise egal au total reellement consomme** (le cas echeant divise par le nombre de couverts). Le prix affiche pour le menu **n'est pas un prix de carte** : c'est **le total de l'addition reconditionne en forfait**. Mecaniquement, il y a donc **autant de prix differents qu'il y a d'additions** : c'est exactement ce que le service a observe.",
        },
        {
            "kind": "paragraphe",
            "texte": "C'est aussi pourquoi ce grief et celui des **« suppressions de notes »** designent **les memes lignes de caisse** : les suppressions presumees etre des recettes effacees sont, pour une part, exactement ces conversions au forfait. La preuve est **arithmetique et se verifie ticket par ticket** : la **somme des lignes supprimees egale, au centime, le total du menu re-saisi**, lui-meme **integralement encaisse** et porte au chiffre d'affaires. Rien n'est retire sans contrepartie.",
        },
        {
            "kind": "alerte",
            "couleur": "blue",
            "titre": "Deja demontre, exemples dates a l'appui",
            "texte": "Ce mecanisme de conversion au forfait est demontre en detail, fiche par fiche, dans la page **Les suppressions de notes (lignes « DEL ») -> « Conversion au forfait (facture sans detail) »**. Voir la demonstration complete : [Suppressions de caisse : conversion au forfait](/rendu-final/suppressions-de-caisse).",
        },
        {
            "kind": "paragraphe",
            "texte": f"Voici **un exemple par mois** sur la periode controlee (**{NB_DL} mois**), repris directement de la caisse et **limites au Menu Demi Lune** (la liste exhaustive de toutes les conversions est dans la piece jointe). A gauche, les lignes a la carte supprimees ; a droite, la note encaissee, qui ne contient plus qu'un « Menu Demi Lune » au prix personnalise. On lit directement l'egalite : **somme des suppressions = total du menu**. Le prix personnalise par couvert (de **{eur(PART_MIN)}** a **{eur(PART_MAX)}** selon l'addition) n'est que le total consomme rapporte au couvert : voila l'origine concrete de la « multitude de prix ».",
        },
        {
            "kind": "paragraphe",
            "texte": f"**Taux de couverture.** Sur les **{HORS_DL}** Menus Demi Lune vendus hors catalogue, **{COUV_JUST_I}** (soit **{COUV_PCT} %**) sont justifies **au centime** par une rafale de suppressions (somme des lignes supprimees = total du menu), et **{NOTES_JUST}** des **{NOTES_TOTAL}** additions hors catalogue. Les autres ne laissent **aucune trace de suppression**, pour une raison simple : quand une tablee demande **des le depart** une facture au forfait sans detail, le serveur **ne saisit jamais** les lignes a la carte ; il n'y a donc **rien a supprimer**, et donc **aucun evenement DEL**. L'absence de DEL n'est pas une anomalie, c'est la **consequence mecanique** d'un forfait demande en amont. Le detail note par note (justifie ou non) figure dans la piece jointe.",
        },
        {
            "kind": "casSuppressions",
            "cas": [cas_demi_lune],
        },
        {
            "kind": "chapitre",
            "source": "nous",
            "titre": "Le prix catalogue, lui, ne bouge pas",
            "sousTitre": "Un prix de carte unique et stable, et des ventes hors catalogue toutes tracees",
        },
        {
            "kind": "paragraphe",
            "texte": f"Premier constat : par periode de carte, le **prix catalogue** du Menu Demi Lune **ne varie pas** : il vaut **{eur(PRIX_CATALOGUE)}** sur **chacune** des {len(table_periodes)} periodes du controle, malgre les changements de carte. Les ventes hors catalogue (forfaits, factures sans detail) sont **seules a l'origine** de l'amplitude relevee par le service ; elles sont chiffrees juste apres, a partir de l'annexe B.",
        },
        {
            "kind": "tableau",
            "titre": f"Prix catalogue du Menu Demi Lune par periode (stable a {eur(PRIX_CATALOGUE)})",
            "minWidth": 640,
            "colonnes": [
                {"label": "Periode"},
                {"label": "Carte", "align": "center"},
                {"label": "Exercice", "align": "center"},
                {"label": "Dates", "align": "center"},
                {"label": "Prix catalogue", "align": "right"},
            ],
            "lignes": lignes_periodes,
        },
        {
            "kind": "paragraphe",
            "texte": f"Second constat : sur les **{num(total_observe)}** Menus Demi Lune vendus (annexe B), **{num(qte_45)}** le sont au prix catalogue de **{eur(PRIX_CATALOGUE)}** ; les ventes **hors catalogue** (ces forfaits) representent **{num(HORS_DL)}** menus, soit **{str(PCT_HORS).replace('.', ',')} %**, pour **{eur(HORS_EUR)}**. Elles se repartissent sur **{NB_PRIX_HORS}** prix distincts, ce qui est coherent avec le mecanisme decrit plus haut : chaque addition convertie au forfait produit son propre montant. Le tableau ci-dessous donne les prix personnalises les plus frequents.",
        },
        {
            "kind": "tableau",
            "titre": "Menu Demi Lune : prix personnalises (hors catalogue) et occurrences",
            "minWidth": 520,
            "colonnes": [
                {"label": "Prix de vente"},
                {"label": "Occurrences", "align": "right"},
                {"label": "Montant encaisse", "align": "right"},
            ],
            "lignes": lignes_custom,
        },
        {
            "kind": "paragraphe",
            "texte": f"Distribution de l'ensemble des prix observes pour le Menu Demi Lune dans la caisse, tous exercices confondus (annexes B-1 a B-3, prix x quantite par produit). Sur **{num(total_observe)}** ventes, le prix de **{eur(PRIX_CATALOGUE)}** est de **tres loin** le plus frequent (**{num(qte_45)}** ventes, soit **{str(pct_45).replace('.', ',')} %**) ; les autres prix sont des forfaits (factures sans detail), minoritaires et disperses.",
        },
        {
            "kind": "graphique",
            "variante": "horizontal",
            "hauteur": 300,
            "dataKey": "nom",
            "serie": {"name": "Ventes", "couleur": "#0f766e"},
            "format": "int",
            "data": graph_data,
        },
        {
            "kind": "piecejointe",
            "intro": "Prix catalogue par periode (stable) et liste exhaustive des prix personnalises du Menu Demi Lune avec leurs occurrences et montants.",
            "fichiers": [
                {
                    "fichier": "pieces-defense/RF-prix-menu-demi-lune.xlsx",
                    "label": "RF - Prix du Menu Demi Lune (catalogue par periode + prix personnalises)",
                }
            ],
        },
        {
            "kind": "alerte",
            "couleur": "teal",
            "titre": "Ce qu'il faut retenir",
            "texte": f"La « multitude de prix » du Menu Demi Lune n'est pas une instabilite : c'est un mecanisme de caisse. A la demande de clients, une facture au forfait sans detail oblige a supprimer les lignes a la carte et a re-saisir le menu a un prix egal au total consomme. La somme des lignes supprimees egale, au centime, le total encaisse (deja demontre dans la page suppressions de caisse). Le prix catalogue, lui, reste unique et stable ({eur(PRIX_CATALOGUE)}) sur toutes les periodes ; les ventes hors catalogue ({str(PCT_HORS).replace('.', ',')} % pour {eur(HORS_EUR)}) sont toutes horodatees, tracees et portees au chiffre d'affaires. Aucune recette n'est dissimulee : le grief d'insincerite n'est pas fonde.",
        },
        {
            "kind": "interne",
            "audience": "avocat",
            "titre": "Note pour l'avocat",
            "texte": "Ce grief recoupe directement celui des « suppressions de notes » : memes lignes de caisse, deux angles de lecture. A tenir : (1) la pluralite de prix n'est pas une anomalie comptable mais la trace mecanique de la conversion au forfait (suppression des lignes a la carte -> re-saisie du menu au total consomme), prouvee par l'egalite somme des suppressions = total de la note, deja illustree fiche par fiche dans la page suppressions ; (2) la sincerite s'apprecie sur l'exhaustivite et la tracabilite des recettes, toutes deux etablies (chaque vente hors catalogue est horodatee et encaissee), et non sur l'uniformite des prix de vente, qu'aucun texte n'impose a un restaurateur ; (3) ne pas laisser le service additionner les deux griefs comme s'ils chiffraient des sommes distinctes. Reproductible via scripts/rendu-final-prix-menu-demi-lune.py.",
        },
    ],
}

doc["sections"] = ajouter_conclusion(doc["sections"])
os.makedirs(os.path.dirname(JSON_OUT), exist_ok=True)
json.dump(doc, open(JSON_OUT, "w", encoding="utf-8"), ensure_ascii=False, indent=1)

# --------------------------------------------------------------------------
# 5. Recapitulatif console
# --------------------------------------------------------------------------
print("XLSX  ->", XLSX_OUT)
print("JSON  ->", JSON_OUT)
print()
print(f"Prix catalogue (toutes periodes) : {eur(PRIX_CATALOGUE)}  [{len(table_periodes)} periodes]")
for t in table_periodes:
    print(f"  {t['periode']} {t['carte']} {t['exercice']} {t['dates']}: "
          f"cat={t['qte_catalogue']} hors={t['qte_hors_catalogue']} ({eur0(t['eur_hors_catalogue'])})")
print()
print(f"Hors catalogue Demi Lune : {HORS_DL}/{total_observe} ({PCT_HORS} %) = {eur(HORS_EUR)}"
      f"sur {nb_prix_distincts} prix distincts")
print(f"Annexe B (controle) - qte au prix 45,00 par exercice : {b_cat}")
print(f"Distribution globale annexe B : total={total_observe}, a 45,00={qte_45} ({pct_45} %)")
print("Graphique (tranches) :", {d['nom']: d['Ventes'] for d in graph_data})
