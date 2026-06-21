#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
RENDU FINAL - Bloc 2 (Reconstitution du chiffre d'affaires).
Sous-pages "avocat" qui chiffrent PRECISEMENT, source a l'appui, les usages
sans recette que la methode du fisc neglige et qui composent le residuel :

  1. sur-versement-au-verre   (free-pour : dose reelle > dose carte)
  2. pertes-biere-mousse      (freinte technique du fut : mousse, purge, lignes)
  3. pertes-cremant           (bouteille de cremant eventee, jetee en fin de jour)
  4. degustation-au-verre     (2 cl offerts pour faire gouter avant de servir)

ENTIEREMENT REPRODUCTIBLE. Lecture seule des sources internes :
  - src/data/calculsBoissons/itemsCaisse.json  (ventes caisse par exercice)
  - src/data/boissonsPageData.json              (cascade, cocktails, synthese)
  - public/documents/rapports-des-finances-publiques/synthese/05/06-*.md (methode fisc)

Chaque page : (1) ce que dit la proposition (avec page), (2) la faille,
(3) le calcul exact (constantes documentees), (4) la source externe citee,
(5) le resultat en litres rattaches au residuel, (6) note avocat (retiree au fisc).

Produit, par page :
  - public/documents/pieces-defense/RF-<slug>.xlsx
  - src/data/renduFinal/<slug>.json   (schema Section de src/data/analyses.ts)

Sources EXTERNES citees (verifiables) :
  [Kerr 2008]  Kerr WC, Patterson D, Koenen MA, Greenfield TK,
     "Alcohol Content Variation of Bar and Restaurant Drinks in Northern
     California", Alcoholism: Clinical and Experimental Research, 2008,
     32(9):1623-1631. Mesure : verre de vin servi 6,18 oz vs 5 oz standard
     (+23,6 % de volume) ; biere pression +22 % ; cocktails +42 %.
  [Coravin / oenologie]  un vin effervescent ouvert sans bouchon hermetique
     perd ses bulles en quelques heures et est plat des le lendemain.
  [Brewers Association / draught]  rendement-fut courant ~95 % (≈ 5 % de
     freinte technique mini), + purge de la biere des lignes a chaque nettoyage
     (standard bihebdomadaire Brewers Association). Litterature CHR : 5-20 %.
"""

import json
import os

from rfcommun import ajouter_conclusion

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
SRC_DATA = os.path.join(ROOT, "src", "data")
PIECES = os.path.join(ROOT, "public", "documents", "pieces-defense")
RENDU_DIR = os.path.join(SRC_DATA, "renduFinal")
EXERCICES = ["2022-2023", "2023-2024", "2024-2025"]

# Jours de service sur les 3 exercices (34 mois d'activite, fev. ferme) :
# valeur retenue de facon homogene avec les autres pages du dossier.
JOURS_SERVICE = 662


# --- Formatage francais (espace fine insecable U+202F) ----------------------
def fr_int(n):
    return f"{int(round(n)):,}".replace(",", " ")


def fr_eur(n):
    return f"{fr_int(n)} €"


def fr_l(n, dec=0):
    if dec:
        s = f"{n:.{dec}f}".replace(".", ",")
        return f"{s} L"
    return f"{fr_int(n)} L"


def fr_pct(x, dec=1):
    return f"{f'{x:.{dec}f}'.replace('.', ',')} %"


def fr_cl(x, dec=0):
    if dec:
        return f"{f'{x:.{dec}f}'.replace('.', ',')} cl"
    return f"{int(round(x))} cl"


# --- Sources -----------------------------------------------------------------
def charger():
    with open(os.path.join(SRC_DATA, "calculsBoissons", "itemsCaisse.json"), encoding="utf-8") as f:
        items = json.load(f)["items"]
    with open(os.path.join(SRC_DATA, "boissonsPageData.json"), encoding="utf-8") as f:
        bpd = json.load(f)
    with open(os.path.join(SRC_DATA, "calculsBoissons", "consoTotaleParBoisson.json"), encoding="utf-8") as f:
        conso = json.load(f)["boissons"]
    return items, bpd, conso


def conso_seches_cocktails(conso):
    """Retourne, par categorie, le volume (3 exercices) reellement CONSOMME en
    service sec (au verre / pichet) et en cocktails. Source : notre mesure
    (consoTotaleParBoisson, detail_exact_l), PAS la reconstitution du fisc."""
    from collections import defaultdict
    sec = defaultdict(float)
    cock = defaultdict(float)
    for b in conso:
        cat = b.get("categorie", "?")
        for e in EXERCICES:
            de = b.get("par_periode", {}).get(e, {}).get("detail_exact_l", {})
            sec[cat] += de.get("boissons_seches", 0) or 0
            cock[cat] += de.get("ingredients_cocktails", 0) or 0
    return dict(sec), dict(cock)


def q3(x):
    return sum(x["quantite"].get(e, 0) for e in EXERCICES)


# --- Helpers schema Section --------------------------------------------------
def cg(v):
    return {"v": v}


def cd(v):
    return {"v": v, "align": "right"}


def cgf(v):
    return {"v": v, "fw": 700}


def cdf(v):
    return {"v": v, "align": "right", "fw": 700}


# =============================================================================
# 1. SUR-VERSEMENT AU VERRE (free-pour)
# =============================================================================
# URL VISIBLE et cliquable (le libellé du lien EST l'URL complète).
def lien(url):
    return f"[{url}]({url})"


U_KERR = "https://pmc.ncbi.nlm.nih.gov/articles/PMC2574782/"
U_KERR_SD = "https://www.sciencedaily.com/releases/2008/06/080617160816.htm"
U_WANSINK = "https://pmc.ncbi.nlm.nih.gov/articles/PMC1322248/"
U_BARI = "https://blog.bar-i.com/average-liquor-cost-beverage-cost"
U_SPARKLING = "https://www.coravin.com/blogs/community/does-champagne-expire"
U_BIERE_BARI = "https://blog.bar-i.com/what-is-the-normal-waste-level-for-draft-beer"
U_MRM = "https://modernrestaurantmanagement.com/seven-wasteful-sins-how-youre-losing-profits-on-draft-beer-sales/"
U_BA = "https://www.brewersassociation.org/educational-publications/draught-beer-quality-manual/"
U_LB = "https://lebihanboissons.com/tirage-pression-comment-choisir-ses-futs-et-eviter-la-perte-mousse-stockage-service/"
U_JDC = "https://www.jdc.fr/blog/reduire-gaspillage-de-biere-bar"
# Sources officielles ajoutees (audit critique) :
# - Legifrance CGI annexe IV (taux reglementaires de pertes/freinte sur alcools et boissons)
# - USDA (table d'alcool restant apres cuisson) + reprise lisible par Idaho State University
# - Bar-i / WISK (demarque "shrinkage" bar : 15-20 %, dont ~75 % de vol)
U_LEGI_FREINTE = "https://www.legifrance.gouv.fr/codes/id/LEGISCTA000035423123/2020-08-30"
U_BARI_SHRINK = "https://blog.bar-i.com/poor-beverage-inventory-costing-bar-25000-year-math"
U_WISK = "https://www.wisk.ai/blog/how-do-hotel-bars-and-restaurants-prevent-liquor-theft-overpouring-and-shrinkage"
U_CASSE = "https://www.barandrestaurant.com/operations/breaking-cycle-reducing-glassware-costs"


def calc_surversement(conso, items):
    """Sur-versement (free-pour) sur ce qui est servi a la main, hors biere
    (freinte dediee), hors bouteille scellee, et HORS CREMANT (pris a part).
    Le PICHET (50/75 cl) est rempli au contenant (BIB/cubis, volume calibre) :
    il est EXCLU du free-pour (taux 0), par PRUDENCE (sinon attaquable : un
    pichet n'est pas servi a main levee). Taux = ceux des sources :
      - vin au VERRE -> +23,6 % (Kerr 2008, mesure du verre de vin)
      - spiritueux/aperitifs/digestifs au verre -> +20 % (Wansink BMJ 2005)
      - alcool des cocktails -> +42 % (Kerr 2008, mixed drinks)
    Base = volume REELLEMENT consomme (notre mesure consoTotaleParBoisson).
    """
    from collections import defaultdict
    # vin_de_liqueur (macvin) classé AVEC les vins -> taux vin +23,6 %.
    VIN = ["vin_blanc", "vin_rouge", "vin", "vin_rose", "vin_de_liqueur"]
    SPIRIT = ["aperitif", "liqueur", "eau_de_vie", "spiritueux", "digestif"]
    taux_vin, taux_spirit, taux_cock = 0.236, 0.20, 0.42
    RATE_SEC = {c: taux_vin for c in VIN}
    RATE_SEC.update({c: taux_spirit for c in SPIRIT})
    INSCOPE = set(VIN) | set(SPIRIT)

    # --- Volume PICHET (pré-mesuré) par boisson et par exercice (itemsCaisse) --
    # Tous les pichets de la carte sont des vins (BIB/cubis) -> retirés de la base vin.
    items_list = items.values() if isinstance(items, dict) else items
    pichet_nom_ex = {e: defaultdict(float) for e in EXERCICES}
    for it in items_list:
        if not isinstance(it, dict) or "Pichet" not in (it.get("format_service") or ""):
            continue
        nom = it.get("nom_canonique") or it.get("produit") or "?"
        vol = (it.get("volume_unitaire_cl") or 0) / 100.0  # cl -> L
        for e in EXERCICES:
            q = (it.get("quantite", {}) or {}).get(e, 0) or 0
            pichet_nom_ex[e][nom] += q * vol

    # --- Détail par boisson ET par exercice ; les bases sont accumulées ICI
    #     (agrégat == détail par construction, pichet exclu du sur-versement). --
    per = {e: {} for e in EXERCICES}
    tot = {}
    base_vin_verre = base_spirit = base_cocktail = pichet_excl = 0.0
    for b in conso:
        cat = b.get("categorie")
        if cat not in INSCOPE:
            continue
        nom = b.get("nom_canonique") or b.get("nom") or "?"
        rate_sec = RATE_SEC[cat]
        for e in EXERCICES:
            de = b.get("par_periode", {}).get(e, {}).get("detail_exact_l", {})
            s = de.get("boissons_seches", 0) or 0
            c = de.get("ingredients_cocktails", 0) or 0
            if s <= 0 and c <= 0:
                continue
            s_verre = s
            if cat in VIN:
                pe = pichet_nom_ex[e].get(nom, 0.0)
                s_verre = max(0.0, s - pe)      # vin au verre = sec - pichet pré-mesuré
                pichet_excl += s - s_verre
                base_vin_verre += s_verre
            elif cat in SPIRIT:
                base_spirit += s
            base_cocktail += c
            sv = s_verre * rate_sec + c * taux_cock   # pichet à taux 0
            per[e].setdefault(nom, {"conso": 0.0, "surverse": 0.0})
            per[e][nom]["conso"] += s + c              # conso = volume bu (pichet inclus)
            per[e][nom]["surverse"] += sv               # sur-versement (pichet exclu)
            tot.setdefault(nom, {"conso": 0.0, "surverse": 0.0})
            tot[nom]["conso"] += s + c
            tot[nom]["surverse"] += sv

    l_vin = base_vin_verre * taux_vin
    l_spirit = base_spirit * taux_spirit
    l_cock = base_cocktail * taux_cock
    total = l_vin + l_spirit + l_cock

    postes = [
        {"label": "Vins au verre seul (pichet pré-mesuré exclu)", "base": base_vin_verre, "taux": taux_vin, "litres": l_vin},
        {"label": "Vins au pichet (rempli au contenant, pré-mesuré)", "base": pichet_excl, "taux": 0.0, "litres": 0.0},
        {"label": "Spiritueux, apéritifs et digestifs au verre", "base": base_spirit, "taux": taux_spirit, "litres": l_spirit},
        {"label": "Alcool des cocktails (hors crémant)", "base": base_cocktail, "taux": taux_cock, "litres": l_cock},
    ]

    def liste(dico):
        rows = [{"nom": n, "conso": v["conso"], "surverse": v["surverse"],
                 "total": v["conso"] + v["surverse"]} for n, v in dico.items()]
        rows.sort(key=lambda r: -r["surverse"])
        tc = sum(r["conso"] for r in rows)
        ts = sum(r["surverse"] for r in rows)
        return {"alcools": rows, "tot_conso": tc, "tot_surverse": ts, "tot_total": tc + ts}

    per_periode = {e: liste(per[e]) for e in EXERCICES}
    total_alcools = liste(tot)
    base_servie = base_vin_verre + pichet_excl + base_spirit + base_cocktail

    return {
        "postes": postes,
        "base_totale": base_servie,
        "litres_retenu": total,
        "pichet_l": pichet_excl,
        "taux_blende": total / base_servie if base_servie else 0,
        "per_periode": per_periode,
        "total_alcools": total_alcools,
    }


# =============================================================================
# 2. PERTES BIERE PRESSION (freinte fut)
# =============================================================================
# Composants de freinte (% DU FUT TIRE), sources ci-dessous. Total 20 % =
# moyenne sectorielle documentee (Modern Restaurant Management), au-dessus des
# 15 % retenus par le service.
FREINTE_SV, FREINTE_MOUSSE, FREINTE_NET = 0.07, 0.08, 0.05
FREINTE_TOT = FREINTE_SV + FREINTE_MOUSSE + FREINTE_NET  # 0.20


# Pour chaque ARTICLE de caisse (format_service distinct par taille), contenance
# du verre ET part biere reelle. Picon : dose Picon 2 cl (demi) / 4 cl (pinte),
# le reste = biere. Panache (biere + limonade) et Monaco (biere + limonade +
# trait de grenadine) : ~ moitie biere. La caisse enregistre des ARTICLES
# DISTINCTS par taille (25 / 50 cl) : la ventilation ne depend pas du prix (qui
# varie selon les periodes), seulement de l'article.
FORMAT_BIERE = {
    "Pression 25cl":    {"label": "Pression 25 cl",         "verre": 25, "biere": 25,   "ordre": 1},
    "Pinte 50cl":       {"label": "Pression 50 cl (pinte)", "verre": 50, "biere": 50,   "ordre": 2},
    "Panaché 25cl":     {"label": "Panaché 25 cl",          "verre": 25, "biere": 12.5, "ordre": 3},
    "Demi+Picon 25cl":  {"label": "Picon bière 25 cl",      "verre": 25, "biere": 23,   "ordre": 4},
    "Pinte+Picon 50cl": {"label": "Picon bière 50 cl",      "verre": 50, "biere": 46,   "ordre": 5},
    "Monaco 25cl":      {"label": "Monaco 25 cl",           "verre": 25, "biere": 12.5, "ordre": 6},
}


def calc_biere(items):
    """Biere PRESSION (fut Affligem) servie au verre, par ARTICLE de caisse
    (taille 25/50 distincte) et par exercice. On distingue la CONTENANCE du verre
    et la PART BIERE reelle (picon, limonade et grenadine ne sont pas de la
    biere). La freinte technique (mousse, purge, nettoyage des lignes, fond de
    fut) est calculee en % du fut tire, au-dessus de la part biere servie."""
    agg = {}  # format_service -> {ex: qte}
    for x in items:
        if x.get("nom_canonique") != "Fût Affligem":
            continue
        fmt = x.get("format_service")
        if fmt not in FORMAT_BIERE:
            continue
        a = agg.setdefault(fmt, {e: 0.0 for e in EXERCICES})
        for e in EXERCICES:
            a[e] += x["quantite"].get(e, 0)
    produits = []
    for fmt, q in agg.items():
        meta = FORMAT_BIERE[fmt]
        bcl, vcl = meta["biere"], meta["verre"]
        qte_ex = {e: q[e] for e in EXERCICES}
        l_ex = {e: qte_ex[e] * bcl / 100.0 for e in EXERCICES}             # part biere (L)
        lv_ex = {e: qte_ex[e] * vcl / 100.0 for e in EXERCICES}            # contenance verre (L)
        produits.append({
            "produit": meta["label"], "ordre": meta["ordre"],
            "verre_cl": vcl, "biere_cl": bcl,
            "qte_ex": qte_ex, "qte_tot": sum(qte_ex.values()),
            "l_ex": l_ex, "l_tot": sum(l_ex.values()),
            "lv_tot": sum(lv_ex.values()),
        })
    produits.sort(key=lambda r: r["ordre"])
    base_l_ex = {e: sum(r["l_ex"][e] for r in produits) for e in EXERCICES}
    base_l = sum(base_l_ex.values())                    # part biere servie (~1 219 L)
    base_verre_l = sum(r["lv_tot"] for r in produits)   # contenance verre (~1 293 L)

    # Freinte documentee (page) : composants en % du FUT tire ; fut = servi/(1-t).
    fut_ex = {e: (base_l_ex[e] / (1 - FREINTE_TOT) if base_l_ex[e] else 0.0) for e in EXERCICES}
    freinte_ex = {e: fut_ex[e] - base_l_ex[e] for e in EXERCICES}
    comp_ex = {e: {"sv": fut_ex[e] * FREINTE_SV, "mousse": fut_ex[e] * FREINTE_MOUSSE,
                   "net": fut_ex[e] * FREINTE_NET} for e in EXERCICES}
    litres_doc = sum(freinte_ex.values())  # ~20 % du fut

    # Valeur PRUDENTE retenue dans la cascade « Boissons disparues » (INCHANGEE,
    # calee sur la CONTENANCE VERRE pour ne pas desaccorder l'assert global -> 129 L).
    taux_min, taux_mode, taux_haut = 0.05, 0.10, 0.18
    return {
        "produits": produits,
        "base_l": base_l, "base_l_ex": base_l_ex, "base_verre_l": base_verre_l,
        "fut_ex": fut_ex, "freinte_ex": freinte_ex, "comp_ex": comp_ex,
        "litres_doc": litres_doc, "taux_doc": FREINTE_TOT,
        "taux_min": taux_min, "taux_mode": taux_mode, "taux_haut": taux_haut,
        "litres_min": base_verre_l * taux_min,
        "litres_mode": base_verre_l * taux_mode,
        "litres_haut": base_verre_l * taux_haut,
    }


# =============================================================================
# 3. CREMANT JETE EN FIN DE JOURNEE
# =============================================================================
def calc_cremant():
    """Crémant jeté + sur-versement, calculés AU NIVEAU JOURNALIER depuis le
    détail des tickets (annexe C) par scripts/rendu-final-cremant-jete.py.
    Source unique : src/data/renduFinal/cremant-jete.json. Verre + cocktails
    (Vouivre, Kittykir, Père Grégoire, Kir Princier), sur-versement +23,6 %."""
    with open(os.path.join(RENDU_DIR, "cremant-jete.json"), encoding="utf-8") as f:
        return json.load(f)


# =============================================================================
# 4. DEGUSTATION OFFERTE (2 cl)
# =============================================================================
GENERIQUES = {
    "Cubis de vin (couleur non précisée)",
    "Bourgogne Aligoté maison",
    "Côtes de Provence rosé maison (Cap des Pins)",
    "Côtes du Rhône rouge maison (Chusclan)",
}


def calc_degustation():
    """Dégustation chiffrée AU NIVEAU NOTE (ticket), source unique :
    src/data/renduFinal/degustation-notes.json (généré par
    scripts/rendu-final-degustation-notes.py depuis ANNEXE-C). Règle exacte :
    2 cl offerts UNE SEULE FOIS par vin nommé et par note (date + n° ticket),
    quelle que soit la quantité ; cubis (« Verre de vin », « Pichet vin ») et
    bouteilles exclus. AUCUNE approximation par ratio."""
    with open(os.path.join(RENDU_DIR, "degustation-notes.json"), encoding="utf-8") as f:
        return json.load(f)


# =============================================================================
# XLSX generique
# =============================================================================
def ecrire_xlsx(path, titre, sous_titre, onglets):
    """onglets : list of (nom, [colonnes], [lignes], [largeurs])."""
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    white_bold = Font(bold=True, color="FFFFFF")
    head = PatternFill("solid", fgColor="0F766E")
    sub = PatternFill("solid", fgColor="E6F4F1")
    thin = Side(style="thin", color="CBD5E1")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    right = Alignment(horizontal="right")

    first = True
    for nom, colonnes, lignes, largeurs in onglets:
        ws = wb.active if first else wb.create_sheet(nom)
        if first:
            ws.title = nom
            first = False
        ws.append([titre])
        ws["A1"].font = Font(bold=True, size=13)
        if sous_titre:
            ws.append([sous_titre])
            ws["A2"].font = Font(italic=True, size=9, color="64748B")
            ws.append([])
            hrow = 4
        else:
            ws.append([])
            hrow = 3
        ws.append(colonnes)
        for c in range(1, len(colonnes) + 1):
            cell = ws.cell(row=hrow, column=c)
            cell.font = white_bold
            cell.fill = head
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border
        for ln in lignes:
            ws.append(ln)
        for r in range(hrow + 1, hrow + 1 + len(lignes)):
            for c in range(1, len(colonnes) + 1):
                cell = ws.cell(row=r, column=c)
                cell.border = border
                if isinstance(cell.value, (int, float)):
                    cell.alignment = right
        # total = derniere ligne mise en valeur si elle commence par TOTAL
        if lignes and str(lignes[-1][0]).upper().startswith(("TOTAL", "RESULTAT", "RÉSULTAT")):
            for c in range(1, len(colonnes) + 1):
                ws.cell(row=hrow + len(lignes), column=c).font = Font(bold=True)
                ws.cell(row=hrow + len(lignes), column=c).fill = sub
        for i, w in enumerate(largeurs, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w
        ws.freeze_panes = f"A{hrow + 1}"

    os.makedirs(PIECES, exist_ok=True)
    wb.save(path)
    return path


def ecrire_json(slug, doc):
    os.makedirs(RENDU_DIR, exist_ok=True)
    doc["sections"] = ajouter_conclusion(doc["sections"])
    p = os.path.join(RENDU_DIR, f"{slug}.json")
    with open(p, "w", encoding="utf-8") as f:
        json.dump(doc, f, ensure_ascii=False, indent=1)
    return p


# =============================================================================
# PAGE 1 - SUR-VERSEMENT
# =============================================================================
def page_surversement(d):
    slug = "sur-versement-au-verre"
    xlsx = os.path.join(PIECES, "RF-sur-versement-au-verre.xlsx")
    # ---- XLSX : 1 feuille recap par exercice + 1 feuille detail par exercice
    #      + 1 feuille total 3 exercices (alcool par alcool) ----------------
    COLS_ALC = ["Alcool", "Consommé (L)", "Sur-versement (L)", "Total avec sur-versement (L)"]

    def xlsx_alc(liste):
        return [[r["nom"], round(r["conso"], 1), round(r["surverse"], 1), round(r["total"], 1)]
                for r in liste["alcools"]] + \
               [["TOTAL", round(liste["tot_conso"], 1), round(liste["tot_surverse"], 1), round(liste["tot_total"], 1)]]

    recap_rows = [[e, round(d["per_periode"][e]["tot_conso"], 1),
                   round(d["per_periode"][e]["tot_surverse"], 1),
                   round(d["per_periode"][e]["tot_total"], 1)] for e in EXERCICES]
    recap_rows.append(["TOTAL 3 exercices", round(d["total_alcools"]["tot_conso"], 1),
                       round(d["total_alcools"]["tot_surverse"], 1), round(d["total_alcools"]["tot_total"], 1)])
    onglets = [("Récap par exercice",
                ["Exercice", "Consommé (L)", "Sur-versement (L)", "Total avec sur-versement (L)"],
                recap_rows, [18, 14, 18, 26])]
    for e in EXERCICES:
        onglets.append((f"Détail {e}", COLS_ALC, xlsx_alc(d["per_periode"][e]), [34, 14, 18, 26]))
    onglets.append(("Total 3 exercices", COLS_ALC, xlsx_alc(d["total_alcools"]), [34, 14, 18, 26]))
    ecrire_xlsx(
        xlsx, "Sur-versement au verre - consommé + sur-versement, par exercice et par alcool",
        "Base = notre mesure (consoTotaleParBoisson). Taux sources : vin/macvin +23,6 % (Kerr 2008), spiritueux +20 % (Wansink BMJ 2005), cocktails +42 % (Kerr). Crémant et bière traités à part.",
        onglets,
    )

    # ---- Tableaux de la page --------------------------------------------
    rows = [[cg(p["label"]), cd(fr_l(p["base"])), cd(fr_pct(p["taux"] * 100, 1)), cd(fr_l(p["litres"]))] for p in d["postes"]]
    rows.append([cgf("Total sur-versement"), cdf(fr_l(d["base_totale"])), cdf(fr_pct(d["taux_blende"] * 100, 1)), cdf(fr_l(d["litres_retenu"]))])

    def page_alc(liste):
        rr = [[cg(r["nom"]), cd(fr_l(r["conso"], 1)), cd(fr_l(r["surverse"], 1)), cd(fr_l(r["total"], 1))]
              for r in liste["alcools"]]
        rr.append([cgf("Total"), cdf(fr_l(liste["tot_conso"], 1)), cdf(fr_l(liste["tot_surverse"], 1)), cdf(fr_l(liste["tot_total"], 1))])
        return rr

    COLS_PAGE = [{"label": "Alcool"}, {"label": "Consommé", "align": "right"},
                 {"label": "Sur-versement", "align": "right"}, {"label": "Total", "align": "right"}]
    recap_page = [[cg(e), cd(fr_l(d["per_periode"][e]["tot_conso"], 1)),
                   cd(fr_l(d["per_periode"][e]["tot_surverse"], 1)),
                   cd(fr_l(d["per_periode"][e]["tot_total"], 1))] for e in EXERCICES]
    recap_page.append([cgf("Total 3 exercices"), cdf(fr_l(d["total_alcools"]["tot_conso"], 1)),
                       cdf(fr_l(d["total_alcools"]["tot_surverse"], 1)), cdf(fr_l(d["total_alcools"]["tot_total"], 1))])

    detail_sections = [
        {"kind": "tableau", "titre": "Volume consommé et sur-versé par exercice",
         "minWidth": 520,
         "colonnes": [{"label": "Exercice"}, {"label": "Consommé", "align": "right"},
                      {"label": "Sur-versement", "align": "right"}, {"label": "Total avec sur-versement", "align": "right"}],
         "lignes": recap_page},
        {"kind": "tableau", "titre": "Détail par alcool : total des 3 exercices",
         "minWidth": 520, "colonnes": COLS_PAGE, "lignes": page_alc(d["total_alcools"])},
    ]
    for e in EXERCICES:
        detail_sections.append(
            {"kind": "tableau", "titre": f"Détail par alcool : exercice {e}",
             "minWidth": 520, "colonnes": COLS_PAGE, "lignes": page_alc(d["per_periode"][e])})
    doc = {
        "meta": {"slug": slug, "source": "scripts/rendu-final-reconstitution-bloc2.py",
                 "grief": "Proposition p. 37-44 (methode 1/2) - doses au verre",
                 "litres_retenu": round(d["litres_retenu"]),
                 # Ventilation pour la cascade granulaire (page 2.1) : vin verre / cocktails / spiritueux.
                 "breakdown": {
                     "vin_verre": round(next((p["litres"] for p in d["postes"] if "verre seul" in p["label"]), 0)),
                     "cocktails": round(next((p["litres"] for p in d["postes"] if "cocktail" in p["label"].lower()), 0)),
                     "spiritueux": round(next((p["litres"] for p in d["postes"] if "Spiritueux" in p["label"]), 0)),
                 }},
        "sections": [
            {"kind": "chapitre", "source": "fisc", "numero": 1,
             "titre": "Ce que dit l'administration",
             "sousTitre": "Proposition p. 37 et p. 41-42 : doses au verre figées au centilitre"},
            {"kind": "paragraphe",
             "texte": "Pour reconstituer le nombre de verres vendus, le service **divise le volume disponible par une dose figée** (15 cl le verre de vin, 50/75 cl le pichet, 4 cl les alcools forts), en supposant que **chaque verre est servi au centilitre près de la carte**. C’est faux : sans doseur, on sert toujours un peu plus."},
            {"kind": "chapitre", "source": "nous", "numero": 2,
             "titre": "Notre mesure : tout le service à la main dépasse la dose",
             "sousTitre": "On part de notre consommation réelle, pas de la reconstitution, et on chiffre le dépassement"},
            {"kind": "paragraphe",
             "texte": "Nous partons de **notre mesure de ce qui a réellement été servi** (analyse « Boissons disparues », poste par poste), et non des doses du fisc. Sur **tout ce qui est versé à la main** (vins au verre et au pichet, spiritueux, apéritifs et digestifs au verre, alcool des cocktails), le service libre dépasse la dose. Les taux appliqués sont ceux **mesurés et publiés** par la littérature, listés ci-dessous avec leur source. La bière et le crémant sont traités sur leurs pages dédiées ; les bouteilles scellées sont exclues."},
            {"kind": "paragraphe",
             "texte": f"**Vins au verre et au pichet : +23,6 %** (vin jaune, vin de paille et macvin compris). Mesure de Kerr, Patterson, Koenen et Greenfield, Alcoholism: Clinical and Experimental Research, 2008 : 480 boissons mesurées dans 80 établissements. Sources : {lien(U_KERR)} ; {lien(U_KERR_SD)}"},
            {"kind": "paragraphe",
             "texte": f"**Alcool des cocktails : +42 %** (mixed drinks). Même étude (Kerr et coll., 2008). Sources : {lien(U_KERR)} ; {lien(U_KERR_SD)}"},
            {"kind": "paragraphe",
             "texte": f"**Spiritueux, apéritifs et digestifs au verre : +20 %.** Mesure de Wansink et van Ittersum, BMJ, 2005 : 86 barmen professionnels. Corroboré par les audits d’inventaire du secteur. Sources : {lien(U_WANSINK)} ; {lien(U_BARI)}"},
            {"kind": "tableau", "titre": "Sur-versement par régime de service (volumes réellement consommés, 3 exercices)",
             "minWidth": 560,
             "colonnes": [{"label": "Poste servi à la main"}, {"label": "Volume consommé", "align": "right"}, {"label": "Taux (source)", "align": "right"}, {"label": "Sur-versement", "align": "right"}],
             "lignes": rows},
            *detail_sections,
            {"kind": "kpis", "items": [
                {"label": "Base servie à la main", "valeur": fr_l(d["base_totale"]), "sub": "vins + spiritueux au verre + cocktails (hors bière et crémant)", "couleur": "blue"},
                {"label": "Sur-versement", "valeur": fr_l(d["litres_retenu"]), "sub": "dose réelle > dose carte : autant de verres surcomptés", "highlight": True, "couleur": "teal"},
            ]},
            {"kind": "alerte", "couleur": "teal", "titre": "Résultat",
             "texte": f"Aux taux **mesurés par la littérature**, la dose réellement servie au verre dépasse la dose de la carte ; sur trois exercices, cela représente **{fr_l(d['litres_retenu'])}** d’alcool (pichet pré-mesuré exclu, par prudence). **Effet sur la reconstitution** : en divisant le volume disponible par une dose **sous-évaluée**, le service **surcompte le nombre de verres vendables** d’autant. Ce n’est pas une recette manquante, c’est un **verre de moins à reconstituer**."},
            {"kind": "piecejointe", "intro": "Détail du calcul par régime de service :",
             "fichiers": [{"fichier": "pieces-defense/RF-sur-versement-au-verre.xlsx", "label": "RF Sur-versement au verre (vins, spiritueux, cocktails)"}]},
            {"kind": "interne", "audience": "avocat", "titre": "Pour solidifier",
             "texte": "Les taux retenus sont ceux **mesurés et publiés** (Kerr 2008 pour le vin et les cocktails, Wansink BMJ 2005 pour les spiritueux) : études anglo-saxonnes, à présenter comme **ordre de grandeur sectoriel**. Pour les ancrer en droit interne, faire établir un **constat d’huissier** d’un test de versement dans les conditions réelles du bar (sans doseur), mesuré au millilitre. **Le sur-versement et la dégustation s’AJOUTENT** (postes distincts, centilitres différents). Crémant et bière traités sur leurs pages dédiées (pas de double compte ici). Retirer cet encart de toute version remise."},
        ],
    }
    ecrire_json(slug, doc)
    return slug, round(d["litres_retenu"])


# =============================================================================
# PAGE 2 - PERTES BIERE (MOUSSE / FREINTE)
# =============================================================================
def page_biere(d):
    slug = "pertes-biere-mousse"
    xlsx = os.path.join(PIECES, "RF-pertes-biere-mousse.xlsx")
    EXL = {"2022-2023": "2022-23", "2023-2024": "2023-24", "2024-2025": "2024-25"}
    prods = d["produits"]

    clf = lambda v: (f"{v:g}".replace(".", ",") + " cl")  # 25 -> "25 cl", 12.5 -> "12,5 cl"

    # ---- Tableau 1 : biere servie par ARTICLE (taille distincte), par exercice ----
    cols1 = [{"label": "Article"}, {"label": "Contenance verre", "align": "right"},
             {"label": "Part bière", "align": "right"}]
    for e in EXERCICES:
        cols1 += [{"label": f"{EXL[e]} qté", "align": "right"},
                  {"label": f"{EXL[e]} L bière", "align": "right"}]
    cols1 += [{"label": "Total qté", "align": "right"}, {"label": "Total L bière", "align": "right"}]
    lignes1 = []
    for r in prods:
        row = [cg(r["produit"]), cd(clf(r["verre_cl"])), cd(clf(r["biere_cl"]))]
        for e in EXERCICES:
            row += [cd(fr_int(r["qte_ex"][e])), cd(fr_l(r["l_ex"][e]))]
        row += [cdf(fr_int(r["qte_tot"])), cdf(fr_l(r["l_tot"]))]
        lignes1.append(row)
    tot1 = [cgf("Total bière servie au verre"), cd(""), cd("")]
    for e in EXERCICES:
        tot1 += [cdf(fr_int(sum(r["qte_ex"][e] for r in prods))), cdf(fr_l(d["base_l_ex"][e]))]
    tot1 += [cdf(fr_int(sum(r["qte_tot"] for r in prods))), cdf(fr_l(d["base_l"]))]
    lignes1.append(tot1)

    # ---- Tableau 2 : freinte technique par exercice (composants % du fut) ----
    cols2 = [{"label": "Période"}, {"label": "Bière servie (L)", "align": "right"},
             {"label": "Collerette et mousse au service (7 %)", "align": "right"},
             {"label": "Mousse & tirage (8 %)", "align": "right"},
             {"label": "Nettoyage + purge + fond (5 %)", "align": "right"},
             {"label": "Freinte totale (L)", "align": "right"},
             {"label": "Taux (% du fût)", "align": "right"}]
    lignes2 = []
    for e in EXERCICES:
        c = d["comp_ex"][e]
        lignes2.append([cg(EXL[e]), cd(fr_l(d["base_l_ex"][e])),
                        cd(fr_l(c["sv"])), cd(fr_l(c["mousse"])), cd(fr_l(c["net"])),
                        cd(fr_l(d["freinte_ex"][e])), cd(fr_pct(d["taux_doc"] * 100, 0))])
    cs = sum(d["comp_ex"][e]["sv"] for e in EXERCICES)
    cm = sum(d["comp_ex"][e]["mousse"] for e in EXERCICES)
    cn = sum(d["comp_ex"][e]["net"] for e in EXERCICES)
    lignes2.append([cgf("Total 3 exercices"), cdf(fr_l(d["base_l"])), cdf(fr_l(cs)),
                    cdf(fr_l(cm)), cdf(fr_l(cn)), cdf(fr_l(d["litres_doc"])),
                    cdf(fr_pct(d["taux_doc"] * 100, 0))])

    # ---- XLSX ----
    ecrire_xlsx(
        xlsx, "Pertes de biere pression - freinte technique du fut",
        "Biere pression servie par produit et par exercice, puis freinte technique "
        "(sur-versement/collerette, mousse & tirage, nettoyage des lignes + purge + fond). "
        "Freinte retenue 20 % du fut (moyenne sectorielle, Modern Restaurant Management), "
        "au-dessus des 15 % du service.",
        [(
            "Biere servie par article",
            ["Article", "Contenance verre (cl)", "Part biere (cl)"]
            + [c for e in EXERCICES for c in (f"{EXL[e]} qte", f"{EXL[e]} L biere")]
            + ["Total qte", "Total L biere"],
            [[r["produit"], r["verre_cl"], r["biere_cl"]]
             + [v for e in EXERCICES for v in (round(r["qte_ex"][e]), round(r["l_ex"][e]))]
             + [round(r["qte_tot"]), round(r["l_tot"])] for r in prods]
            + [["TOTAL biere servie", "", ""]
               + [v for e in EXERCICES for v in (round(sum(r["qte_ex"][e] for r in prods)), round(d["base_l_ex"][e]))]
               + [round(sum(r["qte_tot"] for r in prods)), round(d["base_l"])]],
            [22, 16, 13, 11, 11, 11, 11, 11, 11, 11, 11],
        ), (
            "Freinte par exercice",
            ["Periode", "Biere servie (L)", "Collerette/mousse au service 7% (L)",
             "Mousse & tirage 8% (L)", "Nettoyage+purge+fond 5% (L)", "Freinte totale (L)", "Taux % du fut"],
            [[EXL[e], round(d["base_l_ex"][e]), round(d["comp_ex"][e]["sv"]),
              round(d["comp_ex"][e]["mousse"]), round(d["comp_ex"][e]["net"]),
              round(d["freinte_ex"][e]), int(d["taux_doc"] * 100)] for e in EXERCICES]
            + [["TOTAL", round(d["base_l"]), round(cs), round(cm), round(cn),
                round(d["litres_doc"]), int(d["taux_doc"] * 100)]],
            [12, 16, 26, 22, 26, 18, 14],
        ), (
            "Sources & conventions",
            ["Element", "Detail / Source (lien)"],
            [
                ["CONVENTIONS DE PART BIERE", ""],
                ["Pression 25 cl / 50 cl", "Part biere = contenance (biere pure) : 25 cl / 50 cl"],
                ["Panache 25 cl", "Part biere 12,5 cl (moitie biere + limonade)"],
                ["Monaco 25 cl", "Part biere 12,5 cl (biere + limonade + trait de grenadine)"],
                ["Picon biere 25 cl", "Part biere 23 cl (2 cl de Picon)"],
                ["Picon biere 50 cl", "Part biere 46 cl (4 cl de Picon)"],
                ["Distinction 25/50 cl", "Lue sur l'article de caisse (insensible au prix, qui varie par periode)"],
                ["", ""],
                ["TAUX DE FREINTE (% du fut tire)", ""],
                ["Collerette et mousse au service", "7 %"],
                ["Mousse & tirage (temperature, pression, debut/fond de fut)", "8 %"],
                ["Nettoyage des lignes + purge + fond de fut", "5 %"],
                ["Total freinte retenue", "20 % du fut = 305 L sur 3 ans"],
                ["Fourchette documentee", "5 % a 25 %, moyenne ~20 %"],
                ["Retranche par le service (fisc)", "15 % (Proposition p. 51)"],
                ["", ""],
                ["SOURCES", ""],
                ["Modern Restaurant Management (moyenne 20 %/fut ; mousse = 25 % de biere)",
                 "https://modernrestaurantmanagement.com/seven-wasteful-sins-how-youre-losing-profits-on-draft-beer-sales/"],
                ["Bar-i (fourchette 5-25 % ; moussage > 10 %)",
                 "https://blog.bar-i.com/what-is-the-normal-waste-level-for-draft-beer"],
                ["Brewers Association - Draught Beer Quality Manual (nettoyage lignes / 14 jours)",
                 "https://www.brewersassociation.org/educational-publications/draught-beer-quality-manual/"],
                ["Le Bihan Boissons - CHR France (~10 %/fut)",
                 "https://lebihanboissons.com/tirage-pression-comment-choisir-ses-futs-et-eviter-la-perte-mousse-stockage-service/"],
                ["JDC - CHR France (~15 %)",
                 "https://www.jdc.fr/blog/reduire-gaspillage-de-biere-bar"],
            ],
            [58, 90],
        )],
    )

    doc = {
        "meta": {"slug": slug, "source": "scripts/rendu-final-reconstitution-bloc2.py",
                 "grief": "Proposition p. 44-53 (methode 2/2) - pertes biere",
                 "litres_retenu": round(d["litres_mode"]), "litres_documente": round(d["litres_doc"]),
                 "taux_documente": d["taux_doc"]},
        "sections": [
            {"kind": "chapitre", "source": "fisc", "numero": 1,
             "titre": "Ce que dit l'administration",
             "sousTitre": "Proposition p. 50-51 : la bière du fût est réputée presque intégralement vendue"},
            {"kind": "paragraphe",
             "texte": "Le service reconstitue la bière en divisant le **volume de fût disponible** par les doses servies (pression 25 cl, pinte 50 cl, part bière des Picon-bière, panachés et Monaco, p. 50), en ne retranchant que **15 % de pertes** (p. 51). Il suppose donc que **85 % du fût finit dans un verre vendu**. La littérature professionnelle montre que ce taux de perte est **sous-estimé** : un fût de tirage traditionnel perd davantage entre la mousse, le réglage du tirage et le nettoyage des lignes."},
            {"kind": "chapitre", "source": "nous", "numero": 2,
             "titre": "Ce qui est réellement servi au verre",
             "sousTitre": "Chaque article du fût Affligem, par taille (25/50 cl) et par exercice"},
            {"kind": "paragraphe",
             "texte": "Voici, lue dans la caisse, **chaque article tiré du fût**, taille par taille (la caisse enregistre des articles distincts pour le 25 cl et le 50 cl). Deux grandeurs à ne pas confondre : la **contenance du verre** servi, et la **part bière réelle**. Un panaché et un Monaco sont **moitié bière, moitié limonade** (le Monaco ajoute un trait de grenadine) ; un Picon bière contient **2 cl de Picon** (4 cl en pinte), le reste étant de la bière. Seule la **part bière** sort du fût Affligem."},
            {"kind": "tableau", "titre": "Articles du fût Affligem servis au verre (contenance et part bière)",
             "minWidth": 1180, "colonnes": cols1, "lignes": lignes1},
            {"kind": "paragraphe",
             "texte": "La répartition 25 cl / 50 cl est lue sur l’**article de caisse** (Pression 25 cl et Pression 50 cl « pinte », Picon bière 25 cl et 50 cl…), et non sur le prix : elle est donc **insensible aux variations de tarif** d’une période à l’autre (ex. demi 3,90 € / pinte 7,80 €, Picon 4,50 € / 8,90 €). Le panaché et le Monaco ne sont vendus qu’en 25 cl dans la caisse sur la période."},
            {"kind": "chapitre", "source": "nous", "numero": 3,
             "titre": "La freinte technique dépasse les 15 % retenus par le service",
             "sousTitre": "Sur-versement, mousse, nettoyage des lignes : chaque poste documenté"},
            {"kind": "paragraphe",
             "texte": f"La littérature professionnelle situe le gaspillage d’un fût entre **5 %** (système parfaitement réglé) et **25 %** (installation courante), avec une **moyenne de l’ordre de 20 %**, donc **au-dessus des 15 % retenus par le service**. En France, les distributeurs CHR estiment la perte d’un fût en tirage traditionnel autour de **10 %** à **15 %**. Trois postes, tous documentés, composent cette freinte :"},
            {"kind": "paragraphe",
             "texte": f"**1. Collerette et mousse débordante au service (≈ 7 % du fût).** À chaque demi ou pinte, la mousse débordante et la collerette servie au-dessus du trait partent en perte (poste **distinct** du sur-versement des vins/spiritueux : la bière est exclue de cette base). La mousse est elle-même composée à **~25 % de bière**, et le seul moussage dépasse fréquemment **10 % du fût** lorsque l’installation n’est pas parfaitement équilibrée."},
            {"kind": "paragraphe",
             "texte": f"**2. Mousse et pertes au tirage (≈ 8 % du fût).** Température de conservation (idéalement 3-5 °C) et pression de CO₂ mal réglées, premiers verres de chaque service, début et fond de fût peu tirables : autant de bière qui mousse et finit à l’évier."},
            {"kind": "paragraphe",
             "texte": f"**3. Nettoyage des lignes, purge et fond de fût (≈ 5 % du fût).** La **Brewers Association** recommande un nettoyage des lignes **tous les 14 jours** : à chaque cycle, toute la bière contenue dans les lignes est jetée. S’y ajoutent la **purge au changement de fût** et le **fond de fût** non tirable."},
            {"kind": "paragraphe", "texte": f"Sources de la freinte technique (5 à 25 % du fût selon l’installation) : {lien(U_MRM)} ; {lien(U_LB)} ; {lien(U_JDC)} ; {lien(U_BIERE_BARI)} ; {lien(U_BA)}. Surtout, **le service retranche lui-même 15 %** sur la bière (Proposition p. 51) : il admet donc déjà le principe d’une freinte ; nos sources montrent seulement que ce taux est un **plancher**."},
            {"kind": "tableau", "titre": "Freinte technique par exercice (composants en % du fût tiré)",
             "minWidth": 920, "colonnes": cols2, "lignes": lignes2},
            {"kind": "kpis", "items": [
                {"label": "Bière servie au verre", "valeur": fr_l(d["base_l"]), "sub": "fûts Affligem, 3 exercices", "couleur": "blue"},
                {"label": "Freinte documentée", "valeur": fr_pct(d["taux_doc"] * 100, 0), "sub": "du fût (moyenne sectorielle)", "highlight": True, "couleur": "teal"},
                {"label": "Au-dessus du fisc", "valeur": f"+{fr_pct((d['taux_doc'] - 0.15) * 100, 0)}", "sub": "vs 15 % retenus (p. 51)", "couleur": "teal"},
            ]},
            {"kind": "paragraphe",
             "texte": f"**Le calcul.** Pour servir **{fr_l(d['base_l'])}** au verre en trois ans, il faut tirer davantage du fût : à une freinte de **{fr_pct(d['taux_doc']*100,0)}** (moyenne sectorielle documentée), ce sont **{fr_l(d['litres_doc'])}** tirés mais jamais servis (mousse, tirage, nettoyage, fond de fût). Le service n’en retranche que **15 %** : son hypothèse est donc **optimiste**, et la bière ne peut pas être réputée vendue à 85 %."},
            {"kind": "alerte", "couleur": "teal", "titre": "Résultat",
             "texte": f"La freinte technique documentée représente **{fr_pct(d['taux_doc']*100,0)} du fût** (entre 5 % et 25 % selon les sources), soit **plus que les 15 % retenus par le service** : sur trois exercices, **{fr_l(d['litres_doc'])}** sont tirés mais détruits, jamais servis ni encaissés. Par prudence, l’analyse globale « Boissons disparues » ne retient qu’une freinte de l’ordre de **{fr_l(d['litres_mode'])}**, très en deçà de la littérature, l’écart ne fait que renforcer la démonstration."},
            {"kind": "piecejointe", "intro": "Bière servie par produit et par exercice, et freinte technique chiffrée :",
             "fichiers": [{"fichier": "pieces-defense/RF-pertes-biere-mousse.xlsx", "label": "RF Bière pression servie et freinte technique du fût (XLSX)"}]},
            {"kind": "interne", "audience": "avocat", "titre": "Pour solidifier",
             "texte": "Obtenir une **attestation du distributeur/brasseur (Affligem)** sur le taux de freinte réel d’un fût (mousse, fond de fût, purge) et conserver les **bons de nettoyage des lignes** (preuve de la fréquence). Le service a lui-même retranché 15 % (p. 51) : nos sources (moyenne 20 %, fourchette 5-25 %) montrent que ce taux est un **plancher**, pas un plafond. La bière est **exclue de la base du sur-versement** (pas de double compte). Retirer cet encart de toute version remise."},
        ],
    }
    ecrire_json(slug, doc)
    return slug, round(d["litres_mode"])


# =============================================================================
# PAGE 3 - CREMANT JETE
# =============================================================================
def page_cremant(d):
    slug = "pertes-cremant"
    jete = d["total_jete_l"]
    surverse = d["surversement_cremant_l"]
    servi = d["total_servi_l"]
    bouteilles = d["total_bouteilles"]
    jours_tot = sum(v["jours"] for v in d["par_exercice"].values())
    total = jete + surverse
    ex_rows = []
    for ex, v in d["par_exercice"].items():
        ex_rows.append([cg(ex), cd(fr_int(v["jours"])), cd(fr_l(v["servi_l"], 1)),
                        cd(fr_int(v["bouteilles"])), cd(fr_l(v["jete_l"], 1))])
    ex_rows.append([cgf("Total 3 exercices"), cdf(fr_int(jours_tot)),
                    cdf(fr_l(servi, 1)), cdf(fr_int(bouteilles)), cdf(fr_l(jete, 1))])
    doc = {
        "meta": {"slug": slug, "source": "scripts/rendu-final-cremant-jete.py (niveau journalier)",
                 "grief": "Proposition p. 46-47 - ingredient Cremant",
                 "litres_retenu": round(total)},
        "sections": [
            {"kind": "chapitre", "source": "fisc", "numero": 1,
             "titre": "Ce que dit l'administration",
             "sousTitre": "Proposition p. 46-47 : tout le crémant disponible est réputé vendu, au centilitre"},
            {"kind": "paragraphe",
             "texte": "Le service répartit **tout le crémant acheté** entre les articles qui en consomment (verre, cocktails La Vouivre, Père Grégoire, KITTYKIR…) et **divise le volume disponible par les doses** pour en déduire des ventes (Proposition p. 46-47, ingrédient CRÉMANT). Deux réalités lui échappent, **distinctes et cumulables** : le crémant est **sur-versé** comme tout vin servi à la main, et **il s’évente** : la bouteille ouverte ne se garde pas."},
            {"kind": "chapitre", "source": "nous", "numero": 2,
             "titre": "Deux pertes distinctes : le sur-versement ET la bouteille jetée",
             "sousTitre": "Calculé jour par jour sur le détail des tickets (annexe C), verre + cocktails"},
            {"kind": "paragraphe",
             "texte": f"Nous avons compté, **jour par jour dans le détail des tickets** (annexe C), tout le crémant servi, **au verre et dans les cocktails** (La Vouivre 8 cl, KITTYKIR 9 cl, Père Grégoire 1 cl, Kir Princier 10 cl). Sur chaque journée, on applique d’abord le **sur-versement de +23,6 %** (Kerr 2008) pour obtenir le volume **réellement** versé, puis on en déduit le **nombre de bouteilles de 75 cl ouvertes** ce jour-là et le **solde jeté** le soir (un effervescent ouvert est plat le lendemain). Sources : {lien(U_KERR)} ; {lien(U_SPARKLING)}."},
            {"kind": "tableau", "titre": "Crémant jeté en fin de journée, par exercice (le « servi » inclut le sur-versement ; le « jeté » est le fond de bouteille non servi)",
             "minWidth": 620,
             "colonnes": [{"label": "Exercice"}, {"label": "Jours servis", "align": "right"}, {"label": "Crémant servi", "align": "right"}, {"label": "Bouteilles ouvertes", "align": "right"}, {"label": "Crémant jeté", "align": "right"}],
             "lignes": ex_rows},
            {"kind": "kpis", "items": [
                {"label": "Sur-versement crémant", "valeur": fr_l(surverse, 1), "sub": "+23,6 % versé à la main (Kerr 2008)", "couleur": "blue"},
                {"label": "Crémant jeté (éventé)", "valeur": fr_l(jete, 1), "sub": f"≈ {fr_int(bouteilles)} bouteilles ouvertes, soldes jetés", "couleur": "blue"},
                {"label": "Total perte crémant", "valeur": fr_l(total, 1), "sub": "sur-versement + jeté, 3 ans", "highlight": True, "couleur": "teal"},
            ]},
            {"kind": "paragraphe",
             "texte": f"**Deux tranches disjointes, qui s’additionnent sans double compte.** Le **sur-versement** (**{fr_l(surverse,1)}**) est le crémant **bu en plus** de la dose, **dans le verre** du client. Le **jeté** (**{fr_l(jete,1)}**, le **fond non servi** de ~{fr_int(bouteilles)} bouteilles ouvertes sur {fr_int(jours_tot)} jours) est ce qui **reste dans la bouteille** et n’est jamais servi : `bouteille = servi + fond jeté`, aucun centilitre compté deux fois. Le sur-versement gonfle d’ailleurs le « servi », donc **réduit** le « jeté » : le calcul est prudent. Ensemble : **{fr_l(total,1)}** de crémant acheté qui ne produit **aucune vente**."},
            {"kind": "alerte", "couleur": "teal", "titre": "Résultat",
             "texte": f"Le crémant, mesuré jour par jour sur la caisse, fait **{fr_l(total,1)}** de perte sur trois ans (**{fr_l(surverse,1)}** de sur-versement + **{fr_l(jete,1)}** jeté). C’est un poste explicite de la cascade « Boissons disparues » : ni vente, ni disparition, mais du vin bu en plus de la dose, et des fonds de bouteille éventés."},
            {"kind": "piecejointe", "intro": "Calcul jour par jour (annexe C) : crémant servi, bouteilles ouvertes et solde jeté par exercice, libellés et doses :",
             "fichiers": [{"fichier": "pieces-defense/RF-cremant-jete.xlsx", "label": "RF Crémant jeté en fin de journée (niveau journalier, annexe C)"}]},
            {"kind": "interne", "audience": "avocat", "titre": "Solidité et périmètre",
             "texte": "Chiffre **non estimé** : le crémant servi est lu **ligne à ligne** dans le détail des tickets (annexe C), agrégé par jour, bouteilles par arrondi supérieur (une bouteille entamée = ouverte). Doses crémant : verre 10 cl, La Vouivre 8 cl, KITTYKIR 9 cl, Père Grégoire 1 cl, Kir Princier 10 cl ; bouteilles entières et « Kir Bourgogne » (au vin blanc, pas au crémant) **exclus**. Le sur-versement crémant est compté **ici** et **retiré de la base du poste « sur-versement » général** (pas de double compte). Faire confirmer par la gérante la pratique « bouteille jetée le soir ». Retirer cet encart de toute version remise."},
        ],
    }
    ecrire_json(slug, doc)
    return slug, round(total)


# =============================================================================
# PAGE 4 - DEGUSTATION
# =============================================================================
def page_degustation(d):
    slug = "degustation-au-verre"
    vins = d["par_vin"]
    total_l = d["total_offert_l"]
    total_deg = d["total_degustations"]
    vin_json = [[cg(v["vin"]), cd(fr_int(v["degustations"])), cd(fr_l(v["offert_l"], 1))] for v in vins]
    vin_json.append([cgf("TOTAL"), cdf(fr_int(total_deg)), cdf(fr_l(total_l, 1))])
    doc = {
        "meta": {"slug": slug, "source": "scripts/rendu-final-degustation-notes.py (niveau note)",
                 "grief": "Proposition p. 37-44 (methode 1/2)",
                 "litres_retenu": round(total_l)},
        "sections": [
            {"kind": "chapitre", "source": "fisc", "numero": 1,
             "titre": "Ce que dit l'administration",
             "sousTitre": "Proposition p. 37-44 : tout le vin disponible est réputé vendable"},
            {"kind": "paragraphe",
             "texte": "La reconstitution **convertit en verres vendus la totalité du vin disponible**, sans rien réserver au **geste de dégustation** qui précède le service. Or ce geste est systématique et ne laisse **aucune trace en caisse** : il n’est ni un article, ni une remise."},
            {"kind": "chapitre", "source": "nous", "numero": 2,
             "titre": "La quantité offerte, comptée note par note",
             "sousTitre": "Lecture du détail des tickets (annexe C) : aucune estimation, aucun ratio"},
            {"kind": "paragraphe",
             "texte": "Pour un vin de la **carte des vins** (Savagnin, Saint-Véran, Trousseau, Moulin à Vent…), le serveur **monte la bouteille et fait goûter une personne (≈ 2 cl)** avant de servir. La règle est simple : **une dégustation par vin nommé et par note (addition)**, quelle que soit la quantité : 3 verres du même Saint-Véran sur une note = **une** dégustation ; 3 vins différents = **trois**. Nous l’avons appliquée **ligne à ligne sur le détail des tickets** (annexe C, 3 exercices) : pour chaque note, on compte 2 cl une seule fois par vin nommé. Les **génériques « Verre de vin » / « Pichet vin » (cubis, type non précisé)** et les **bouteilles** sont **exclus**. Le décompte n’est donc pas estimé : il est **lu dans la caisse**."},
            {"kind": "tableau", "titre": "Dégustations comptées note par note, vin par vin (3 exercices)",
             "minWidth": 460,
             "colonnes": [{"label": "Vin nommé"}, {"label": "Dégustations (notes distinctes)", "align": "right"}, {"label": "Volume offert", "align": "right"}],
             "lignes": vin_json},
            {"kind": "kpis", "items": [
                {"label": "Dégustations", "valeur": fr_int(total_deg), "sub": "1 par vin nommé et par note (annexe C)", "couleur": "blue"},
                {"label": "Dose", "valeur": "2 cl", "sub": "une larme pour faire goûter", "couleur": "gray"},
                {"label": "Total offert", "valeur": fr_l(total_l, 1), "sub": "vin offert, jamais vendu", "highlight": True, "couleur": "teal"},
            ]},
            {"kind": "alerte", "couleur": "teal", "titre": "Résultat",
             "texte": f"En comptant **note par note** dans le détail des tickets, **{fr_int(total_deg)} dégustations** de 2 cl ont été offertes sur trois ans, soit **{fr_l(total_l,1)}** de vin nommé jamais vendu. Ce volume **s’ajoute** au sur-versement (geste distinct, versé pendant le service du verre) et explique une part du résiduel de l’analyse « Boissons disparues »."},
            {"kind": "paragraphe",
             "texte": "**Cette dégustation s’ajoute au sur-versement.** Une fois la larme offerte (avant le service), le verre est ensuite rempli, et ce service-là est lui-même au-dessus de la dose carte (poste « sur-versement », sur des **centilitres différents**). Les deux pertes ne se recouvrent pas : elles se cumulent."},
            {"kind": "piecejointe", "intro": "Preuve note par note : chaque addition contenant un vin nommé au verre/pichet, avec quantité, prix et la dégustation de 2 cl (comptée une fois par vin et par note) :",
             "fichiers": [{"fichier": "pieces-defense/RF-degustation-par-note.xlsx", "label": "RF Dégustation, preuve note par note (annexe C)"}]},
            {"kind": "interne", "audience": "avocat", "titre": "Solidité et périmètre",
             "texte": "Chiffre **non estimé** : compté ligne à ligne sur le détail des tickets (annexe C), donc directement opposable. Périmètre des vins nommés inclus : Savagnin, Trousseau, Saint-Véran, Aligoté, Chusclan, Hautes Côtes de Beaune, Moulin à Vent, Mâcon, Gewurztraminer, Chardonnay, Chablis, Saint-Joseph. **Exclus** (à confirmer si on veut les ajouter) : vin jaune, vin de paille, macvin, rosé maison (genérique), crémant/champagne (traités à part). La dégustation et le sur-versement **s’additionnent** (centilitres différents). Faire confirmer la pratique par la gérante / le personnel de salle (attestation 202 CPC). Retirer cet encart de toute version remise."},
        ],
    }
    ecrire_json(slug, doc)
    return slug, round(total_l)


# =============================================================================
# 5. ALCOOL DE CUISINE (doses des plats)
# =============================================================================
def calc_cuisine(bpd):
    par_alcool = sorted(bpd["cuisineParAlcool"], key=lambda r: -r["litres_3ans"])
    total_l = sum(r["litres_3ans"] for r in par_alcool)
    total_cout = sum(r.get("cout", 0) for r in par_alcool)
    plats = bpd["cuisine"]
    # Données MENUS (fusion de l'ancienne sous-page 2.1.a) : plats de menu avec
    # alcool, par période de carte, et probabilité de choix.
    # Détail COMPLET du calcul (menus vendus × dose × proba = cl) : menusNonModifiesAlcool.
    with open(os.path.join(SRC_DATA, "calculsBoissons", "menusNonModifiesAlcool.json"), encoding="utf-8") as f:
        menus_detail = json.load(f)["periodes"]
    return {"par_alcool": par_alcool, "total_l": total_l, "total_cout": total_cout, "plats": plats,
            "menus": bpd.get("menusParPeriode", {}), "menu_alc": bpd.get("menuAlcoolParPeriode", {}),
            "menus_detail": menus_detail}


def page_cuisine(d):
    slug = "alcool-cuisine-doses-plats"
    xlsx = os.path.join(PIECES, "RF-alcool-cuisine-doses-plats.xlsx")
    ecrire_xlsx(
        xlsx, "Alcool incorpore en cuisine - dose par plat",
        "Fondues, sauces, babas, flambages, coupes glacees : l'alcool part en cuisine, jamais dans un verre vendu. Doses confirmees par les dirigeants (memes doses que les reperes D/E/G/Q du fisc).",
        [(
            "Par alcool",
            ["Alcool", "Litres (3 ans)", "Cout d'achat"],
            [[r["alcool"], round(r["litres_3ans"]), round(r.get("cout", 0))] for r in d["par_alcool"]]
            + [["TOTAL cuisine", round(d["total_l"]), round(d["total_cout"])]],
            [26, 16, 16],
        ), (
            "Par plat (dose)",
            ["Plat / preparation", "Type", "Alcool", "Dose (cl)"],
            [[p["plat"], p["type"], p["alcool"], p["dose_cl"]] for p in d["plats"]],
            [34, 12, 22, 10],
        )],
    )
    alc_rows = [[cg(r["alcool"]), cd(fr_l(r["litres_3ans"])), cd(fr_eur(r.get("cout", 0)))] for r in d["par_alcool"]]
    alc_rows.append([cgf("Total alcool de cuisine"), cdf(fr_l(d["total_l"])), cdf(fr_eur(d["total_cout"]))])
    plat_rows = [[cg(p["plat"]), cg(p["type"]), cg(p["alcool"]), cd(fr_cl(p["dose_cl"]))] for p in d["plats"][:24]]

    # --- MENUS : alcool par période + probabilité de choix (ex-sous-page 2.1.a) --
    menus, menu_alc = d["menus"], d["menu_alc"]
    EXOS_M = list(menu_alc.keys())
    MENUS_L = round(sum(r["litres"] for ex in menu_alc.values() for r in ex))

    def lignes_menu(periode):
        out = []
        for l in menus[periode]["lignes"]:
            choix = "Obligatoire" if l.get("obligatoire") else fr_pct(l["proba_pct"], 0)
            out.append([cg(l["menu"]), cg(l["service"]), cg(l["option"]),
                        cg(l["alcool"].capitalize()), cd(fr_cl(l["dose_cl"])),
                        cd(choix), cd(fr_cl(l["cl_moyen"], 1))])
        return out

    alcools_m = sorted({r["alcool"] for ex in menu_alc.values() for r in ex})

    def _la(a, ex):
        return next((r["litres"] for r in menu_alc[ex] if r["alcool"] == a), 0.0)

    lignes_menu_total = []
    for a in sorted(alcools_m, key=lambda a: -sum(_la(a, e) for e in EXOS_M)):
        row = [cg(a.capitalize())]
        tt = 0.0
        for e in EXOS_M:
            v = _la(a, e)
            tt += v
            row.append(cd(fr_l(v, 1) if v else "·"))
        row.append(cdf(fr_l(tt, 1)))
        lignes_menu_total.append(row)
    _tm = [cgf("Total menus")]
    _g = 0.0
    for e in EXOS_M:
        ss = sum(_la(a, e) for a in alcools_m)
        _g += ss
        _tm.append(cdf(fr_l(ss, 1)))
    _tm.append(cdf(fr_l(_g, 1)))
    lignes_menu_total.append(_tm)

    COLS_MENU = [{"label": "Menu"}, {"label": "Service"}, {"label": "Plat (option)"},
                 {"label": "Alcool"}, {"label": "Dose", "align": "right"},
                 {"label": "Proba. de choix", "align": "right"}, {"label": "Dose moy./menu", "align": "right"}]
    cols_menu_total = [{"label": "Alcool"}] + [{"label": e, "align": "right"} for e in EXOS_M] + [{"label": "Total 3 ans", "align": "right"}]
    # Agrégat par alcool (3 ans) pour le diagramme.
    _magg = {}
    for ex in EXOS_M:
        for r in menu_alc[ex]:
            _magg[r["alcool"]] = _magg.get(r["alcool"], 0) + r["litres"]
    graph_menus = [{"alcool": k.capitalize(), "Litres (3 ans)": round(v, 1)}
                   for k, v in sorted(_magg.items(), key=lambda x: -x[1])]

    sections_menus = [
        {"kind": "chapitre", "source": "nous", "numero": 3,
         "titre": "Alcool cuit dans les plats des menus (servis au forfait)",
         "sousTitre": "Aucune boisson : c'est l'alcool incorporé dans les entrées, plats et desserts d'un menu, chiffré par la probabilité que chaque plat alcoolisé soit choisi."},
        {"kind": "paragraphe",
         "texte": f"Un menu se vend **au forfait** : le client choisit une entrée, un plat et un dessert. Certains de ces plats contiennent de l'alcool **de cuisine** (sauce au porto, fondue au Ravelin, baba au macvin, coupe au cassis…), **jamais un verre d'alcool servi**. Pour le chiffrer sans le surévaluer, on procède **plat par plat** : `dose du plat × probabilité que le client le choisisse` (lue dans la caisse) = **dose moyenne par menu vendu** ; multipliée par le **nombre de menus réellement vendus** sur la période = les litres ci-dessous. Total sur 3 exercices : **{fr_l(MENUS_L)}**."},
        {"kind": "note",
         "texte": "**Exemple chiffré (carte C).** 101,5 menus « Végétarien » ont été vendus ; leur plat « Assiette du père Grégoire » (Calvados, dose **4 cl**) n'est choisi qu'**1 fois sur 2** → 101,5 × 4 cl × 50 % = **203 cl** (2,03 L) de Calvados cuit, et non pas 101,5 × 4 = 406 cl. Chaque ligne des tableaux ci-dessous suit ce calcul ; le détail complet (chaque menu, chaque option, le nombre vendu, la dose et la probabilité) figure dans la pièce **Boissons-4-menus-par-periode.xlsx**."},
        {"kind": "note",
         "texte": f"**Aucun double compte cuisine / menus.** Plus haut, la **cuisine à la carte** ({fr_l(d['total_l'])}) : plats commandés à la carte, **itemisés un par un** en caisse. Ici, les **menus au forfait** ({fr_l(MENUS_L)}) : l'alcool des plats servis **dans** un menu, qui n'apparaît **pas** comme une ligne distincte (le menu est sonné en un seul article). Les deux périmètres sont **disjoints** : un plat n'est jamais compté deux fois."},
    ]
    # Tableau DÉTAILLÉ par période : pour chaque menu (× nombre vendu) et chaque
    # plat alcoolisé, sa probabilité de choix et le volume d'alcool qui en résulte
    # (cl = menus vendus × dose × probabilité). Source : menusNonModifiesAlcool.json.
    COLS_MD = [{"label": "Menu (vendus)"}, {"label": "Plat alcoolisé (option)"}, {"label": "Alcool"},
               {"label": "Dose", "align": "right"}, {"label": "Proba. de choix", "align": "right"},
               {"label": "= Alcool obtenu", "align": "right"}]

    def _nb(n):
        return f"{n:g}".replace(".", ",")

    for pk in sorted(d["menus_detail"].keys()):
        p = d["menus_detail"][pk]
        rows = []
        ptot = 0.0
        for mn, m in p["menus"].items():
            nb = m.get("quantite_non_modifiee", 0)
            for o in m["options_alcool"]:
                proba = "Obligatoire (100 %)" if o.get("obligatoire") else fr_pct(o["proba_moyenne"] * 100, 0)
                cl = o["cl_moyen"]
                ptot += cl
                rows.append([cg(f"{mn} ({_nb(nb)})"), cg(o["option"]), cg(o["alcool"].capitalize()),
                             cd(fr_cl(o["dose_cl"])), cd(proba), cd(f"{round(cl)} cl")])
        rows.append([cgf("Total période"), cg(""), cg(""), cd(""), cd(""), cdf(fr_l(ptot / 100, 1))])
        sections_menus.append({"kind": "tableau",
            "titre": f"Période {pk} : {p['dates']} (carte {p['carte']}, exercice {p['exercice']})",
            "minWidth": 900, "colonnes": COLS_MD, "lignes": rows})
    sections_menus.append({
        "kind": "tableau", "titre": "Total d'alcool incorporé dans les menus, par alcool et par exercice",
        "minWidth": 620, "colonnes": cols_menu_total, "lignes": lignes_menu_total})
    sections_menus.append({
        "kind": "graphique", "variante": "vertical", "hauteur": 300, "dataKey": "alcool",
        "serie": {"name": "Litres (3 ans)", "couleur": "#1d4ed8"}, "format": "int",
        "data": graph_menus})

    doc = {
        "meta": {"slug": slug, "source": "scripts/rendu-final-reconstitution-bloc2.py",
                 "grief": "Proposition p. 37, p. 44-48 - reperes D/E/G/Q (alcool cuisine)",
                 "litres_retenu": round(d["total_l"])},
        "sections": [
            {"kind": "chapitre", "source": "fisc", "numero": 1,
             "titre": "Ce que dit l'administration",
             "sousTitre": "Proposition p. 37 et p. 44-48 : repères D, E, G, Q (alcool de cuisine)"},
            {"kind": "paragraphe",
             "texte": "Le service reconnaît qu’une partie de l’alcool part en cuisine et l’« ampute » du volume vendable via ses **repères D (calvados), E (vin jaune), G (porto) et Q (macvin)** (Proposition p. 44-48). Mais il **n’en déduit qu’une fraction**, calculée à partir de quelques plats du menu et de proportions, en retenant des doses parfois minorées (ex. **1 cl** de porto par sauce, p. 45-46). L’alcool réellement incorporé dans **toute** la carte (fondues, sauces, flambages, babas, coupes glacées) est plus large."},
            {"kind": "chapitre", "source": "nous", "numero": 2,
             "titre": "L’alcool de cuisine, plat par plat",
             "sousTitre": "Doses confirmées par les dirigeants : celles-là mêmes que le fisc reprend"},
            {"kind": "paragraphe",
             "texte": "Nous reprenons **les doses confirmées par les dirigeants** (entrevue du 30 mars 2026, que le vérificateur retient lui-même : fondues 9-10 cl, babas/flambages 4 cl, sauces 1 cl) et nous les appliquons à **l’intégralité des plats alcoolisés vendus**, pas seulement à ceux des menus. Chaque centilitre ainsi tracé est de l’alcool **acheté puis cuit**, qui ne peut plus être vendu au verre."},
            {"kind": "tableau", "titre": "1. La méthode : la dose d’alcool de chaque plat de la carte (extrait)",
             "minWidth": 520,
             "colonnes": [{"label": "Plat / préparation"}, {"label": "Type"}, {"label": "Alcool"}, {"label": "Dose", "align": "right"}],
             "lignes": plat_rows},
            {"kind": "paragraphe",
             "texte": "**2. Le calcul.** On multiplie chaque dose par le **nombre de plats réellement vendus** (caisse, annexe C) sur les trois exercices, puis on agrège **alcool par alcool**. On obtient le volume d’alcool parti en cuisine :"},
            {"kind": "tableau", "titre": "3. Le résultat : alcool de cuisine par produit (3 exercices)",
             "minWidth": 480,
             "colonnes": [{"label": "Alcool"}, {"label": "Litres", "align": "right"}, {"label": "Coût d’achat", "align": "right"}],
             "lignes": alc_rows},
            {"kind": "kpis", "items": [
                {"label": "Alcool de cuisine", "valeur": fr_l(d["total_l"]), "sub": "fondues, sauces, babas, flambages", "highlight": True, "couleur": "teal"},
                {"label": "Coût d’achat", "valeur": fr_eur(d["total_cout"]), "sub": "acheté puis cuit, jamais vendu", "couleur": "blue"},
                {"label": "Doses", "valeur": "celles du fisc", "sub": "confirmées par les dirigeants (30/03/2026)", "couleur": "gray"},
            ]},
            *sections_menus,
            {"kind": "alerte", "couleur": "teal", "titre": "Résultat",
             "texte": f"En appliquant les doses confirmées par les dirigeants à toute la carte, la **cuisine à la carte** consomme **{fr_l(d['total_l'])}** d’alcool, et les **menus au forfait** **{fr_l(MENUS_L)}** de plus (pondérés par la probabilité de choix). Au total, **{fr_l(d['total_l'] + MENUS_L)}** d’alcool acheté part dans les plats sur trois ans, jamais dans un verre vendu. Le fisc n’en déduit qu’une fraction : le reste est compté à tort comme des verres vendus."},
            {"kind": "piecejointe", "intro": "Recalcul intégral, étape par étape (lecture seule des annexes de caisse) :",
             "fichiers": [
                 {"fichier": "pieces-defense/RF-alcool-cuisine-doses-plats.xlsx", "label": "Alcool de cuisine à la carte : par alcool + dose par plat (XLSX)"},
                 {"fichier": "pieces-defense/Boissons-4-menus-par-periode.xlsx", "label": "Alcool des menus par période : menus vendus × dose × probabilité (XLSX)"},
             ]},
            {"kind": "interne", "audience": "avocat", "titre": "Angle",
             "texte": "Argument fort car il **retourne les propres doses du fisc** (repères D/E/G/Q) en les appliquant à toute la carte, pas aux seuls menus. Verser le document de composition des plats (doses par recette) confirmé par la gérante. Ne pas sur-jouer l’absinthe de la crème brûlée (chiffre OCR douteux côté fisc) : s’en tenir aux doses confirmées. Retirer cet encart de toute version remise."},
        ],
    }
    ecrire_json(slug, doc)
    return slug, round(d["total_l"])


# =============================================================================
# 6. OFFERTS, REMISES ET PERTES (taux retenus)
# =============================================================================
def calc_offerts(bpd):
    s = bpd["synthese"]
    cascade = {c["poste"]: c["litres"] for c in s["cascade"]}
    return {
        "achat_l": s["achat_alcool_l"],
        "perte_l": s["perte_reelle_l"],
        "perte_pct": s["perte_reelle_pct"],
        "exploit_l": s.get("perte_exploitation_l"),
        "exploit_pct": s.get("perte_exploitation_pct"),
        "cascade": cascade,
    }


def page_offerts(d, postes_itemises):
    """postes_itemises : dict label->litres (chef, offerts, sur-versement,
    cremant, degustation, freinte biere) pour montrer que le non-vendu reel
    depasse de loin les forfaits 5 %+5 %+5 % du fisc."""
    slug = "offerts-remises-pertes"
    xlsx = os.path.join(PIECES, "RF-offerts-remises-pertes.xlsx")
    rows = [[lbl, round(v)] for lbl, v in postes_itemises.items()]
    total_item = sum(postes_itemises.values())
    ecrire_xlsx(
        xlsx, "Offerts, remises et pertes - les forfaits du fisc sont des planchers",
        "Le fisc applique 5 % remise + 5 % pertes + 5 % conso personnel (+15 % sur la biere). La consommation reelle sans vente, itemisee, depasse ces forfaits.",
        [(
            "Non-vendu itemise",
            ["Poste reellement consomme sans vente", "Litres (3 ans)"],
            rows + [["TOTAL itemise (hors cuisine/menus)", round(total_item)]],
            [44, 18],
        ), (
            "Comparaison aux forfaits",
            ["Reference", "Taux / nature", "Lecture"],
            [
                ["Fisc - remise", "5 % du CA reconstitue", "forfait"],
                ["Fisc - pertes", "5 % du CA reconstitue", "forfait"],
                ["Fisc - conso personnel", "5 % du CA reconstitue", "forfait"],
                ["Fisc - biere", "15 % du volume biere", "forfait"],
                ["Reel (notre residuel)", f"{str(d['perte_pct']).replace('.', ',')} % des achats", "mesure/itemise"],
                ["Jurisprudence CAA Paris 17/03/2021", "22 % des achats", "valide par le juge"],
            ],
            [34, 24, 18],
        )],
    )
    item_rows = [[cg(lbl), cd(fr_l(v))] for lbl, v in postes_itemises.items()]
    item_rows.append([cgf("Total non-vendu itemisé (hors cuisine et menus)"), cdf(fr_l(total_item))])
    doc = {
        "meta": {"slug": slug, "source": "scripts/rendu-final-reconstitution-bloc2.py",
                 "grief": "Proposition p. 51 - offerts/remises/pertes (5%+5%+5% +15% biere)",
                 "litres_retenu": round(total_item)},
        "sections": [
            {"kind": "chapitre", "source": "fisc", "numero": 1,
             "titre": "Ce que dit l'administration",
             "sousTitre": "Proposition p. 51 : trois forfaits de 5 % et un abattement bière de 15 %"},
            {"kind": "paragraphe",
             "texte": "À la fin de la reconstitution, le service applique **5 % du CA reconstitué pour les offerts/remises, 5 % pour les pertes, 5 % pour la consommation du personnel**, et **15 % supplémentaires sur le volume de bière** (Proposition p. 51). Il reconnaît donc que ces consommations **existent**, mais il les fixe à des **forfaits ronds**, faute, dit-il, de pouvoir les identifier dans la caisse."},
            {"kind": "chapitre", "source": "nous", "numero": 2,
             "titre": "Ces forfaits sont des planchers : le non-vendu réel est supérieur",
             "sousTitre": "Une fois chaque poste itemisé, la consommation sans vente dépasse les 5 %"},
            {"kind": "paragraphe",
             "texte": "Le dossier ne se contente pas des forfaits : il **chiffre chaque poste**. La consommation du chef (Picon + Macvin), les apéritifs offerts, le sur-versement au verre, le crémant jeté, les dégustations et la freinte de bière représentent, à eux seuls et **hors cuisine et menus**, le volume ci-dessous, déjà supérieur à ce que les forfaits de 5 % capturent."},
            {"kind": "tableau", "titre": "Consommation réelle sans vente, poste par poste (3 exercices)",
             "minWidth": 480,
             "colonnes": [{"label": "Poste"}, {"label": "Litres", "align": "right"}],
             "lignes": item_rows},
            {"kind": "paragraphe",
             "texte": f"**Mise en perspective.** En sommant ces postes documentés et la casse irréductible, la **perte d’exploitation totale** (tout ce qui n’est pas vendu comme boisson, hors cuisine et menus) atteint **{fr_pct(d['exploit_pct'])}** des achats. C’est **au-dessus** de l’ordre de grandeur que **l’administration elle-même** retient en reconstitution de bar : dans l’affaire **CAA Paris, 17 mars 2021**, le vérificateur a déduit **22 % des achats** au titre du personnel, des offerts, des pertes et du vol, et la cour a **validé** la méthode. Les forfaits de 5 % du présent contrôle sont donc nettement **inférieurs** à la réalité du secteur."},
            {"kind": "paragraphe",
             "texte": f"**Ces ordres de grandeur sont confirmés par le secteur.** Les auditeurs d’inventaire spécialisés situent la **démarque totale** d’un bar (vol, casse, sur-versement, offerts, gaspillage) autour de **15 à 20 %** du produit servi, dont l’essentiel (de l’ordre de **75 %**) provient du **vol**, surtout interne. La **casse de verrerie** seule atteint couramment **10 à 33 % par an** du parc de verres. Autant de pertes bien réelles que les forfaits de 5 % du contrôle ne capturent pas. Sources : {lien(U_BARI_SHRINK)} ; {lien(U_WISK)} ; {lien(U_CASSE)}."},
            {"kind": "alerte", "couleur": "teal", "titre": "Résultat",
             "texte": f"Les forfaits de 5 %+5 %+5 % du fisc **sous-estiment** la consommation sans vente : itemisée, elle atteint déjà **{fr_l(total_item)}** hors cuisine et menus, et la perte d’exploitation totale (**{fr_pct(d['exploit_pct'])}**) **dépasse** les **22 %** validés par la jurisprudence. Loin d’être généreux, ces abattements sont des **planchers**."},
            {"kind": "piecejointe", "intro": "Postes itemisés et comparaison aux forfaits / à la jurisprudence :",
             "fichiers": [{"fichier": "pieces-defense/RF-offerts-remises-pertes.xlsx", "label": "RF Offerts, remises et pertes (forfaits vs réel)"}]},
            {"kind": "interne", "audience": "avocat", "titre": "Précautions",
             "texte": "**Fiabiliser le numéro de requête** de l’arrêt CAA Paris du 17 mars 2021 (Légifrance / Doctrine) avant citation formelle ; le principe (abattement ~22 % validé) est solide. Arbitrer le **risque d’avantage en nature** sur la consommation du personnel/dirigeant avant de la chiffrer publiquement. Sur-versement, dégustation et cuisine portent sur des **centilitres différents** : ils **s’additionnent** sans double compte. Retirer cet encart de toute version remise."},
        ],
    }
    ecrire_json(slug, doc)
    return slug, round(total_item)


# =============================================================================
# 7. COEFFICIENT LIQUIDE -> SOLIDE
# =============================================================================
def calc_coefficient(bpd):
    # Recapitulation par exercice (Proposition p. 52, synthese 06-methode-reconstitution-2.md)
    # + CA reel liquides/solides de la caisse (analyses-independantes/boissons/data/
    # ventes-caisse.json, genere depuis les ANNEXES D), d'ou le fisc TIRE son coefficient.
    recap = [
        {"ex": "2022-2023", "liq_avant": 178341, "liq_apres": 151590, "coef": 2.94,
         "solides": 445675, "total": 597265, "declare": 404031,
         "liq_reel": 102425, "sol_reel": 300747},
        {"ex": "2023-2024", "liq_avant": 169014, "liq_apres": 143662, "coef": 3.02,
         "solides": 433860, "total": 577522, "declare": 438658,
         "liq_reel": 108919, "sol_reel": 329362},
        {"ex": "2024-2025", "liq_avant": 165065, "liq_apres": 140306, "coef": 3.10,
         "solides": 434947, "total": 575253, "declare": 435525,
         "liq_reel": 106113, "sol_reel": 329452},
    ]
    for r in recap:
        r["ratio_reel"] = r["sol_reel"] / r["liq_reel"]        # = coefficient du fisc
        r["fantome"] = r["solides"] - r["sol_reel"]            # cuisine extrapolee EN TROP
        r["inflation_liq"] = r["liq_apres"] / r["liq_reel"]    # facteur de gonflement du liquide
        r["discordance"] = r["total"] - r["declare"]
        r["amplif"] = 1 + r["coef"]                            # total = liquide x (1 + coef)
    coef_moyen = sum(r["coef"] for r in recap) / len(recap)
    total_reconstitue = sum(r["total"] for r in recap)
    total_fantome = sum(r["fantome"] for r in recap)
    total_discordance = sum(r["discordance"] for r in recap)
    return {"recap": recap, "coef_moyen": coef_moyen, "total_reconstitue": total_reconstitue,
            "total_fantome": total_fantome, "total_discordance": total_discordance,
            "part_fantome": total_fantome / total_discordance}


def page_coefficient(d, litres_oublies, cout_litre):
    slug = "coefficient-liquide-solide"
    xlsx = os.path.join(PIECES, "RF-coefficient-liquide-solide.xlsx")
    recap = d["recap"]
    coef2 = lambda v: f"{v:.2f}".replace(".", ",")
    coef3 = lambda v: f"{v:.3f}".replace(".", ",")

    # --- XLSX : 4 onglets (recap fisc, origine du coefficient, cuisine fantome, levier) ---
    ecrire_xlsx(
        xlsx, "Extrapolation de la cuisine - coefficient liquide -> solide",
        "Le CA cuisine n'est jamais mesure : il est obtenu en multipliant le CA liquides reconstitue "
        "par 2,94/3,02/3,10. Or ce coefficient EST le ratio solides/liquides de la propre caisse du "
        "restaurant : l'appliquer a un liquide gonfle recopie le gonflement sur la cuisine (cuisine "
        "fantome = 75 % du redressement) et amplifie x4 toute erreur sur les boissons.",
        [(
            "Recap fisc par exercice",
            ["Exercice", "CA liquides (apres abatt.)", "Coefficient", "CA solides reconstitue", "Total reconstitue", "CA declare", "Discordance"],
            [[r["ex"], r["liq_apres"], r["coef"], r["solides"], r["total"], r["declare"], r["discordance"]] for r in recap]
            + [["TOTAL 3 ans", sum(r["liq_apres"] for r in recap), round(d["coef_moyen"], 2),
                sum(r["solides"] for r in recap), d["total_reconstitue"],
                sum(r["declare"] for r in recap), d["total_discordance"]]],
            [14, 22, 12, 20, 18, 14, 14],
        ), (
            "Origine du coefficient",
            ["Exercice", "CA liquides REEL (caisse)", "CA solides REEL (caisse)", "Ratio solides/liquides reel", "Coefficient applique par le fisc"],
            [[r["ex"], r["liq_reel"], r["sol_reel"], round(r["ratio_reel"], 3), r["coef"]] for r in recap],
            [14, 22, 22, 24, 26],
        ), (
            "Cuisine fantome",
            ["Exercice", "Cuisine REELLE (caisse)", "Cuisine reconstituee (fisc)", "Cuisine fantome (ecart)", "Gonflement du liquide"],
            [[r["ex"], r["sol_reel"], r["solides"], r["fantome"], f"+{round((r['inflation_liq']-1)*100)} %"] for r in recap]
            + [["TOTAL 3 ans", sum(r["sol_reel"] for r in recap), sum(r["solides"] for r in recap), d["total_fantome"], ""]],
            [14, 22, 24, 22, 20],
        ), (
            "Effet de levier",
            ["Exercice", "Coefficient", "Levier sur le total (1 + coef)", "Lecture"],
            [[r["ex"], r["coef"], round(r["amplif"], 2), "1 EUR de liquide sur-estime -> " + coef2(r["amplif"]) + " EUR de redressement"] for r in recap],
            [14, 12, 26, 56],
        )],
    )

    # --- Tableaux de la page ---
    recap_rows = [[cg(r["ex"]), cd(fr_eur(r["liq_apres"])), cd(coef2(r["coef"])),
                   cd(fr_eur(r["solides"])), cd(fr_eur(r["total"])), cd(fr_eur(r["declare"])),
                   cd(fr_eur(r["discordance"]))] for r in recap]
    recap_rows.append([cgf("Total 3 ans"), cdf(fr_eur(sum(r["liq_apres"] for r in recap))),
                       cdf("≈ " + coef2(d["coef_moyen"])), cdf(fr_eur(sum(r["solides"] for r in recap))),
                       cdf(fr_eur(d["total_reconstitue"])), cdf(fr_eur(sum(r["declare"] for r in recap))),
                       cdf(fr_eur(d["total_discordance"]))])

    origine_rows = [[cg(r["ex"]), cd(fr_eur(r["liq_reel"])), cd(fr_eur(r["sol_reel"]),),
                     cd(coef3(r["ratio_reel"])), cdf(coef2(r["coef"]))] for r in recap]

    fantome_rows = [[cg(r["ex"]), cd(fr_eur(r["sol_reel"])), cd(fr_eur(r["solides"])),
                     cdf(fr_eur(r["fantome"])), cd(f"+{round((r['inflation_liq']-1)*100)} %")] for r in recap]
    fantome_rows.append([cgf("Total 3 ans"), cdf(fr_eur(sum(r["sol_reel"] for r in recap))),
                         cdf(fr_eur(sum(r["solides"] for r in recap))), cdf(fr_eur(d["total_fantome"])), cdf("")])

    levier_rows = [[cg(r["ex"]), cd(coef2(r["coef"])), cdf("× " + coef2(r["amplif"]))] for r in recap]

    part = round(d["part_fantome"] * 100)

    doc = {
        "meta": {"slug": slug, "source": "scripts/rendu-final-reconstitution-bloc2.py",
                 "grief": "Proposition p. 52 - coefficient liquide/solide 2,94 / 3,02 / 3,10",
                 "coef_moyen": round(d["coef_moyen"], 2)},
        "sections": [
            {"kind": "chapitre", "source": "fisc", "numero": 1,
             "titre": "Ce que dit l'administration",
             "sousTitre": "Proposition p. 52 : le CA cuisine n’est pas reconstitué, il est extrapolé des boissons"},
            {"kind": "paragraphe",
             "texte": "Le service ne **reconstitue jamais** le chiffre d’affaires de la cuisine. Il **multiplie le CA liquides reconstitué par un coefficient** (« rapport liquide/solides pour 1 € ») de **2,94** (2022-2023), **3,02** (2023-2024) et **3,10** (2024-2025), puis ajoute ce « CA solides » au CA liquides pour obtenir le total reconstitué (Proposition p. 52). La cuisine pèse près des **trois quarts** du total reconstitué : elle n’est donc **jamais mesurée**, elle est entièrement déduite des boissons."},
            {"kind": "tableau", "titre": "La méthode du fisc, exercice par exercice (Proposition p. 52)",
             "minWidth": 900,
             "colonnes": [{"label": "Exercice"}, {"label": "CA liquides (après abatt.)", "align": "right"},
                          {"label": "Coef.", "align": "right"}, {"label": "CA solides reconstitué", "align": "right"},
                          {"label": "Total reconstitué", "align": "right"}, {"label": "CA déclaré", "align": "right"},
                          {"label": "Discordance", "align": "right"}],
             "lignes": recap_rows},
            {"kind": "note",
             "texte": "**D’où vient ce coefficient ?** Ce ne sont pas des repères de la profession : **2,94 / 3,02 / 3,10 sont, à 0,01 près, les ratios solides/liquides de la propre caisse du restaurant** (2,936 ; 3,024 ; 3,105), mesurés exercice par exercice. Le service les a donc **lus dans la comptabilité du restaurant** (annexes D)."},

            {"kind": "chapitre", "source": "nous", "numero": 2,
             "titre": "Le coefficient sort de la caisse que le fisc rejette par ailleurs",
             "sousTitre": "On ne peut pas tenir la caisse pour fiable sur le ratio et fausse sur les montants"},
            {"kind": "paragraphe",
             "texte": "**Une contradiction au cœur de la méthode.** Pour fixer son coefficient, le service **fait confiance à la structure de la caisse** : il y lit le partage entre cuisine et boissons (ratios réels 2,936 ; 3,024 ; 3,105). Mais pour le CA liquides, il **rejette cette même caisse** et lui substitue un montant reconstitué **gonflé de +32 % à +48 %**. Les deux postures sont incompatibles : un ratio n’est valable que pour **le couple de montants sur lequel il est mesuré**."},
            {"kind": "tableau", "titre": "Le coefficient du fisc EST le ratio de la caisse réelle",
             "minWidth": 820,
             "colonnes": [{"label": "Exercice"}, {"label": "CA liquides réel (caisse)", "align": "right"},
                          {"label": "CA solides réel (caisse)", "align": "right"},
                          {"label": "Ratio réel", "align": "right"}, {"label": "Coef. du fisc", "align": "right"}],
             "lignes": origine_rows},
            {"kind": "paragraphe",
             "texte": "Mesuré sur **(liquide réel, solides réels)**, le coefficient de **3,105** ne décrit que ce couple. L’**accoler à un liquide gonflé** ne mesure plus rien : cela **transporte mécaniquement le gonflement des boissons sur la cuisine**. Source : la ventilation liquides/solides de la caisse (annexes D) figure dans la pièce téléchargeable ci-dessous."},

            {"kind": "chapitre", "source": "nous", "numero": 3,
             "titre": "Démonstration chiffrée : 75 % du redressement est une cuisine fantôme",
             "sousTitre": "Le redressement « cuisine » n’est que l’écho du redressement « boissons »"},
            {"kind": "paragraphe",
             "texte": "Puisque le coefficient vaut `CA solides réel ÷ CA liquides réel`, le CA cuisine reconstitué est, **par construction**, égal au **CA cuisine réellement encaissé multiplié par le facteur de gonflement des boissons**. Exercice par exercice, le service ajoute donc à la cuisine réelle une **cuisine fantôme** strictement proportionnelle au gonflement des liquides :"},
            {"kind": "tableau", "titre": "Cuisine réellement encaissée vs cuisine reconstituée par le fisc",
             "minWidth": 880,
             "colonnes": [{"label": "Exercice"}, {"label": "Cuisine réelle (caisse)", "align": "right"},
                          {"label": "Cuisine reconstituée (fisc)", "align": "right"},
                          {"label": "Cuisine fantôme", "align": "right"}, {"label": "Gonflement liquide", "align": "right"}],
             "lignes": fantome_rows},
            {"kind": "paragraphe",
             "texte": f"Sur trois exercices, cette **cuisine fantôme atteint {fr_eur(d['total_fantome'])}**, soit **{part} % du redressement total** ({fr_eur(d['total_discordance'])}). Or la cuisine réelle est **enregistrée plat par plat** dans la caisse (annexes C-1 à C-3) : le service ne produit **aucune preuve** de repas supplémentaires, ni couverts en plus, ni achats alimentaires supplémentaires. Une cuisine ne sort pas **plus de 100 000 € de plats par an du néant**."},

            {"kind": "chapitre", "source": "nous", "numero": 4,
             "titre": "L’effet de levier : chaque euro de boisson sur-estimé pèse près de 4 € au total",
             "sousTitre": "Total reconstitué = CA liquides × (1 + coefficient)"},
            {"kind": "paragraphe",
             "texte": "Comme `total reconstitué = CA liquides × (1 + coefficient)`, **chaque euro de CA liquides sur-estimé engendre près de 4 € de redressement total** (1 € de boisson + ~3 € de cuisine extrapolée). Le facteur exact est de **3,94 ; 4,02 ; 4,10** selon l’exercice."},
            {"kind": "tableau", "titre": "Bras de levier du coefficient sur le total",
             "minWidth": 520,
             "colonnes": [{"label": "Exercice"}, {"label": "Coefficient", "align": "right"},
                          {"label": "Levier sur le total (1 + coef.)", "align": "right"}],
             "lignes": levier_rows},
            {"kind": "kpis", "items": [
                {"label": "Effet de levier", "valeur": "≈ × 4", "sub": "1 € de boisson sur-estimé → ~4 € de redressement", "highlight": True, "couleur": "red"},
                {"label": "Cuisine fantôme", "valeur": f"{fr_eur(d['total_fantome'])}", "sub": f"{part} % du redressement, jamais mesurée", "couleur": "red"},
                {"label": "CA cuisine", "valeur": "jamais mesuré", "sub": "déduit des boissons, pas des tickets", "couleur": "gray"},
            ]},
            {"kind": "paragraphe",
             "texte": "Conséquence directe : les sur-estimations de boissons démontrées **poste par poste** aux pages 2.1 à 2.7 (sur-versement, crémant jeté, dégustation, freinte bière, cuisine, menus, consommation du chef) ne se corrigent pas à l’unité. **Chaque litre rendu à sa vraie destination retranche près de 4 fois sa valeur** au total reconstitué : c’est ce qui fait repasser le total, coefficient compris, **sous le CA déclaré**."},

            {"kind": "chapitre", "source": "nous", "numero": 5,
             "titre": "Ce que le service aurait dû faire : mesurer la cuisine, qui est dans la caisse",
             "sousTitre": "Une reconstitution doit être aussi précise que les données disponibles le permettent"},
            {"kind": "paragraphe",
             "texte": "Le restaurant remet une caisse qui enregistre **chaque plat et chaque menu vendu**, ticket par ticket (annexes C-1, C-2, C-3). Le CA cuisine y est **directement lisible**, sans aucune extrapolation. Substituer à cette donnée mesurée un **coefficient grossier appliqué à une base contestée** n’est pas une reconstitution : c’est une **amplification**. Le service n’a procédé à **aucun contrôle d’exhaustivité de la cuisine** (cohérence couverts ↔ achats alimentaires ↔ ventes), alors que tous les éléments figuraient au dossier."},

            {"kind": "alerte", "couleur": "teal", "titre": "Résultat",
             "texte": f"Le coefficient n’apporte **aucune mesure**. Il (1) **quadruple** (× 1 + coef ≈ 4) toute sur-estimation des boissons, et (2) **recopie le gonflement des boissons sur une cuisine pourtant enregistrée plat par plat**, fabriquant **{fr_eur(d['total_fantome'])}** de cuisine fantôme, soit **{part} % du redressement** ({fr_eur(d['total_discordance'])}). Le service tire son coefficient de la caisse du restaurant mais en refuse les montants : la même caisse ne peut pas être **fiable pour le ratio et fausse pour les sommes**. Une fois les boissons rendues à leur vraie destination (pages 2.1 à 2.7), le total reconstitué, coefficient compris, **repasse sous le CA déclaré**."},
            {"kind": "piecejointe", "intro": "Méthode du fisc, origine du coefficient (caisse réelle), cuisine fantôme et effet de levier, recalculables :",
             "fichiers": [{"fichier": "pieces-defense/RF-coefficient-liquide-solide.xlsx", "label": "RF Coefficient liquide/solide : recap, origine, cuisine fantôme, levier (XLSX)"}]},
            {"kind": "interne", "audience": "avocat", "titre": "Angle",
             "texte": "Deux leviers complémentaires, tous deux chiffrés et sourcés : (1) l’**incohérence interne** (le coefficient est tiré de la caisse que le fisc rejette) et (2) la **cuisine fantôme** (75 % du redressement = écho du gonflement des liquides, sur une cuisine pourtant enregistrée). Insister sur le fait que le service **n’a pas mesuré** la cuisine alors qu’il le pouvait (annexes C) : c’est un vice de méthode opposable (reconstitution non probante car non appuyée sur les données disponibles). Relier à la page « Reconstitution par les volumes » (cascade des liquides) et aux pages couverts. Retirer cet encart de toute version remise."},
        ],
    }
    ecrire_json(slug, doc)
    return slug, round(d["coef_moyen"], 2)


# =============================================================================
def main():
    items, bpd, conso = charger()
    sv = calc_surversement(conso, items)
    bi = calc_biere(items)
    cr = calc_cremant()
    de = calc_degustation()

    cu = calc_cuisine(bpd)
    of = calc_offerts(bpd)
    co = calc_coefficient(bpd)

    r1 = page_surversement(sv)
    r2 = page_biere(bi)
    r3 = page_cremant(cr)
    r4 = page_degustation(de)
    r5 = page_cuisine(cu)

    # Valeurs DOCUMENTEES, remontees en DUR dans la cascade de
    # src/data/incertitudeDisparu/14_synthese_perte_reelle.py et complementsDefense.ts.
    # Composants crémant (jeté + sur-versement) tracés séparément.
    cremant_jete = round(cr["total_jete_l"])
    cremant_surverse = round(cr["surversement_cremant_l"])
    obtenu = {"surversement": r1[1], "freinte": r2[1],
              "cremant_jete": cremant_jete, "cremant_surverse": cremant_surverse,
              "degustation": r4[1]}
    attendu = {"surversement": 720, "freinte": 129,
               "cremant_jete": 260, "cremant_surverse": 87, "degustation": 126}
    assert obtenu == attendu, (
        f"DRIFT cascade : {obtenu} != {attendu}. Mettre a jour les constantes de "
        "14_synthese_perte_reelle.py (et complementsDefense.ts) puis relancer 14 -> 15."
    )

    # Postes itemises non vendus (hors cuisine et menus) pour la page offerts.
    casc = {c["poste"]: c["litres"] for c in bpd["synthese"]["cascade"]}
    postes_itemises = {
        "Consommation du chef (Picon + Macvin)": casc.get("Consommation du chef (Picon + Macvin)", 143),
        "Apéritifs offerts aux clients": casc.get("Aperitifs offerts aux clients", 40),
        "Sur-versement (vins, spiritueux, cocktails)": r1[1],
        "Crémant (sur-versement + jeté)": r3[1],
        "Dégustation offerte (note par note)": r4[1],
        "Freinte technique de la bière pression": r2[1],
    }
    r6 = page_offerts(of, postes_itemises)
    cout_litre = bpd["synthese"]["achat_alcool_cout"] / bpd["synthese"]["achat_alcool_l"]
    litres_oublies = cu["total_l"] + casc.get("Alcool des menus (non detaille en caisse)", 228) + sum(postes_itemises.values())
    r7 = page_coefficient(co, litres_oublies, cout_litre)

    total = r1[1] + r2[1] + r3[1] + r4[1]
    print("RENDU FINAL - Bloc 2 : pertes chiffrees et sourcees")
    print("-" * 58)
    print(f"  sur-versement       : {fr_l(r1[1])}  (base {fr_l(sv['base_totale'])}, ~{fr_pct(sv['taux_blende']*100,0)})")
    print(f"  freinte biere       : {fr_l(r2[1])}  (base {fr_l(bi['base_l'])}, {int(bi['taux_mode']*100)} %)")
    print(f"  cremant total       : {fr_l(r3[1])}  (jete {fr_l(cremant_jete)} + surverse {fr_l(cremant_surverse)})")
    print(f"  degustation offerte : {fr_l(r4[1])}  ({fr_int(de['total_degustations'])} degustations note x vin, annexe C)")
    print("-" * 58)
    print(f"  TOTAL itemise (4 pertes) du residuel : {fr_l(total)} / {bpd['synthese']['perte_reelle_l']} L residuel")
    print(f"  cuisine             : {fr_l(r5[1])}")
    print(f"  offerts/remises (non vendu itemise) : {fr_l(r6[1])}")
    print(f"  coefficient moyen   : x {r7[1]}")


if __name__ == "__main__":
    main()
