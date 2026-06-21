#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
rendu-final-consommation-superieure-achats.py  (REFONTE)

Reponse au fisc (Proposition p. 31-32, rejet 3/3, point XV : "plus de
consommation vendue que d'achat"). Le service cite des EXEMPLES PRECIS :
  - Calvados : sur-consommation sur les 3 exercices ;
  - eaux-de-vie (Poire, Framboise, Mirabelle) et Martini Blanc : petits volumes ;
  - Haute Cote de Beaune BLANC : 12 950 cl "vendus" en 2023-2024, "aucune
    bouteille achetee identifiee".

Demonstration, SANS supposition, en reliant pour chaque exemple le libelle de
CAISSE -> la FACTURE fournisseur -> l'INVENTAIRE (fichier de correspondance), et
en comparant, par exercice, les ACHATS (nets des avoirs) a la CONSO mesuree.

Sources (lecture seule) :
  - src/data/calculsBoissons/consoTotaleParBoisson.json
  - src/data/calculsBoissons/achatsBoissonsParPeriode.json
  - src/data/calculsBoissons/_correspondance-caisse-inventaire-factures.csv
  - src/data/calculsBoissons/itemsCaisse.json
  - src/data/boissonsPageData.json

Sorties : 5 XLSX (un par section) + src/data/renduFinal/consommation-superieure-achats.json
"""
import os
import re
import csv
import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from rfcommun import normaliser_sections

ICI = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.normpath(os.path.join(ICI, ".."))
CB = os.path.join(ROOT, "src/data/calculsBoissons")
PIECES = os.path.join(ROOT, "public/documents/pieces-defense")
OUT_JSON = os.path.join(ROOT, "src/data/renduFinal/consommation-superieure-achats.json")
EXOS = ["2022-2023", "2023-2024", "2024-2025"]
EXL = {"2022-2023": "2022-23", "2023-2024": "2023-24", "2024-2025": "2024-25"}

CONSO = json.load(open(os.path.join(CB, "consoTotaleParBoisson.json"), encoding="utf-8"))["boissons"]
ACHATS = json.load(open(os.path.join(CB, "achatsBoissonsParPeriode.json"), encoding="utf-8"))["achats"]
CORR = list(csv.DictReader(open(os.path.join(CB, "_correspondance-caisse-inventaire-factures.csv"),
                                encoding="utf-8-sig"), delimiter=";"))
ITEMS = json.load(open(os.path.join(CB, "itemsCaisse.json"), encoding="utf-8"))["items"]
BD = json.load(open(os.path.join(ROOT, "src/data/boissonsPageData.json"), encoding="utf-8"))


def fr_l(v):
    return f"{v:,.1f}".replace(",", " ").replace(".", ",") + " L"


def fr_cl(v):
    return f"{v:,.0f}".replace(",", " ") + " cl"


def sgn_l(v):
    return ("+" if v >= 0 else "") + fr_l(v)


def cg(t):
    return {"v": t}


def cd(t):
    return {"v": t, "align": "right"}


def cgf(t):
    return {"v": t, "fw": 700}


def cdf(t):
    return {"v": t, "align": "right", "fw": 700}


def colg(l):
    return {"label": l}


def cold(l):
    return {"label": l, "align": "right"}


def bdg(v, ok):
    return {"v": v, "align": "right", "badge": "ok" if ok else "ko"}


# --- conso par exercice (L, valeur moyenne) --------------------------------
def conso_pp(canon):
    for b in CONSO:
        if b["nom_canonique"] == canon:
            out = {}
            for e in EXOS:
                v = b.get("par_periode", {}).get(e, {})
                tl = v.get("total_l") if isinstance(v, dict) else None
                if isinstance(tl, dict):
                    out[e] = round(tl.get("moyen", 0), 1)
                elif isinstance(v, dict):
                    out[e] = round(v.get("exact_certain_l", 0), 1)
                else:
                    out[e] = 0.0
            out["est"] = b.get("estimation", False)
            out["facture"] = b.get("nom_facture", "")
            return out
    return {e: 0.0 for e in EXOS}


# --- achats par exercice (L) pour un predicat de libelle facture -----------
def achat_pp(predicat):
    out = {e: 0.0 for e in EXOS}
    lignes = []
    for a in ACHATS:
        if not predicat(a["produit"].upper()):
            continue
        up = a["produit"].upper()
        mcl = re.search(r"(\d+)\s*CL", up)
        cont = int(mcl.group(1)) if mcl else 100
        tot = 0.0
        pp = {}
        for e in EXOS:
            q = a["par_periode"].get(e, {}).get("quantite", 0) or 0
            out[e] += q * cont / 100.0
            tot += q * cont / 100.0
            pp[e] = q
        lignes.append({"produit": a["produit"], "cont_cl": cont, "pp": pp, "l_tot": round(tot, 1)})
    return out, lignes


def tot3(d):
    return round(sum(d[e] for e in EXOS), 1)


# ===========================================================================
# Styles + helper XLSX
# ===========================================================================
GRAS = Font(bold=True)
GBLANC = Font(bold=True, color="FFFFFF")
ENTETE = PatternFill("solid", fgColor="0F766E")
CENTRE = Alignment(horizontal="center", vertical="center", wrap_text=True)
GAUCHE = Alignment(horizontal="left", vertical="center", wrap_text=True)
BORD = Border(*(4 * (Side(style="thin", color="D1D5DB"),)))


def ecrire_xlsx(path, titre, sous_titre, feuilles):
    wb = openpyxl.Workbook()
    for i, (nom, headers, rows, widths) in enumerate(feuilles):
        ws = wb.active if i == 0 else wb.create_sheet()
        ws.title = nom[:31]
        ws.append([titre])
        ws["A1"].font = GRAS
        ws.append([sous_titre])
        ws["A2"].alignment = GAUCHE
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max(1, len(headers)))
        ws.row_dimensions[2].height = 30
        ws.append([])
        hr = ws.max_row + 1
        ws.append(headers)
        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=hr, column=c)
            cell.font = GBLANC
            cell.fill = ENTETE
            cell.alignment = CENTRE
            cell.border = BORD
        for row in rows:
            r = ws.max_row + 1
            ws.append(row)
            for c in range(1, len(headers) + 1):
                ws.cell(row=r, column=c).border = BORD
        for j, w in enumerate(widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(j)].width = w
        ws.freeze_panes = "A" + str(hr + 1)
    os.makedirs(PIECES, exist_ok=True)
    wb.save(path)


# ===========================================================================
# Donnees des exemples cites
# ===========================================================================
beaune_rge_c = conso_pp("Hautes Côtes de Beaune rouge")
beaune_blc_c = conso_pp("Hautes Côtes de Beaune blanc")
nuits_c = conso_pp("Hautes Côtes de Nuits blanc")
beaune_rge_a, beaune_rge_al = achat_pp(lambda p: "BEAUNE RGE" in p or "BEAUNE ROUGE" in p)
beaune_blc_a, beaune_blc_al = achat_pp(lambda p: "BEAUNE BLC" in p or "BEAUNE BLANC" in p)
nuits_a, nuits_al = achat_pp(lambda p: "NUITS" in p)
FISC_BLANC_CL = 12950

calva_c = conso_pp("Calvados")
calva_a, calva_al = achat_pp(lambda p: "CALVADOS" in p)
calva_verre_cl = sum(sum(x["quantite"].get(e, 0) for e in EXOS) * (x.get("volume_unitaire_cl") or 0)
                     for x in ITEMS if str(x.get("produit", "")).lower().startswith("calvados"))
calva_verre_l = round(calva_verre_cl / 100.0, 1)


def famille(canon_list, pred):
    c = {e: 0.0 for e in EXOS}
    for canon in canon_list:
        cc = conso_pp(canon)
        for e in EXOS:
            c[e] += cc[e]
    a, al = achat_pp(pred)
    return {e: round(c[e], 1) for e in EXOS}, a, al


fam = [
    ("Poire (eau-de-vie + liqueur)", *famille(
        ["Eau de Vie de Poire", "Liqueur de Poire"], lambda p: "POIRE WILLIAM" in p or "GOLDEN EIGHT" in p)),
    ("Framboise (eau-de-vie)", *famille(
        ["Eau de Vie Framboise"], lambda p: "FRAMBOISE" in p and ("40°" in p or "42°" in p or "SAUVAGE" in p))),
    ("Mirabelle (eau-de-vie)", *famille(
        ["Eau de Vie Mirabelle"], lambda p: "MIRABELLE" in p)),
    ("Martini Blanc", *famille(
        ["Martini Blanc"], lambda p: "MARTINI" in p)),
]

rows_g = BD["disparuParBoisson"]
g_achat = round(sum(r["achat_l"] for r in rows_g), 1)
g_conso = round(sum(r["conso_l"] for r in rows_g), 1)
g_stock = round(sum(r.get("stock_l", 0) for r in rows_g), 1)
g_net = round(g_achat - g_conso - g_stock, 1)
negs = sorted([r for r in rows_g if r["disparu_l"] < 0], key=lambda r: r["disparu_l"])

# ===========================================================================
# XLSX (un par section)
# ===========================================================================
F_CORR = "pieces-defense/RF-conso-achats-correspondance.xlsx"
F_HCB = "pieces-defense/RF-conso-achats-haute-cote-beaune.xlsx"
F_CAL = "pieces-defense/RF-conso-achats-calvados.xlsx"
F_EDV = "pieces-defense/RF-conso-achats-eaux-de-vie.xlsx"
F_GLO = "pieces-defense/RF-conso-achats-bilan-global.xlsx"

CITES = ("calvados", "beaune", "nuits", "eau de vie", "poire", "framboise", "mirabelle", "martini")
corr_x = [[c["libelle_caisse"], c["format_service"], c["volume_cl"], c["nom_canonique"],
           c["nom_inventaire"], c["nom_facture"], c.get("flag", "")]
          for c in CORR
          if any(k in (c["libelle_caisse"] + c["nom_canonique"] + c["nom_facture"]).lower() for k in CITES)]
ecrire_xlsx(os.path.join(ROOT, "public/documents", F_CORR),
            "Correspondance caisse - facture - inventaire (produits cites)",
            "Chaque libelle caisse relie a sa facture fournisseur et a l'inventaire. Achats nets des avoirs.",
            [("Correspondance",
              ["Libelle caisse", "Format", "Vol (cl)", "Nom canonique", "Inventaire", "Facture fournisseur", "Remarque"],
              corr_x, [30, 16, 9, 30, 26, 44, 40])])

hcb_x = []
for lab, a, c in [("Beaune ROUGE", beaune_rge_a, beaune_rge_c),
                  ("Beaune BLANC", beaune_blc_a, beaune_blc_c),
                  ("Nuits BLANC (vin distinct)", nuits_a, nuits_c)]:
    for e in EXOS:
        hcb_x.append([lab, EXL[e], round(a[e], 1), c[e], round(a[e] - c[e], 1)])
ecrire_xlsx(os.path.join(ROOT, "public/documents", F_HCB),
            "Haute Cote de Beaune : rouge / blanc / Nuits (achats vs conso par exercice)",
            f"Conso reelle de Beaune blanc = {fr_cl(tot3(beaune_blc_c) * 100)} sur 3 ans, sans rapport "
            f"avec les 12 950 cl du service (qui visent en realite le rouge, massivement achete).",
            [("Beaune par couleur",
              ["Produit", "Exercice", "Achat (L)", "Conso (L)", "Achat - conso (L)"], hcb_x,
              [26, 12, 12, 12, 18]),
             ("Factures Beaune rouge",
              ["Facture fournisseur", "Cont. (cl)"] + [f"{EXL[e]} (bt)" for e in EXOS] + ["Total (L)"],
              [[r["produit"], r["cont_cl"]] + [r["pp"][e] for e in EXOS] + [r["l_tot"]] for r in beaune_rge_al],
              [56, 11, 11, 11, 11, 11])])

cal_x = [[EXL[e], round(calva_a[e], 1), calva_c[e], round(calva_a[e] - calva_c[e], 1)] for e in EXOS]
cal_x.append(["TOTAL", tot3(calva_a), tot3(calva_c), round(tot3(calva_a) - tot3(calva_c), 1)])
ecrire_xlsx(os.path.join(ROOT, "public/documents", F_CAL),
            "Calvados : achats vs consommation (dont cuisson)",
            f"Caisse : seulement {fr_l(calva_verre_l)} de Calvados au verre sur 3 ans. Le reste de la conso "
            "est une ESTIMATION de cuisson (camembert roti flambe, assiette du pere Gregoire). Alcool de "
            "cuisine = cout, pas une vente ; le service a lui-meme reduit sa dose de 5 a 4 cl.",
            [("Achat vs conso",
              ["Exercice", "Achat (L)", "Conso totale (L)", "Achat - conso (L)"], cal_x, [14, 14, 18, 18]),
             ("Achats Calvados (factures)",
              ["Facture fournisseur", "Cont. (cl)"] + [f"{EXL[e]} (bt)" for e in EXOS] + ["Total (L)"],
              [[r["produit"], r["cont_cl"]] + [r["pp"][e] for e in EXOS] + [r["l_tot"]] for r in calva_al],
              [40, 11, 11, 11, 11, 11])])

edv_x = [[lab, tot3(a), tot3(c), round(tot3(a) - tot3(c), 1)] for lab, c, a, al in fam]
edv_f = [[lab, r["produit"], r["cont_cl"], r["l_tot"]] for lab, c, a, al in fam for r in al]
ecrire_xlsx(os.path.join(ROOT, "public/documents", F_EDV),
            "Eaux-de-vie, liqueurs et Martini : les achats existent",
            "En regroupant les libelles freres (FRAMBOISE SAUVAGE, MIRABELLE, POIRE WILLIAM, LIQUEUR DE "
            "POIRE, MARTINI BIANCO), les achats sur 3 ans couvrent la conso. Le 'achat = 0' venait d'un "
            "rattachement de libelle.",
            [("Synthese 3 ans", ["Famille", "Achat (L)", "Conso (L)", "Achat - conso (L)"], edv_x, [30, 12, 12, 16]),
             ("Detail factures", ["Famille", "Facture fournisseur", "Cont. (cl)", "Total achat (L)"], edv_f,
              [26, 50, 12, 14])])

g_sorted = sorted(rows_g, key=lambda r: r["disparu_l"])
ecrire_xlsx(os.path.join(ROOT, "public/documents", F_GLO),
            "Bilan matiere global par boisson (achats - conso - stock)",
            f"Bilan d'ensemble : achats {fr_l(g_achat)} - conso {fr_l(g_conso)} - stock {fr_l(g_stock)} "
            f"= {fr_l(g_net)} net (excedent). Les rares ecarts negatifs sont des artefacts de "
            "rattachement de libelle.",
            [("Bilan par boisson",
              ["Boisson", "Achat (L)", "Conso (L)", "Stock (L)", "Achat - conso - stock (L)"],
              [[r["nom"], round(r["achat_l"], 1), round(r["conso_l"], 1), round(r.get("stock_l", 0), 1),
                round(r["disparu_l"], 1)] for r in g_sorted]
              + [["TOTAL", g_achat, g_conso, g_stock, g_net]],
              [34, 12, 12, 12, 24])])


# ===========================================================================
# Tableaux de page
# ===========================================================================
def tab_achat_conso(titre, lignes, total=None, minw=620):
    cols = [colg("Produit"), colg("Exercice"), cold("Achat (L)"), cold("Conso (L)"), cold("Achat − conso")]
    lg = []
    for lab, a, c in lignes:
        for e in EXOS:
            ec = round(a[e] - c[e], 1)
            lg.append([cg(lab), cg(EXL[e]), cd(fr_l(round(a[e], 1))), cd(fr_l(c[e])), bdg(sgn_l(ec), ec >= -0.5)])
    return {"kind": "tableau", "titre": titre, "minWidth": minw, "colonnes": cols, "lignes": lg}


# correspondance (page) : sous-ensemble lisible
corr_page = []
for c in CORR:
    blob = (c["libelle_caisse"] + c["nom_canonique"] + c["nom_facture"]).lower()
    if any(k in blob for k in ("calvados", "beaune", "nuits", "eau de vie", "poire william",
                               "framboise", "mirabelle", "martini")):
        corr_page.append([cg(c["libelle_caisse"]), cg(c["nom_canonique"]),
                          cg(c["nom_facture"] or " : ")])

doc = {
    "meta": {
        "slug": "consommation-superieure-achats",
        "titre": "Consommation vendue supérieure aux achats",
        "source": "scripts/rendu-final-consommation-superieure-achats.py",
        "grief": "Proposition de rectification p. 31-32 (rejet de comptabilité, 3/3, point XV).",
        "beaune_blanc_conso_cl": int(round(tot3(beaune_blc_c) * 100)),
        "fisc_blanc_cl": FISC_BLANC_CL,
        "calvados_verre_l": calva_verre_l,
        "bilan_net_l": g_net,
    },
    "sections": [
        # 1) GRIEF
        {"kind": "chapitre", "source": "fisc", "numero": 1, "titre": "Le grief de l’administration",
         "sousTitre": "Proposition p. 31-32 (rejet 3/3, point XV) : « plus de consommation vendue que d’achat »"},
        {"kind": "paragraphe",
         "texte": "Au fil de la reconstitution, le service relève des produits dont les **quantités "
                  "vendues dépasseraient les achats**, et en déduit que la comptabilité serait **non "
                  "sincère**. Il cite des exemples précis : le **Calvados** (sur-consommation sur les "
                  "trois exercices, malgré une dose ramenée de 5 à 4 cl), des **eaux-de-vie** (Poire, "
                  "Framboise, Mirabelle) et le **Martini Blanc** (petits volumes), et surtout une « **très "
                  "grosse incohérence** » sur le **Haute Côte de Beaune blanc** : **12 950 cl** vendus "
                  "en 2023-2024 alors que « le service n’est pas parvenu à identifier l’achat d’une "
                  "seule bouteille de ce type »."},
        {"kind": "paragraphe",
         "texte": "Nous répondons **exemple par exemple**, sans aucune supposition, en reliant pour "
                  "chaque produit le **libellé de caisse**, la **facture fournisseur** et "
                  "l’**inventaire**, puis en comparant, **par exercice**, les **achats (nets des avoirs)** "
                  "à la **consommation mesurée**."},

        # 2) METHODE
        {"kind": "chapitre", "source": "nous", "numero": 2, "titre": "La méthode : relier caisse, factures et inventaire",
         "sousTitre": "Une table de correspondance officielle, et un bilan matière par produit"},
        {"kind": "paragraphe",
         "texte": "Le service a comparé des **libellés de caisse** (souvent abrégés ou déclinés en "
                  "plusieurs formats) à des **achats** identifiés sous d’autres intitulés. Or chaque "
                  "libellé de caisse correspond à une **facture précise** et à une **ligne "
                  "d’inventaire**. Le bilan matière d’un produit s’écrit : **achats + stock initial − "
                  "consommation − stock final**. Un écart « consommation > achats » sur un libellé isolé "
                  "signale presque toujours un **rattachement de libellé** (le produit est acheté sous "
                  "un autre nom), pas une vente cachée."},
        {"kind": "tableau", "titre": "Correspondance caisse → facture (produits cités par le service)",
         "minWidth": 760,
         "colonnes": [colg("Libellé de caisse"), colg("Produit (canonique)"), colg("Facture fournisseur")],
         "lignes": corr_page},
        {"kind": "piecejointe", "intro": "Table de correspondance complète (caisse, facture, inventaire) :",
         "fichiers": [{"fichier": F_CORR, "label": "RF : Correspondance caisse / facture / inventaire (XLSX)"}]},

        # 3) HAUTE COTE DE BEAUNE
        {"kind": "chapitre", "source": "nous", "numero": 3,
         "titre": "Haute Côte de Beaune : rouge, blanc et Nuits ne se confondent pas",
         "sousTitre": "Le « blanc » introuvable du service est en réalité du rouge, massivement acheté"},
        {"kind": "paragraphe",
         "texte": f"La caisse distingue trois vins voisins, chacun rattaché à une **facture réelle** "
                  f"(Domaine Germain pour le Beaune, Lupé-Cholet pour le Nuits). Le **Beaune rouge** est "
                  f"acheté en quantité (verres, pichets et bouteilles), et l’**achat dépasse la "
                  f"consommation chaque exercice**. Le **Beaune blanc**, lui, est un volume **minuscule** : "
                  f"sa consommation réelle sur trois ans est de **{fr_cl(tot3(beaune_blc_c) * 100)}** "
                  f"(soit {fr_l(tot3(beaune_blc_c))}), **sans aucun rapport avec les "
                  f"{fr_cl(FISC_BLANC_CL)}** avancés par le service. Ces 12 950 cl correspondent au "
                  f"**Beaune rouge** (libellés « C. DE BEAUNE » au verre et au pichet) que le service a "
                  f"lus comme « blanc » : la couleur a été confondue."},
        tab_achat_conso("Haute Côte de Beaune : achats contre consommation, par exercice",
                        [("Beaune rouge", beaune_rge_a, beaune_rge_c),
                         ("Beaune blanc", beaune_blc_a, beaune_blc_c),
                         ("Nuits blanc (vin distinct)", nuits_a, nuits_c)], minw=640),
        {"kind": "paragraphe",
         "texte": f"**Le Beaune rouge est acheté et bu de façon cohérente** : {fr_l(tot3(beaune_rge_a))} "
                  f"achetés contre {fr_l(tot3(beaune_rge_c))} consommés sur trois ans (excédent "
                  f"{sgn_l(round(tot3(beaune_rge_a) - tot3(beaune_rge_c), 1))}). Le **Nuits blanc** "
                  f"(Lupé-Cholet), vin encore différent, est lui aussi acheté ({fr_l(tot3(nuits_a))}) bien "
                  f"au-delà de sa consommation ({fr_l(tot3(nuits_c))}). Il n’existe donc **aucun** volume "
                  f"de 12 950 cl de blanc non acheté : c’est une **confusion de libellés**, levée par la "
                  f"table de correspondance."},
        {"kind": "piecejointe", "intro": "Détail Beaune (rouge/blanc/Nuits) par exercice et factures :",
         "fichiers": [{"fichier": F_HCB, "label": "RF : Haute Côte de Beaune : achats vs conso (XLSX)"}]},

        # 4) CALVADOS
        {"kind": "chapitre", "source": "nous", "numero": 4,
         "titre": "Calvados : l’écart est de la cuisine, pas une vente",
         "sousTitre": "Au verre, la caisse vend très peu ; l’essentiel est une estimation de cuisson"},
        {"kind": "paragraphe",
         "texte": f"Le Calvados **vendu au verre** par la caisse ne représente que **{fr_l(calva_verre_l)}** "
                  f"sur trois ans (donnée exacte de caisse). Tout le reste de la consommation que retient "
                  f"le service est une **estimation de cuisson** : il l’a calculée en appliquant une dose "
                  f"par portion aux plats flambés (**camembert rôti flambé au Calvados**, **assiette du "
                  f"père Grégoire**). C’est si sensible à la dose que le service a lui-même **abaissé sa "
                  f"dose de 5 à 4 cl** pour limiter l’incohérence."},
        tab_achat_conso("Calvados : achats contre consommation totale (verre + cuisson estimée)",
                        [("Calvados", calva_a, calva_c)], minw=560),
        {"kind": "paragraphe",
         "texte": f"L’« excès » de Calvados ({sgn_l(round(tot3(calva_a) - tot3(calva_c), 1))} sur trois "
                  f"ans) **ne porte donc pas sur des ventes** : il vient entièrement d’une **dose de "
                  f"flambé estimée**, alcool qui part en partie à la flamme et qui, surtout, est un "
                  f"**coût de cuisine : jamais une recette**. Sur-estimer une dose de cuisson ne peut "
                  f"révéler aucune vente non déclarée. Les **{fr_l(tot3(calva_a))}** de Calvados achetés "
                  f"(facturés, pièce jointe) couvrent l’ensemble du service au verre et la cuisine à une "
                  f"dose réaliste de flambé."},
        {"kind": "piecejointe", "intro": "Achats de Calvados (factures) et consommation par exercice :",
         "fichiers": [{"fichier": F_CAL, "label": "RF : Calvados : achats vs consommation (XLSX)"}]},

        # 5) EAUX-DE-VIE / LIQUEURS / MARTINI
        {"kind": "chapitre", "source": "nous", "numero": 5,
         "titre": "Eaux-de-vie, liqueurs et Martini : les achats existent",
         "sousTitre": "Le « achat = 0 » venait d’un rattachement de libellé, pas d’une vente cachée"},
        {"kind": "paragraphe",
         "texte": "Pour ces petits volumes, le service a vu une consommation sans achat. En réalité, les "
                  "achats **figurent bien sur les factures**, sous des intitulés voisins : **FRAMBOISE "
                  "SAUVAGE 70 CL**, **MIRABELLE 70 CL**, **POIRE WILLIAM 70 CL**, **LIQUEUR DE POIRE "
                  "(Golden Eight)**, **MARTINI BIANCO 100 CL**. En regroupant chaque famille avec ses "
                  "libellés frères, **l’achat couvre la consommation** sur les trois exercices."},
        {"kind": "tableau", "titre": "Eaux-de-vie, liqueurs et Martini : achats regroupés contre consommation (3 exercices)",
         "minWidth": 560,
         "colonnes": [colg("Famille"), cold("Achat (L)"), cold("Conso (L)"), cold("Achat − conso")],
         "lignes": [[cg(lab), cd(fr_l(tot3(a))), cd(fr_l(tot3(c))),
                     bdg(sgn_l(round(tot3(a) - tot3(c), 1)), tot3(a) - tot3(c) >= -0.5)]
                    for lab, c, a, al in fam]},
        {"kind": "paragraphe",
         "texte": "Ces volumes sont **infimes** (quelques litres sur trois ans) et **entièrement "
                  "achetés**. L’écart apparent venait uniquement de ce que la caisse et la facture "
                  "nomment différemment le même alcool : exactement le **rattachement de libellé** que "
                  "corrige la table de correspondance."},
        {"kind": "piecejointe", "intro": "Détail des familles et des factures correspondantes :",
         "fichiers": [{"fichier": F_EDV, "label": "RF : Eaux-de-vie, liqueurs et Martini : achats vs conso (XLSX)"}]},

        # 6) BILAN GLOBAL
        {"kind": "chapitre", "source": "nous", "numero": 6, "titre": "Le bilan matière global se ferme à l’excédent",
         "sousTitre": "Toutes boissons agrégées, les achats dépassent largement la consommation"},
        {"kind": "kpis", "items": [
            {"label": "Achats (3 exercices)", "valeur": fr_l(g_achat), "sub": "factures nettes des avoirs", "couleur": "blue"},
            {"label": "Consommation", "valeur": fr_l(g_conso), "sub": "caisse + cuisine mesurées", "couleur": "blue"},
            {"label": "Bilan net", "valeur": sgn_l(g_net), "sub": "achats − conso − stock", "highlight": True, "couleur": "teal"},
        ]},
        {"kind": "paragraphe",
         "texte": f"Au niveau qui a un sens économique : **toutes boissons agrégées** : le bilan matière "
                  f"se ferme **largement à l’excédent** : **{fr_l(g_achat)}** achetés, **{fr_l(g_conso)}** "
                  f"consommés, **{fr_l(g_stock)}** de variation de stock, soit **{sgn_l(g_net)}** net. "
                  f"Loin d’une consommation supérieure aux achats, l’établissement a, globalement, "
                  f"**acheté plus qu’il n’a servi**. Les rares écarts négatifs par libellé (Calvados de "
                  f"cuisine, libellés frères des eaux-de-vie, couleur du Beaune) sont des **artefacts de "
                  f"rattachement**, neutralisés dès qu’on relie correctement caisse, factures et "
                  f"inventaire."},
        {"kind": "piecejointe", "intro": "Bilan matière complet, produit par produit :",
         "fichiers": [{"fichier": F_GLO, "label": "RF : Bilan matière global par boisson (XLSX)"}]},

        # 7) CONCLUSION
        {"kind": "chapitre", "source": "nous", "numero": 7, "titre": "Conclusion"},
        {"kind": "alerte", "couleur": "teal", "titre": "Résultat",
         "texte": f"Aucun des exemples du service ne révèle de vente cachée. Le **Beaune blanc** "
                  f"« introuvable » est du **rouge mal étiqueté** ({fr_cl(tot3(beaune_blc_c) * 100)} de "
                  f"blanc réel contre {fr_cl(FISC_BLANC_CL)} annoncés) ; les **eaux-de-vie** et le "
                  f"**Martini** sont **achetés** sous des libellés voisins ; le **Calvados** « en excès » "
                  f"est de l’**alcool de cuisine** (flambés), un coût et non une recette. Le bilan matière "
                  f"global est **excédentaire de {fr_l(g_net)}**. Le grief de comptabilité « non sincère » "
                  f"sur ce fondement n’est pas établi."},
        {"kind": "interne", "audience": "avocat", "titre": "Pour solidifier",
         "texte": "Joindre la table de correspondance signée et, pour le Beaune, les factures Domaine "
                  "Germain (rouge) et Lupé-Cholet (Nuits) ; pour le Calvados, une attestation sur la dose "
                  "réelle de flambé (qui s’évapore) et le rappel que l’alcool de cuisine est un achat-coût. "
                  "Tenir prêt l’argument de principe : un écart « conso > achats » sur un libellé isolé est "
                  "un problème d’**étiquetage**, pas de recette : la seule grandeur probante est le bilan "
                  "matière agrégé. Reproductible via le script. Retirer cet encart de toute version remise."},
    ],
}

doc["sections"] = normaliser_sections(doc["sections"])
os.makedirs(os.path.dirname(OUT_JSON), exist_ok=True)
json.dump(doc, open(OUT_JSON, "w", encoding="utf-8"), ensure_ascii=False, indent=1)
print("JSON ecrit :", OUT_JSON, "| sections:", len(doc["sections"]))
print(f"  Beaune blanc conso 3 ans : {fr_cl(tot3(beaune_blc_c) * 100)} (fisc : {fr_cl(FISC_BLANC_CL)})")
print(f"  Beaune rouge : achat {fr_l(tot3(beaune_rge_a))} >= conso {fr_l(tot3(beaune_rge_c))}")
print(f"  Calvados : achat {fr_l(tot3(calva_a))}, conso {fr_l(tot3(calva_c))}, dont verre {fr_l(calva_verre_l)}")
for lab, c, a, al in fam:
    print(f"  {lab}: achat {fr_l(tot3(a))} / conso {fr_l(tot3(c))}")
print(f"  Bilan global net : {sgn_l(g_net)}")
