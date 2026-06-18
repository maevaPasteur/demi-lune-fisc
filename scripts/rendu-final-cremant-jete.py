#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
RENDU FINAL - Crémant jeté en fin de journée (reste de bouteille perdu).

Demande de la cliente : recalculer, au NIVEAU JOURNALIER et depuis le détail
certifié des tickets de caisse, le crémant jeté chaque soir. Une bouteille
d'effervescent ouverte ne se garde pas au lendemain (elle devient plate) : le
reste de la dernière bouteille ouverte chaque jour est perdu.

Le crémant est servi DE DEUX façons :
  - au verre (10 cl) ;
  - dans des cocktails (La Vouivre 8 cl, Kittykir 9 cl, Kir Princier 10 cl,
    Apéritif du Père Grégoire 1 cl).
Les bouteilles vendues entières (Crémant du Jura, Crémant Rosé...) sont EXCLUES :
une bouteille vendue scellée n'est pas ouverte puis partiellement jetée.

Deux scénarios de sur-versement :
  - +23.6 % (OVERPOUR = 1.236), sur-versement mesuré du vin d'après
    Kerr, Patterson, Koenen & Greenfield, "Alcohol Content Variation of Bar and
    Restaurant Drinks in Northern California", Alcoholism: Clinical and
    Experimental Research, 2008 ;
  - nominal (OVERPOUR = 1.0) pour comparaison.

Le sur-versement sur le crémant (volume servi en plus) est une perte DISTINCTE
du reste jeté : les deux sont réels et indépendants.

Source (LECTURE SEULE, reproductible) :
  public/documents/caisse-enregistreuse/ANNEXE-C{1,2,3}_detail-tickets_{exo}.xls
  col 0 = date ticket ("YYYY-MM-DD..."), col 2 = n° note, col 10 = libellé,
  col 11 = quantité (float), col 13 = prix unitaire.
Sorties :
  src/data/renduFinal/cremant-jete.json
  public/documents/pieces-defense/RF-cremant-jete.xlsx
"""
import os
import json
import math
import collections
import xlrd

ICI = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.normpath(os.path.join(ICI, ".."))
CAISSE = os.path.join(ROOT, "public/documents/caisse-enregistreuse/")
PIECES = os.path.join(ROOT, "public/documents/pieces-defense/")
RENDU = os.path.join(ROOT, "src/data/renduFinal/")
EXOS = ["2022-2023", "2023-2024", "2024-2025"]
C = {e: CAISSE + f"ANNEXE-C{i}_detail-tickets_{e}.xls" for i, e in enumerate(EXOS, 1)}
JSON_OUT = os.path.join(RENDU, "cremant-jete.json")
XLSX_OUT = os.path.join(PIECES, "RF-cremant-jete.xlsx")

BOTTLE = 75.0
OVERPOUR = 1.236  # +23.6 % (Kerr et al. 2008)

# --- Libellé caisse -> dose de crémant (cl) -------------------------------- #
# Verre : 10 cl ; cocktails au crémant : Vouivre 8, Kittykir 9, Kir Princier 10,
# Apéritif du Père Grégoire 1 cl (la cliente insiste pour l'inclure).
# Les libellés "Crémant du Jura", "Crémant Rosé", "Crément du Jura" sont des
# BOUTEILLES vendues entières -> EXCLUS (jamais dans cette map).
CREMANT_DOSES = {
    # verre de crémant
    "Crémant": 10,
    "Crément / Jura VERRE": 10,
    # cocktail La Vouivre
    "La Vouivre": 8,
    "VOUIVRE": 8,
    # cocktail Kittykir
    "KITTYKIR": 9,
    "Kittykir": 9,
    # cocktail Kir Princier
    "Kir Princier": 10,
    # Apéritif du Père Grégoire (libellé caisse : "Père Gregoire")
    "Père Gregoire": 1,
    "Père Grégoire": 1,
    "Apéritif du Père Grégoire": 1,
}

# Libellés explicitement EXCLUS (bouteilles vendues entières), pour la trace.
EXCLUS_BOUTEILLES = ["Crémant du Jura", "Crément du Jura", "Crémant Rosé"]


def sheet(fn):
    return xlrd.open_workbook(fn).sheet_by_index(0)


# --------------------------------------------------------------------------- #
# 0. SCAN : tout libellé évoquant crémant / cocktails au crémant (contrôle).
# --------------------------------------------------------------------------- #
SCAN_KEYS = ["rémant", "rement", "crémant", "cremant", "vouivre", "kitty",
             "kir", "grégoire", "gregoire", "gregoir"]


def scan_libelles():
    cnt = collections.Counter()
    qte = collections.defaultdict(float)
    for ex, fn in C.items():
        sh = sheet(fn)
        for r in range(1, sh.nrows):
            lib = str(sh.cell_value(r, 10)).strip()
            if any(k in lib.lower() for k in SCAN_KEYS):
                try:
                    q = float(sh.cell_value(r, 11))
                except ValueError:
                    q = 0.0
                cnt[lib] += 1
                qte[lib] += q
    return [{"libelle": lib, "lignes": c, "qte": round(qte[lib], 1),
             "dose_cl": CREMANT_DOSES.get(lib),
             "retenu": lib in CREMANT_DOSES}
            for lib, c in cnt.most_common()]


# --------------------------------------------------------------------------- #
# 1. CRÉMANT JETÉ — niveau journalier, par scénario de sur-versement.
# --------------------------------------------------------------------------- #
def cremant_jete(overpour):
    """Retourne (par_exercice, qte_par_libelle, consomme_total_cl)."""
    per = {}
    qte_lib = collections.defaultdict(float)  # quantité 3 ans par libellé retenu
    consomme_total = 0.0
    for ex, fn in C.items():
        sh = sheet(fn)
        day = collections.defaultdict(float)  # date -> crémant consommé (cl)
        for r in range(1, sh.nrows):
            lib = str(sh.cell_value(r, 10)).strip()
            dose = CREMANT_DOSES.get(lib)
            if dose is None:
                continue
            try:
                q = float(sh.cell_value(r, 11))
            except ValueError:
                q = 0.0
            if q > 0:
                day[str(sh.cell_value(r, 0))[:10]] += q * dose * overpour
                qte_lib[lib] += q
        served = sum(day.values())
        bottles = sum(math.ceil(v / BOTTLE) for v in day.values() if v > 0)
        thrown = bottles * BOTTLE - served
        consomme_total += served
        per[ex] = {
            "jours": sum(1 for v in day.values() if v > 0),
            "servi_l": round(served / 100, 1),
            "bouteilles": bottles,
            "jete_l": round(thrown / 100, 1),
        }
    return per, qte_lib, consomme_total


def totaux(per):
    return {
        "total_servi_l": round(sum(v["servi_l"] for v in per.values()), 1),
        "total_jete_l": round(sum(v["jete_l"] for v in per.values()), 1),
        "total_bouteilles": sum(v["bouteilles"] for v in per.values()),
    }


# --------------------------------------------------------------------------- #
# 2. SORTIE XLSX (pièce de défense).
# --------------------------------------------------------------------------- #
def fr_num(x, dec=1):
    return f"{x:,.{dec}f}".replace(",", " ").replace(".", ",")


def ecrire_xlsx(per_236, tot_236, qte_lib_236, surv_l):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    white = Font(bold=True, color="FFFFFF")
    bold = Font(bold=True)
    head = PatternFill("solid", fgColor="0F766E")  # teal
    sub = PatternFill("solid", fgColor="E6F4F1")
    tot_fill = PatternFill("solid", fgColor="CCE7E2")
    thin = Side(style="thin", color="D8DEE4")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    right = Alignment(horizontal="right")

    # ---- Feuille 1 : synthèse par exercice -------------------------------
    ws = wb.active
    ws.title = "Crémant jeté"
    ws.append(["Crémant jeté en fin de journée - reste de bouteille perdu (sur-versement +23,6 %)"])
    ws["A1"].font = Font(bold=True, size=13)
    ws.append(["Une bouteille effervescente ouverte ne se garde pas au lendemain : "
               "le reste de la dernière bouteille ouverte chaque jour est jeté. "
               "Crémant servi au verre (10 cl) + cocktails (Vouivre 8, Kittykir 9, "
               "Kir Princier 10, Père Grégoire 1). Sur-versement +23,6 % (Kerr et al., 2008). Bouteille = 75 cl."])
    ws["A2"].font = Font(italic=True, size=9, color="64748B")
    ws.append([])
    hrow = 4
    cols = ["Exercice", "Jours servis", "Crémant servi (L)",
            "Bouteilles ouvertes", "Crémant jeté (L)"]
    ws.append(cols)
    for c in range(1, len(cols) + 1):
        cell = ws.cell(row=hrow, column=c)
        cell.font = white
        cell.fill = head
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    for ex in EXOS:
        v = per_236[ex]
        ws.append([ex, v["jours"], v["servi_l"], v["bouteilles"], v["jete_l"]])
        r = ws.max_row
        for c in range(1, len(cols) + 1):
            ws.cell(row=r, column=c).border = border
        ws.cell(row=r, column=3).number_format = '#,##0.0 "L"'
        ws.cell(row=r, column=5).number_format = '#,##0.0 "L"'
        for c in (2, 3, 4, 5):
            ws.cell(row=r, column=c).alignment = right

    # ligne TOTAL
    ws.append(["TOTAL (3 exercices)",
               sum(per_236[e]["jours"] for e in EXOS),
               tot_236["total_servi_l"], tot_236["total_bouteilles"],
               tot_236["total_jete_l"]])
    r = ws.max_row
    for c in range(1, len(cols) + 1):
        ws.cell(row=r, column=c).font = bold
        ws.cell(row=r, column=c).fill = tot_fill
        ws.cell(row=r, column=c).border = border
    ws.cell(row=r, column=3).number_format = '#,##0.0 "L"'
    ws.cell(row=r, column=5).number_format = '#,##0.0 "L"'
    for c in (2, 3, 4, 5):
        ws.cell(row=r, column=c).alignment = right

    # ligne sur-versement (perte distincte)
    ws.append([])
    rr = ws.max_row + 1
    ws.cell(row=rr, column=1, value="Sur-versement crémant (perte distincte, +23,6 %)")
    ws.cell(row=rr, column=5, value=round(surv_l, 1))
    ws.cell(row=rr, column=5).number_format = '#,##0.0 "L"'
    for c in range(1, len(cols) + 1):
        ws.cell(row=rr, column=c).font = bold
        ws.cell(row=rr, column=c).fill = sub
    ws.cell(row=rr, column=5).alignment = right

    widths = [40, 14, 18, 20, 16]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A5"

    # ---- Feuille 2 : libellés retenus + doses ----------------------------
    ws2 = wb.create_sheet("Libellés crémant")
    ws2.append(["Libellés crémant retenus - dose et quantité vendue (3 exercices)"])
    ws2["A1"].font = Font(bold=True, size=13)
    ws2.append(["Bouteilles vendues entières exclues : " + ", ".join(EXCLUS_BOUTEILLES)])
    ws2["A2"].font = Font(italic=True, size=9, color="64748B")
    ws2.append([])
    cols2 = ["Libellé caisse", "Dose crémant (cl)", "Quantité vendue (3 ans)"]
    ws2.append(cols2)
    for c in range(1, len(cols2) + 1):
        cell = ws2.cell(row=4, column=c)
        cell.font = white
        cell.fill = head
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    # n'affiche que les libellés effectivement vendus
    for lib in sorted(qte_lib_236, key=lambda k: -qte_lib_236[k]):
        ws2.append([lib, CREMANT_DOSES[lib], round(qte_lib_236[lib], 1)])
        r = ws2.max_row
        for c in range(1, len(cols2) + 1):
            ws2.cell(row=r, column=c).border = border
        ws2.cell(row=r, column=2).alignment = right
        ws2.cell(row=r, column=3).alignment = right
    for i, w in enumerate([28, 18, 22], 1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    os.makedirs(PIECES, exist_ok=True)
    wb.save(XLSX_OUT)


# --------------------------------------------------------------------------- #
def main():
    scan = scan_libelles()
    per_236, qte_lib_236, cons_236 = cremant_jete(OVERPOUR)
    per_nom, qte_lib_nom, cons_nom = cremant_jete(1.0)
    tot_236 = totaux(per_236)
    tot_nom = totaux(per_nom)
    surv_l = round((cons_236 - cons_nom) / 100, 1)  # sur-versement crémant (L)

    # doses réellement utilisées (libellés vendus)
    doses_used = {lib: CREMANT_DOSES[lib] for lib in qte_lib_236}

    doc = {
        "_source": "scripts/rendu-final-cremant-jete.py (ANNEXE-C, niveau journalier)",
        "methode": (
            "Crémant jeté = reste de la dernière bouteille (75 cl) ouverte chaque jour, "
            "perdu à la fermeture (un effervescent ouvert ne se garde pas). "
            "Par jour et par libellé : crémant consommé += quantité x dose x sur-versement ; "
            "bouteilles ouvertes = arrondi supérieur(consommé/75) ; "
            "jeté = bouteilles x 75 - consommé. Servi au verre (10 cl) et en cocktails "
            "(Vouivre 8, Kittykir 9, Kir Princier 10, Apéritif du Père Grégoire 1). "
            "Bouteilles vendues entières exclues. Sur-versement crémant = consommé(1,236) - consommé(1,0), "
            "perte distincte du reste jeté."),
        "overpour": OVERPOUR,
        "overpour_source": ("Kerr, Patterson, Koenen & Greenfield (2008), Alcohol Content "
                            "Variation of Bar and Restaurant Drinks in Northern California, "
                            "Alcoholism: Clinical and Experimental Research (+23,6 %)."),
        "bottle_cl": BOTTLE,
        "doses": doses_used,
        "libelles_exclus_bouteilles": EXCLUS_BOUTEILLES,
        "scan_libelles": scan,
        "par_exercice": per_236,
        "total_servi_l": tot_236["total_servi_l"],
        "total_jete_l": tot_236["total_jete_l"],
        "total_bouteilles": tot_236["total_bouteilles"],
        "nominal": {
            "par_exercice": per_nom,
            "total_servi_l": tot_nom["total_servi_l"],
            "total_jete_l": tot_nom["total_jete_l"],
            "total_bouteilles": tot_nom["total_bouteilles"],
        },
        "surversement_cremant_l": surv_l,
        "par_libelle": {lib: {"qte_3ans": round(qte_lib_236[lib], 1),
                              "dose_cl": CREMANT_DOSES[lib]}
                        for lib in sorted(qte_lib_236, key=lambda k: -qte_lib_236[k])},
    }
    os.makedirs(RENDU, exist_ok=True)
    with open(JSON_OUT, "w", encoding="utf-8") as f:
        json.dump(doc, f, ensure_ascii=False, indent=1)

    ecrire_xlsx(per_236, tot_236, qte_lib_236, surv_l)

    # ---- Résumé console (français) ---------------------------------------
    print("RENDU FINAL - Crémant jeté en fin de journée")
    print("=" * 62)
    print("\nScan des libellés (crémant / cocktails au crémant) :")
    for s in scan:
        flag = "RETENU dose=%s cl" % s["dose_cl"] if s["retenu"] else "exclu"
        print(f"  {s['lignes']:5d} lignes  qte={fr_num(s['qte'])}  | {s['libelle']!r}  [{flag}]")

    print("\nDoses retenues (cl de crémant par unité) :")
    for lib in sorted(qte_lib_236, key=lambda k: -qte_lib_236[k]):
        print(f"  {CREMANT_DOSES[lib]:3d} cl  | {lib}")
    print("  (exclus, bouteilles entières : " + ", ".join(EXCLUS_BOUTEILLES) + ")")

    for nom, per, tot in [("+23,6 % (Kerr et al. 2008)", per_236, tot_236),
                          ("nominal (1,0)", per_nom, tot_nom)]:
        print(f"\nScénario {nom} :")
        print(f"  {'Exercice':<12}{'Jours':>8}{'Servi (L)':>12}{'Bouteilles':>12}{'Jeté (L)':>12}")
        for ex in EXOS:
            v = per[ex]
            print(f"  {ex:<12}{v['jours']:>8}{fr_num(v['servi_l']):>12}"
                  f"{v['bouteilles']:>12}{fr_num(v['jete_l']):>12}")
        print(f"  {'TOTAL':<12}{sum(per[e]['jours'] for e in EXOS):>8}"
              f"{fr_num(tot['total_servi_l']):>12}{tot['total_bouteilles']:>12}"
              f"{fr_num(tot['total_jete_l']):>12}")

    print(f"\nSur-versement crémant (perte distincte) : {fr_num(surv_l)} L")
    print("-" * 62)
    print("JSON :", JSON_OUT)
    print("XLSX :", XLSX_OUT)


if __name__ == "__main__":
    main()
