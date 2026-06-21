#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
rendu-final-tables-trop-nombreuses.py

Grief du fisc (Proposition p. 11 + section V « suppression de note », p. 18-19) :
le logiciel comporte des tables « virtuelles » (n° de note ouverts pour le règlement
séparé) ET autorise la suppression pure et simple d'une table et de tous ses articles.
Le service en conclut que les modifications ne sont pas vérifiables -> comptabilité
non probante (vecteur d'occultation supposé).

Réponse construite, fichiers à l'appui :
  (A) LE NUMÉRO N'EST PAS UNE PLACE. La salle réelle compte 14 tables intérieures
      (n° 1 à 12, 14, 15) ; les n° 13, 16, 17, 18 ne désignent AUCUNE table physique.
      Le numéro de note de la caisse va de 1 à 57 : c'est un rang de note quotidien,
      jamais un numéro de table (les numéros de terrasse 101-125 n'apparaissent JAMAIS
      dans l'export). On ne peut donc pas déduire des couverts d'un numéro.
  (B) ENCAISSEMENTS PAR NUMÉRO DE NOTE. Pour chaque rang de note (1 à 57), nombre
      d'encaissements, par saison et par exercice. La queue (rangs élevés) ne survient
      que les jours de pointe et correspond aux notes de règlement séparé.
  (C) LE MÉCANISME ET SON RETOURNEMENT. Un règlement = une note. Pour un paiement
      séparé, on ressaisit les articles sur une 2e note et on supprime la table
      d'origine : cette suppression est ARITHMÉTIQUEMENT NÉCESSAIRE pour ne pas
      DOUBLER le CA. Elle protège le CA, elle ne le réduit pas. Signature mécanique :
      1 906 notes (11,5 %) à quantité fractionnaire ; exemple daté de table scindée.

Le mécanisme est décrit par le vérificateur lui-même (synthèse 02-rejet-comptabilite-1.md,
p. 11 : « quatre tables virtuelles figurent à l'écran (n° 15 à n° 18, PAR EXEMPLE).
Cette modalité permet le traitement des notes multiples pour une table, au moment du
règlement. »).

Lecture SEULE des sources. Sorties :
  - public/documents/pieces-defense/RF-tables-trop-nombreuses.xlsx
  - src/data/renduFinal/tables-trop-nombreuses.json

Python : /tmp/xlsenv/bin/python (xlrd + openpyxl).
"""
import os
import re
import json
import datetime
import collections

import xlrd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from rfcommun import normaliser_sections

ICI = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.normpath(os.path.join(ICI, ".."))
CAISSE = os.path.join(ROOT, "public/documents/caisse-enregistreuse/")
CALCULS = os.path.join(ROOT, "src/data/renduFinalCalculs.json")
CASSUP = os.path.join(ROOT, "src/data/renduFinalCasSuppressions.json")
OUT_XLSX = os.path.join(ROOT, "public/documents/pieces-defense/RF-tables-trop-nombreuses.xlsx")
OUT_JSON = os.path.join(ROOT, "src/data/renduFinal/tables-trop-nombreuses.json")
# Section graphique des couverts/semaine, produite par rendu-final-couverts.py
HEBDO_JSON = os.path.join(ROOT, "src/data/renduFinal/couverts-hebdomadaires.json")

EXOS = ["2022-2023", "2023-2024", "2024-2025"]
C = {e: CAISSE + f"ANNEXE-C{i}_detail-tickets_{e}.xls" for i, e in enumerate(EXOS, 1)}
F = {e: CAISSE + f"ANNEXE-F{i}_reglements_{e}.xls" for i, e in enumerate(EXOS, 1)}

# Plan de salle RÉEL (confirmé par la gérante) : 14 tables intérieures seulement.
# Les numéros 13, 16, 17, 18 ne désignent AUCUNE table physique.
INT_REELLES = [1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 14, 15]
INT_INEXISTANTES = [13, 16, 17, 18]
NB_INT_REELLES = len(INT_REELLES)  # 14
SAISONS = ["Printemps", "Été", "Automne", "Hiver"]
# Une couleur par exercice (mêmes teintes que le graphique couverts/semaine).
COULEURS_EXO = {"2022-2023": "#0f766e", "2023-2024": "#b45309", "2024-2025": "#1d4ed8"}
# Étiquettes de mois pour l'axe d'un exercice (avril -> mars).
MOIS_LABELS = [(4, "Avr"), (5, "Mai"), (6, "Juin"), (7, "Juil"), (8, "Août"),
               (9, "Sep"), (10, "Oct"), (11, "Nov"), (12, "Déc"),
               (1, "Jan"), (2, "Fév"), (3, "Mars")]
MOIS_SAISON = {3: "Printemps", 4: "Printemps", 5: "Printemps",
               6: "Été", 7: "Été", 8: "Été",
               9: "Automne", 10: "Automne", 11: "Automne",
               12: "Hiver", 1: "Hiver", 2: "Hiver"}

DATE_RE = re.compile(r"\d{4}-\d{2}-\d{2}")


def euro(v):
    return f"{v:,.2f}".replace(",", " ").replace(".", ",") + " €"


def fr_int(v):
    return f"{int(round(v)):,}".replace(",", " ")


# ---------------------------------------------------------------------------
# 1) Notes fractionnaires (fichier C) : notes contenant >= 1 article à qté non entière
# ---------------------------------------------------------------------------
def notes_fractionnaires(ex):
    sh = xlrd.open_workbook(C[ex]).sheet_by_index(0)
    frac = set()
    for r in range(1, sh.nrows):
        date = str(sh.cell_value(r, 0))[:10]
        if not DATE_RE.fullmatch(date):
            continue
        no = str(sh.cell_value(r, 2)).replace(".0", "")
        try:
            q = float(sh.cell_value(r, 11))
        except (ValueError, TypeError):
            q = 0.0
        if q and abs(q - round(q)) > 1e-9:
            frac.add((date, no))
    return frac


# ---------------------------------------------------------------------------
# 2) Encaissements par NUMÉRO DE NOTE (fichier F = règlements)
#    Un encaissement = une note distincte = (date, no_ticket).
#    Le no_ticket est un rang quotidien (1..57), réattribué chaque jour : seule
#    granularité présente dans l'export (le n° de table physique n'y est jamais).
# ---------------------------------------------------------------------------
def encaissements():
    """
    Retourne :
      par_note[no] = {ex: count, "total": count}
      matrice[no][(ex, saison)] = count
      totaux_ys[(ex, saison)] = count
      total_global = int
    """
    par_note = collections.defaultdict(lambda: {e: 0 for e in EXOS})
    matrice = collections.defaultdict(lambda: collections.defaultdict(int))
    totaux_ys = collections.defaultdict(int)
    total_global = 0
    for ex in EXOS:
        sh = xlrd.open_workbook(F[ex]).sheet_by_index(0)
        vues = set()
        for r in range(1, sh.nrows):
            date = str(sh.cell_value(r, 0))[:10]
            nt = sh.cell_value(r, 2)
            if not DATE_RE.fullmatch(date) or not isinstance(nt, float):
                continue
            key = (date, int(nt))
            if key in vues:
                continue
            vues.add(key)
            no = int(nt)
            mois = int(date[5:7])
            sa = MOIS_SAISON[mois]
            par_note[no][ex] += 1
            matrice[no][(ex, sa)] += 1
            totaux_ys[(ex, sa)] += 1
            total_global += 1
    out = {}
    for k in sorted(par_note):
        row = dict(par_note[k])
        row["total"] = sum(par_note[k][e] for e in EXOS)
        out[k] = row
    return out, matrice, totaux_ys, total_global


# ---------------------------------------------------------------------------
# 2 bis) Encaissements par JOUR (fichier F = règlements).
#   Un encaissement = une note distincte = (date, no_ticket). Pour chaque date
#   d'ouverture, on compte le nombre de notes distinctes encaissées ce jour-là.
#   Sert au graphique « encaissements par jour travaillé, moyenne par semaine ».
# ---------------------------------------------------------------------------
def encaissements_par_jour():
    """Retourne {exercice: {date 'YYYY-MM-DD': nb_encaissements}}."""
    par_jour = {e: collections.defaultdict(int) for e in EXOS}
    for ex in EXOS:
        sh = xlrd.open_workbook(F[ex]).sheet_by_index(0)
        vues = set()
        for r in range(1, sh.nrows):
            date = str(sh.cell_value(r, 0))[:10]
            nt = sh.cell_value(r, 2)
            if not DATE_RE.fullmatch(date) or not isinstance(nt, float):
                continue
            key = (date, int(nt))
            if key in vues:
                continue
            vues.add(key)
            par_jour[ex][date] += 1
    return par_jour


def _semaine_exercice(date_str, an_debut):
    """Index de semaine 1..52 dans l'exercice (1er avril de an_debut -> mars)."""
    d = datetime.date(int(date_str[:4]), int(date_str[5:7]), int(date_str[8:10]))
    sem = (d - datetime.date(an_debut, 4, 1)).days // 7 + 1
    return max(1, min(52, sem))


def graphique_encaissements_hebdo(par_jour):
    """Section graphiqueLignes : nombre MOYEN d'encaissements par jour travaillé,
    pour chaque semaine de l'exercice (avril -> mars), une ligne par exercice.
    Retourne (section, somme_par_semaine, jours_par_semaine) pour la pièce XLSX."""
    NB_SEM = 52
    somme = {e: collections.defaultdict(int) for e in EXOS}   # encaissements
    jours = {e: collections.defaultdict(int) for e in EXOS}   # jours travaillés
    for e in EXOS:
        an = int(e[:4])
        for date, cnt in par_jour[e].items():
            sem = _semaine_exercice(date, an)
            somme[e][sem] += cnt
            jours[e][sem] += 1

    # Période calendaire représentative de chaque semaine d'exercice (base
    # 1er avril, comme l'axe des mois) : « 28/10 - 03/11 ».
    base = datetime.date(2022, 4, 1)

    def periode_label(sem):
        debut = base + datetime.timedelta(days=(sem - 1) * 7)
        fin = debut + datetime.timedelta(days=6)
        return f"{debut.day:02d}/{debut.month:02d} - {fin.day:02d}/{fin.month:02d}"

    data = []
    for sem in range(1, NB_SEM + 1):
        ligne = {"semaine": sem, "periode": periode_label(sem)}
        for e in EXOS:
            j = jours[e].get(sem, 0)
            ligne[e] = int(round(somme[e][sem] / j)) if j else 0
        data.append(ligne)

    mois_ticks = []
    for m, lib in MOIS_LABELS:
        an = 2022 if m >= 4 else 2023
        sem = (datetime.date(an, m, 1) - datetime.date(2022, 4, 1)).days // 7 + 1
        mois_ticks.append({"tick": max(1, min(NB_SEM, sem)), "label": lib})

    section = {
        "kind": "graphiqueLignes",
        "titre": "Encaissements par jour travaillé, moyenne par semaine (avril → mars)",
        "sousTitre": ("Une ligne par exercice. Pour chaque semaine, nombre moyen de notes encaissées "
                      "par jour d'ouverture (total des encaissements de la semaine ÷ nombre de jours "
                      "travaillés). Les trois années se superposent : même montée au printemps, même "
                      "pic estival, même creux hivernal (fermeture janvier-février)."),
        "hauteur": 340,
        "dataKey": "semaine",
        "series": [{"name": e, "couleur": COULEURS_EXO[e]} for e in EXOS],
        "data": data,
        "format": "int",
        "moisTicks": mois_ticks,
        "tooltipSousTitre": "Moyenne de tables encaissées cette semaine",
    }
    return section, somme, jours


# ---------------------------------------------------------------------------
# XLSX dédié : toutes les données du graphique (pour le fisc).
#   Onglet 1 : moyenne/semaine (ce qui est tracé) + jours travaillés.
#   Onglet 2 : détail par jour (date, exercice, semaine, nb encaissements).
# ---------------------------------------------------------------------------
OUT_XLSX_ENC = os.path.join(ROOT, "public/documents/pieces-defense/RF-encaissements-par-jour.xlsx")


def construire_xlsx_encaissements(par_jour, somme, jours):
    wb = openpyxl.Workbook()

    # --- Onglet 1 : moyenne par semaine (données tracées) ----------------
    ws = wb.active
    ws.title = "Moyenne par semaine"
    ws.append(["Encaissements par jour travaillé : moyenne par semaine d'exercice (avril → mars)"])
    ws["A1"].font = GRAS
    ws.append([("Pour chaque semaine et chaque exercice : total des encaissements (notes distinctes), "
                "nombre de jours travaillés, et moyenne d'encaissements par jour travaillé (= la "
                "valeur tracée sur le graphique).")])
    ws["A2"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells("A2:J2")
    ws.row_dimensions[2].height = 45
    ws.append([])
    head = ["Semaine"]
    for e in EXOS:
        head += [f"{e} total", f"{e} jours", f"{e} moy/jour"]
    h = ws.max_row + 1
    ws.append(head)
    styler_entete(ws, len(head), h)
    for sem in range(1, 53):
        r = ws.max_row + 1
        ligne = [sem]
        for e in EXOS:
            j = jours[e].get(sem, 0)
            tot = somme[e].get(sem, 0)
            moy = round(tot / j, 1) if j else 0
            ligne += [tot, j, moy]
        ws.append(ligne)
        for c in range(1, len(head) + 1):
            ws.cell(row=r, column=c).border = BORD
    largeurs(ws, [9] + [11, 9, 11] * len(EXOS))
    ws.freeze_panes = "B" + str(h + 1)

    # --- Onglet 2 : détail par jour --------------------------------------
    ws2 = wb.create_sheet("Détail par jour")
    ws2.append(["Encaissements par jour (notes distinctes encaissées) : source annexes F"])
    ws2["A1"].font = GRAS
    ws2.append([])
    h2 = ws2.max_row + 1
    ws2.append(["Date", "Exercice", "Semaine d'exercice", "Nb encaissements"])
    styler_entete(ws2, 4, h2)
    for e in EXOS:
        an = int(e[:4])
        for date in sorted(par_jour[e]):
            r = ws2.max_row + 1
            ws2.append([date, e, _semaine_exercice(date, an), par_jour[e][date]])
            for c in range(1, 5):
                ws2.cell(row=r, column=c).border = BORD
    largeurs(ws2, [14, 14, 18, 18])
    ws2.freeze_panes = "A" + str(h2 + 1)

    os.makedirs(os.path.dirname(OUT_XLSX_ENC), exist_ok=True)
    wb.save(OUT_XLSX_ENC)
    return OUT_XLSX_ENC


# ---------------------------------------------------------------------------
# 3) Données déjà mesurées (paiements fractionnés / bancarisation)
# ---------------------------------------------------------------------------
CALC = json.load(open(CALCULS, encoding="utf-8"))
BANC = CALC["bancarisation"]
PF = CALC["paiements_fractionnes"]

NOTES_TOTAL = BANC["notes_total"]                 # 16608
NOTES_FRAC = BANC["notes_fractionnees"]            # 1906
NOTES_FRAC_PCT = BANC["notes_fractionnees_pct"]    # 11.5
BANC_PCT = BANC["bancarise_pct"]                   # 98.64
ESP_PCT = BANC["especes_pct"]                      # 1.36
PARTAGES_3ANS_EST = 880


# ---------------------------------------------------------------------------
# 4) Exemple daté de table réelle scindée (cas "partage")
# ---------------------------------------------------------------------------
CAS = json.load(open(CASSUP, encoding="utf-8"))
PARTAGE = next(c for c in CAS["cas"] if c["id"] == "partage")
EX = next(e for e in PARTAGE["exemples"] if e["date"] == "2022-05-14")
NB_EXEMPLES = len(PARTAGE["exemples"])


# ---------------------------------------------------------------------------
# 5) XLSX (pièce jointe)
# ---------------------------------------------------------------------------
GRAS = Font(bold=True)
GRAS_BLANC = Font(bold=True, color="FFFFFF")
ENTETE = PatternFill("solid", fgColor="0F766E")
SURLIGNE = PatternFill("solid", fgColor="CCFBF1")
CENTRE = Alignment(horizontal="center", vertical="center", wrap_text=True)
DROITE = Alignment(horizontal="right", vertical="center")
BORD = Border(*(4 * (Side(style="thin", color="D1D5DB"),)))


def styler_entete(ws, ncols, ligne=1):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=ligne, column=c)
        cell.font = GRAS_BLANC
        cell.fill = ENTETE
        cell.alignment = CENTRE
        cell.border = BORD


def largeurs(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w


def construire_xlsx(par_note, matrice, totaux_ys, total_global):
    wb = openpyxl.Workbook()

    # --- Onglet 1 : Encaissements par note (total + par exercice) --------
    ws = wb.active
    ws.title = "Encaissements par note"
    ws.append(["Nombre d'encaissements par numéro de note (rang quotidien, 1 à 57)"])
    ws["A1"].font = GRAS
    ws.append([("Le numéro de note est un rang quotidien réattribué chaque jour (1, 2, 3 ...). "
                "C'est la seule granularité de l'export : le numéro de table physique n'y figure pas. "
                "Un encaissement = une note distincte (date + n° de note).")])
    ws["A2"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells("A2:E2")
    ws.row_dimensions[2].height = 45
    ws.append([])
    h = ws.max_row + 1
    ws.append(["N° de note"] + EXOS + ["Total"])
    styler_entete(ws, 5, h)
    for no in sorted(par_note):
        v = par_note[no]
        r = ws.max_row + 1
        ws.append([f"n°{no}", v[EXOS[0]], v[EXOS[1]], v[EXOS[2]], v["total"]])
        for c in range(1, 6):
            ws.cell(row=r, column=c).border = BORD
    rt = ws.max_row + 1
    ws.append(["Total",
               sum(v[EXOS[0]] for v in par_note.values()),
               sum(v[EXOS[1]] for v in par_note.values()),
               sum(v[EXOS[2]] for v in par_note.values()),
               total_global])
    for c in range(1, 6):
        ws.cell(row=rt, column=c).font = GRAS
        ws.cell(row=rt, column=c).border = BORD
    largeurs(ws, [14, 14, 14, 14, 12])
    ws.freeze_panes = "A" + str(h + 1)

    # --- Onglet 2 : Note x saison x exercice -----------------------------
    ws2 = wb.create_sheet("Note x saison x exercice")
    ws2.append(["Encaissements par numéro de note, par saison et par exercice"])
    ws2["A1"].font = GRAS
    ws2.append([])
    cols = [("ex", e, s) for e in EXOS for s in SAISONS]
    head = ["N° de note"] + [f"{e}\n{s}" for (_, e, s) in cols] + ["Total"]
    h2 = ws2.max_row + 1
    ws2.append(head)
    styler_entete(ws2, len(head), h2)
    for no in sorted(par_note):
        r = ws2.max_row + 1
        ligne = [f"n°{no}"]
        tot = 0
        for (_, e, s) in cols:
            val = matrice[no].get((e, s), 0)
            tot += val
            ligne.append(val)
        ligne.append(tot)
        ws2.append(ligne)
        for c in range(1, len(head) + 1):
            ws2.cell(row=r, column=c).border = BORD
    # ligne totaux
    rt2 = ws2.max_row + 1
    ligne = ["Total"]
    grand = 0
    for (_, e, s) in cols:
        val = totaux_ys.get((e, s), 0)
        grand += val
        ligne.append(val)
    ligne.append(grand)
    ws2.append(ligne)
    for c in range(1, len(head) + 1):
        ws2.cell(row=rt2, column=c).font = GRAS
        ws2.cell(row=rt2, column=c).border = BORD
    largeurs(ws2, [14] + [11] * len(cols) + [12])
    ws2.freeze_panes = "B" + str(h2 + 1)

    # --- Onglet 3 : Plan de salle reel -----------------------------------
    ws3 = wb.create_sheet("Plan de salle reel")
    ws3.append(["Plan de salle RÉEL (confirmé par la gérante) contre numérotation de la caisse"])
    ws3["A1"].font = GRAS
    ws3.append([])
    h3 = ws3.max_row + 1
    ws3.append(["N° dans la caisse", "Table physique ?", "Précision"])
    styler_entete(ws3, 3, h3)
    lignes3 = []
    for n in range(1, 19):
        if n in INT_REELLES:
            lignes3.append([f"n°{n}", "Oui (intérieur)", ""])
        else:
            note = "Numéro de note (règlement séparé) / non attribué"
            lignes3.append([f"n°{n}", "NON, aucune table", note])
    lignes3.append(["n°101 à 117 (sauf 102)", "Oui (terrasse)", ""])
    lignes3.append(["n°102", "NON, aucune table", "N'existe pas"])
    lignes3.append(["n°125", "NON, aucune table", "N'existe pas"])
    lignes3.append(["n°116 et 117", "Terrasse, occasionnelle", "Jours de forte affluence"])
    for row in lignes3:
        r = ws3.max_row + 1
        ws3.append(row)
        for c in range(1, 4):
            ws3.cell(row=r, column=c).border = BORD
    largeurs(ws3, [18, 38, 36])
    ws3.freeze_panes = "A" + str(h3 + 1)

    # --- Onglet 4 : Notes fractionnees (synthèse mesurée) ----------------
    ws4 = wb.create_sheet("Notes fractionnees")
    ws4.append(["Le partage d'addition, déjà mesuré sur les fichiers de caisse"])
    ws4["A1"].font = GRAS
    ws4.append([])
    rows4 = [
        ("Notes (règlements) sur 3 exercices", fr_int(NOTES_TOTAL)),
        ("Notes portant une quantité fractionnaire (0,5 / 0,33...)", fr_int(NOTES_FRAC)),
        ("Part des notes fractionnaires", f"{str(NOTES_FRAC_PCT).replace('.', ',')} %"),
        ("Partages estimés sur 3 ans (~1 par service)", "~" + fr_int(PARTAGES_3ANS_EST)),
        ("Règlements bancarisés (CB, TR, CHV, CHQ)", f"{str(BANC_PCT).replace('.', ',')} %"),
        ("Règlements en espèces", f"{str(ESP_PCT).replace('.', ',')} %"),
    ]
    h4 = ws4.max_row + 1
    ws4.append(["Indicateur", "Valeur"])
    styler_entete(ws4, 2, h4)
    for lib, val in rows4:
        r = ws4.max_row + 1
        ws4.append([lib, val])
        ws4.cell(row=r, column=1).border = BORD
        ws4.cell(row=r, column=2).border = BORD
        ws4.cell(row=r, column=2).alignment = DROITE
    largeurs(ws4, [56, 24])
    ws4.freeze_panes = "A" + str(h4 + 1)

    # --- Onglet 5 : Exemple table scindee --------------------------------
    ws5 = wb.create_sheet("Exemple table scindee")
    ws5.append([f"Exemple daté : table réelle scindée en plusieurs notes : {EX['date']}"])
    ws5["A1"].font = GRAS
    ws5.append(["Deux notes de MÊME TOTAL, encaissées à la même minute, articles divisés en 0,5. "
                "Supprimer la table d'origine évite de compter ces articles deux fois."])
    ws5.append([])
    h5 = ws5.max_row + 1
    ws5.append(["Note", "Heure", "Article", "Quantité", "Prix unitaire", "Montant",
                "Physiquement insécable ?"])
    styler_entete(ws5, 7, h5)
    insecables = {"truite jurassienne", "croûtes morilles", "salade franc-comptoi", "cancoine"}
    for n in EX["notes"]:
        for it in n["items"]:
            r = ws5.max_row + 1
            lib = it["lib"]
            est_insec = "Oui" if (it.get("highlight") or lib.lower() in insecables) else ""
            ws5.append([n["label"], n["heure"], lib,
                        str(it["qte"]).replace(".", ","),
                        euro(it["pu"]), euro(it["montant"]), est_insec])
            for c in range(1, 8):
                ws5.cell(row=r, column=c).border = BORD
            if it.get("highlight"):
                for c in range(1, 8):
                    ws5.cell(row=r, column=c).fill = SURLIGNE
        rt = ws5.max_row + 1
        ws5.append([n["label"], n["heure"], "TOTAL NOTE", "", "", euro(n["total"]), ""])
        for c in range(1, 8):
            ws5.cell(row=rt, column=c).font = GRAS
            ws5.cell(row=rt, column=c).border = BORD
    largeurs(ws5, [12, 8, 24, 10, 14, 12, 24])
    ws5.freeze_panes = "A" + str(h5 + 1)

    os.makedirs(os.path.dirname(OUT_XLSX), exist_ok=True)
    wb.save(OUT_XLSX)
    return OUT_XLSX


# ---------------------------------------------------------------------------
# 6) JSON (page recodée)
# ---------------------------------------------------------------------------
def cg(t):
    return {"v": t}


def cd(t):
    return {"v": t, "align": "right"}


_HEBDO_CACHE = []


def charge_section_hebdo():
    """Section graphiqueLignes (couverts/semaine) produite par rendu-final-couverts.py.
    Retourne None si le fichier n'a pas encore été généré."""
    if _HEBDO_CACHE:
        return _HEBDO_CACHE[0]
    if not os.path.exists(HEBDO_JSON):
        _HEBDO_CACHE.append(None)
        return None
    with open(HEBDO_JSON, encoding="utf-8") as f:
        _HEBDO_CACHE.append(json.load(f))
    return _HEBDO_CACHE[0]


SAISONS_JSON = os.path.join(ROOT, "src/data/renduFinal/couverts-saisons.json")


def charge_sections_saisons():
    """Liste des 4 sections 'saisonCouverts' (couverts/jour, midi/soir, par saison),
    produites par rendu-final-couverts.py. Liste vide si non encore généré."""
    if not os.path.exists(SAISONS_JSON):
        return []
    with open(SAISONS_JSON, encoding="utf-8") as f:
        return json.load(f)


CAPACITE_JSON = os.path.join(ROOT, "src/data/renduFinal/couverts-capacite.json")


def charge_section_capacite():
    """Section 'couvertsCapacite' (couverts/jour par semaine + bandes 58 %),
    produite par rendu-final-couverts.py. None si non encore généré."""
    if not os.path.exists(CAPACITE_JSON):
        return None
    with open(CAPACITE_JSON, encoding="utf-8") as f:
        return json.load(f)


# Lien URL visible et cliquable (le libellé EST l'URL complète).
def _lien(u):
    return f"[{u}]({u})"


U_FIDUCIAL = "https://www.sudradio.fr/sud-radio/lobservatoire-fiducial-2025-le-metier-de-la-restauration-a-la-loupe"
U_PENNYLANE = "https://www.pennylane.com/fr/fiches-pratiques/comptabilite/taux-de-remplissage-dun-restaurant-tout-savoir"
U_CREER = "https://www.creerentreprise.fr/taux-remplissage-occupation-hotel-restaurant-calcul-moyen/"


def construire_json(par_note, matrice, totaux_ys, total_global, graph_enc):
    max_note = max(par_note)

    # Section graphique capacité (couverts/jour vs 58 %) + valeurs dérivées.
    _cap_section = charge_section_capacite()
    _cap_int = _cap_section["capaciteInterieur"] if _cap_section else 44
    _cap_ext = _cap_section["capaciteTerrasse"] if _cap_section else 43
    _cap_int_58 = _cap_section["seuilInterieur"] if _cap_section else 26
    _cap_ext_58 = _cap_section["seuilTerrasse"] if _cap_section else 25

    # --- Card A : plan de salle réel INTÉRIEUR (numéro != place) ----------
    # Couverts par table (places assises), confirmés par la gérante.
    COUV_INT = {1: 2, 2: 4, 3: 2, 4: 4, 5: 4, 6: 4, 7: 2, 8: 2, 9: 4,
                10: 2, 11: 2, 12: 4, 14: 2, 15: 6}
    tables_int = []
    for n in range(1, 19):
        if n in INT_REELLES:
            tables_int.append({"numero": f"n°{n}", "type": "reelle", "couverts": COUV_INT[n]})
        else:
            tables_int.append({"numero": f"n°{n}", "type": "inexistante", "note": "Aucune table"})
    grille_int = {
        "kind": "grilleTables",
        "variante": "plan",
        "titre": "Plan de salle réel à l'intérieur (14 tables, n° 1 à 15 sauf 13)",
        "sousTitre": ("Sur les 18 numéros possibles à l'intérieur, seuls 14 désignent une table. "
                      "Les n° 13, 16, 17 et 18 ne correspondent à AUCUNE table physique : ce sont "
                      "des numéros de note (règlement séparé) ou des numéros non attribués."),
        "tables": tables_int,
        "resume": {"tables": len(COUV_INT), "couverts": sum(COUV_INT.values())},
    }

    # --- Card A bis : plan de salle réel EXTÉRIEUR (terrasse) ------------
    # Terrasse réelle = n° 101 à 117 (la n° 102 n'existe pas). Les n° 116 et 117
    # ne sont montées que les jours de forte affluence -> "rare", exclues du
    # décompte des places (météo). Aucun de ces numéros n'apparaît dans l'export.
    COUV_EXT = {101: 2, 103: 2, 104: 4, 105: 4, 106: 4, 107: 2, 108: 4,
                109: 2, 110: 3, 111: 2, 112: 4, 113: 2, 114: 4, 115: 4}
    METEO_EXT = {116: 2, 117: 2}
    tables_ext = []
    for n in range(101, 118):  # 101 à 117
        if n == 102:
            tables_ext.append({"numero": "n°102", "type": "inexistante", "note": "N'existe pas"})
        elif n in METEO_EXT:
            tables_ext.append({"numero": f"n°{n}", "type": "rare", "couverts": METEO_EXT[n],
                               "note": "Unique pour jours exceptionnels comme 14 juillet et week-end du Chat Gourmand"})
        else:
            tables_ext.append({"numero": f"n°{n}", "type": "reelle", "couverts": COUV_EXT[n]})
    # La n° 125 est citée par le service mais n'existe pas davantage que la 102.
    tables_ext.append({"numero": "n°125", "type": "inexistante", "note": "N'existe pas"})
    grille_ext = {
        "kind": "grilleTables",
        "variante": "plan",
        "titre": "Plan de salle réel à l'extérieur (terrasse)",
        "sousTitre": ("La terrasse porte les numéros 101 à 117 (les n° 102 et 125 n'existent pas). "
                      "Les n° 116 et 117 ne sont montées que les jours de forte affluence : elles "
                      "sont donc exclues du décompte des places. Aucun de ces numéros de terrasse "
                      "n'apparaît dans l'export de caisse."),
        "tables": tables_ext,
        "resume": {"tables": len(COUV_EXT), "couverts": sum(COUV_EXT.values())},
    }

    # --- Card B : encaissements par numéro de note -----------------------
    # --- Tableau : partage mesuré ----------------------------------------
    tab_partage = {
        "kind": "tableau",
        "titre": "Le partage d'addition, déjà mesuré sur les fichiers de caisse (annexes C)",
        "minWidth": 560,
        "colonnes": [{"label": "Indicateur"}, {"label": "Valeur", "align": "right"}],
        "lignes": [
            [cg("Notes (règlements) sur 3 exercices"), cd(fr_int(NOTES_TOTAL))],
            [cg("Notes portant une quantité fractionnaire (0,5 / 0,33...)"),
             cd(fr_int(NOTES_FRAC) + " (" + str(NOTES_FRAC_PCT).replace(".", ",") + " %)")],
            [cg("Partages estimés sur 3 ans (environ 1 par service)"),
             cd("~" + fr_int(PARTAGES_3ANS_EST))],
            [cg("Règlements bancarisés (CB, TR, CHV, CHQ)"),
             cd(str(BANC_PCT).replace(".", ",") + " %")],
            [cg("Règlements en espèces"), cd(str(ESP_PCT).replace(".", ",") + " %")],
        ],
    }

    # --- Tableau : exemple table scindée ---------------------------------
    lignes_ex = []
    for n in EX["notes"]:
        for it in n["items"]:
            lignes_ex.append([
                cg(n["label"]), cg(n["heure"]), cg(it["lib"]),
                cd(str(it["qte"]).replace(".", ",")),
                cd(euro(it["pu"])), cd(euro(it["montant"])),
            ])
        lignes_ex.append([
            cg(n["label"]), cg(n["heure"]), cg("TOTAL NOTE"),
            cd(""), cd(""), cd(euro(n["total"])),
        ])
    tab_exemple = {
        "kind": "tableau",
        "titre": f"Exemple daté ({EX['date']}) : une table scindée en deux notes de même total",
        "minWidth": 640,
        "colonnes": [
            {"label": "Note"}, {"label": "Heure"}, {"label": "Article"},
            {"label": "Quantité", "align": "right"},
            {"label": "Prix unitaire", "align": "right"},
            {"label": "Montant", "align": "right"},
        ],
        "lignes": lignes_ex,
    }

    pct_fr = str(NOTES_FRAC_PCT).replace(".", ",")

    sections = [
        # ============ 1) CE QUE DIT L'ADMINISTRATION ============
        {"kind": "chapitre", "source": "fisc", "numero": 1,
         "titre": "Ce que dit l'administration",
         "sousTitre": "Proposition de rectification p. 11 et section « suppression de note » (p. 18-19)."},
        {"kind": "paragraphe",
         "texte": ("Le vérificateur relève que la caisse comporte des **tables « virtuelles »** et "
                   "qu'elle **autorise la suppression d'une table et de tous ses articles**. Il en "
                   "conclut qu'« *il n'est pas possible de confirmer que ces modifications soient "
                   "légitimes* » et que la comptabilité serait, de ce fait, non probante (vecteur "
                   "supposé de couverts ou de recettes non déclarés).")},
        {"kind": "paragraphe",
         "texte": ("Le vérificateur décrit pourtant lui-même la fonction (p. 11) : « *quatre tables "
                   "virtuelles figurent à l'écran (n° 15 à n° 18, **par exemple**). Cette modalité "
                   "permet le traitement des notes multiples pour une table, au moment du "
                   "règlement.* » Une table « virtuelle » n'est donc pas une table en plus : c'est "
                   "une **note supplémentaire** ouverte pour encaisser séparément les convives "
                   "d'une **même** table. Nous le démontrons ci-dessous, fichiers de caisse à "
                   "l'appui.")},

        # ============ 2) CE QUE DÉMONTRENT LES FICHIERS ============
        {"kind": "chapitre", "source": "nous", "numero": 2,
         "titre": "Ce que démontrent les fichiers",
         "sousTitre": "Un numéro de note n'est pas une place ; la suppression protège le CA, elle ne le réduit pas."},

        # --- A. Le numéro n'est pas une place ---
        {"kind": "titre", "numero": "2.1", "texte": "Premier fait : un numéro n'est pas une table"},
        {"kind": "paragraphe",
         "texte": ("Avant tout, **les fichiers d'export de la caisse ne mentionnent jamais les "
                   "tables.** Ils ne listent que des **encaissements (les notes), numérotés de 1 à "
                   f"{max_note}** au fil de la journée. Ce numéro **ne reflète pas le nombre de "
                   "tables réelles** : beaucoup de clients **règlent séparément**, si bien que "
                   "**plusieurs notes, donc plusieurs numéros, correspondent en réalité à une seule "
                   "table**. Compter les numéros revient donc à compter des règlements, pas des "
                   "tables.")},
        {"kind": "paragraphe",
         "texte": ("La salle compte en réalité "
                   f"**{NB_INT_REELLES} tables à l'intérieur** (n° 1 à 12, 14 et 15). Les numéros "
                   "**13, 16, 17 et 18 ne désignent aucune table physique** : ce sont des numéros "
                   "de note (règlement séparé) ou des numéros non attribués. C'est vérifiable sur "
                   "place : il suffit de compter les tables de la salle.")},
        grille_int,
        {"kind": "paragraphe",
         "texte": ("Même constat à l'extérieur. La terrasse porte les numéros **101 à 117** (les "
                   "**tables n° 102 et 125 n'existent pas**). Les **tables n° 116 et 117** ne sont "
                   "montées que les **jours de forte affluence** ; le reste de l'année, elles ne "
                   "servent pas, et elles sont exclues du décompte des places. **Preuve décisive "
                   "dans les fichiers** : si le numéro de la caisse était un numéro de table, on "
                   "verrait apparaître les numéros "
                   f"de terrasse (101 à 125). Or le numéro de note de l'export ne va **jamais "
                   f"au-delà de {max_note}** : c'est un **rang de note quotidien** (1ʳᵉ note "
                   "encaissée, 2ᵉ, etc.), réattribué chaque jour, et non un numéro de table. On ne "
                   "peut donc rien déduire de « couverts » à partir d'un numéro.")},
        grille_ext,

        # --- B. La capacité de la salle et le taux d'occupation de 58 % ---
        {"kind": "titre", "numero": "2.2",
         "texte": "Deuxième fait : la capacité de la salle borne le nombre de couverts"},
        {"kind": "paragraphe",
         "texte": ("Le "
                   "taux de remplissage moyen d'un restaurant traditionnel en France est de "
                   "**58 %** (Observatoire FIDUCIAL 2025). Appliqué aux places réelles, cela "
                   f"représente **{_cap_int_58} couverts par service à l'intérieur** ({_cap_int} "
                   f"places) et **{_cap_ext_58} couverts en terrasse** ({_cap_ext} places). Surtout, "
                   "**la cuisine est trop petite pour servir à la fois en salle et en terrasse** : "
                   "un seul espace est ouvert à chaque service. **Du début mai au début septembre**, "
                   "seule la **terrasse** est utilisée, sauf les jours de pluie où le service se fait "
                   "à l'intérieur ; **de mi-septembre à fin avril**, seul l'**intérieur** est "
                   "utilisé.")},
        {"kind": "paragraphe",
         "texte": ("Sources du taux de remplissage de 58 % (restauration traditionnelle, France) : "
                   + _lien(U_FIDUCIAL) + " ; " + _lien(U_PENNYLANE) + " ; " + _lien(U_CREER) + ".")},

        # --- Méthodologie : compter les couverts depuis la caisse ---
        {"kind": "titre", "texte": "Compter les couverts à partir de la caisse"},
        {"kind": "paragraphe",
         "texte": ("Le détail des tickets (annexes "
                   "C) permet d'identifier **tous les encaissements contenant un plat ou un menu** "
                   "(un couvert = un plat ou un menu servi). On estime ainsi, de façon assez juste, "
                   "le **nombre de personnes servies par jour**, et même **par service** : l'**heure** "
                   "de chaque encaissement le classe au **midi** (avant 18 h) ou au **soir** (18 h à "
                   "minuit). Les fichiers ci-dessous contiennent ces données, recalculables ligne à "
                   "ligne.")},
        {"kind": "piecejointe",
         "intro": "Données de couverts (lecture des annexes C, reproductible) :",
         "fichiers": [
             {"fichier": "pieces-defense/RF-items-classification.xlsx",
              "label": "Items de la caisse classés « plat/menu » (ce qui compte comme un couvert) (XLSX)"},
             {"fichier": "pieces-defense/RF-encaissements-couverts.xlsx",
              "label": "Couverts par encaissement : date, heure, nombre de couverts (XLSX)"},
             {"fichier": "pieces-defense/RF-couverts-par-service.xlsx",
              "label": "Couverts par jour et par service (midi avant 18 h / soir 18 h à minuit) (XLSX)"},
         ]},

        # --- Graphique : couverts/jour par semaine vs capacité à 58 % ---
        *([_cap_section] if _cap_section else []),

        # --- C. Le mécanisme et son retournement ---
        {"kind": "titre", "numero": "2.3",
         "texte": "Troisième fait : la suppression de la table d'origine est nécessaire et protège le CA"},
        {"kind": "paragraphe",
         "texte": ("Le logiciel n'accepte **qu'un seul règlement par note**. Pour "
                   "un paiement séparé, le serveur ressaisit les articles d'un convive sur une "
                   "**deuxième note** (la table « virtuelle ») et l'encaisse. Les articles existent "
                   "alors **en double** : sur la table d'origine et sur la note ressaisie. "
                   "Supprimer la table d'origine est donc **arithmétiquement obligatoire pour ne "
                   "pas compter le repas deux fois**. Sans cette suppression, le chiffre d'affaires "
                   "serait **gonflé**, pas réduit. Le raisonnement de l'administration est inversé : "
                   "cette suppression **garantit l'exactitude** du CA, elle ne sert pas à le "
                   "diminuer.")},
        {"kind": "paragraphe",
         "texte": (f"L'exemple ci-dessous (**{EX['date']}**) est typique. Une même table est scindée "
                   f"en **deux notes (n°22 et n°23)** de **total identique "
                   f"({euro(EX['notes'][0]['total'])})**, encaissées **à la même minute**. Chaque "
                   "article y figure en **quantité 0,5**, y compris des plats **physiquement "
                   "insécables** (une **TRUITE JURASSIENNE**, des **Croûtes Morilles**) : on ne "
                   "sert pas une demi-truite à chaque convive, c'est **un seul plat payé à deux**. "
                   "Une fois ces deux notes encaissées, la table d'origine est supprimée pour "
                   f"éviter le double comptage. {fr_int(NB_EXEMPLES)} fiches de ce type (une par "
                   "mois actif) figurent dans la pièce jointe.")},
        tab_exemple,

        # --- Pièce jointe ---
        {"kind": "piecejointe",
         "intro": ("Encaissements par numéro de note (total et matrice saison × exercice), plan de "
                   "salle réel, notes fractionnaires et exemple daté de table scindée."),
         "fichiers": [
             {"fichier": "pieces-defense/RF-tables-trop-nombreuses.xlsx",
              "label": "RF - Tables et numéros de note : encaissements par note, saison et exercice (XLSX)"},
         ]},

        # ============ 3) CONCLUSION ============
        {"kind": "chapitre", "source": "nous", "numero": 3, "titre": "Conclusion"},
        {"kind": "paragraphe",
         "texte": ("Les tables « virtuelles » sont une **fonction de règlement**, et la suppression "
                   "qui les accompagne est une **opération anti-double-comptage**. Trois faits "
                   "reproductibles le démontrent : (1) le numéro de note de l'export va de 1 à "
                   f"{max_note} et ne désigne **jamais** une table (la salle compte "
                   f"{NB_INT_REELLES} tables intérieures, et les n° 13/16/17/18 n'existent pas) ; "
                   "(2) le rythme des encaissements par jour travaillé se superpose d'un exercice à "
                   "l'autre (activité stable, sans à-coup inexpliqué) ; (3) supprimer la table "
                   "d'origine après ressaisie **évite de doubler le CA**. Le grief ne révèle "
                   "**aucun couvert ni aucune recette non déclarée**.")},
        {"kind": "paragraphe",
         "texte": ("Cet argument se retourne contre l'administration : le vérificateur décrit "
                   "lui-même (p. 11) les tables virtuelles comme servant « au traitement des notes "
                   "multiples pour une table, au moment du règlement ». Il ne peut pas, ensuite, "
                   "présenter cette même fonction comme la preuve d'une dissimulation. La "
                   "réconciliation complète des suppressions (le CA déclaré égale les encaissements) "
                   "est démontrée au grief « suppressions de caisse ».")},
    ]

    meta = {
        "slug": "tables-trop-nombreuses",
        "titre": "Trop de tables / tables « virtuelles »",
        "refRapport": "Proposition p. 11 et p. 18-19",
        "griefFisc": ("La caisse comporte des tables « virtuelles » et autorise la suppression "
                      "d'une table et de tous ses articles ; ces modifications ne seraient pas "
                      "vérifiables, rendant la comptabilité non probante."),
        "reponseCourte": ("Un numéro de note n'est pas une table (la salle a 14 tables intérieures, "
                          "les n° 13/16/17/18 n'existent pas). La table « virtuelle » est une note "
                          "de règlement séparé, et la suppression de la table d'origine évite de "
                          "compter le repas deux fois : elle protège le CA, elle ne le réduit pas."),
        "tables_interieures_reelles": NB_INT_REELLES,
        "numeros_inexistants_int": INT_INEXISTANTES,
        "max_note": max_note,
        "notes_total": NOTES_TOTAL,
        "notes_fractionnees": NOTES_FRAC,
        "notes_fractionnees_pct": NOTES_FRAC_PCT,
        "encaissements_total": total_global,
        "_source": {
            "fichiers_caisse": [os.path.relpath(F[e], ROOT) for e in EXOS],
            "calculs": "src/data/renduFinalCalculs.json (paiements_fractionnes, bancarisation)",
            "cas": "src/data/renduFinalCasSuppressions.json (cas partage)",
            "plan_salle_reel": "confirmé par la gérante (mémoire tables-physiques-reelles)",
            "script": "scripts/rendu-final-tables-trop-nombreuses.py",
        },
    }

    sections = normaliser_sections(sections)
    os.makedirs(os.path.dirname(OUT_JSON), exist_ok=True)
    with open(OUT_JSON, "w", encoding="utf-8") as f:
        json.dump({"meta": meta, "sections": sections}, f, ensure_ascii=False, indent=2)
    return OUT_JSON


# ---------------------------------------------------------------------------
def main():
    par_note, matrice, totaux_ys, total_global = encaissements()
    par_jour = encaissements_par_jour()
    graph_enc, somme_sem, jours_sem = graphique_encaissements_hebdo(par_jour)
    chemin_xlsx = construire_xlsx(par_note, matrice, totaux_ys, total_global)
    chemin_xlsx_enc = construire_xlsx_encaissements(par_jour, somme_sem, jours_sem)
    chemin_json = construire_json(par_note, matrice, totaux_ys, total_global, graph_enc)

    print("=== Tables trop nombreuses : synthèse ===")
    print(f"Tables intérieures réelles : {NB_INT_REELLES} (n° 13/16/17/18 inexistants)")
    print(f"Encaissements totaux (notes distinctes) : {fr_int(total_global)}")
    print(f"Max numéro de note : {max(par_note)}")
    for e in EXOS:
        nj = len(par_jour[e])
        ne = sum(par_jour[e].values())
        moy = round(ne / nj, 1) if nj else 0
        print(f"  {e} : {fr_int(nj)} jours travaillés, {fr_int(ne)} encaissements, "
              f"moyenne {moy}/jour")
    for ex in EXOS:
        tot = sum(matrice[no].get((ex, s), 0) for no in par_note for s in SAISONS)
        print(f"  {ex} : {fr_int(tot)} encaissements")
    print("Totaux par saison x exercice :")
    for e in EXOS:
        for s in SAISONS:
            print(f"  {e} {s:9s} : {totaux_ys.get((e, s), 0)}")
    print(f"Notes fractionnaires : {fr_int(NOTES_FRAC)} / {fr_int(NOTES_TOTAL)} = {NOTES_FRAC_PCT} %")
    print(f"Exemple table scindée : {EX['date']}, total {euro(EX['notes'][0]['total'])} x{len(EX['notes'])}")
    print(f"XLSX -> {chemin_xlsx}")
    print(f"XLSX (encaissements/jour) -> {chemin_xlsx_enc}")
    print(f"JSON -> {chemin_json}")


if __name__ == "__main__":
    main()
