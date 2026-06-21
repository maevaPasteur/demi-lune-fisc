#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
rendu-final-couverts.py

Outils préparatoires pour ESTIMER le nombre de couverts par encaissement.
Trois fichiers XLSX (lecture seule des annexes de caisse) :

  1) RF-items-classification.xlsx
     Tous les items vendus en caisse (annexe D, par famille) + une colonne
     « Plat/Menu (couvert) » Oui/Non. Règle : un couvert = un plat OU un menu.
       - Menus / formules / plat du jour            -> Oui
       - Plats (poulet, truite, entrecôte, fondues, assiettes, galettes...) -> Oui
       - GALETTES (salées) = des plats              -> Oui
       - SALADES = des plats, SAUF « salade chèvre chaud » et
         « salade verte aux noix » qui sont des entrées -> Non
       - Entrées, desserts (glaces, crêpes sucrées, babas, coupes...),
         accompagnements / suppléments, fromages, boissons -> Non
     Une colonne « Confiance » signale les noms fantaisie (galette vs crêpe)
     à vérifier sur la carte.

  2) RF-encaissements-couverts.xlsx
     Un encaissement = une note = (date, n° de note). Colonnes : date, heure,
     id (s'il existe), nombre de couverts = somme des quantités des lignes
     classées « plat/menu ».

  3) RF-couverts-par-service.xlsx
     Couverts par jour et par service : MIDI (avant 18h, ~10h-17h) et SOIR
     (18h à minuit), + total.

Sources : annexes C (détail tickets) et D (synthèse produit) ;
référentiel public/documents/vins-boissons/alcool-plats-cocktails-menus.json
(ancre quelques galettes/crêpes/coupes).

Python : /tmp/xlsenv/bin/python (xlrd + openpyxl).
"""
import os
import json
import datetime
import unicodedata
import collections

import xlrd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

ICI = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.normpath(os.path.join(ICI, ".."))
CAISSE = os.path.join(ROOT, "public/documents/caisse-enregistreuse/")
OUT = os.path.join(ROOT, "public/documents/pieces-defense/")
OUT_JSON_DIR = os.path.join(ROOT, "src/data/renduFinal/")
EXOS = ["2022-2023", "2023-2024", "2024-2025"]
C = {e: CAISSE + f"ANNEXE-C{i}_detail-tickets_{e}.xls" for i, e in enumerate(EXOS, 1)}
D = {e: CAISSE + f"ANNEXE-D{i}_synthese-produit_{e}.xls" for i, e in enumerate(EXOS, 1)}
COULEURS_EXO = {"2022-2023": "#0f766e", "2023-2024": "#b45309", "2024-2025": "#1d4ed8"}
MOIS_LABELS = [(4, "Avr"), (5, "Mai"), (6, "Juin"), (7, "Juil"), (8, "Août"),
               (9, "Sep"), (10, "Oct"), (11, "Nov"), (12, "Déc"),
               (1, "Jan"), (2, "Fév"), (3, "Mars")]


def norm(s):
    s = str(s).replace("œ", "oe").replace("Œ", "OE").replace("æ", "ae").replace("Æ", "AE")
    s = unicodedata.normalize("NFKD", s)
    s = "".join(c for c in s if not unicodedata.combining(c))
    return " ".join(s.upper().split())


# ---------------------------------------------------------------------------
# Famille de chaque libellé (annexe D). Famille != SOLIDE => boisson => Non.
# ---------------------------------------------------------------------------
def familles():
    fam_of = {}
    for e in EXOS:
        sh = xlrd.open_workbook(D[e]).sheet_by_index(0)
        fam = None
        for r in range(sh.nrows):
            c0 = str(sh.cell_value(r, 0)).strip()
            c1 = str(sh.cell_value(r, 1)).strip()
            if (not c0 and not c1) or c0 == "Ref_prd":
                continue
            is_ref = c0[:1].isdigit() and len(c0) <= 5
            if c1 == "" and not is_ref:
                fam = c0
                continue
            if is_ref:
                fam_of.setdefault(c1, fam)
    return fam_of


# ---------------------------------------------------------------------------
# Classification d'un item alimentaire (famille SOLIDE).
# Retour : (couvert: bool, categorie: str, confiance: str, remarque: str)
# ---------------------------------------------------------------------------
# Overrides CONFIRMÉS par la gérante (priorité absolue, confiance haute).
CONFIRME = {
    "DAMES VERTES": (False, "dessert", "crêpe/dessert (confirmé gérante)"),
    "TANTE ARIE": (False, "dessert", "crêpe/dessert (confirmé gérante)"),
    "FLAMME": (False, "dessert", "crêpe/dessert (confirmé gérante)"),
    "VOUIVRE": (False, "dessert", "crêpe/dessert (confirmé gérante)"),
    "PAYSANNE": (True, "plat", "plat (confirmé gérante)"),
    "MACHON GOURMET": (False, "entree", "entrée, pas un plat (confirmé gérante)"),
    "SALADE VERTE": (False, "accompagnement", "salade verte nature, pas un plat (confirmé gérante)"),
    "FROMAGES REGIONAUX": (False, "fromage", "pas un plat (confirmé gérante)"),
}
# Galettes salées (= plats). Noms régionaux / comtois. Confiance moyenne.
GALETTES = {
    "FRELEE", "FRELE", "CANCOINE", "PEUTE", "BARBEULENET", "CABRICHOU", "TANTELIN",
    "LIGOUSE", "DADOULET", "LOURNETTE", "ECRESSI",
    "FOULETOT", "NOIRAUD", "NIOLERIES", "DAHU", "VERDELET",
    "COMPLETE", "GALETTE FORESTIERE",
}
# Crêpes sucrées / coupes glacées (= desserts). Ancrés par le référentiel
# (Ivresse, Folie, Coquette, Dijonnaise, Grappins, Vogeotte, Basilic) + noms « tendres ».
DESSERTS_NOMMES = {
    "IVRESSE", "FOLIE", "COQUETTE", "DIJONNAISE", "GRAPPINS", "VOGEOTTE", "BASILIC",
    "DOUCEUR", "PASSION", "TENDRESSE", "CALIN", "BEGUIN", "FLIRT", "CAPRICE",
    "VALENTINE", "TOQUADE", "BAGATELLE", "JALOUSIE", "BLUETTE", "CONQUETE",
    "CUPIDON", "CHOUKA", "NORMANDINE", "BEGIN",
}
# Entrées explicites (jamais un couvert "plat").
ENTREES = {
    "OEUF", "CHEVRE CHAUD", "SALADE VERTE AUX NOI", "SALADE VERTE AUX NOIX",
    "CROUTES MORILLES", "FEUILLETE FORESTIER", "ASSORTIMENT CHARCUTE",
    "ASSORTIMENT CHARCUTERIE", "JAMBON PERSILLE", "JAMHON PERSILLE",
}
# Items à vérifier sur la carte (conflit ou ambiguïté forte).
A_VERIFIER = {
    "CHOUKA": ("galette ou crêpe ? classé dessert", False),
    "SPRINTZ": ("probable boisson (Spritz) mal rangée : non-couvert", False),
}

ACCOMP_KW = ("POMME", "FRITES", "CHAMPIGNON", "COMTE", "MORBIER", "CANCOILLOTTE",
             "BLEU DE GEX", "RATATOUILLE", "CREME FRAICHE", "AMANDES", "BEURRE",
             "OIGNON", "TOMATE", "JUS DE CITRON", "JAMBON BLANC", "JAMBON SEC",
             "BACON", "CHAMPIGNONS", "FROMAGE DE CHEVRE", "MORBIER")
DESSERT_KW = ("GLACE", "CREME BRULEE", "PAVLOVA", "BABA", "CHANTILLY", "COUPE GLAC",
              "CREPE", "CREPES", "DESSERT", "COULIS", "SALIDOU", "BANANE", "CHOCOLAT",
              "NUTELLA", "CONFITURE", "CARAMEL", "MIEL", "MARRON", "CRUMBLE", "CHURROS",
              "POIRE AU SIROP", "SMARTIES", "NOISETTE", "PIGNONS", "NOIX DE COCO",
              "GRIOTTINES", "FLAMBEE", "SAUCE", "1 BOULE GLACE", "ECLATS", "NOIX",
              "CHOCOLAT FONDU")
PLAT_KW = ("POULET", "BURGER", "TRUITE", "ENTRECOTE", "FONDUE", "CAMEMBERT ROTI",
           "ASSIETTE", "MACHON", "CREMEUX DU JURA", "MORTEAU", "SAUCISSE DE MORTEAU",
           "STEAK HACHE", "GIGOT", "GALETTE")


def classer(lib, fam):
    n = norm(lib)
    # 1) Boissons : toute famille hors SOLIDE.
    if fam and not norm(fam).startswith("SOLIDE"):
        return (False, "boisson", "haute", "")
    if fam is None:
        # items de C absents de D
        if n.startswith("FORMULE") or n.startswith("MENU"):
            return (True, "menu", "haute", "")
        if "VERRE" in n or "PICHET" in n:
            return (False, "boisson", "haute", "")
        if "CANCOILLOTTE" in n:
            return (False, "accompagnement", "haute", "")
    # 1bis) Overrides confirmés par la gérante (priorité absolue)
    if n in CONFIRME:
        couvert, cat, rem = CONFIRME[n]
        return (couvert, cat, "haute", rem)
    # 2) Items à vérifier (priorité haute, conflits connus)
    if n in A_VERIFIER:
        rem, couvert = A_VERIFIER[n]
        cat = "plat (à vérifier)" if couvert else "non-couvert (à vérifier)"
        return (couvert, cat, "à vérifier", rem)
    # 3) Menus / formules
    if n.startswith("MENU") or n.startswith("FORMULE") or n in ("PLAT DU JOUR", "PLAT"):
        return (True, "menu", "haute", "")
    # 4) Entrées explicites
    if n in ENTREES or n.startswith("ESCARGOT") or n.startswith("JAMBON PERSILLE") \
            or n.startswith("JAMHON PERSILLE") or n.startswith("CROUTE"):
        return (False, "entree", "haute", "")
    # 5) Galettes nommées (plats)
    if n in GALETTES:
        return (True, "plat (galette)", "à vérifier", "nom de galette : à confirmer sur la carte")
    # 6) Desserts nommés (crêpes sucrées / coupes)
    if n in DESSERTS_NOMMES:
        return (False, "dessert", "à vérifier", "nom de crêpe/coupe sucrée : à confirmer sur la carte")
    # 7) Desserts par mot-clé (avant les plats : "Crêpe ..." = dessert)
    if any(k in n for k in DESSERT_KW):
        return (False, "dessert", "haute", "")
    # 8) Accompagnements / suppléments / fromages
    if any(k in n for k in ACCOMP_KW):
        return (False, "accompagnement", "haute", "")
    # 9) Plats par mot-clé (galette comprise)
    if any(k in n for k in PLAT_KW):
        return (True, "plat", "haute", "")
    # 10) Salades : des plats sauf exceptions (déjà en ENTREES)
    if "SALADE" in n:
        return (True, "plat (salade)", "haute", "")
    # 11) Fallback : item SOLIDE non reconnu -> à vérifier (par défaut non-couvert)
    return (False, "non-couvert (à vérifier)", "à vérifier", "item non reconnu : à classer manuellement")


# ---------------------------------------------------------------------------
# Quantités 3 ans par libellé (annexe D) pour info
# ---------------------------------------------------------------------------
def quantites():
    q = collections.Counter()
    for e in EXOS:
        sh = xlrd.open_workbook(D[e]).sheet_by_index(0)
        for r in range(sh.nrows):
            c0 = str(sh.cell_value(r, 0)).strip()
            c1 = str(sh.cell_value(r, 1)).strip()
            if not (c0[:1].isdigit() and len(c0) <= 5) or not c1:
                continue
            try:
                q[c1] += float(sh.cell_value(r, 4))
            except (ValueError, TypeError):
                pass
    return q


# ---------------------------------------------------------------------------
# Styles XLSX
# ---------------------------------------------------------------------------
GRAS = Font(bold=True)
GRAS_BLANC = Font(bold=True, color="FFFFFF")
ENTETE = PatternFill("solid", fgColor="0F766E")
OUI_FILL = PatternFill("solid", fgColor="DCFCE7")
VERIF_FILL = PatternFill("solid", fgColor="FEF9C3")
CENTRE = Alignment(horizontal="center", vertical="center", wrap_text=True)
DROITE = Alignment(horizontal="right", vertical="center")
GAUCHE = Alignment(horizontal="left", vertical="center", wrap_text=True)
BORD = Border(*(4 * (Side(style="thin", color="D1D5DB"),)))


def entete(ws, ncols, ligne=1):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=ligne, column=c)
        cell.font = GRAS_BLANC
        cell.fill = ENTETE
        cell.alignment = CENTRE
        cell.border = BORD


def largeurs(ws, ws_widths):
    for i, w in enumerate(ws_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w


# ---------------------------------------------------------------------------
# 1) Fichier classification des items
# ---------------------------------------------------------------------------
def fichier_items(fam_of, qte, classif):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Classification items"
    ws.append(["Items vendus en caisse : un couvert = un plat OU un menu"])
    ws["A1"].font = GRAS
    ws.append(["Oui = compte comme un couvert (plat/menu). Galettes = plats ; salades = plats "
               "sauf salade chèvre chaud et salade verte aux noix (entrées). « À vérifier » = "
               "nom fantaisie galette/crêpe à confirmer sur la carte."])
    ws["A2"].alignment = GAUCHE
    ws.merge_cells("A2:G2")
    ws.row_dimensions[2].height = 45
    ws.append([])
    h = ws.max_row + 1
    ws.append(["Libellé", "Famille", "Catégorie déduite", "Plat/Menu (couvert)",
               "Confiance", "Remarque", "Qté 3 ans"])
    entete(ws, 7, h)
    for lib in sorted(classif, key=lambda x: (classif[x][1] != "menu", -qte.get(x, 0))):
        couvert, cat, conf, rem = classif[lib]
        r = ws.max_row + 1
        ws.append([lib, fam_of.get(lib, "(hors annexe D)"), cat,
                   "Oui" if couvert else "Non", conf, rem, int(qte.get(lib, 0))])
        for c in range(1, 8):
            ws.cell(row=r, column=c).border = BORD
        ws.cell(row=r, column=6).alignment = GAUCHE
        if couvert:
            ws.cell(row=r, column=4).fill = OUI_FILL
            ws.cell(row=r, column=4).font = GRAS
        if conf == "à vérifier":
            ws.cell(row=r, column=5).fill = VERIF_FILL
    largeurs(ws, [26, 26, 22, 18, 12, 40, 10])
    ws.freeze_panes = "A" + str(h + 1)
    os.makedirs(OUT, exist_ok=True)
    p = OUT + "RF-items-classification.xlsx"
    wb.save(p)
    return p


# ---------------------------------------------------------------------------
# 2) + 3) Encaissements -> couverts, et couverts par service
# ---------------------------------------------------------------------------
def lire_encaissements(classif):
    """
    Retourne la liste des encaissements :
      [ {date, heure, id, exo, couverts} ] (un par (date, no_ticket))
    couverts = somme des quantités des lignes classées plat/menu.
    """
    enc = {}
    for e in EXOS:
        sh = xlrd.open_workbook(C[e]).sheet_by_index(0)
        for r in range(1, sh.nrows):
            date = str(sh.cell_value(r, 0))[:10]
            if len(date) != 10 or date[4] != "-":
                continue
            no = sh.cell_value(r, 2)
            if not isinstance(no, float):
                continue
            key = (date, int(no))
            heure = str(sh.cell_value(r, 1))[:5]
            idv = sh.cell_value(r, 8)
            lib = str(sh.cell_value(r, 10)).strip()
            try:
                q = float(sh.cell_value(r, 11))
            except (ValueError, TypeError):
                q = 0.0
            rec = enc.get(key)
            if rec is None:
                rec = {"date": date, "heure": heure,
                       "id": (int(idv) if isinstance(idv, float) and idv else ""),
                       "exo": e, "couverts": 0.0}
                enc[key] = rec
            else:
                if heure and heure < rec["heure"]:
                    rec["heure"] = heure
            couvert = classif.get(lib)
            if couvert and couvert[0]:
                rec["couverts"] += q
    # tri par date puis heure
    out = [enc[k] for k in sorted(enc, key=lambda k: (k[0], enc[k]["heure"], k[1]))]
    return out


def fichier_encaissements(encs):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Encaissements couverts"
    ws.append(["Encaissements (notes) et nombre de couverts estimé"])
    ws["A1"].font = GRAS
    ws.append(["Un encaissement = une note = (date, n° de note). Couverts = somme des quantités "
               "des lignes classées plat/menu (les demi-parts 0,5 d'un partage se somment à 1)."])
    ws["A2"].alignment = GAUCHE
    ws.merge_cells("A2:E2")
    ws.row_dimensions[2].height = 30
    ws.append([])
    h = ws.max_row + 1
    ws.append(["Date", "Heure", "Id", "Exercice", "Couverts"])
    entete(ws, 5, h)
    for rec in encs:
        r = ws.max_row + 1
        cv = rec["couverts"]
        cv = int(cv) if abs(cv - round(cv)) < 1e-9 else round(cv, 2)
        ws.append([rec["date"], rec["heure"], rec["id"], rec["exo"], cv])
        for c in range(1, 6):
            ws.cell(row=r, column=c).border = BORD
    largeurs(ws, [14, 10, 12, 14, 12])
    ws.freeze_panes = "A" + str(h + 1)
    p = OUT + "RF-encaissements-couverts.xlsx"
    wb.save(p)
    return p


def fichier_services(encs):
    """
    Couverts par jour et par service. MIDI = heure < 18:00, SOIR = heure >= 18:00.
    """
    jours = collections.defaultdict(lambda: {"midi": 0.0, "soir": 0.0,
                                             "n_midi": 0, "n_soir": 0, "exo": ""})
    for rec in encs:
        j = jours[rec["date"]]
        j["exo"] = rec["exo"]
        soir = rec["heure"] >= "18:00"
        if soir:
            j["soir"] += rec["couverts"]
            j["n_soir"] += 1
        else:
            j["midi"] += rec["couverts"]
            j["n_midi"] += 1

    def fmt(v):
        return int(v) if abs(v - round(v)) < 1e-9 else round(v, 2)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Couverts par service"
    ws.append(["Couverts par jour et par service"])
    ws["A1"].font = GRAS
    ws.append(["MIDI = encaissements avant 18h (service ~10h-17h) ; SOIR = 18h à minuit. "
               "Couverts = somme des quantités plat/menu."])
    ws["A2"].alignment = GAUCHE
    ws.merge_cells("A2:F2")
    ws.row_dimensions[2].height = 30
    ws.append([])
    h = ws.max_row + 1
    ws.append(["Date", "Exercice", "Couverts midi (10h-17h)", "Couverts soir (18h-minuit)",
               "Total couverts", "Notes midi / soir"])
    entete(ws, 6, h)
    tm = ts = 0.0
    for date in sorted(jours):
        j = jours[date]
        tm += j["midi"]
        ts += j["soir"]
        r = ws.max_row + 1
        ws.append([date, j["exo"], fmt(j["midi"]), fmt(j["soir"]),
                   fmt(j["midi"] + j["soir"]), f"{j['n_midi']} / {j['n_soir']}"])
        for c in range(1, 7):
            ws.cell(row=r, column=c).border = BORD
    rt = ws.max_row + 1
    ws.append(["TOTAL", "", fmt(tm), fmt(ts), fmt(tm + ts), ""])
    for c in range(1, 7):
        ws.cell(row=rt, column=c).font = GRAS
        ws.cell(row=rt, column=c).border = BORD
    largeurs(ws, [14, 14, 24, 26, 16, 18])
    ws.freeze_panes = "A" + str(h + 1)
    p = OUT + "RF-couverts-par-service.xlsx"
    wb.save(p)
    return p, fmt(tm), fmt(ts)


# ---------------------------------------------------------------------------
# JSON : couverts par semaine d'exercice (1 série par exercice), pour le
# graphique multi-lignes de la page tables-trop-nombreuses.
# ---------------------------------------------------------------------------
def fichier_hebdo(encs):
    NB_SEM = 52
    par_exo = {e: collections.defaultdict(float) for e in EXOS}
    for rec in encs:
        e = rec["exo"]
        an = int(e[:4])
        d = datetime.date(int(rec["date"][:4]), int(rec["date"][5:7]), int(rec["date"][8:10]))
        sem = (d - datetime.date(an, 4, 1)).days // 7 + 1
        sem = max(1, min(NB_SEM, sem))
        par_exo[e][sem] += rec["couverts"]

    def rnd(v):
        return int(round(v))

    data = []
    for sem in range(1, NB_SEM + 1):
        ligne = {"semaine": sem}
        for e in EXOS:
            ligne[e] = rnd(par_exo[e].get(sem, 0))
        data.append(ligne)

    mois_ticks = []
    for m, lib in MOIS_LABELS:
        an = 2022 if m >= 4 else 2023
        sem = (datetime.date(an, m, 1) - datetime.date(2022, 4, 1)).days // 7 + 1
        mois_ticks.append({"tick": max(1, min(NB_SEM, sem)), "label": lib})

    section = {
        "kind": "graphiqueLignes",
        "titre": "Couverts par semaine : superposition des trois exercices (avril → mars)",
        "sousTitre": ("Une ligne par exercice. Saisonnalité marquée (pic l'été, creux hivernal : "
                      "le restaurant ferme janvier et février), et profils quasi superposables "
                      "d'une année à l'autre."),
        "hauteur": 340,
        "dataKey": "semaine",
        "series": [{"name": e, "couleur": COULEURS_EXO[e]} for e in EXOS],
        "data": data,
        "format": "int",
        "moisTicks": mois_ticks,
    }
    os.makedirs(OUT_JSON_DIR, exist_ok=True)
    p = OUT_JSON_DIR + "couverts-hebdomadaires.json"
    with open(p, "w", encoding="utf-8") as f:
        json.dump(section, f, ensure_ascii=False, indent=2)
    return p


# ---------------------------------------------------------------------------
# JSON : couverts par JOUR, par saison (4 trimestres), service midi / soir
# separes, + stats par mois (services AVEC couverts uniquement, 0 exclus).
# Une section "saisonCouverts" par saison.
# ---------------------------------------------------------------------------
SAISONS = [
    ("Mars, avril, mai", [(3, "Mars"), (4, "Avril"), (5, "Mai")]),
    ("Juin, juillet, août", [(6, "Juin"), (7, "Juillet"), (8, "Août")]),
    ("Septembre, octobre, novembre", [(9, "Septembre"), (10, "Octobre"), (11, "Novembre")]),
    ("Décembre, janvier, février", [(12, "Décembre"), (1, "Janvier"), (2, "Février")]),
]
JOURS_MOIS = {1: 31, 2: 29, 3: 31, 4: 30, 5: 31, 6: 30, 7: 31, 8: 31, 9: 30, 10: 31, 11: 30, 12: 31}


def fichier_saisons(encs):
    # Agrege par date : couverts midi / soir (split a 18h), avec l'exercice.
    jours = {}
    for rec in encs:
        j = jours.setdefault(rec["date"], {"exo": rec["exo"], "midi": 0.0, "soir": 0.0})
        if rec["heure"] >= "18:00":
            j["soir"] += rec["couverts"]
        else:
            j["midi"] += rec["couverts"]

    sections = []
    for titre, mois in SAISONS:
        offset, acc = {}, 0
        for m, _ in mois:
            offset[m] = acc
            acc += JOURS_MOIS[m]
        total = acc
        mois_set = {m for m, _ in mois}

        soir = {i: {"jour": i} for i in range(1, total + 1)}
        midi = {i: {"jour": i} for i in range(1, total + 1)}
        # Stats par mois ET par exercice (une moyenne par année). Seuls les
        # services AVEC couverts comptent (un service à 0 est exclu du dénominateur).
        stat = {m: {e: {"midis": 0, "soirs": 0, "sm": 0.0, "ss": 0.0} for e in EXOS}
                for m, _ in mois}

        for date, j in jours.items():
            m, d = int(date[5:7]), int(date[8:10])
            if m not in mois_set:
                continue
            idx = offset[m] + d
            e = j["exo"]
            if j["soir"] > 0:
                soir[idx][e] = round(j["soir"])
            if j["midi"] > 0:
                midi[idx][e] = round(j["midi"])
            s = stat[m][e]
            if j["midi"] > 0:
                s["midis"] += 1
                s["sm"] += j["midi"]
            if j["soir"] > 0:
                s["soirs"] += 1
                s["ss"] += j["soir"]

        stats = [{
            "mois": lib,
            "parExercice": [{
                "exo": e,
                "midis": stat[m][e]["midis"],
                "soirs": stat[m][e]["soirs"],
                "moyMidi": round(stat[m][e]["sm"] / stat[m][e]["midis"]) if stat[m][e]["midis"] else 0,
                "moySoir": round(stat[m][e]["ss"] / stat[m][e]["soirs"]) if stat[m][e]["soirs"] else 0,
            } for e in EXOS],
        } for m, lib in mois]

        sections.append({
            "kind": "saisonCouverts",
            "titre": titre,
            "sousTitre": ("Couverts par jour : service du soir (en haut) et du midi (en bas), "
                          "une ligne par exercice. À droite, la moyenne de couverts par midi et "
                          "par soir, mois par mois et exercice par exercice. Seuls les services "
                          "ayant eu des couverts comptent dans la moyenne (les services à 0, ou "
                          "non travaillés, sont exclus)."),
            "dataKey": "jour",
            "series": [{"name": e, "couleur": COULEURS_EXO[e]} for e in EXOS],
            "soir": [soir[i] for i in range(1, total + 1)],
            "midi": [midi[i] for i in range(1, total + 1)],
            "moisTicks": [{"tick": offset[m] + 1, "label": lib} for m, lib in mois],
            "stats": stats,
        })

    os.makedirs(OUT_JSON_DIR, exist_ok=True)
    p = OUT_JSON_DIR + "couverts-saisons.json"
    with open(p, "w", encoding="utf-8") as f:
        json.dump(sections, f, ensure_ascii=False, indent=2)
    return p


# ---------------------------------------------------------------------------
# JSON : couverts moyens par jour travaille, par semaine (cycle mars -> fevrier),
# 1 ligne par exercice, + bandes de capacite a 58 % (taux d'occupation FIDUCIAL) :
# zone TERRASSE (debut mai -> 1re semaine de septembre) et zone INTERIEUR (reste).
# ---------------------------------------------------------------------------
CAP_INT = 44          # places assises en salle interieure (somme des couverts/table)
CAP_EXT = 43          # places en terrasse (hors 116/117 occasionnelles)
TAUX_OCC = 0.58       # taux de remplissage moyen restauration traditionnelle (FIDUCIAL 2025)
SEUIL_INT = round(CAP_INT * TAUX_OCC)   # 26
SEUIL_EXT = round(CAP_EXT * TAUX_OCC)   # 25
MOIS_CYCLE = [(3, "Mars"), (4, "Avr"), (5, "Mai"), (6, "Juin"), (7, "Juil"), (8, "Août"),
              (9, "Sep"), (10, "Oct"), (11, "Nov"), (12, "Déc"), (1, "Jan"), (2, "Fév")]


def _semaine_cycle(m, d):
    """Semaine (1..52) dans le cycle mars -> fevrier, alignee par mois calendaire."""
    y = 2001 if m >= 3 else 2002
    return (datetime.date(y, m, d) - datetime.date(2001, 3, 1)).days // 7 + 1


def fichier_capacite(encs):
    # Couverts par (date, service) : split midi / soir a 18h.
    jours = {}
    for rec in encs:
        j = jours.setdefault(rec["date"], {"exo": rec["exo"], "midi": 0.0, "soir": 0.0})
        if rec["heure"] >= "18:00":
            j["soir"] += rec["couverts"]
        else:
            j["midi"] += rec["couverts"]
    # moyenne par service AVEC couverts (0 exclu), par (exo, semaine).
    aggm = {e: collections.defaultdict(lambda: [0.0, 0]) for e in EXOS}
    aggs = {e: collections.defaultdict(lambda: [0.0, 0]) for e in EXOS}
    for date, j in jours.items():
        wk = _semaine_cycle(int(date[5:7]), int(date[8:10]))
        if j["midi"] > 0:
            a = aggm[j["exo"]][wk]
            a[0] += j["midi"]
            a[1] += 1
        if j["soir"] > 0:
            a = aggs[j["exo"]][wk]
            a[0] += j["soir"]
            a[1] += 1
    NB_WK = 52

    def build(agg):
        out = []
        for wk in range(1, NB_WK + 1):
            row = {"semaine": wk}
            for e in EXOS:
                a = agg[e].get(wk)
                if a and a[1] > 0:
                    row[e] = round(a[0] / a[1])
            out.append(row)
        return out

    # --- Taux d'occupation par mois (couverts moyens/service / places ouvertes) ---
    # Espace ouvert : terrasse de mai a aout (43 places), interieur le reste (44).
    def _cap_mois(m):
        return (CAP_EXT, "terrasse") if m in (5, 6, 7, 8) else (CAP_INT, "intérieur")

    ORDRE_MOIS = [(3, "Mars"), (4, "Avril"), (5, "Mai"), (6, "Juin"), (7, "Juillet"),
                  (8, "Août"), (9, "Sept."), (10, "Oct."), (11, "Nov."), (12, "Déc."),
                  (1, "Janv."), (2, "Févr.")]
    magg = {m: {"sm": 0.0, "nm": 0, "ss": 0.0, "ns": 0} for m in range(1, 13)}
    for date, j in jours.items():
        a = magg[int(date[5:7])]
        if j["midi"] > 0:
            a["sm"] += j["midi"]
            a["nm"] += 1
        if j["soir"] > 0:
            a["ss"] += j["soir"]
            a["ns"] += 1
    occupation = []
    for m, lib in ORDRE_MOIS:
        a = magg[m]
        cap, esp = _cap_mois(m)
        tot_c, tot_s = a["sm"] + a["ss"], a["nm"] + a["ns"]
        occupation.append({
            "mois": lib,
            "espace": esp,
            "pct": round(100 * (tot_c / tot_s) / cap) if tot_s else None,
            "pctMidi": round(100 * (a["sm"] / a["nm"]) / cap) if a["nm"] else None,
            "pctSoir": round(100 * (a["ss"] / a["ns"]) / cap) if a["ns"] else None,
        })
    # Taux d'occupation moyen PONDERE : total couverts / total capacite servie
    # (chaque service compte pour la capacite de l'espace ouvert ce mois-la).
    tot_couv = sum(magg[m]["sm"] + magg[m]["ss"] for m in range(1, 13))
    tot_capa = sum((magg[m]["nm"] + magg[m]["ns"]) * _cap_mois(m)[0] for m in range(1, 13))
    occ_moy_exact = 100 * tot_couv / tot_capa if tot_capa else 0
    MOY_FR = 58  # taux d'occupation moyen restauration traditionnelle (FIDUCIAL 2025)

    w_mai = _semaine_cycle(5, 1)     # debut mai
    w_sep = _semaine_cycle(9, 7)     # fin de la 1re semaine de septembre
    zones = [
        {"x1": 1, "x2": w_mai, "seuil": SEUIL_INT, "espace": "interieur"},
        {"x1": w_mai, "x2": w_sep, "seuil": SEUIL_EXT, "espace": "terrasse"},
        {"x1": w_sep, "x2": NB_WK, "seuil": SEUIL_INT, "espace": "interieur"},
    ]
    section = {
        "kind": "couvertsCapacite",
        "titre": "Couverts moyens par service et taux d'occupation de 58 %",
        "sousTitre": ("Moyenne des couverts par service, par semaine (cycle mars → février), une "
                      "ligne par exercice (les services sans couvert sont exclus de la moyenne). "
                      "La ligne pointillée marque 58 % de remplissage, le taux d'occupation moyen "
                      "d'un restaurant traditionnel (FIDUCIAL 2025), appliqué à l'espace ouvert : "
                      "intérieur de mi-septembre à fin avril, terrasse de début mai à début "
                      "septembre (un seul espace ouvert à la fois)."),
        "dataKey": "semaine",
        "series": [{"name": e, "couleur": COULEURS_EXO[e]} for e in EXOS],
        "midi": build(aggm),
        "soir": build(aggs),
        "moisTicks": [{"tick": _semaine_cycle(m, 1), "label": lib} for m, lib in MOIS_CYCLE],
        "zones": zones,
        "seuilInterieur": SEUIL_INT,
        "seuilTerrasse": SEUIL_EXT,
        "capaciteInterieur": CAP_INT,
        "capaciteTerrasse": CAP_EXT,
        "occupationMois": occupation,
        "occupationMoyenne": round(occ_moy_exact),
        "moyenneFrance": MOY_FR,
        "ecartMoyenneFrancePct": round((occ_moy_exact / MOY_FR - 1) * 100),
    }
    os.makedirs(OUT_JSON_DIR, exist_ok=True)
    p = OUT_JSON_DIR + "couverts-capacite.json"
    with open(p, "w", encoding="utf-8") as f:
        json.dump(section, f, ensure_ascii=False, indent=2)
    return p


# ---------------------------------------------------------------------------
def main():
    fam_of = familles()
    qte = quantites()
    # tous les libellés vus en C + D
    libs = set(fam_of)
    for e in EXOS:
        sh = xlrd.open_workbook(C[e]).sheet_by_index(0)
        for r in range(1, sh.nrows):
            lib = str(sh.cell_value(r, 10)).strip()
            if lib:
                libs.add(lib)
    classif = {lib: classer(lib, fam_of.get(lib)) for lib in libs}

    p1 = fichier_items(fam_of, qte, classif)
    encs = lire_encaissements(classif)
    p2 = fichier_encaissements(encs)
    p3, tm, ts = fichier_services(encs)
    p4 = fichier_hebdo(encs)
    p5 = fichier_saisons(encs)
    p6 = fichier_capacite(encs)

    # bilan console
    nb_oui = sum(1 for v in classif.values() if v[0])
    nb_verif = sum(1 for v in classif.values() if v[2] == "à vérifier")
    tot_couv = sum(rec["couverts"] for rec in encs)
    print("=== Couverts : synthèse ===")
    print(f"Items classés : {len(classif)}  (dont {nb_oui} = plat/menu, {nb_verif} à vérifier)")
    print(f"Encaissements (notes) : {len(encs)}")
    print(f"Couverts estimés (total) : {tot_couv:.0f}  | midi {tm} / soir {ts}")
    print("Items 'à vérifier' (galette vs crêpe / conflits) :")
    for lib in sorted(l for l, v in classif.items() if v[2] == "à vérifier"):
        v = classif[lib]
        print(f"   [{'Oui' if v[0] else 'Non'}] {lib}  -> {v[1]} ({v[3]})")
    print(f"XLSX -> {p1}")
    print(f"XLSX -> {p2}")
    print(f"XLSX -> {p3}")
    print(f"JSON -> {p4}")
    print(f"JSON -> {p5}")
    print(f"JSON -> {p6}")


if __name__ == "__main__":
    main()
